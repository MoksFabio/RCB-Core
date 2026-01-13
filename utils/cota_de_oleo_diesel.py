import tkinter as tk
from tkinter import filedialog, messagebox, Listbox, PhotoImage, ttk
import os
import re
import sys
import threading
import datetime
import subprocess
import traceback
import copy
import math
import pandas as pd
import openpyxl
import json
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from tkinterdnd2 import DND_FILES, TkinterDnD
from PIL import Image, ImageTk
import sv_ttk

class Config:
    RENDIMENTO_DISPEL = {
        "Micro Urbano  s/ar": 2.610,
        "MIDI MEDIO URBANO S/AR": 2.610,
        "Básico Méd. urb s/ar": 2.610,
        "Básico Méd. urb c/ar": 2.219,
        "Padron 12  Pés s/ar": 2.610,
        "Padron 12  Pés c/ar": 2.2185,
        "Padron 13 Pés s /ar": 2.420,
        "Padron 13 Pés c /ar": 2.057,
        "Padron 14 Pés s /ar": 2.381,
        "Padron 14 Pés c /ar": 2.218,
        "Padron 15 Pés s /ar": 2.381,
        "Padron 15 Pés C /ar": 1.832,
        "Padron 15 Pés c/ar": 1.832,
        "Artic.Ext.Pés s/ar": 1.750,
        "Artic.Ext.Pés c/ar": 1.200,
        "BRT 1 Art.Ext Pés C/ar": 1.200,
        "BRT 1 Art.Ext Pés c/ ar": 1.200,
        "RODVIÁRIOS P. 13 C/AR": 1.750,
        "RODVIÁRIOS S/AR": 0,
        "RODOVIÁRIOS C/AR": 1.750,
        "RODVIÁRIO C/AR": 1.750,
        "RODVIÁRIOS C/AR": 1.750,
        "RODOVIÁRIOS P. 13 C/AR": 1.750,
        "RODOVIÁRIO P. 13 C/AR": 1.750,
        "ARTICULADO COM AR E CÂMBIO": 1.200,
        "ART.EXT PES URBANO S/AR": 1.750,
        "Midi Urbano  s/ar": 2.610,
        "Mini Urbano  c/ar": 2.088
    }
    NOMES_EMPRESAS = {
        "BOA": "BOA - Borborema Imperial Transportes Ltda", "CAX": "CAX - Caxangá Empresa de Transporte Coletivo Ltda",
        "CSR": "CSR - Consórcio Recife de Transporte", "CNO": "CNO – CONSÓRCIO CONORTE",
        "EME": "EME - Metropolitana Empresa de Transporte Coletivo Ltda", "GLO": "GLO - Transportadora Globo Ltda",
        "MOB": "MOB – MobiBrasil Expresso S.A", "SJT": "SJT - José Faustino e Companhia Ltda", "VML": "VML - Viação Mirim Ltda",
        "CTC": "CTC - Companhia de Transp. e Comunicação"
    }
    FONT_TITULO_PRINCIPAL = Font(bold=True, size=14)
    FONT_HEADER_EMPRESA = Font(bold=True, color="FFFFFF", size=12)
    FILL_HEADER_EMPRESA = PatternFill(start_color="000000", fill_type="solid")
    FONT_HEADER_TABELA = Font(bold=True)
    FILL_HEADER_TABELA = PatternFill(start_color="FFA500", fill_type="solid")
    FONT_TOTAL = Font(bold=True)
    FILL_TOTAL = PatternFill(start_color="F5F5F5", fill_type="solid")
    FONT_MEDIA_GERAL = Font(bold=True, size=28)
    FILL_MEDIA_GERAL = PatternFill(start_color="FFFF00", fill_type="solid")
    FONT_TITULO_CALC = Font(name='Calibri', size=22, italic=True)
    FONT_SUBTITULO_CALC = Font(name='Calibri', bold=True)
    FONT_CABECALHO_CALC = Font(name='Calibri', bold=True, size=10)
    FILL_CABECALHO_CALC = PatternFill(start_color="FFA500", fill_type="solid")
    FONT_TOTAL_GRUPO_CALC = Font(name='Calibri', bold=True, size=11)
    FONT_DADOS_CALC = Font(name='Calibri', size=11)
    FONT_TOTAL_GERAL_CALC = Font(name='Calibri', bold=True, size=11, color="FFFFFF")
    FILL_YELLOW = PatternFill(start_color="FFFF00", fill_type="solid")
    FILL_BLUE_TOTAL_GERAL = PatternFill(start_color="FFA500", fill_type="solid")
    ESTILOS_SUMARIO = {
        'titulo_principal': {'font': Font(bold=True, size=16, color="FFFFFF"), 'fill': PatternFill(start_color="FFA500", fill_type="solid")},
        'titulo_bloco': {'font': Font(bold=True, size=12, color="FFFFFF"), 'fill': PatternFill(start_color="000000", fill_type="solid")},
        'cabecalho_tabela': {'font': Font(bold=True, color="000000"), 'fill': PatternFill(start_color="FFA500", fill_type="solid")},
        'final_label': {'font': Font(bold=True, size=12, color="FFFFFF"), 'fill': PatternFill(start_color="FFA500", fill_type="solid")},
        'final_valor': {'font': Font(bold=True, size=12, color="FFFFFF"), 'fill': PatternFill(start_color="FFA500", fill_type="solid")},
        'status_ok': Font(color="006100", bold=True), 'status_diff': Font(color="9C0006", bold=True),
        'dxf_diff': DifferentialStyle(font=Font(color="FF0000")), 'zebra': PatternFill(start_color="F5F5F5", fill_type="solid"),
        'zebra_removida': PatternFill(start_color="FFC7CE", fill_type="solid"),
        'sumario_total': PatternFill(start_color="F5F5F5", fill_type="solid"),
        'borda': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'font_sumario_label': Font(bold=True), 'font_sumario_valor': Font(bold=True),
    }

def adicionar_marca_dagua(ws, anexo_celula='P1'):
    caminho_logo = "static/images/rcb_logo.png"
    try:
        img = OpenpyxlImage(caminho_logo)
        img.height = 60
        img.width = 76
        ws.add_image(img, anexo_celula)
    except FileNotFoundError:
        print(f"AVISO: Arquivo de imagem para marca d'água não encontrado em '{caminho_logo}'.")
    except Exception as e:
        print(f"AVISO: Não foi possível adicionar a marca d'água. Erro: {e}")

def auto_ajustar_colunas(ws):
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

def aplicar_wrap_text_em_todas_abas(workbook):
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                new_alignment = copy.copy(cell.alignment)
                new_alignment.wrap_text = True
                cell.alignment = new_alignment

def encontrar_mes_no_arquivo(caminho_arquivo):
    try:
        import xlrd
        book = xlrd.open_workbook(caminho_arquivo, on_demand=True)
        sheet = book.sheet_by_index(0)
        for row_idx in range(min(sheet.nrows, 50)):
            cell = sheet.cell(row_idx, 0)
            mes_num = None
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    date_tuple = xlrd.xldate_as_tuple(cell.value, book.datemode)
                    mes_num = date_tuple[1]
                except Exception: continue
            elif cell.ctype == xlrd.XL_CELL_TEXT:
                match = re.search(r'\d{2}/(\d{2})/\d{4}', str(cell.value))
                if match: mes_num = int(match.group(1))
            
            if mes_num:
                meses_map = {1: 'JANEIRO', 2: 'FEVEREIRO', 3: 'MARÇO', 4: 'ABRIL', 5: 'MAIO', 6: 'JUNHO', 7: 'JULHO', 8: 'AGOSTO', 9: 'SETEMBRO', 10: 'OUTUBRO', 11: 'NOVEMBRO', 12: 'DEZEMBRO'}
                book.release_resources()
                return meses_map.get(mes_num, "MÊS")
    except Exception: pass
    finally:
        if 'book' in locals() and 'release_resources' in dir(book): book.release_resources()
    return "MÊS DESCONHECIDO"

def copiar_dados_planilha(caminho_arquivo, ws_destino, offset_linha, offset_coluna):
    try:
        import xlrd
        book_origem = xlrd.open_workbook(caminho_arquivo, formatting_info=True)
        sheet_origem = book_origem.sheet_by_index(0)
        
        merged_cells_map = {}
        for rlo, rhi, clo, chi in sheet_origem.merged_cells:
            for rowx in range(rlo, rhi):
                for colx in range(clo, chi):
                    if (rowx, colx) != (rlo, clo): merged_cells_map[(rowx, colx)] = (rlo, clo)

        max_cols = max((sheet_origem.row_len(r) for r in range(sheet_origem.nrows)), default=0)
        for r_idx in range(sheet_origem.nrows):
            for c_idx in range(sheet_origem.row_len(r_idx)):
                source_r, source_c = r_idx, c_idx
                is_merged_slave = (r_idx, c_idx) in merged_cells_map
                if is_merged_slave: source_r, source_c = merged_cells_map[(r_idx, c_idx)]
                
                source_cell = sheet_origem.cell(source_r, source_c)
                cell_value = source_cell.value
                
                if source_cell.ctype == xlrd.XL_CELL_DATE:
                    cell_value = xlrd.xldate.xldate_as_datetime(source_cell.value, book_origem.datemode)
                if is_merged_slave: cell_value = ""

                dest_cell = ws_destino.cell(row=r_idx + offset_linha + 1, column=c_idx + offset_coluna + 1, value=cell_value)
                if isinstance(cell_value, (datetime.date, datetime.datetime)): dest_cell.number_format = 'dd/mm/yyyy'

        for rlo, rhi, clo, chi in sheet_origem.merged_cells:
            ws_destino.merge_cells(start_row=rlo + offset_linha + 1, start_column=clo + offset_coluna + 1, end_row=rhi + offset_linha, end_column=chi + offset_coluna)
        
        for c_idx_rel in range(2): 
            col_real = c_idx_rel + offset_coluna + 1
            start_merge_row = -1
            
            for r_idx in range(sheet_origem.nrows):
                row_real = r_idx + offset_linha + 1
                cell = ws_destino.cell(row=row_real, column=col_real)

                if cell.value is not None and str(cell.value).strip() != "":
                    if start_merge_row != -1 and row_real - 1 > start_merge_row:
                        ws_destino.merge_cells(start_row=start_merge_row, start_column=col_real, end_row=row_real - 1, end_column=col_real)
                        ws_destino.cell(start_merge_row, col_real).alignment = Alignment(vertical='center')
                    start_merge_row = row_real
                elif r_idx == sheet_origem.nrows - 1:
                    if start_merge_row != -1 and row_real > start_merge_row:
                        ws_destino.merge_cells(start_row=start_merge_row, start_column=col_real, end_row=row_real, end_column=col_real)
                        ws_destino.cell(start_merge_row, col_real).alignment = Alignment(vertical='center')

        return sheet_origem.nrows, max_cols
    except Exception as e:
        print(f"Erro ao copiar o arquivo {caminho_arquivo}: {e}")
        return 0, 0

def obter_chave_ordenacao(filepath):
    filename = os.path.basename(filepath).upper()
    if "DUT" in filename: return 0
    if "SÁB" in filename or "SAB" in filename: return 1
    if "DOM" in filename: return 2
    return 3

def encontrar_indices_colunas(sheet, linha_cabecalho, col_inicio, col_fim):
    indices, chaves_necessarias = {}, {'Linha', 'Km Total', 'Empresa', 'Dia Tipo'}
    for r in range(linha_cabecalho, min(linha_cabecalho + 10, sheet.max_row + 1)):
        for col in range(col_inicio, col_fim + 1):
            valor_cabecalho = sheet.cell(row=r, column=col).value
            if isinstance(valor_cabecalho, str):
                header = valor_cabecalho.strip()
                if header in chaves_necessarias and header not in indices:
                    indices[header] = col
                    indices['linha_header_real'] = r
            if all(k in indices for k in chaves_necessarias): break
    return indices

def processar_bloco_para_totais(sheet, linha_inicio, linha_fim, col_inicio, col_fim):
    indices = encontrar_indices_colunas(sheet, linha_inicio, col_inicio, col_fim)
    if not all(k in indices for k in ['Empresa', 'Km Total']): return {'empresas': {}, 'geral': 0.0}

    dados_empresas, valor_geral_bloco = {}, 0.0
    linha_dados_inicio = indices.get('linha_header_real', linha_inicio) + 1
    for linha in range(linha_dados_inicio, linha_fim + 1):
        cell_values = [str(sheet.cell(row=linha, column=c).value or '').upper() for c in range(col_inicio, col_fim + 1)]
        is_total_row = any('TOTAL' in v for v in cell_values)
        is_geral_row = any('GERAL' in v for v in cell_values)
        try:
            km_total = float(sheet.cell(row=linha, column=indices['Km Total']).value or 0.0)
            if is_geral_row:
                valor_geral_bloco = km_total
                continue
            empresa_cell_val = sheet.cell(row=linha, column=indices['Empresa']).value
            if not empresa_cell_val and not is_total_row:
                continue
            
            empresa = str(empresa_cell_val).strip()

            dados_empresas.setdefault(empresa, {'soma_calculada': 0.0, 'valor_declarado': 0.0, 'dia_tipo': "N/D"})
            if is_total_row:
                empresa_no_total = next((e for e in dados_empresas.keys() if e in empresa and e), None)
                if empresa_no_total:
                    dados_empresas[empresa_no_total]['valor_declarado'] = km_total
            else:
                if not empresa: continue
                dados_empresas[empresa]['soma_calculada'] += km_total
                if dados_empresas[empresa]['dia_tipo'] == "N/D" and 'Dia Tipo' in indices:
                    dia_tipo_val = sheet.cell(row=linha, column=indices['Dia Tipo']).value
                    if dia_tipo_val: dados_empresas[empresa]['dia_tipo'] = dia_tipo_val
        except (ValueError, TypeError, KeyError): continue
    return {'empresas': dados_empresas, 'geral': valor_geral_bloco}

def processar_quinzena(caminho_arquivo, mes, quinzena):
    intervalo = {'1ª': (7, 19, 8), '2ª': (30, 42, 30)}
    linha_inicio, linha_fim, linha_cabecalho = intervalo[quinzena]

    try:
        df_mes = pd.read_excel(caminho_arquivo, sheet_name=mes, header=None)
        bloco = df_mes.iloc[linha_inicio:linha_fim, :21].copy()
        headers = bloco.iloc[linha_cabecalho - linha_inicio].tolist()
        headers[0] = "EMPRESA"

        for c in ["Total CCT", "Total Geladinho"]:
            if c in headers:
                idx = headers.index(c)
                bloco.drop(bloco.columns[idx], axis=1, inplace=True)
                headers.pop(idx)

        bloco.columns = headers
        bloco = bloco.drop(bloco.index[linha_cabecalho - linha_inicio])
        bloco = bloco.dropna(how='all')

        dados_finais = []
        for _, linha in bloco.iterrows():
            empresa = str(linha['EMPRESA']).strip()
            if empresa.upper() in ["TOTAL STPP", "TOTAL", "", "EMPRESA"] or pd.isna(empresa):
                continue
            for disel, valor in linha.items():
                if disel == 'EMPRESA' or pd.isna(valor):
                    continue
                try:
                    valor_num = int(float(valor))
                    rendimento = Config.RENDIMENTO_DISPEL.get(disel, "")
                    dados_finais.append({
                        'EMPRESA': empresa,
                        'Disel': disel,
                        'Valor': valor_num,
                        'Rendimento pela Licitação': rendimento,
                        'Mês': mes,
                        'Quinzena': quinzena
                    })
                except (ValueError, TypeError):
                    continue

        df = pd.DataFrame(dados_finais)
        if not df.empty:
            df["Qte X Rend"] = pd.to_numeric(df["Valor"]) * pd.to_numeric(df["Rendimento pela Licitação"])
        return df

    except Exception as e:
        print(f"ERRO DETALHADO em 'processar_quinzena' na aba '{mes}':")
        traceback.print_exc() 
        messagebox.showerror("Erro ao processar planilha", f"Erro na aba '{mes}':\n{e}\n\nVerifique o console para mais detalhes.")
        return pd.DataFrame()
        
def adicionar_aba_rendimento_formatada(wb, df):
    ws = wb.create_sheet("Rendimento PCO")
    linha_atual = 1
    borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    rendimentos_finais = {}

    for empresa in df["EMPRESA"].unique():
        df_emp_original = df[df["EMPRESA"] == empresa].copy()
        
        disel_upper = df_emp_original["Disel"].astype(str).str.strip().str.upper()
        mascara_total = disel_upper.str.contains("TOTAL")
        mascara_idade_media = disel_upper.isin(["IDADE MÉDIA ATÉ VIDA ÚTIL", "IDADE MÉDIA TOTAL"])
        
        df_emp = df_emp_original[~mascara_total & ~mascara_idade_media].copy() 
        df_total_original = df_emp_original[mascara_total]

        if df_emp.empty: continue

        df_emp["Qte X Rend"] = pd.to_numeric(df_emp["Qte X Rend"], errors="coerce").fillna(0.0)
        df_emp["Valor"] = pd.to_numeric(df_emp["Valor"], errors="coerce").fillna(0).astype(int)
        
        total_qte_rend_calculado = df_emp["Qte X Rend"].sum()
        total_valor_calculado = df_emp["Valor"].sum()
        
        rend_total = total_qte_rend_calculado / total_valor_calculado if total_valor_calculado else 0
        df_emp["Porção(%)"] = df_emp["Valor"] / total_valor_calculado if total_valor_calculado else 0
        
        rendimentos_finais[empresa] = rend_total
        
        ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=6)
        cell_header_empresa = ws.cell(row=linha_atual, column=1, value=Config.NOMES_EMPRESAS.get(empresa, empresa))
        cell_header_empresa.font = Config.FONT_HEADER_EMPRESA
        cell_header_empresa.fill = Config.FILL_HEADER_EMPRESA
        cell_header_empresa.alignment = align_center
        linha_atual += 1

        headers = ['Categoria', 'Qtde Veic. PCO', 'Proporção (%)', 'Rendimento pela Licitação', 'Qte X Rend', 'Média Geral de Rendimento da Empresa']
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=linha_atual, column=col, value=text)
            cell.font = Config.FONT_HEADER_TABELA
            cell.fill = Config.FILL_HEADER_TABELA
            cell.alignment = align_center
            cell.border = borda_fina
        
        linha_inicio_dados = linha_atual + 1
        
        for _, row_data in df_emp.iterrows():
            linha_atual += 1
            data_to_write = [row_data["Disel"], row_data["Valor"], row_data["Porção(%)"], row_data["Rendimento pela Licitação"], row_data["Qte X Rend"]]
            for col, value in enumerate(data_to_write, 1):
                cell = ws.cell(row=linha_atual, column=col, value=value)
                cell.border = borda_fina
                cell.alignment = align_center if col > 1 else align_left
                if col == 2: cell.number_format = '0'
                if col == 3: cell.number_format = '0.00%'
                if col == 4: cell.number_format = '#,##0.000'
                if col == 5: cell.number_format = '#,##0.00'

        linha_total_inicio = linha_atual + 1
        
        if not df_total_original.empty:
            linha_total_final = linha_total_inicio
            label_total = df_total_original.iloc[0]["Disel"]
            rend_total_label = df_total_original.iloc[0]["Rendimento pela Licitação"]
        else:
            linha_total_final = linha_total_inicio
            label_total = "TOTAL CADASTRO"
            rend_total_label = ""

        linha_atual += 1
        data_to_write = [
            label_total, 
            total_valor_calculado,
            1,
            rend_total_label,
            total_qte_rend_calculado
        ]
        
        for col, value in enumerate(data_to_write, 1):
            cell = ws.cell(row=linha_atual, column=col, value=value)
            cell.font = Config.FONT_TOTAL
            cell.fill = Config.FILL_TOTAL
            cell.border = borda_fina
            cell.alignment = align_center if col > 1 else align_left
            if col == 2: cell.number_format = '0'
            if col == 3: cell.number_format = '100.00%'
            if col == 4: cell.number_format = '#,##0.000'
            if col == 5: cell.number_format = '#,##0.00'

        if not df_total_original.empty:
             linha_atual = linha_total_inicio + len(df_total_original) -1
        
        if linha_total_final < linha_total_inicio:
             linha_total_final = linha_total_inicio


        ws.merge_cells(start_row=linha_inicio_dados, start_column=6, end_row=linha_total_final, end_column=6)
        cell_media_geral = ws.cell(row=linha_inicio_dados, column=6, value=rend_total)
        cell_media_geral.font = Config.FONT_MEDIA_GERAL
        cell_media_geral.fill = Config.FILL_MEDIA_GERAL
        cell_media_geral.alignment = align_center
        cell_media_geral.border = borda_fina
        cell_media_geral.number_format = '0.00'
        
        linha_atual = linha_total_final + 2
        
    adicionar_marca_dagua(ws, anexo_celula='H1')
    return rendimentos_finais

def get_rendimento(vehicle_type, empresa):
    if not isinstance(vehicle_type, str): return 0
    v_type = vehicle_type.upper()
    rules_mob = { "MICRO": 2.2185, "PADRON COM AR": 1.8315 }
    rules_cno = {} 
    common_rules = { "ARTICULADO": 1.200, "PESADO COM AR": 2.2185, "BRT": 1.2, "MIDI": 2.61, "PESADO": 2.61 }
    empresa_rules = rules_mob if empresa.upper() == "MOB" else rules_cno
    for key, value in {**empresa_rules, **common_rules}.items():
        if key in v_type:
            return value
    return 0

def _adicionar_aba_calculo_consolidado(wb_destino, caminho_arquivo, df_km_prog_atual, contagem_dias, nome_aba, empresa_id):
    if not caminho_arquivo: return None
    try:
        df_source = pd.read_excel(caminho_arquivo, header=None)
        colunas_a_manter = min(df_source.shape[1], 12)
        df_source = df_source.iloc[:, :colunas_a_manter]

        mapa_km = {}
        if not df_km_prog_atual.empty:
            df_km_prog_atual['Dia Tipo'] = df_km_prog_atual['Dia Tipo'].str.upper().str.strip()
            df_km_prog_atual['Linha'] = df_km_prog_atual['Linha'].astype(str).str.strip()
            dia_tipo_map = {'DIA ÚTIL': 'DUT', 'SABADO': 'SAB', 'DOMINGO': 'DOM'}
            df_km_prog_atual['Dia Tipo'] = df_km_prog_atual['Dia Tipo'].replace(dia_tipo_map)
            for _, row in df_km_prog_atual.iterrows():
                chave = (row['Linha'], row['Dia Tipo'])
                mapa_km[chave] = row['Km Total']
        
        col_cod, col_km_dut, col_km_sab, col_km_dom = 0, 9, 10, 11
        for i in range(4, len(df_source)):
            try:
                cod_val = df_source.iloc[i, col_cod]
                if pd.isna(cod_val) or not str(cod_val).strip(): continue
                cod_str = str(int(cod_val))
                df_source.iat[i, col_km_dut] = mapa_km.get((cod_str, 'DUT'), df_source.iloc[i, col_km_dut])
                df_source.iat[i, col_km_sab] = mapa_km.get((cod_str, 'SAB'), df_source.iloc[i, col_km_sab])
                df_source.iat[i, col_km_dom] = mapa_km.get((cod_str, 'DOM'), df_source.iloc[i, col_km_dom])
            except (ValueError, TypeError, IndexError): continue
        
        ws = wb_destino.create_sheet(nome_aba)
        align_center_vcenter = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left_vcenter = Alignment(horizontal='left', vertical='center')
        borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        df_header = df_source.iloc[0:4].reindex(columns=range(16)).copy().astype(object)
        df_header.iat[2, 9], df_header.iat[3, 9], df_header.iat[3, 10], df_header.iat[3, 11] = "QUILOMETRAGEM", "DUT", "SAB", "DOM"
        df_header.iat[2, 12], df_header.iat[2, 13], df_header.iat[3, 13], df_header.iat[3, 14], df_header.iat[3, 15] = "RENDIMENTO", "COTA CONSIDERADA (L)", "DUT", "SAB", "DOM"

        for _, row_data in df_header.iterrows():
            ws.append(row_data.where(pd.notna(row_data), None).tolist())
        
        ws.merge_cells('A1:P1'); ws.merge_cells('A2:P2'); ws.merge_cells('A3:A4'); ws.merge_cells('B3:B4'); ws.merge_cells('C3:C4'); ws.merge_cells('D3:F3')
        ws.merge_cells('G3:I3'); ws.merge_cells('J3:L3'); ws.merge_cells('M3:M4'); ws.merge_cells('N3:P3')

        ws.row_dimensions[1].height = 35.15; ws['A1'].font = Config.FONT_TITULO_CALC; ws['A1'].alignment = align_left_vcenter
        ws.row_dimensions[2].height = 20; ws['A2'].font = Config.FONT_SUBTITULO_CALC; ws['A2'].alignment = align_left_vcenter
        ws.row_dimensions[3].height = 39.7
        for row in ws.iter_rows(min_row=3, max_row=4):
            for cell in row:
                cell.font, cell.fill, cell.border, cell.alignment = Config.FONT_CABECALHO_CALC, Config.FILL_CABECALHO_CALC, borda_fina, align_center_vcenter

        df_data = df_source.iloc[4:].reindex(columns=range(16)).copy()
        df_data.columns = range(df_data.shape[1])
        for col in df_data.columns[3:12]: df_data[col] = pd.to_numeric(df_data[col], errors='coerce')
        
        condition = (df_data[0].astype(str).str.strip().replace('nan', '') == '') & (df_data[1].astype(str).str.strip().replace('nan', '') == '') & (df_data[2].astype(str).str.strip().replace('nan', '') == '')
        total_geral_indices = df_data[condition].index
        if not total_geral_indices.empty: df_data = df_data.drop(index=total_geral_indices)
        if not df_data.empty: df_data = df_data.iloc[:-1].copy()

        df_data[12] = df_data[2].apply(lambda x: get_rendimento(x, empresa_id))
        df_data.loc[:, 0] = df_data[0].ffill().infer_objects(copy=False)
        df_data.loc[:, 1] = df_data[1].ffill().infer_objects(copy=False)

        dados_processados = pd.DataFrame()
        for _, group in df_data.groupby(0, dropna=True):
            group_modificado = group.copy()
            unique_vehicle_type_count = group[2].dropna().astype(str).str.strip().replace('', pd.NA).nunique()
            is_multi_vehicle_group = unique_vehicle_type_count > 1

            if is_multi_vehicle_group:
                col_veic_dut, col_veic_sab, col_veic_dom = 6, 7, 8
                total_km_dut, total_km_sab, total_km_dom = group.iloc[0, col_km_dut], group.iloc[0, col_km_sab], group.iloc[0, col_km_dom]
                total_veic_dut, total_veic_sab, total_veic_dom = group[col_veic_dut].sum(), group[col_veic_sab].sum(), group[col_veic_dom].sum()
                group_modificado[col_km_dut] = group.apply(lambda r: (total_km_dut / total_veic_dut * r[col_veic_dut]) if total_veic_dut else 0, axis=1)
                group_modificado[col_km_sab] = group.apply(lambda r: (total_km_sab / total_veic_sab * r[col_veic_sab]) if total_veic_sab else 0, axis=1)
                group_modificado[col_km_dom] = group.apply(lambda r: (total_km_dom / total_veic_dom * r[col_veic_dom]) if total_veic_dom else 0, axis=1)
            dados_processados = pd.concat([dados_processados, group_modificado])

        for i, cota_col in enumerate([13, 14, 15]):
            dados_processados[cota_col] = dados_processados.apply(lambda r: r[col_km_dut + i] / r[12] if r[12] else 0, axis=1)

        linha_destino_atual = 5
        for _, group in dados_processados.groupby(0, dropna=True):
            start_merge_row = linha_destino_atual
            unique_vehicle_type_count = group[2].dropna().astype(str).str.strip().replace('', pd.NA).nunique()
            is_multi_vehicle_group = unique_vehicle_type_count > 1

            for i, (_, data_row) in enumerate(group.iterrows()):
                row_to_write = data_row.where(pd.notna(data_row), None).tolist()
                if i > 0: row_to_write[0], row_to_write[1] = None, None
                ws.append(row_to_write)
                new_row_obj = ws[linha_destino_atual + i]
                for c_idx, cell in enumerate(new_row_obj):
                    cell.font, cell.border, cell.alignment = Config.FONT_DADOS_CALC, borda_fina, align_center_vcenter
                    if c_idx == 0: cell.number_format = '0'; continue
                    if isinstance(cell.value, (int, float)):
                        if col_km_dut <= c_idx <= col_km_dom: cell.number_format = '#,##0.00'
                        elif c_idx == 12: cell.number_format = '#,##0.0000'
                        elif 13 <= c_idx <= 15: cell.number_format = '#,##0.00'
                        else: cell.number_format = '#,##0'
                    if c_idx >= 9 and is_multi_vehicle_group: cell.fill = Config.FILL_YELLOW
            
            linha_destino_atual += len(group)
            if is_multi_vehicle_group:
                end_merge_row = linha_destino_atual - 1
                ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=end_merge_row, end_column=1)
                ws.cell(start_merge_row, 1).alignment = align_center_vcenter
                ws.merge_cells(start_row=start_merge_row, start_column=2, end_row=end_merge_row, end_column=2)
                ws.cell(start_merge_row, 2).alignment = align_center_vcenter

                total_row_list = [''] * group.shape[1]; total_row_list[2] = "TOTAL"
                for col_idx in range(3, group.shape[1]):
                    if col_idx not in [12, 13, 14, 15]:
                        col_letra = get_column_letter(col_idx + 1)
                        total_row_list[col_idx] = f"=SUM({col_letra}{start_merge_row}:{col_letra}{end_merge_row})"
                ws.append(total_row_list)
                total_row_obj = ws[linha_destino_atual]
                for c_idx, cell in enumerate(total_row_obj):
                    cell.font, cell.border, cell.alignment, cell.fill = Config.FONT_TOTAL_GRUPO_CALC, borda_fina, align_center_vcenter, Config.FILL_YELLOW
                    if (isinstance(cell.value, (int, float))) or (isinstance(cell.value, str) and cell.value.startswith('=')):
                        cell.number_format = '#,##0.00' if 13 <= c_idx <= 15 else '#,##0'
                linha_destino_atual += 1
        
        total_litros_final = None
        if not dados_processados.empty:
            total_geral_series = pd.Series([''] * 16, index=range(16)); total_geral_series[0] = "TOTAL GERAL"
            for col_idx in range(3, 16):
                if col_idx != 12: total_geral_series[col_idx] = dados_processados[col_idx].sum()
            ws.append(total_geral_series.where(pd.notna(total_geral_series), None).tolist())
            total_geral_row_idx = ws.max_row
            ws.merge_cells(start_row=total_geral_row_idx, start_column=1, end_row=total_geral_row_idx, end_column=3)
            for cell in ws[total_geral_row_idx]:
                cell.font, cell.border, cell.fill, cell.alignment = Config.FONT_TOTAL_GERAL_CALC, borda_fina, Config.FILL_BLUE_TOTAL_GERAL, align_center_vcenter
                if isinstance(cell.value, (int, float)): cell.number_format = '#,##0'

            dias = { 'DUT': contagem_dias.get('DUT', 0), 'SAB': contagem_dias.get('SAB', 0), 'DOM': contagem_dias.get('DOM', 0) }
            linha_inicio_resumo = ws.max_row + 3
            ws.cell(row=linha_inicio_resumo - 1, column=11, value="Qtd. Dias").font = Font(bold=True)
            ws.cell(row=linha_inicio_resumo - 1, column=12, value="LITROS").font = Font(bold=True)

            litros_diario = { 'DUT': total_geral_series[13] or 0, 'SAB': total_geral_series[14] or 0, 'DOM': total_geral_series[15] or 0 }
            litros_total = { k: dias[k] * litros_diario[k] for k in dias }
            
            dados_resumo = [("DUT", dias['DUT'], litros_total['DUT']), ("SAB", dias['SAB'], litros_total['SAB']), ("DOM", dias['DOM'], litros_total['DOM'])]
            for i, (label, dia, litro) in enumerate(dados_resumo):
                linha = linha_inicio_resumo + i
                ws.cell(row=linha, column=10, value=label); ws.cell(row=linha, column=11, value=dia).number_format = '0'; ws.cell(row=linha, column=12, value=litro).number_format = '#,##0.00'

            linha_total_resumo = linha_inicio_resumo + len(dados_resumo)
            ws.cell(row=linha_total_resumo, column=10, value="TOTAL").font = Font(bold=True)
            ws.cell(row=linha_total_resumo, column=11, value=f"=SUM(K{linha_inicio_resumo}:K{linha_total_resumo-1})").font = Font(bold=True); ws.cell(row=linha_total_resumo, column=11).number_format = '0'; ws.cell(row=linha_total_resumo, column=11).fill = Config.FILL_YELLOW
            ws.cell(row=linha_total_resumo, column=12, value=f"=SUM(L{linha_inicio_resumo}:L{linha_total_resumo-1})").font = Font(bold=True); ws.cell(row=linha_total_resumo, column=12).number_format = '#,##0.00'; ws.cell(row=linha_total_resumo, column=12).fill = Config.FILL_YELLOW
            
            total_litros_final = sum(litros_total.values())
            valor_arredondado = math.ceil(total_litros_final / 5000) * 5000
            ws.cell(row=linha_total_resumo, column=13, value=valor_arredondado).font = Font(bold=True); ws.cell(row=linha_total_resumo, column=13).number_format = '#,##0'

            for row in ws.iter_rows(min_row=linha_inicio_resumo, max_row=linha_total_resumo, min_col=10, max_col=13):
                for cell in row:
                    if not (cell.column == 13 and cell.row != linha_total_resumo): cell.border = borda_fina
        
        ws.column_dimensions['A'].width = 5.57
        adicionar_marca_dagua(ws, anexo_celula='Q1')
        return total_litros_final
    
    except Exception as e:
        messagebox.showerror(f"Erro no {nome_aba}", f"Não foi possível processar a aba '{nome_aba}'.\n\nDetalhes: {e}")
        return None

def _aplicar_estilos_e_criar_sumario(ws_destino, blocos_info_atual, start_col_atual, max_cols_atual, linhas_alteradas_df, mes_passado_str, mes_atual_str):
    estilos = Config.ESTILOS_SUMARIO
    col_inicio_totais, num_cols_tabela = 1, 6
    ws_destino.merge_cells(start_row=1, start_column=col_inicio_totais, end_row=1, end_column=col_inicio_totais + num_cols_tabela - 1)
    cell_cut_titulo = ws_destino.cell(row=1, column=col_inicio_totais, value="CUT - CONFERÊNCIA DE QUILOMETRAGEM")
    cell_cut_titulo.font, cell_cut_titulo.fill, cell_cut_titulo.alignment = estilos['titulo_principal']['font'], estilos['titulo_principal']['fill'], Alignment(horizontal='center', vertical='center')

    linha_atual_totais = 2
    todos_os_dados_blocos = {}

    for bloco in sorted(blocos_info_atual, key=lambda item: obter_chave_ordenacao(item['arquivo'])):
        resultados = processar_bloco_para_totais(ws_destino, bloco['inicio'], bloco['fim'], start_col_atual + 1, start_col_atual + max_cols_atual)
        dados_bloco, valor_geral_dia = resultados['empresas'], resultados['geral']
        
        nome_arquivo = os.path.basename(bloco['arquivo']).upper()
        if "DUT" in nome_arquivo: tipo_dia_chave = 'DUT'
        elif "SÁB" in nome_arquivo or "SAB" in nome_arquivo: tipo_dia_chave = 'SAB'
        elif "DOM" in nome_arquivo: tipo_dia_chave = 'DOM'
        else: tipo_dia_chave = 'OUTRO'
        
        todos_os_dados_blocos[tipo_dia_chave] = dados_bloco
        tipo_dia_titulo = "SÁBADO" if tipo_dia_chave == 'SAB' else "DOMINGO" if tipo_dia_chave == 'DOM' else "DIA ÚTIL"

        ws_destino.merge_cells(start_row=linha_atual_totais, start_column=col_inicio_totais, end_row=linha_atual_totais, end_column=col_inicio_totais + num_cols_tabela - 1)
        cell_bloco_titulo = ws_destino.cell(row=linha_atual_totais, column=col_inicio_totais, value=f"CONFERÊNCIA - {tipo_dia_titulo}")
        cell_bloco_titulo.font, cell_bloco_titulo.fill, cell_bloco_titulo.alignment = estilos['titulo_bloco']['font'], estilos['titulo_bloco']['fill'], Alignment(horizontal='center', vertical='center')
        linha_atual_totais += 1

        headers = ["Status", "Empresa", "Dia Tipo", "KM Declarado", "KM Calculado", "Diferença"]
        for col, text in enumerate(headers):
            cell = ws_destino.cell(row=linha_atual_totais, column=col_inicio_totais + col, value=text)
            cell.font, cell.fill, cell.alignment, cell.border = estilos['cabecalho_tabela']['font'], estilos['cabecalho_tabela']['fill'], Alignment(horizontal='center'), estilos['borda']
        linha_atual_totais += 1
        
        linha_inicio_dados_bloco = linha_atual_totais
        total_dia_tipo_soma = 0
        empresas_desejadas = ['BOA', 'CAX', 'CNO', 'CSR', 'EME', 'GLO', 'MOB', 'SJT', 'VML']
        for idx, empresa in enumerate(empresas_desejadas):
            if empresa in dados_bloco:
                dados = dados_bloco[empresa]
                diferenca = round(dados.get('valor_declarado', 0) - dados.get('soma_calculada', 0), 2)
                total_dia_tipo_soma += dados.get('soma_calculada', 0)
                is_ok = abs(diferenca) < 0.01
                
                row_data = ["✓ Ok" if is_ok else "✗ Diferença", empresa, dados.get('dia_tipo', 'N/D'), dados.get('valor_declarado', 0), dados.get('soma_calculada', 0), diferenca]
                for col, value in enumerate(row_data):
                    cell = ws_destino.cell(row=linha_atual_totais, column=col_inicio_totais + col, value=value)
                    cell.border = estilos['borda']
                    if idx % 2 != 0: cell.fill = estilos['zebra']
                
                ws_destino.cell(row=linha_atual_totais, column=col_inicio_totais).font = estilos['status_ok'] if is_ok else estilos['status_diff']
                for c_idx in range(3, 6): ws_destino.cell(row=linha_atual_totais, column=col_inicio_totais + c_idx).number_format = '#,##0.00'
                linha_atual_totais += 1

        if linha_atual_totais > linha_inicio_dados_bloco:
            col_letra = get_column_letter(col_inicio_totais + 5)
            range_str = f"{col_letra}{linha_inicio_dados_bloco}:{col_letra}{linha_atual_totais - 1}"
            regra = Rule(type="expression", dxf=estilos['dxf_diff'], formula=[f"ABS({col_letra}{linha_inicio_dados_bloco})>=0.01"])
            ws_destino.conditional_formatting.add(range_str, regra)

        valor_ctc_dia = dados_bloco.get('CTC', {}).get('valor_declarado', 0) or dados_bloco.get('CTC', {}).get('soma_calculada', 0)
        total_da_ctc = total_dia_tipo_soma + valor_ctc_dia
        dif_geral = round(valor_geral_dia - total_da_ctc, 2)

        resumo_data = [("Total do Dia Tipo (Calculado)", total_dia_tipo_soma), ("CTC", valor_ctc_dia), ("Total da CTC (Soma + CTC)", total_da_ctc), ("Valor Geral Declarado (do Bloco)", valor_geral_dia)]
        for label, value in resumo_data:
            ws_destino.merge_cells(start_row=linha_atual_totais, start_column=col_inicio_totais, end_row=linha_atual_totais, end_column=col_inicio_totais + 4)
            cell_label = ws_destino.cell(row=linha_atual_totais, column=col_inicio_totais, value=label)
            cell_label.alignment, cell_label.font = Alignment(horizontal='right', vertical='center'), estilos['font_sumario_label']
            cell_value = ws_destino.cell(row=linha_atual_totais, column=col_inicio_totais + 5, value=value)
            cell_value.number_format, cell_value.font, cell_value.alignment = '#,##0.00', estilos['font_sumario_valor'], Alignment(horizontal='right')
            if "Total da CTC" in label or "Valor Geral" in label:
                for c in range(col_inicio_totais, col_inicio_totais + 6): ws_destino.cell(linha_atual_totais, c).fill = estilos['sumario_total']
            for c in range(col_inicio_totais, col_inicio_totais + 6): ws_destino.cell(linha_atual_totais, c).border = estilos['borda']
            linha_atual_totais += 1

        ws_destino.merge_cells(start_row=linha_atual_totais, start_column=col_inicio_totais, end_row=linha_atual_totais, end_column=col_inicio_totais + 4)
        cell_label_final = ws_destino.cell(row=linha_atual_totais, column=col_inicio_totais, value="DIFERENÇA GERAL")
        cell_label_final.alignment, cell_label_final.font, cell_label_final.fill = Alignment(horizontal='right'), estilos['final_label']['font'], estilos['final_label']['fill']
        cell_value_final = ws_destino.cell(row=linha_atual_totais, column=col_inicio_totais + 5, value=dif_geral)
        cell_value_final.number_format, cell_value_final.font, cell_value_final.fill, cell_value_final.alignment = '#,##0.00', estilos['final_valor']['font'], estilos['final_valor']['fill'], Alignment(horizontal='right')
        for c in range(col_inicio_totais, col_inicio_totais + 6): ws_destino.cell(linha_atual_totais, c).border = estilos['borda']
        linha_atual_totais += 2

    if not linhas_alteradas_df.empty:
        linha_atual_totais += 2
        ws_destino.merge_cells(start_row=linha_atual_totais, start_column=col_inicio_totais, end_row=linha_atual_totais, end_column=col_inicio_totais + 4)
        cell_titulo_excluidas = ws_destino.cell(row=linha_atual_totais, column=col_inicio_totais, value="Linhas Novas e Excluídas")
        cell_titulo_excluidas.font = estilos['titulo_bloco']['font']
        cell_titulo_excluidas.fill = estilos['titulo_bloco']['fill']
        cell_titulo_excluidas.alignment = Alignment(horizontal='center', vertical='center')
        linha_atual_totais += 1

        headers_excluidas = ["Status", "Empresa", "Dia Tipo", "Linha", "Km"]
        for col, text in enumerate(headers_excluidas):
            cell = ws_destino.cell(row=linha_atual_totais, column=col_inicio_totais + col, value=text)
            cell.font, cell.fill, cell.alignment, cell.border = estilos['cabecalho_tabela']['font'], estilos['cabecalho_tabela']['fill'], Alignment(horizontal='center'), estilos['borda']
        linha_atual_totais += 1

        for idx, row_data in linhas_alteradas_df.iterrows():
            dados_linha = [
                row_data.get('Status', 'N/A'),
                row_data.get('Empresa', 'N/A'), 
                row_data.get('Dia Tipo', 'N/A'), 
                row_data.get('Linha', 'N/A'),
                row_data.get('Km Total', 0)
            ]
            for col, value in enumerate(dados_linha):
                cell = ws_destino.cell(row=linha_atual_totais, column=col_inicio_totais + col, value=value)
                cell.border = estilos['borda']
                if col == 4:
                    cell.number_format = '#,##0.00'
                if idx % 2 != 0:
                    cell.fill = estilos['zebra']
            linha_atual_totais += 1

    return todos_os_dados_blocos

def extrair_dados_para_dataframe(lista_arquivos):
    dados_completos = []
    chaves_necessarias = {'Linha', 'Km Total', 'Empresa', 'Dia Tipo'}
    for arquivo in lista_arquivos:
        try:
            import xlrd
            book = xlrd.open_workbook(arquivo)
            sheet = book.sheet_by_index(0)
            raw_data = [sheet.row_values(i) for i in range(sheet.nrows)]
            indices, linha_header = {}, -1
            for r in range(min(15, len(raw_data))):
                for c, valor_cabecalho in enumerate(raw_data[r]):
                    if isinstance(valor_cabecalho, str):
                        header = valor_cabecalho.strip()
                        if header in chaves_necessarias and header not in indices:
                            indices[header] = c
                if len(indices) == len(chaves_necessarias):
                    linha_header = r
                    break
            if linha_header == -1: continue
            for r in range(linha_header + 1, len(raw_data)):
                row_data = raw_data[r]
                try:
                    linha_val = row_data[indices['Linha']]
                    if 'total' in str(linha_val).lower(): continue
                    try: linha_str = str(int(float(linha_val)))
                    except (ValueError, TypeError): linha_str = str(linha_val).strip()
                    dados_completos.append({
                        'Linha': linha_str,
                        'Empresa': str(row_data[indices['Empresa']]),
                        'Dia Tipo': str(row_data[indices['Dia Tipo']]),
                        'Km Total': float(row_data[indices['Km Total']])
                    })
                except (ValueError, TypeError, IndexError): continue
        except Exception: continue
    if not dados_completos: return pd.DataFrame()
    df = pd.DataFrame(dados_completos)
    df.drop_duplicates(subset=['Linha', 'Empresa', 'Dia Tipo'], keep='last', inplace=True)
    return df

def _criar_aba_cota_mes(wb, km_a_usar_por_empresa, rendimentos_por_empresa, contagem_dias, mes_ref_str, info_pco_str, cota_cno_especial=None, cota_mob_especial=None, dados_rateamento=None, reducoes_dict=None):
    ws = wb.create_sheet("Cota do Mês")
    
    font_titulo_principal = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
    font_subtitulo = Font(name='Calibri', size=11, bold=True, color="404040")
    font_header_grupo = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    font_header_coluna = Font(name='Calibri', size=11, bold=True, color="000000")
    font_dados_empresa = Font(name='Calibri', size=11, bold=True)
    font_dados_normal = Font(name='Calibri', size=11)
    font_dados_destaque = Font(name='Calibri', size=11, bold=True, color="BF5B00")
    font_total = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    font_total_secundario = Font(name='Calibri', size=11, bold=True, color="000000")

    fill_titulo_principal = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    fill_header_grupo = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    fill_header_coluna = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    fill_total_principal = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    fill_total_secundario = PatternFill(start_color="D9D9D9", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8F8F8", fill_type="solid")
    fill_destaque = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_especial_cno_mob = PatternFill(start_color="FFEB9C", fill_type="solid")

    borda_fina = Border(left=Side(style='thin', color="BFBFBF"), right=Side(style='thin', color="BFBFBF"), top=Side(style='thin', color="BFBFBF"), bottom=Side(style='thin', color="BFBFBF"))
    borda_total = Border(top=Side(style='double', color='000000'), bottom=Side(style='double', color='000000'))

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('A1:M1')
    titulo1 = ws.cell(row=1, column=1, value="QUANTIDADE MÁXIMA DE ÓLEO DIESEL A SER ADQUIRIDO POR CRÉDITO PRESUMIDO DO ICMS NOS TERMOS DO CONVÊNIO 21/2023")
    titulo1.font = font_titulo_principal
    titulo1.fill = fill_titulo_principal
    titulo1.alignment = align_center
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:O2')
    data_elaborado = datetime.date.today().strftime('%d/%m/%Y')
    titulo2_text = f"REFERÊNCIA: {mes_ref_str}  (Elaborado em {data_elaborado} – Fonte: {info_pco_str})"
    titulo2 = ws.cell(row=2, column=1, value=titulo2_text)
    titulo2.font = font_subtitulo
    titulo2.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    headers_grupo = [
        ("EMPRESA", 'A4:A5'), ("QUILOMETRAGEM DIÁRIA (CALCULADA)", 'B4:D4'), 
        ("CÁLCULO BASE", 'E4:G4'), ("AJUSTES E REDUÇÃO", 'H4:I4'),
        ("COTA RESULTANTE", 'J4:K4'), ("COMPARAÇÃO (OPCIONAL)", 'L4:O4')
    ]
    
    for text, merge_range in headers_grupo:
        ws.merge_cells(merge_range)
        cell = ws[merge_range.split(':')[0]]
        cell.value = text
        cell.font = font_header_grupo
        cell.fill = fill_header_grupo
        cell.alignment = align_center
        
    sub_headers = [
        "DIA ÚTIL", "SÁBADO", "DOMINGO", "TOTAL KM MENSAL", "RENDIMENTO", "COTA CALCULADA (L)",
        "REDUÇÃO / ACRÉSCIMO", "COTA COM REDUÇÃO",
        "MÚLTIPLO 5.000 (L)", "COTA CONSIDERADA (L)", "VALOR SOLICITADO (L)", "DIFERENÇA", "% DIFERENÇA", "STATUS"
    ]
    
    for i, text in enumerate(sub_headers, 2):
        cell = ws.cell(row=5, column=i, value=text)
        cell.font = font_header_coluna
        cell.fill = fill_header_coluna
        cell.alignment = align_center

    for row in ws['A4:O5']:
        for cell in row:
            cell.border = borda_fina

    empresas_ordem = ['BOA', 'CAX', 'CNO', 'CSR', 'EME', 'GLO', 'MOB', 'SJT', 'VML']
    linha_atual = 6
    cotas_consideradas = {}

    solicitado_por_empresa = {}
    if dados_rateamento:
        dados_rateamento_local = dados_rateamento.copy()
        if 'MOBI' in dados_rateamento_local:
            dados_rateamento_local['MOB'] = dados_rateamento_local.pop('MOBI')
        
        for empresa, dados in dados_rateamento_local.items():
            solicitado_por_empresa[empresa] = sum(d[0] for d in dados if d[0] is not None)

    dias_mult = {'DUT': contagem_dias.get('DUT', 0), 'SAB': contagem_dias.get('SAB', 0), 'DOM': contagem_dias.get('DOM', 0)}

    for i, empresa_cod in enumerate(empresas_ordem):
        km_dut = km_a_usar_por_empresa.get((empresa_cod, 'DUT'), 0)
        km_sab = km_a_usar_por_empresa.get((empresa_cod, 'SAB'), 0)
        km_dom = km_a_usar_por_empresa.get((empresa_cod, 'DOM'), 0)
        
        total_km = (km_dut * dias_mult['DUT']) + (km_sab * dias_mult['SAB']) + (km_dom * dias_mult['DOM'])
        rendimento = rendimentos_por_empresa.get(empresa_cod, 0)
        
        cota_calculada = (total_km / rendimento) if rendimento else 0
        if empresa_cod == 'CNO' and cota_cno_especial is not None: cota_calculada = cota_cno_especial
        if empresa_cod == 'MOB' and cota_mob_especial is not None: cota_calculada = cota_mob_especial

        rendimento_para_exibir = "" if empresa_cod in ['CNO', 'MOB'] else rendimento

        # --- LÓGICA DE REDUÇÃO ---
        reducao_pct = 0.0
        if reducoes_dict:
            reducao_pct = reducoes_dict.get(empresa_cod, 0.0)
        
        # Cota com Redução = Cota Calc + (Cota Calc * %Redução)
        cota_com_reducao = cota_calculada + (cota_calculada * reducao_pct)

        # O Múltiplo agora usa a cota COM redução
        multiplo_5000 = (math.floor(cota_com_reducao / 5000) * 5000 if (cota_com_reducao % 5000) < 2500 else math.ceil(cota_com_reducao / 5000) * 5000) if cota_com_reducao > 0 else 0
        
        # --- INÍCIO DA CORREÇÃO 2 ---
        # Pega o valor bruto
        valor_solicitado_bruto = solicitado_por_empresa.get(empresa_cod, 0)
        # Aplica a MESMA regra de arredondamento
        valor_solicitado = (math.floor(valor_solicitado_bruto / 5000) * 5000 if (valor_solicitado_bruto % 5000) < 2500 else math.ceil(valor_solicitado_bruto / 5000) * 5000) if valor_solicitado_bruto > 0 else 0

        cota_considerada = multiplo_5000
        # A verificação agora usa o valor JÁ arredondado
        if 0 < valor_solicitado < cota_considerada:
            cota_considerada = valor_solicitado
        # --- FIM DA CORREÇÃO 2 ---
            
        cotas_consideradas[empresa_cod] = cota_considerada    

        valores_linha = [
            Config.NOMES_EMPRESAS.get(empresa_cod, empresa_cod),
            km_dut, km_sab, km_dom, total_km, rendimento_para_exibir, cota_calculada,
            reducao_pct, cota_com_reducao,
            multiplo_5000, cota_considerada, valor_solicitado
        ]

        fill_linha = fill_zebra if i % 2 != 0 else None
        
        for col, valor in enumerate(valores_linha, 1):
            cell = ws.cell(row=linha_atual, column=col, value=valor)
            cell.alignment = align_center
            cell.font = font_dados_normal
            if fill_linha: cell.fill = fill_linha
            if col == 1: cell.font = font_dados_empresa; cell.alignment = align_left
            if col in [2,3,4,5]: cell.number_format = '#,##0.00'
            if col == 6: cell.number_format = '0.00'
            if col == 7: cell.number_format = '#,##0.00' # Cota Calc
            if col == 8: cell.number_format = '0.00%' # Redução
            if col == 9: cell.number_format = '#,##0.00' # Cota Com Redução
            if col in [10,11,12]: cell.number_format = '#,##0' # Multiplo, Cota Cons, Solicitado
            if col == 11: cell.font = font_dados_destaque; cell.fill = fill_destaque
        
        # Fórmulas de Diferença e Status (Colunas 13, 14, 15)
        # Diferença = Solicitado (L) - Cota Considerada (K) -> M = L - K
        ws.cell(row=linha_atual, column=13, value=f"=L{linha_atual}-K{linha_atual}").number_format = '#,##0'
        # % Diferença = Diferença (M) / Solicitado (L)? Não, a lógica anterior era: if solicitado != 0, diff/solicitado.
        # Mas espere, a lógica antiga era J-I (Solicitado - Considerada). E a % era K/J (Diff / Solicitado).
        # Agora: Solicitado=L, Dif=M. Então =M/L.
        ws.cell(row=linha_atual, column=14, value=f'=IF(L{linha_atual}<>0, M{linha_atual}/L{linha_atual}, 0)').number_format = '0.00%'
        
        # Status
        status_cell = ws.cell(row=linha_atual, column=15, value=f'=IF(L{linha_atual}>0, IF(ABS(N{linha_atual})>0.05, "ATENÇÃO", "OK"), "N/A")')
        status_cell.font = font_dados_normal
        status_cell.alignment = align_center

        for c in range(1, 16):
            ws.cell(row=linha_atual, column=c).border = borda_fina
            if empresa_cod in ['CNO', 'MOB'] and (cota_cno_especial is not None or cota_mob_especial is not None):
                ws.cell(row=linha_atual, column=c).fill = fill_especial_cno_mob

        linha_atual += 1

    linha_total = linha_atual
    cell_total_label = ws.cell(row=linha_total, column=1, value="TOTAL GERAL")
    
    colunas_soma = ['B', 'C', 'D', 'E', 'G', 'I', 'J', 'K', 'L']
    for col_letra in colunas_soma:
        col_idx = ord(col_letra) - 64
        formula = f"=SUM({col_letra}{6}:{col_letra}{linha_total-1})"
        cell = ws.cell(row=linha_total, column=col_idx, value=formula)
        if col_idx <= 5: cell.number_format = '#,##0.00'
        elif col_idx == 7: cell.number_format = '#,##0.00' # Cota Calc (G)
        elif col_idx == 9: cell.number_format = '#,##0.00' # Cota Red (I)
        else: cell.number_format = '#,##0' # J, K, L

    formula_rend_medio = f"=AVERAGE(F6:F{linha_total-1})"
    cell_rend_medio = ws.cell(row=linha_total, column=6, value=formula_rend_medio)
    cell_rend_medio.number_format = '0.00'
    
    for c in range(1, 16):
        cell = ws.cell(row=linha_total, column=c)
        cell.fill = fill_total_principal
        cell.font = font_total
        cell.border = borda_total
        cell.alignment = align_center
    ws.cell(row=linha_total, column=1).alignment = align_left

    linha_dias = linha_total + 2
    ws.merge_cells(f'E{linha_dias}:G{linha_dias}')
    cell_titulo_dias = ws.cell(row=linha_dias, column=5, value="RESUMO DO PERÍODO")
    cell_titulo_dias.fill = fill_header_coluna
    cell_titulo_dias.font = font_header_coluna
    cell_titulo_dias.alignment = align_center
    cell_titulo_dias.border = borda_fina
    ws.cell(row=linha_dias, column=7).border = borda_fina

    dias_data = {"Dias Úteis (DUT):": dias_mult['DUT'], "Sábados (SAB):": dias_mult['SAB'], "Domingos (DOM):": dias_mult['DOM']}
    
    for i, (label, value) in enumerate(dias_data.items()):
        ws.cell(row=linha_dias + i + 1, column=5, value=label).alignment = align_right
        ws.cell(row=linha_dias + i + 1, column=5).font = font_total_secundario
        ws.cell(row=linha_dias + i + 1, column=6, value=value).alignment = align_center
        ws.cell(row=linha_dias + i + 1, column=6).font = font_dados_normal
    
    total_dias = sum(dias_mult.values())
    ws.cell(row=linha_dias + 4, column=5, value="Total de Dias:").font = font_total_secundario
    ws.cell(row=linha_dias + 4, column=5, value="Total de Dias:").alignment = align_right
    cell_total_dias = ws.cell(row=linha_dias + 4, column=6, value=total_dias)
    cell_total_dias.alignment = align_center
    cell_total_dias.font = font_total_secundario
    cell_total_dias.fill = fill_total_secundario

    for row_idx in range(linha_dias + 1, linha_dias + 5):
        ws.cell(row_idx, 5).border = borda_fina
        ws.cell(row_idx, 6).border = borda_fina

    ws.column_dimensions['A'].width = 45
    for col_letter in ['B', 'C', 'D']: ws.column_dimensions[col_letter].width = 14
    for col_letter in ['E', 'G', 'I', 'J', 'K', 'L', 'M']: ws.column_dimensions[col_letter].width = 18
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['H'].width = 14 # Redução
    ws.column_dimensions['N'].width = 12
    ws.column_dimensions['O'].width = 10
    
    ws.freeze_panes = 'A6'
    adicionar_marca_dagua(ws, 'P1')

    return cotas_consideradas

def _adicionar_aba_rateamento(workbook, dados_empresas, cotas_consideradas=None):
    sheet = workbook.create_sheet('Rateamento das Garagens')

    empresas = ['BOA', 'CAX', 'CSR', 'CNO', 'EME', 'GLO', 'MOBI', 'SJT', 'VML']
    
    borda_fina = Border(left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000'))
    borda_total = Border(top=Side(style='double', color='000000'), 
                         bottom=Side(style='double', color='000000'))

    alinhamento_central = Alignment(horizontal='center', vertical='center', wrap_text=True)

    estilo_titulo = NamedStyle(name='titulo')
    estilo_titulo.font = Font(name='Calibri', size=26, bold=True, color='FFFFFF')
    estilo_titulo.alignment = alinhamento_central
    estilo_titulo.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    estilo_titulo.border = borda_fina

    estilo_cabecalho_principal = NamedStyle(name='cabecalho_principal')
    estilo_cabecalho_principal.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    estilo_cabecalho_principal.alignment = alinhamento_central
    estilo_cabecalho_principal.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    estilo_cabecalho_principal.border = borda_fina

    estilo_sub_cabecalho = NamedStyle(name='sub_cabecalho')
    estilo_sub_cabecalho.font = Font(name='Calibri', size=11, bold=True, color='000000')
    estilo_sub_cabecalho.alignment = alinhamento_central
    estilo_sub_cabecalho.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    estilo_sub_cabecalho.border = borda_fina

    estilo_nome_empresa = NamedStyle(name='nome_empresa')
    estilo_nome_empresa.font = Font(name='Calibri', size=16, bold=True, color='000000')
    estilo_nome_empresa.alignment = alinhamento_central
    estilo_nome_empresa.border = borda_fina

    estilo_cota_empresa = NamedStyle(name='cota_empresa')
    estilo_cota_empresa.font = Font(name='Calibri', size=16, bold=True, color='000000')
    estilo_cota_empresa.alignment = alinhamento_central
    estilo_cota_empresa.border = borda_fina
    estilo_cota_empresa.number_format = '#,##0'

    estilo_dados_padrao = NamedStyle(name='dados_padrao')
    estilo_dados_padrao.border = borda_fina
    estilo_dados_padrao.font = Font(name='Calibri', size=11, color='000000')
    estilo_dados_padrao.alignment = alinhamento_central
    
    estilo_dados_percentual_sem_casas = NamedStyle(name='dados_percentual_sem_casas', number_format='0%')
    estilo_dados_percentual_sem_casas.border = borda_fina
    estilo_dados_percentual_sem_casas.font = Font(name='Calibri', size=11, color='000000')
    estilo_dados_percentual_sem_casas.alignment = alinhamento_central

    estilo_dados_percentual_uma_casa = NamedStyle(name='dados_percentual_uma_casa', number_format='0.0%')
    estilo_dados_percentual_uma_casa.border = borda_fina
    estilo_dados_percentual_uma_casa.font = Font(name='Calibri', size=11, color='000000')
    estilo_dados_percentual_uma_casa.alignment = alinhamento_central

    estilo_dados_percentual_duas_casas = NamedStyle(name='dados_percentual_duas_casas', number_format='0.00%')
    estilo_dados_percentual_duas_casas.border = borda_fina
    estilo_dados_percentual_duas_casas.font = Font(name='Calibri', size=11, color='000000')
    estilo_dados_percentual_duas_casas.alignment = alinhamento_central
    
    estilo_dados_valor = NamedStyle(name='dados_valor', number_format='#,##0')
    estilo_dados_valor.border = borda_fina
    estilo_dados_valor.font = Font(name='Calibri', size=11, color='000000')
    estilo_dados_valor.alignment = alinhamento_central

    estilo_dados_litros = NamedStyle(name='dados_litros', number_format='#,##0')
    estilo_dados_litros.border = borda_fina
    estilo_dados_litros.font = Font(name='Calibri', size=11, color='000000')
    estilo_dados_litros.alignment = alinhamento_central

    estilo_soma_litros = NamedStyle(name='soma_litros', number_format='#,##0')
    estilo_soma_litros.border = borda_fina
    estilo_soma_litros.font = Font(name='Calibri', size=11, bold=True, color='000000')
    estilo_soma_litros.alignment = alinhamento_central
    
    estilo_soma_label = NamedStyle(name='soma_label')
    estilo_soma_label.border = borda_fina
    estilo_soma_label.font = Font(name='Calibri', size=11, bold=True, color='000000')
    estilo_soma_label.alignment = alinhamento_central

    estilo_dados_companhia = NamedStyle(name='dados_companhia', number_format='#,##0')
    estilo_dados_companhia.border = borda_fina
    estilo_dados_companhia.font = Font(name='Calibri', size=11, color='000000')
    estilo_dados_companhia.alignment = alinhamento_central

    estilo_separador = NamedStyle(name='separador')
    estilo_separador.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

    estilo_total_geral_litros = NamedStyle(name='total_geral_litros')
    estilo_total_geral_litros.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    estilo_total_geral_litros.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    estilo_total_geral_litros.border = borda_total
    estilo_total_geral_litros.number_format = '#,##0'
    estilo_total_geral_litros.alignment = alinhamento_central

    estilo_total_geral_companhia = NamedStyle(name='total_geral_companhia')
    estilo_total_geral_companhia.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    estilo_total_geral_companhia.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    estilo_total_geral_companhia.border = borda_total
    estilo_total_geral_companhia.number_format = '#,##0'
    estilo_total_geral_companhia.alignment = alinhamento_central
    
    estilo_soma_percentual_uma_casa = NamedStyle(name='soma_percentual_uma_casa', number_format='0.0%')
    estilo_soma_percentual_uma_casa.border = borda_fina
    estilo_soma_percentual_uma_casa.font = Font(name='Calibri', size=11, bold=True, color='000000')
    estilo_soma_percentual_uma_casa.alignment = alinhamento_central

    estilo_soma_percentual_duas_casas = NamedStyle(name='soma_percentual_duas_casas', number_format='0.00%')
    estilo_soma_percentual_duas_casas.border = borda_fina
    estilo_soma_percentual_duas_casas.font = Font(name='Calibri', size=11, bold=True, color='000000')
    estilo_soma_percentual_duas_casas.alignment = alinhamento_central

    estilo_total_geral_vazio = NamedStyle(name='total_geral_vazio')
    estilo_total_geral_vazio.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    estilo_total_geral_vazio.border = borda_total
    estilo_total_geral_vazio.alignment = alinhamento_central
    
    sheet.merge_cells('A1:O1')
    titulo_cell = sheet['A1']
    titulo_cell.value = 'RATEAMENTO DAS GARAGENS'
    titulo_cell.style = estilo_titulo
    sheet.row_dimensions[1].height = 40

    headers_principais = {
        'A2': ('DISTRIBUIÇÃO DAS EMPRESAS', 'A2:E2'), 
        'G2': ('SOLICITADO', None), 
        'H2': ('VALORES SOLICITADOS', 'H2:J2'),
        'L2': ('VERIFICAÇÃO DA EMPRESA', 'L2:O2')
    }
    for cell_ref, (text, merge_range) in headers_principais.items():
        if merge_range: sheet.merge_cells(merge_range)
        cell = sheet[cell_ref]
        cell.value = text
        cell.style = estilo_cabecalho_principal
    
    sheet['F2'].style = estilo_cabecalho_principal 
    sheet['K2'].style = estilo_cabecalho_principal

    sub_headers = ['EMPRESA', '% PARTICIPAÇÃO', 'CNPJ', 'VALOR CALCULADO', 'VALORES', '', 'TOTAL DOS LITROS', 'COMPANHIA', '% DO TOTAL', 'POSTOS', '', 'VALOR CALCULADO', 'VALOR %', 'VALORES REAIS', 'VALORES']
    for col_idx, text in enumerate(sub_headers, 1):
        cell = sheet.cell(row=3, column=col_idx)
        cell.value = text
        cell.style = estilo_sub_cabecalho

    current_row = 4
    soma_total_geral_refs = {'G': [], 'H': []}

    for empresa_idx, empresa in enumerate(empresas):
        start_row_empresa = current_row
        
        dados_da_empresa_flat = dados_empresas.get(empresa, [])
        if not dados_da_empresa_flat:
            continue

        garagens = []
        current_garage_data = None
        for item in dados_da_empresa_flat:
            total_litros, companhia_litros, posto, cnpj, _ = item
            if total_litros is not None:
                if current_garage_data:
                    garagens.append(current_garage_data)
                current_garage_data = {'total': total_litros, 'cnpj': cnpj, 'suppliers': []}
            if current_garage_data and companhia_litros is not None:
                current_garage_data['suppliers'].append({'litros': companhia_litros, 'posto': posto})
        if current_garage_data:
            garagens.append(current_garage_data)
        
        num_total_data_rows = sum(len(g['suppliers']) for g in garagens)
        
        if num_total_data_rows == 0:
            continue

        linhas_por_empresa = 0
        if empresa in ['BOA', 'CAX', 'MOBI']:
             linhas_por_empresa = num_total_data_rows + (len(garagens) * 2) + 1
        elif empresa == 'CNO':
             linhas_por_empresa = num_total_data_rows + (len(garagens) * 2) + 1
        else:
             linhas_por_empresa = num_total_data_rows + 2

        sum_row_empresa = start_row_empresa + linhas_por_empresa - 1
        
        sheet.merge_cells(f'A{start_row_empresa}:A{start_row_empresa + linhas_por_empresa - 2}')
        empresa_cell = sheet.cell(row=start_row_empresa, column=1, value=empresa)
        empresa_cell.style = estilo_nome_empresa
        
        sheet.merge_cells(f'A{sum_row_empresa}:A{sum_row_empresa}')
        if cotas_consideradas:
            codigo_empresa_map = empresa if empresa != 'MOBI' else 'MOB'
            valor_cota = cotas_consideradas.get(codigo_empresa_map, 0)
            valor_cell = sheet.cell(row=sum_row_empresa, column=1, value=valor_cota)
            valor_cell.style = estilo_cota_empresa
        cota_cell_ref = f'$A${sum_row_empresa}'

        temp_current_row = start_row_empresa
        garage_info_rows = {}
        all_data_rows_for_company = []

        for garage_idx, garage in enumerate(garagens):
            garage_start_row = temp_current_row
            num_suppliers = len(garage['suppliers'])
            
            for i, supplier in enumerate(garage['suppliers']):
                all_data_rows_for_company.append(temp_current_row)
                sheet.cell(row=temp_current_row, column=3, value=garage['cnpj'])
                sheet.cell(row=temp_current_row, column=8, value=supplier['litros'])
                sheet.cell(row=temp_current_row, column=10, value=supplier['posto'])
                
                if i == 0:
                    sheet.cell(row=temp_current_row, column=7, value=garage['total'])
                
                temp_current_row += 1
            
            percent_litros_row = temp_current_row
            vc_rounded_row = temp_current_row + 1

            garage_info_rows[garage_idx] = {
                'data_rows': list(range(garage_start_row, temp_current_row)),
                'total_litros_cell': f'G{garage_start_row}',
                'percent_litros_cell': f'G{percent_litros_row}',
                'vc_unrounded_cell': f'L{garage_start_row}',
                'vc_rounded_cell': f'L{vc_rounded_row}',
            }
            temp_current_row += 2 

        total_litros_empresa_ref = f'G{sum_row_empresa}'
        all_garage_total_refs = [info['total_litros_cell'] for info in garage_info_rows.values()]
        sheet[total_litros_empresa_ref].value = f"=SUM({','.join(all_garage_total_refs)})" if all_garage_total_refs else 0

        for info in garage_info_rows.values():
            sheet[info['percent_litros_cell']].value = f"=IF({total_litros_empresa_ref}<>0, {info['total_litros_cell']}/{total_litros_empresa_ref}, 0)"
            sheet[info['vc_unrounded_cell']].value = f"={info['percent_litros_cell']}*{cota_cell_ref}"
            sheet[info['vc_rounded_cell']].value = f"=ROUND({info['vc_unrounded_cell'].replace('$','')}/5000,0)*5000"
            for row in info['data_rows']:
                sheet.cell(row=row, column=13).value = f"=IF({info['total_litros_cell']}<>0, H{row}/{info['total_litros_cell']}, 0)"
                sheet.cell(row=row, column=14).value = f"=M{row}*{info['vc_rounded_cell']}"
                sheet.cell(row=row, column=15).value = f"=ROUND(N{row}/5000,0)*5000"

        for row in all_data_rows_for_company:
            sheet.cell(row=row, column=9).value = f"=IF({total_litros_empresa_ref}<>0, H{row}/{total_litros_empresa_ref}, 0)"
            sheet.cell(row=row, column=2).value = f'=I{row}'
            sheet.cell(row=row, column=4).value = f'=B{row}*{cota_cell_ref}'
            if empresa in ['EME', 'GLO', 'SJT', 'VML']:
                sheet.cell(row=row, column=5).value = f"=IF(MOD(D{row},10000)>7500,D{row}-MOD(D{row},10000)+10000,IF(MOD(D{row},10000)>2500,D{row}-MOD(D{row},10000)+5000,D{row}-MOD(D{row},10000)))"
            else:
                sheet.cell(row=row, column=5).value = f'=ROUND(D{row}/5000,0)*5000'
        
        for row in range(start_row_empresa, sum_row_empresa + 1):
            for col in range(2, len(sub_headers) + 2):
                cell = sheet.cell(row=row, column=col)
                is_data_row = row in all_data_rows_for_company
                
                if col == 2: cell.style = estilo_dados_percentual_sem_casas
                elif col in [3, 6, 10, 11]: cell.style = estilo_dados_padrao
                elif col in [4, 5, 12, 14, 15]: cell.style = estilo_dados_valor
                elif col == 7:
                    is_percent_row = any(f'G{row}' == info['percent_litros_cell'] for info in garage_info_rows.values())
                    if is_percent_row: cell.style = estilo_dados_percentual_duas_casas
                    else: cell.style = estilo_dados_litros
                elif col == 8: cell.style = estilo_dados_companhia
                elif col == 9: cell.style = estilo_dados_percentual_uma_casa
                elif col == 13: cell.style = estilo_dados_percentual_duas_casas

        sheet.cell(row=sum_row_empresa, column=5).value = f'=SUM(E{start_row_empresa}:E{sum_row_empresa-1})'
        sheet.cell(row=sum_row_empresa, column=5).style = estilo_soma_litros
        sheet.cell(row=sum_row_empresa, column=7).style = estilo_soma_litros
        sheet.cell(row=sum_row_empresa, column=9).value = f'=SUM(I{start_row_empresa}:I{sum_row_empresa-1})'
        sheet.cell(row=sum_row_empresa, column=9).style = estilo_soma_percentual_uma_casa
        sheet.cell(row=sum_row_empresa, column=13).value = f'=SUM(M{start_row_empresa}:M{sum_row_empresa-1})'
        sheet.cell(row=sum_row_empresa, column=13).style = estilo_soma_percentual_duas_casas
        check_cell = sheet.cell(row=sum_row_empresa, column=15)
        check_cell.value = f'=IF(SUM(O{start_row_empresa}:O{sum_row_empresa-1})-A{sum_row_empresa}=0, "Ok", "Erro")'
        check_cell.style = estilo_soma_label

        soma_total_geral_refs['G'].append(f'G{sum_row_empresa}')
        soma_total_geral_refs['H'].append(f'H{sum_row_empresa}')
        
        current_row = sum_row_empresa + 1
        if empresa_idx < len(empresas) - 1:
            for col in range(1, len(sub_headers) + 1):
                sheet.cell(row=current_row, column=col).style = estilo_separador
            current_row += 1

    total_geral_row = current_row
    sheet.cell(row=total_geral_row, column=4, value='TOTAL GERAL').style = estilo_total_geral_vazio
    sheet.cell(row=total_geral_row, column=4).font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    sheet.cell(row=total_geral_row, column=4).alignment = alinhamento_central

    sheet[f'G{total_geral_row}'].value = f"=SUM({','.join(soma_total_geral_refs['G'])})" if soma_total_geral_refs['G'] else 0
    sheet[f'G{total_geral_row}'].style = estilo_total_geral_litros
    
    sheet[f'H{total_geral_row}'].value = f"=SUM({','.join(soma_total_geral_refs['H'])})" if soma_total_geral_refs['H'] else 0
    sheet[f'H{total_geral_row}'].style = estilo_total_geral_companhia
    
    for col in range(1, len(sub_headers) + 1):
        if col not in [4, 7, 8]:
            sheet.cell(row=total_geral_row, column=col).style = estilo_total_geral_vazio

    for col_idx in range(1, len(sub_headers) + 1):
        letter = get_column_letter(col_idx)
        if letter not in ['F', 'K']: 
            sheet.column_dimensions[letter].width = 18
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['F'].width = 8
    sheet.column_dimensions['K'].width = 4
    sheet.freeze_panes = 'A4'

def _criar_aba_sefaz(wb, mes_ref_str, cotas_por_empresa, dados_rateamento):
    try:
        if "SEFAZ" in wb.sheetnames:
            wb.remove(wb["SEFAZ"])
        ws = wb.create_sheet("SEFAZ")

        font_titulo = Font(name='Calibri', size=11, bold=True)
        font_header_tabela = Font(name='Calibri', size=9, bold=True)
        alinhamento_central = Alignment(horizontal='center', vertical='center', wrap_text=True)
        borda_completa = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        font_total_label = Font(name='Calibri', size=10, bold=True)
        font_empresa_principal = Font(name='Calibri', size=8, bold=True)
        font_empresa_sub = Font(name='Calibri', size=8)
        font_dados = Font(name='Calibri', size=9)
        
        cnpj_map = {
            'BOA': { '1-80 BV': '10.882.777/0001-80', '3-42 CD': '10.882.777/0003-42' },
            'CAX': { '1-83 OL': '41.037.250/0001-83', '3-45 OL': '41.037.250/0003-45' },
            'CNO': { '1-39 OL': '70.227.608/0001-39', '1-66 AL': '10.687.226/0001-66', '1-40 OL': '12.790.622/0001-40' },
            'CSR': { '1-09 RE': '36.106.678/0001-09' },
            'EME': { '1-97 RE': '10.407.005/0001-97' },
            'GLO': { '2-00 RE': '12.601.233/0002-00' },
            'MOB': { '1-29 SLM': '18.938.887/0001-29', '2-00 RE': '18.938.887/0002-00' },
            'SJT': { '1-66 CSA': '09.929.134/0001-66' },
            'VML': { '1-00 RE': '08.107.369/0001-00' }
        }
        
        cno_sub_name_map = {
            '1-39 OL': 'CDA - Cidade Alta Transportes e Turismo Ltda',
            '1-66 AL': 'ITA - Transportadora Itamaracá Ltda',
            '1-40 OL': 'ROD - Rodotur Turismo Ltda'
        }
        distribuidora_map = {
            "DISLUB": "Dislub Combustíveis S/A",
            "VIBRA": "VIBRA Energia S/A",
            "IPIRANGA": "Ipiranga Produtos de Petróleo S/A",
            "RAIZEN": "Raízen Combustivéis S/A"
        }

        titulo_texto = (f"QUANTIDADE MÁXIMA DE ÓLEO DIESEL A SER ADQUIRIDO POR CRÉDITO PRESUMIDO DO ICMS NOS TERMOS DO CONVÊNIO 21/2023 "
                        f"- {mes_ref_str.upper()} DAS EMPRESAS PERMISSIONÁRIAS e CONCESSIONÁRIAS")
        ws.merge_cells('A1:G1')
        cell_titulo = ws['A1']
        cell_titulo.value = titulo_texto
        cell_titulo.font = font_titulo
        cell_titulo.alignment = alinhamento_central
        ws.row_dimensions[1].height = 45
        ws.row_dimensions[2].height = 40

        ws.merge_cells('A2:B2')
        ws.cell(row=2, column=1).value = 'EMPRESA'
        ws.cell(row=2, column=3).value = 'INSCRIÇÃO ESTADUAL\n(garagem)'
        ws.cell(row=2, column=4).value = 'CNPJ (garagem)'
        ws.merge_cells('E2:F2')
        ws.cell(row=2, column=5).value = 'COTA MENSAL DE ÓLEO DIESEL\n(EM LITROS)'
        ws.cell(row=2, column=7).value = 'DISTRIBUIDORA DE COMBUSTÍVEL'

        for col_idx in range(1, 8):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = font_header_tabela
            cell.alignment = alinhamento_central
            cell.border = borda_completa

        empresas_ordem = ['BOA', 'CAX', 'CNO', 'CSR', 'EME', 'GLO', 'MOB', 'SJT', 'VML']
        nomes_sefaz = {
            "BOA": "BOA - Borborema Imperial Transportes Ltda", "CAX": "CAX - Caxangá Empresa de Transporte Coletivo Ltda",
            "CNO": "CNO – CONSÓRCIO CONORTE", "CSR": "CSR – Consórcio Recife",
            "EME": "EME - Metropolitana Empresa de Transporte Coletivo Ltda", "GLO": "GLO - Transportadora Globo Ltda",
            "MOB": "MOB – Mobibrasil Expresso S.A.", "SJT": "SJT - José Faustino e Companhia Ltda",
            "VML": "VML – Viação Mirim Ltda"
        }

        current_row = 3
        for empresa_cod in empresas_ordem:
            rateamento_key = 'MOBI' if empresa_cod == 'MOB' else empresa_cod
            company_data = dados_rateamento.get(rateamento_key, [])
            
            cota_total_empresa = cotas_por_empresa.get(empresa_cod, 0)
            
            ### CORREÇÃO 1: Calcular o total da empresa com base nos litros da GARAGEM (d[0]) ###
            total_litros_solicitados_empresa = sum(d[0] for d in company_data if d[0] is not None)

            num_data_rows = len(company_data) if company_data else 1
            start_row = current_row
            data_end_row = start_row + num_data_rows - 1
            total_row_idx = start_row + num_data_rows

            if empresa_cod == 'CNO':
                ws.merge_cells(start_row=start_row, start_column=1, end_row=data_end_row, end_column=1)
                cell_empresa = ws.cell(row=start_row, column=1, value=nomes_sefaz.get(empresa_cod, empresa_cod))
            else:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=data_end_row, end_column=2)
                cell_empresa = ws.cell(row=start_row, column=1, value=nomes_sefaz.get(empresa_cod, empresa_cod))

            cell_empresa.font = font_empresa_principal
            cell_empresa.alignment = alinhamento_central

            grouped_by_ie = {}
            if company_data:
                for data_tuple in company_data:
                    inscricao = data_tuple[4]
                    if inscricao not in grouped_by_ie:
                        grouped_by_ie[inscricao] = []
                    grouped_by_ie[inscricao].append(data_tuple)
            
            temp_row = start_row
            for inscricao, rows_for_this_ie in grouped_by_ie.items():
                if not rows_for_this_ie: continue

                ie_start_row = temp_row
                ie_end_row = ie_start_row + len(rows_for_this_ie) - 1

                if ie_start_row <= ie_end_row:
                    ws.merge_cells(start_row=ie_start_row, start_column=3, end_row=ie_end_row, end_column=3)
                
                cell_insc = ws.cell(row=ie_start_row, column=3, value=inscricao)
                cell_insc.font = font_dados
                cell_insc.alignment = alinhamento_central

                garages_within_ie = []
                if rows_for_this_ie:
                    current_garage_rows = []
                    for data_tuple in rows_for_this_ie:
                        if data_tuple[0] is not None and current_garage_rows:
                            garages_within_ie.append(current_garage_rows)
                            current_garage_rows = []
                        current_garage_rows.append(data_tuple)
                    if current_garage_rows:
                        garages_within_ie.append(current_garage_rows)

                for garage_rows in garages_within_ie:
                    garage_start_row = temp_row
                    garage_end_row = garage_start_row + len(garage_rows) - 1

                    _, _, _, cnpj_short_code, _ = garage_rows[0]
                    lookup_key = str(cnpj_short_code).strip().upper() if cnpj_short_code else ""
                    full_cnpj = cnpj_map.get(empresa_cod, {}).get(lookup_key, cnpj_short_code)

                    if garage_start_row <= garage_end_row:
                        ws.merge_cells(start_row=garage_start_row, start_column=4, end_row=garage_end_row, end_column=4)
                    
                    cell_cnpj = ws.cell(row=garage_start_row, column=4, value=full_cnpj)
                    cell_cnpj.font = font_dados
                    cell_cnpj.alignment = alinhamento_central
                    
                    if empresa_cod == 'CNO':
                        sub_name = cno_sub_name_map.get(lookup_key, "Sub-empresa Desconhecida")
                        ws.merge_cells(start_row=garage_start_row, start_column=2, end_row=garage_end_row, end_column=2)
                        cell_sub_empresa = ws.cell(row=garage_start_row, column=2, value=sub_name)
                        cell_sub_empresa.font = font_empresa_sub
                        cell_sub_empresa.alignment = alinhamento_central

                    ### CORREÇÃO 2: Implementar a lógica de cálculo da Coluna O (Rateamento) ###
                    # 1. Obter Litros_Garagem (d[0] do primeiro item da lista garage_rows)
                    litros_garagem = 0
                    if garage_rows:
                        litros_garagem_val = garage_rows[0][0] 
                        if litros_garagem_val is not None:
                            litros_garagem = litros_garagem_val

                    # 2. Calcular Cota_Garagem_Rounded (replicando a lógica da Coluna L da aba Rateamento)
                    cota_garagem_rounded = 0
                    if total_litros_solicitados_empresa > 0 and litros_garagem > 0:
                        proporcao_garagem = litros_garagem / total_litros_solicitados_empresa
                        cota_garagem_unrounded = proporcao_garagem * cota_total_empresa
                        # Lógica de arredondamento da Coluna L: =ROUND(L{unrounded}/5000,0)*5000
                        cota_garagem_rounded = round(cota_garagem_unrounded / 5000) * 5000

                    for j, data_tuple in enumerate(garage_rows):
                        row_idx = garage_start_row + j
                        _, litros_companhia, posto, _, _ = data_tuple
                        
                        # 3. Calcular Cota_Companhia_Rounded (replicando a lógica das Colunas M, N, O)
                        valor_calculado = 0
                        if litros_garagem > 0 and litros_companhia is not None:
                            # Lógica da Coluna M: =IF(G{garagem}<>0, H{companhia}/G{garagem}, 0)
                            proporcao_companhia = litros_companhia / litros_garagem
                            # Lógica da Coluna N: =M{proporcao}*L{cota_garagem_rounded}
                            cota_companhia_unrounded = proporcao_companhia * cota_garagem_rounded
                            # Lógica da Coluna O: =ROUND(N{unrounded}/5000,0)*5000
                            valor_calculado = round(cota_companhia_unrounded / 5000) * 5000

                        ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=6)
                        cota_cell = ws.cell(row=row_idx, column=5, value=valor_calculado)
                        cota_cell.number_format = '#,##0'
                        cota_cell.font = font_dados
                        cota_cell.alignment = alinhamento_central
                        
                        nome_distribuidora = distribuidora_map.get(str(posto).upper(), posto)
                        ws.cell(row=row_idx, column=7, value=nome_distribuidora).font = font_dados
                        ws.cell(row_idx, column=7).alignment = alinhamento_central

                    temp_row = garage_end_row + 1

            ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=4)
            cell_total_label = ws.cell(row=total_row_idx, column=1)
            cell_total_label.value = f"Total {empresa_cod}"
            cell_total_label.font = font_total_label
            cell_total_label.alignment = Alignment(horizontal='right', vertical='center')

            ws.merge_cells(start_row=total_row_idx, start_column=5, end_row=total_row_idx, end_column=6)
            cell_total_value = ws.cell(row=total_row_idx, column=5)
            cell_total_value.value = cota_total_empresa
            cell_total_value.font = font_total_label
            cell_total_value.number_format = '#,##0'
            cell_total_value.alignment = alinhamento_central

            for r in range(start_row, total_row_idx + 1):
                for c in range(1, 8):
                    ws.cell(row=r, column=c).border = borda_completa
            
            current_row = total_row_idx + 1
        
        total_geral_row = current_row
        total_stpp_rmr = sum(v for v in cotas_por_empresa.values()) if cotas_por_empresa else 0
        
        ws.merge_cells(f'A{total_geral_row}:D{total_geral_row}')
        label_total_geral = ws.cell(row=total_geral_row, column=1, value="TOTAL STPP/RMR")
        label_total_geral.font = font_total_label
        label_total_geral.alignment = alinhamento_central
        
        ws.merge_cells(f'E{total_geral_row}:G{total_geral_row}')
        valor_total_geral = ws.cell(row=total_geral_row, column=5, value=total_stpp_rmr)
        valor_total_geral.font = font_total_label
        valor_total_geral.number_format = '#,##0'
        valor_total_geral.alignment = alinhamento_central

        for c in range(1, 8):
            ws.cell(row=total_geral_row, column=c).border = borda_completa
        
        ws.column_dimensions['A'].width = 17
        ws.column_dimensions['B'].width = 19.43
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 16.57
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 13.14

    except Exception as e:
        import traceback
        print(f"AVISO: Não foi possível criar ou preencher a aba SEFAZ. Erro: {e}\n{traceback.format_exc()}")


def executar_processamento(gui_instance):
    try:
        wb_destino = openpyxl.Workbook()
        ws_destino = wb_destino.active
        ws_destino.title = "Km Prog"

        df_atual = extrair_dados_para_dataframe(gui_instance.arquivos_atual)
        mapa_passado = None
        if gui_instance.arquivos_passado:
            df_passado = extrair_dados_para_dataframe(gui_instance.arquivos_passado)
            if not df_passado.empty:
                df_passado['chave'] = df_passado.apply(lambda row: f"{str(row['Linha']).strip()}|{str(row['Empresa']).strip()}|{str(row['Dia Tipo']).strip()}", axis=1)
                mapa_passado = df_passado.set_index('chave')

        start_col_passado = 8
        col_offset_referencia, max_cols_passado, max_cols_atual = 1, 0, 0
        if gui_instance.arquivos_passado:
            try:
                import xlrd
                temp_book = xlrd.open_workbook(gui_instance.arquivos_passado[0])
                max_cols_passado = max((temp_book.sheet_by_index(0).row_len(r) for r in range(temp_book.sheet_by_index(0).nrows)), default=15)
                col_offset_referencia = start_col_passado + max_cols_passado + 2
            except: col_offset_referencia = start_col_passado + 15 + 2
        else: col_offset_referencia = start_col_passado

        blocos_info_atual = []
        offset_linha = 1
        tipos_de_dia = sorted(list(set([obter_chave_ordenacao(f) for f in gui_instance.arquivos_passado + gui_instance.arquivos_atual])))

        for tipo in tipos_de_dia:
            arq_p = next((f for f in sorted(gui_instance.arquivos_passado, key=obter_chave_ordenacao) if obter_chave_ordenacao(f) == tipo), None)
            arq_a = next((f for f in sorted(gui_instance.arquivos_atual, key=obter_chave_ordenacao) if obter_chave_ordenacao(f) == tipo), None)
            
            linhas_p, linhas_a = 0, 0
            if arq_p: linhas_p, _ = copiar_dados_planilha(arq_p, ws_destino, offset_linha, start_col_passado)
            if arq_a:
                linhas_a, cols_a = copiar_dados_planilha(arq_a, ws_destino, offset_linha, col_offset_referencia)
                blocos_info_atual.append({'inicio': offset_linha + 1, 'fim': offset_linha + linhas_a, 'arquivo': arq_a})
                if cols_a > max_cols_atual: max_cols_atual = cols_a
            offset_linha += max(linhas_p, linhas_a) + 2

        indices_a = encontrar_indices_colunas(ws_destino, 2, col_offset_referencia + 1, col_offset_referencia + max_cols_atual)
        
        col_diferenca = col_offset_referencia + max_cols_atual + 1
        col_conferencia = col_offset_referencia + max_cols_atual + 2
        
        header_row = 2 if 'linha_header_real' not in indices_a else indices_a['linha_header_real']
        ws_destino.cell(row=header_row, column=col_diferenca, value="Diferença")
        ws_destino.cell(row=header_row, column=col_conferencia, value="Status")

        linhas_novas = []
        if 'linha_header_real' in indices_a and mapa_passado is not None:
            for r in range(indices_a['linha_header_real'] + 1, ws_destino.max_row + 1):
                try:
                    linha_a_val = ws_destino.cell(r, indices_a['Linha']).value
                    if not linha_a_val or 'total' in str(linha_a_val).lower(): continue
                    
                    empresa_a_val = ws_destino.cell(r, indices_a['Empresa']).value
                    dia_tipo_a_val = ws_destino.cell(r, indices_a['Dia Tipo']).value
                    km_a_val = float(ws_destino.cell(r, indices_a['Km Total']).value)
                    
                    try: linha_a_str = str(int(float(linha_a_val)))
                    except (ValueError, TypeError): linha_a_str = str(linha_a_val).strip()
                    
                    chave_atual = f"{linha_a_str}|{str(empresa_a_val).strip()}|{str(dia_tipo_a_val).strip()}"
                    
                    km_p_val = 0
                    status = ""

                    if chave_atual in mapa_passado.index:
                        km_p_val = mapa_passado.loc[chave_atual, 'Km Total']
                        diferenca = km_a_val - km_p_val
                        if math.isclose(diferenca, 0):
                            status = "Ok"
                        else:
                            status = "Diferença"
                        mapa_passado.drop(chave_atual, inplace=True)
                    else:
                        diferenca = km_a_val
                        status = "Linha Nova"
                        linhas_novas.append({'Status': status, 'Linha': linha_a_str, 'Empresa': empresa_a_val, 'Dia Tipo': dia_tipo_a_val, 'Km Total': km_a_val})
                    
                    ws_destino.cell(r, col_diferenca, value=diferenca).number_format = '#,##0.00'
                    ws_destino.cell(r, col_conferencia, value=status)

                except (ValueError, TypeError, KeyError): continue
        
        titulo_font, titulo_fill, titulo_align = Font(color="FF0000", bold=True, size=14), PatternFill(start_color="FFFF00", fill_type="solid"), Alignment(horizontal='center', vertical='center')
        
        # --- LÓGICA DE TÍTULO E DATA MODIFICADA ---
        mes_passado_str = "N/A"
        if gui_instance.arquivos_passado:
            mes_passado_str = encontrar_mes_no_arquivo(gui_instance.arquivos_passado[0]) # Busca o mês
            ws_destino.merge_cells(start_row=1, start_column=start_col_passado, end_row=1, end_column=start_col_passado + max_cols_passado)
            cell_titulo_p = ws_destino.cell(row=1, column=start_col_passado, value=f"MÊS PASSADO: {mes_passado_str}")
            cell_titulo_p.font, cell_titulo_p.fill, cell_titulo_p.alignment = titulo_font, titulo_fill, titulo_align
        
        mes_atual_str = "MÊS ATUAL"
        if gui_instance.arquivos_atual:
            mes_atual_str = encontrar_mes_no_arquivo(gui_instance.arquivos_atual[0]) # Busca o mês
            
        if gui_instance.arquivos_atual:
            ws_destino.merge_cells(start_row=1, start_column=col_offset_referencia + 1, end_row=1, end_column=col_offset_referencia + max_cols_atual)
            cell_titulo_a = ws_destino.cell(row=1, column=col_offset_referencia + 1, value=f"MÊS DE REFERÊNCIA: {mes_atual_str}")
            cell_titulo_a.font, cell_titulo_a.fill, cell_titulo_a.alignment = titulo_font, titulo_fill, titulo_align
        # --- FIM DA MODIFICAÇÃO ---
        
        linhas_excluidas_df = pd.DataFrame()
        if mapa_passado is not None and not mapa_passado.empty:
            linhas_excluidas_df = mapa_passado.reset_index()
            linhas_excluidas_df['Status'] = 'Linha Excluída'

        linhas_novas_df = pd.DataFrame(linhas_novas)
        linhas_alteradas_df = pd.concat([linhas_excluidas_df, linhas_novas_df], ignore_index=True)
        
        if not linhas_alteradas_df.empty:
            linhas_alteradas_df['Linha'] = linhas_alteradas_df['Linha'].astype(str).str.strip()
            linhas_alteradas_df = linhas_alteradas_df[linhas_alteradas_df['Linha'] != '']
        
        dados_sumario = _aplicar_estilos_e_criar_sumario(ws_destino, blocos_info_atual, col_offset_referencia, max_cols_atual, linhas_alteradas_df, mes_passado_str, mes_atual_str)
        adicionar_marca_dagua(ws_destino, anexo_celula='W1')
        
        km_a_usar_por_empresa = {}
        for dia_tipo, dados_bloco in dados_sumario.items():
            for empresa, dados in dados_bloco.items():
                chave = (empresa, dia_tipo)
                km_a_usar_por_empresa[chave] = dados.get('soma_calculada', 0)

        rendimentos_por_empresa = {}
        info_pco_str = "N/A"
        if all([gui_instance._caminho_completo_cut2, gui_instance.mes_cut2.get(), gui_instance.quinzena_cut2.get()]):
            df_rendimento = processar_quinzena(gui_instance._caminho_completo_cut2, gui_instance.mes_cut2.get(), gui_instance.quinzena_cut2.get())
            if not df_rendimento.empty:
                
                # --- INÍCIO DA MODIFICAÇÃO (verificação de categorias) ---
                try:
                    disel_upper = df_rendimento["Disel"].astype(str).str.strip().str.upper()
                    mascara_total = disel_upper.str.contains("TOTAL")
                    mascara_idade_media = disel_upper.isin(["IDADE MÉDIA ATÉ VIDA ÚTIL", "IDADE MÉDIA TOTAL"])
                    categorias_no_arquivo = set(df_rendimento[~mascara_total & ~mascara_idade_media]['Disel'].astype(str).str.strip())
                    categorias_conhecidas = set(Config.RENDIMENTO_DISPEL.keys())
                    novas_categorias = categorias_no_arquivo - categorias_conhecidas
                    novas_categorias.discard('')
                    if novas_categorias:
                        mensagem = "AVISO: Novas categorias de 'Rendimento Dispel' foram encontradas no arquivo PCO e não estão na sua lista de configuração (Config.RENDIMENTO_DISPEL).\n\n"
                        mensagem += "Novas categorias encontradas:\n"
                        for categoria in sorted(list(novas_categorias)):
                            mensagem += f"- {categoria}\n"
                        mensagem += "\nElas serão processadas com rendimento 0. Considere atualizar o código-fonte."
                        gui_instance.after(0, lambda: messagebox.showwarning("Novas Categorias de Rendimento", mensagem, parent=gui_instance))
                except Exception as e:
                    print(f"AVISO: Falha ao verificar novas categorias de rendimento: {e}")
                # --- FIM DA MODIFICAÇÃO (verificação de categorias) ---

                rendimentos_por_empresa = adicionar_aba_rendimento_formatada(wb_destino, df_rendimento)
                
                # --- LÓGICA DA FONTE PCO MODIFICADA ---
                # Pega o ano manual. Se estiver vazio, usa o ano atual como padrão.
                ano_ref = gui_instance.ano_pco.get() 
                if not ano_ref.isdigit():
                    ano_ref = str(datetime.date.today().year)
                info_pco_str = f"PCO {gui_instance.quinzena_cut2.get()} {gui_instance.mes_cut2.get()}/{ano_ref}"
                # --- FIM DA MODIFICAÇÃO ---

        cota_cno_final = None
        if gui_instance._caminho_completo_cno:
            cota_cno_final = _adicionar_aba_calculo_consolidado(wb_destino, gui_instance._caminho_completo_cno, df_atual, gui_instance.contagem_dias_dict, "Cálculo CNO", "CNO")

        cota_mob_final = None
        if gui_instance._caminho_completo_mob:
            cota_mob_final = _adicionar_aba_calculo_consolidado(wb_destino, gui_instance._caminho_completo_mob, df_atual, gui_instance.contagem_dias_dict, "Cálculo MOB", "MOB")

        cotas_por_empresa = None
        if km_a_usar_por_empresa:
            cotas_por_empresa = _criar_aba_cota_mes(wb_destino, km_a_usar_por_empresa, rendimentos_por_empresa, gui_instance.contagem_dias_dict, mes_atual_str, info_pco_str, cota_cno_final, cota_mob_final, gui_instance.dados_rateamento, gui_instance.reducoes_processadas)
        
        if gui_instance.dados_rateamento:
            _adicionar_aba_rateamento(wb_destino, gui_instance.dados_rateamento, cotas_por_empresa)

        _criar_aba_sefaz(wb_destino, mes_atual_str, cotas_por_empresa, gui_instance.dados_rateamento)

        for sheet_name in wb_destino.sheetnames:
            if sheet_name != 'SEFAZ':
                auto_ajustar_colunas(wb_destino[sheet_name])
        
        if 'SEFAZ' in wb_destino.sheetnames:
            ws_sefaz = wb_destino['SEFAZ']
            ws_sefaz.column_dimensions['A'].width = 17
            ws_sefaz.column_dimensions['B'].width = 19.43
            ws_sefaz.column_dimensions['C'].width = 10
            ws_sefaz.column_dimensions['D'].width = 16.57
            ws_sefaz.column_dimensions['E'].width = 8
            ws_sefaz.column_dimensions['F'].width = 8
            ws_sefaz.column_dimensions['G'].width = 13.14
        
        aplicar_wrap_text_em_todas_abas(wb_destino)
        
        if "Cota do Mês" in wb_destino.sheetnames:
            wb_destino.active = wb_destino["Cota do Mês"]
        elif "Km Prog" in wb_destino.sheetnames:
            wb_destino.active = wb_destino["Km Prog"]

        wb_destino.save(gui_instance.output_path)
        gui_instance.finalizar_processamento_gui(gui_instance.output_path, sucesso=True)
    except Exception as e:
        gui_instance.finalizar_processamento_gui(f"Erro em executar_processamento: {e}\n{traceback.format_exc()}", sucesso=False)

class CotaDieselGUI(TkinterDnD.Tk):
    def __init__(self, title="Cota de Óleo Diesel", size=(1000, 850)):
        super().__init__()
        self.title(title)
        self.geometry(f'{size[0]}x{size[1]}')
        self.resizable(True, True)

        self.style = ttk.Style(self)
        self.style.configure('Accent.TButton', font=('Segoe UI', 10, 'bold'))
        self.style.configure('TLabel', font=('Segoe UI', 10))
        self.style.configure('TLabelframe.Label', font=('Segoe UI', 11, 'bold'))
        self.style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'))
        self.style.configure('Header.TLabel', font=('Segoe UI', 12, 'bold'))
        
        self.center_window()

        self.LOGO_IMAGE_PATH = "static/images/RCB Logo.png"
        self.BACKGROUND_IMAGE_PATH = "static/images/Cota de Óleo Diesel.png"
        self.arquivos_passado, self.arquivos_atual = [], []
        self._caminho_completo_cut2, self._caminho_completo_cno, self._caminho_completo_mob = "", "", ""
        self.mes_cut2, self.quinzena_cut2 = tk.StringVar(), tk.StringVar()
        self.ano_pco = tk.StringVar() # <-- ADICIONADO
        self.dias_dut, self.dias_sab, self.dias_dom = tk.StringVar(value='0'), tk.StringVar(value='0'), tk.StringVar(value='0')
        self.reducao_boa, self.reducao_cax, self.reducao_cno = tk.StringVar(value="0,00%"), tk.StringVar(value="0,00%"), tk.StringVar(value="0,00%")
        self.reducao_csr, self.reducao_eme, self.reducao_glo = tk.StringVar(value="0,00%"), tk.StringVar(value="0,00%"), tk.StringVar(value="0,00%")
        self.reducao_mob, self.reducao_sjt, self.reducao_vml = tk.StringVar(value="0,00%"), tk.StringVar(value="0,00%"), tk.StringVar(value="0,00%")
        self.reducoes_processadas = {}
        self.pasta_destino, self.nome_arquivo_saida = tk.StringVar(), tk.StringVar()
        self.dados_rateamento = {}
        self.rateamento_widgets = {}
        self.rateamento_empresa_frames = {}
        self.SESSAO_FILE = os.path.join('instance', 'ultima_sessao.json')
        
        self._setar_nome_padrao_arquivo()
        self.create_widgets()
        sv_ttk.set_theme("light")

    def toggle_theme(self):
        sv_ttk.toggle_theme()

    def toggle_size(self):
        if not hasattr(self, 'is_compact') or self.is_compact:
            self.geometry('1000x850')
            self.is_compact = False
        else:
            self.geometry('850x720')
            self.is_compact = True
        self.center_window()

    def create_widgets(self):
        try:
            bg_image_pil = Image.open(self.BACKGROUND_IMAGE_PATH)
            bg_image_pil = bg_image_pil.resize((1000, 850), Image.Resampling.LANCZOS)
            self.background_image = ImageTk.PhotoImage(bg_image_pil)
            background_label = tk.Label(self, image=self.background_image)
            background_label.place(x=0, y=0, relwidth=1, relheight=1)
        except Exception as e:
            print(f"AVISO: Não foi possível carregar a imagem de fundo: {e}.")

        main_frame = ttk.Frame(self, padding=(20, 15))
        main_frame.pack(fill=tk.BOTH, expand=True)

        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 20))

        try:
            logo_original = Image.open(self.LOGO_IMAGE_PATH).convert("RGBA")
            logo_resized = logo_original.resize((int(logo_original.width * 0.4), int(logo_original.height * 0.4)), Image.Resampling.LANCZOS)
            self.logo_image = ImageTk.PhotoImage(logo_resized)
            logo_label = ttk.Label(header_frame, image=self.logo_image)
            logo_label.pack(side=tk.LEFT, padx=(0, 15))
        except Exception:
            print(f"AVISO: Não foi possível carregar a imagem do logo.")
        
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(title_frame, text="Integrador de Cota de Óleo Diesel", style="Title.TLabel").pack(anchor='w')
        ttk.Label(title_frame, text="Ferramenta para consolidação e comparação de quilometragem", style="TLabel").pack(anchor='w')
        
        theme_switch = ttk.Checkbutton(header_frame, text="Tema", style="Switch.TCheckbutton", command=self.toggle_theme)
        theme_switch.pack(side=tk.RIGHT, padx=10)
        
        resize_button = ttk.Button(header_frame, text="Ajustar Janela", command=self.toggle_size)
        resize_button.pack(side=tk.RIGHT, padx=5)

        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=10)

        tab1 = ttk.Frame(notebook)
        tab2 = ttk.Frame(notebook)
        tab3 = ttk.Frame(notebook)
        tab4 = ttk.Frame(notebook)
        notebook.add(tab1, text='     Arquivos Principais     ')
        notebook.add(tab2, text='     Relatórios Adicionais     ')
        notebook.add(tab4, text='     Rateamento de Garagens     ')
        notebook.add(tab3, text='     Saída e Geração     ')
        

        self._criar_aba_arquivos(tab1)
        self._criar_aba_adicionais(tab2)
        self._criar_aba_rateamento(tab4)
        self._criar_aba_saida(tab3)
        
        button_area = ttk.Frame(main_frame, padding=(0, 10, 0, 0))
        button_area.pack(fill=tk.X)

        self.botao_carregar = ttk.Button(button_area, text="Carregar Última Sessão", command=self._carregar_sessao) 
        self.botao_carregar.pack(side=tk.LEFT, padx=(0, 10)) 

        self.botao_limpar = ttk.Button(button_area, text="Limpar Tudo", command=self.limpar_todas_selecoes)
        self.botao_limpar.pack(side=tk.LEFT, padx=(0, 10))
        
        self.botao_gerar = ttk.Button(button_area, text="Processar e Gerar Relatório", style='Accent.TButton', command=self.iniciar_processamento)
        self.botao_gerar.pack(side=tk.RIGHT, fill=tk.X, expand=True)

    def _criar_frame_rolavel(self, parent):
        outer_frame = ttk.Frame(parent)
        outer_frame.pack(fill=tk.BOTH, expand=True)
        
        canvas = tk.Canvas(outer_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas, padding=10)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        return scrollable_frame

    def _criar_aba_arquivos(self, parent):
        scrollable_frame = self._criar_frame_rolavel(parent)
        scrollable_frame.columnconfigure((0, 1), weight=1, uniform="group1")
        scrollable_frame.rowconfigure(0, weight=1)
        
        self.lista_passado = self._criar_painel_arquivos(scrollable_frame, "Arquivos do Mês Passado", self.arquivos_passado, 0)
        self.lista_atual = self._criar_painel_arquivos(scrollable_frame, "Arquivos do Mês de Referência", self.arquivos_atual, 1)

    def _criar_aba_adicionais(self, parent):
        scrollable_frame = self._criar_frame_rolavel(parent)
        scrollable_frame.columnconfigure(1, weight=1)

        frame_pco = ttk.LabelFrame(scrollable_frame, text=" Rendimento PCO (Opcional) ", padding=10)
        frame_pco.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 15))
        frame_pco.columnconfigure(1, weight=1)

        ttk.Button(frame_pco, text="Selecionar Arquivo PCO (.xls)", command=self.selecionar_cut2_file).grid(row=0, column=0, padx=(0, 10), sticky='w')
        self.label_cut2 = ttk.Label(frame_pco, text="Nenhum arquivo selecionado.", style='TLabel', anchor='w')
        self.label_cut2.grid(row=0, column=1, sticky='ew')
        
        pco_options_frame = ttk.Frame(frame_pco)
        pco_options_frame.grid(row=1, column=1, sticky='w', pady=(5,0))
        ttk.Label(pco_options_frame, text="Mês:").pack(side=tk.LEFT, padx=(0,5))
        ttk.Combobox(pco_options_frame, textvariable=self.mes_cut2, values=["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"], state="readonly", width=8).pack(side=tk.LEFT, padx=(0,15))
        ttk.Label(pco_options_frame, text="Quinzena:").pack(side=tk.LEFT, padx=(0,5))
        ttk.Combobox(pco_options_frame, textvariable=self.quinzena_cut2, values=["1ª", "2ª"], state="readonly", width=5).pack(side=tk.LEFT, padx=(0,15))
        
        # --- ADICIONADO ---
        ttk.Label(pco_options_frame, text="Ano:").pack(side=tk.LEFT, padx=(0,5))
        ttk.Entry(pco_options_frame, textvariable=self.ano_pco, width=6).pack(side=tk.LEFT)
        # --- FIM ---

        frame_calc = ttk.LabelFrame(scrollable_frame, text=" Cálculos de Cota (Opcional) ", padding=10)
        frame_calc.grid(row=1, column=0, columnspan=2, sticky="ew")
        frame_calc.columnconfigure(1, weight=1)

        ttk.Button(frame_calc, text="Cálculo CNO (.xls)", command=self.selecionar_cno_file).grid(row=0, column=0, padx=(0,10), pady=5, sticky='w')
        self.label_cno = ttk.Label(frame_calc, text="Nenhum arquivo selecionado.", anchor='w')
        self.label_cno.grid(row=0, column=1, sticky='ew', padx=5)

        ttk.Button(frame_calc, text="Cálculo MOB (.xls)", command=self.selecionar_mob_file).grid(row=1, column=0, padx=(0,10), pady=5, sticky='w')
        self.label_mob = ttk.Label(frame_calc, text="Nenhum arquivo selecionado.", anchor='w')
        self.label_mob.grid(row=1, column=1, sticky='ew', padx=5)
        
        dias_frame = ttk.Frame(frame_calc)
        dias_frame.grid(row=2, column=0, columnspan=2, sticky='w', pady=(10,0))
        ttk.Label(dias_frame, text="Contagem de dias no período:", style='Header.TLabel').pack(side=tk.LEFT, padx=(0,15))
        ttk.Label(dias_frame, text="DUT:").pack(side=tk.LEFT, padx=(0,5)); ttk.Entry(dias_frame, textvariable=self.dias_dut, width=4).pack(side=tk.LEFT, padx=(0,10))
        ttk.Label(dias_frame, text="SAB:").pack(side=tk.LEFT, padx=(0,5)); ttk.Entry(dias_frame, textvariable=self.dias_sab, width=4).pack(side=tk.LEFT, padx=(0,10))
        ttk.Label(dias_frame, text="DOM:").pack(side=tk.LEFT, padx=(0,5)); ttk.Entry(dias_frame, textvariable=self.dias_dom, width=4).pack(side=tk.LEFT)

        frame_reducoes = ttk.LabelFrame(scrollable_frame, text=" Redução / Acréscimo (Opcional - Ex: -5% ou 5%) ", padding=10)
        frame_reducoes.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(15, 0))
        
        companies = [
            ("BOA", self.reducao_boa), ("CAX", self.reducao_cax), ("CNO", self.reducao_cno),
            ("CSR", self.reducao_csr), ("EME", self.reducao_eme), ("GLO", self.reducao_glo),
            ("MOB", self.reducao_mob), ("SJT", self.reducao_sjt), ("VML", self.reducao_vml)
        ]
        
        for i, (label, var) in enumerate(companies):
            row = i // 3
            col = i % 3
            f = ttk.Frame(frame_reducoes)
            f.grid(row=row, column=col, padx=10, pady=5, sticky='w')
            ttk.Label(f, text=f"{label}:").pack(side=tk.LEFT, padx=(0,5))
            ttk.Entry(f, textvariable=var, width=8).pack(side=tk.LEFT)
        

    def _criar_aba_saida(self, parent):
        scrollable_frame = self._criar_frame_rolavel(parent)
        scrollable_frame.columnconfigure(0, weight=1)
        
        frame = ttk.LabelFrame(scrollable_frame, text=" Configuração do Arquivo de Saída ", padding=15)
        frame.pack(fill=tk.BOTH, expand=True)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text="Nome do Arquivo:").grid(row=0, column=0, padx=(0,10), pady=5, sticky='w')
        ttk.Entry(frame, textvariable=self.nome_arquivo_saida).grid(row=0, column=1, pady=5, sticky='ew')
        
        ttk.Button(frame, text="Selecionar Pasta de Destino", command=self.selecionar_pasta_destino).grid(row=1, column=0, padx=(0,10), pady=5, sticky='w')
        self.label_pasta = ttk.Label(frame, text="Nenhuma pasta selecionada.")
        self.label_pasta.grid(row=1, column=1, pady=5, sticky='ew')

    def _criar_aba_rateamento(self, parent):
        scrollable_frame = self._criar_frame_rolavel(parent)

        action_frame = ttk.Frame(scrollable_frame)
        action_frame.pack(fill=tk.X, expand=True, pady=5, padx=5)

        btn_carregar = ttk.Button(action_frame, text="Carregar de Arquivo...", command=self._carregar_dados_rateamento_de_arquivo)
        btn_carregar.pack(side=tk.LEFT, pady=5, padx=(0, 10))

        btn_salvar = ttk.Button(action_frame, text="Salvar Dados do Formulário", command=self._salvar_dados_rateamento, style='Accent.TButton')
        btn_salvar.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=5)
        
        empresas = ['BOA', 'CAX', 'CSR', 'CNO', 'EME', 'GLO', 'MOBI', 'SJT', 'VML']
        self.rateamento_widgets = {emp: {} for emp in empresas}
        self.rateamento_empresa_frames = {}

        for empresa in empresas:
            empresa_frame_container = ttk.LabelFrame(scrollable_frame, text=f" {empresa} ", padding=10)
            empresa_frame_container.pack(fill=tk.X, expand=True, pady=10, padx=5)
            self.rateamento_empresa_frames[empresa] = empresa_frame_container
            self._reconstruir_widgets_empresa(empresa)
            
        bottom_frame = ttk.Frame(scrollable_frame)
        bottom_frame.pack(fill=tk.X, expand=True, pady=20, padx=5)
        self.label_status_rateamento = ttk.Label(bottom_frame, text="Status: Nenhum dado coletado.", font=('Segoe UI', 10, 'italic'), foreground="gray", anchor="center")
        self.label_status_rateamento.pack(pady=(10, 10), fill=tk.X, expand=True)

    def _reconstruir_widgets_empresa(self, empresa, df_empresa=None):
        parent_frame = self.rateamento_empresa_frames[empresa]
        for widget in parent_frame.winfo_children():
            widget.destroy()
        
        self.rateamento_widgets[empresa] = {'garagens': {}}

        if df_empresa is None or df_empresa.empty:
            df_empresa = self._get_default_rateamento_data(empresa)
        
        garagens_ids = sorted(df_empresa['Garagem_ID'].unique())
        
        df_empresa = df_empresa.fillna('')

        for garagem_id in garagens_ids:
            df_garagem = df_empresa[df_empresa['Garagem_ID'] == garagem_id].reset_index(drop=True)
            if df_garagem.empty:
                continue

            garagem_data = df_garagem.iloc[0]
            
            bloco_frame = ttk.LabelFrame(parent_frame, text=f"Garagem {garagem_id}", padding=(10, 5))
            bloco_frame.pack(fill="x", expand=True, padx=5, pady=5)
            
            self.rateamento_widgets[empresa]['garagens'][garagem_id] = {'widgets': {}}
            
            tk.Label(bloco_frame, text="Total Litros:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
            entry_total = ttk.Entry(bloco_frame, width=20)
            entry_total.grid(row=0, column=1, padx=5, pady=2)
            if 'Litros_Garagem' in garagem_data and pd.notna(garagem_data['Litros_Garagem']) and garagem_data['Litros_Garagem'] != '':
                entry_total.insert(0, str(int(garagem_data['Litros_Garagem'])))
            self.rateamento_widgets[empresa]['garagens'][garagem_id]['widgets']['total'] = entry_total
            
            tk.Label(bloco_frame, text="CNPJ:").grid(row=0, column=2, sticky="w", padx=5, pady=2)
            entry_cnpj = ttk.Entry(bloco_frame, width=20)
            entry_cnpj.grid(row=0, column=3, padx=5, pady=2)
            if 'CNPJ' in garagem_data and pd.notna(garagem_data['CNPJ']):
                entry_cnpj.insert(0, garagem_data['CNPJ'])
            self.rateamento_widgets[empresa]['garagens'][garagem_id]['widgets']['cnpj'] = entry_cnpj

            tk.Label(bloco_frame, text="Inscrição Est.:").grid(row=0, column=4, sticky="w", padx=5, pady=2)
            entry_inscricao = ttk.Entry(bloco_frame, width=20)
            entry_inscricao.grid(row=0, column=5, padx=5, pady=2)
            if 'Inscrição_Estadual' in garagem_data and pd.notna(garagem_data['Inscrição_Estadual']):
                entry_inscricao.insert(0, garagem_data['Inscrição_Estadual'])
            self.rateamento_widgets[empresa]['garagens'][garagem_id]['widgets']['inscricao'] = entry_inscricao

            self.rateamento_widgets[empresa]['garagens'][garagem_id]['widgets']['companhias'] = []
            for i, row in df_garagem.iterrows():
                tk.Label(bloco_frame, text=f"Companhia {i+1}:").grid(row=i + 1, column=0, sticky="w", padx=5, pady=2)
                entry_comp = ttk.Entry(bloco_frame, width=20)
                entry_comp.grid(row=i + 1, column=1, padx=5, pady=2)
                if 'Litros_Companhia' in row and pd.notna(row['Litros_Companhia']) and row['Litros_Companhia'] != '':
                    entry_comp.insert(0, str(int(row['Litros_Companhia'])))

                tk.Label(bloco_frame, text=f"Posto {i+1}:").grid(row=i + 1, column=2, sticky="w", padx=5, pady=2)
                entry_posto = ttk.Entry(bloco_frame, width=20)
                entry_posto.grid(row=i + 1, column=3, padx=5, pady=2, columnspan=3)
                if 'Posto' in row and pd.notna(row['Posto']):
                    entry_posto.insert(0, row['Posto'])

                self.rateamento_widgets[empresa]['garagens'][garagem_id]['widgets']['companhias'].append((entry_comp, entry_posto))

    def _get_default_rateamento_data(self, empresa):
        default_data = {
            'BOA': [
                (1, None, '1-80 BV', None, 'DISLUB', ''), (1, None, '1-80 BV', None, 'VIBRA', ''),
                (2, None, '3-42 CD', None, 'IPIRANGA', ''), (2, None, '3-42 CD', None, 'VIBRA', '')
            ],
            'CAX': [
                (1, None, '1-83 OL', None, 'DISLUB', ''), (1, None, '1-83 OL', None, 'RAIZEN', ''),
                (2, None, '3-45 OL', None, 'DISLUB', ''), (2, None, '3-45 OL', None, 'RAIZEN', '')
            ],
            'MOBI': [
                (1, None, '3-27 VARZ', None, 'RAIZEN', ''), (1, None, '3-27 VARZ', None, 'DISLUB', ''),
                (2, None, '1-80 VARZ', None, 'RAIZEN', ''), (2, None, '1-80 VARZ', None, 'DISLUB', '')
            ],
            'CNO': [
                (1, None, '1-50 OL', None, 'RAIZEN', ''), (1, None, '1-50 OL', None, 'IPIRANGA', ''), (1, None, '1-50 OL', None, 'DISLUB', ''),
                (2, None, '1-50 OL', None, 'VIBRA', ''), (2, None, '1-50 OL', None, 'IPIRANGA', ''), (2, None, '1-50 OL', None, 'DISLUB', ''),
                (3, None, '1-50 OL', None, 'DISLUB', '')
            ],
            'CSR': [(1, None, '1-09 RC', None, 'VIBRA', ''), (1, None, '1-09 RC', None, 'DISLUB', ''), (1, None, '1-09 RC', None, 'IPIRANGA', '')],
            'EME': [(1, None, '1-97 RC', None, 'DISLUB', ''), (1, None, '1-97 RC', None, 'RAIZEN', '')],
            'GLO': [(1, None, '2-00 RC', None, 'IPIRANGA', '')],
            'SJT': [(1, None, '1-66 CSA', None, 'VIBRA', ''), (1, None, '1-66 CSA', None, 'DISLUB', ''), (1, None, '1-66 CSA', None, 'IPIRANGA', '')],
            'VML': [(1, None, '1-66 CSA', None, 'RAIZEN', '')]
        }
        data = default_data.get(empresa, [])
        df = pd.DataFrame(data, columns=['Garagem_ID', 'Litros_Garagem', 'CNPJ', 'Litros_Companhia', 'Posto', 'Inscrição_Estadual'])
        return df

    def _criar_painel_arquivos(self, parent, titulo, lista_arquivos, col):
        frame = ttk.LabelFrame(parent, text=f" {titulo} ", padding=10)
        frame.grid(row=0, column=col, padx=10, pady=5, sticky='nsew')
        frame.grid_rowconfigure(1, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        ttk.Label(frame, text="Arraste e solte arquivos .xls aqui ou use os botões abaixo.", justify=tk.CENTER).grid(row=0, column=0, columnspan=2, pady=(0, 10))
        
        tree = ttk.Treeview(frame, columns=('filename',), show='headings', height=6)
        tree.heading('filename', text='Arquivos Selecionados')
        tree.grid(row=1, column=0, columnspan=2, sticky='nsew')
        
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=1, column=2, sticky='ns')

        tree.drop_target_register(DND_FILES)
        tree.dnd_bind('<<Drop>>', lambda e: self._adicionar_arquivos(self.tk.splitlist(e.data), lista_arquivos, tree))
        
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=2, column=0, columnspan=2, sticky='ew', pady=(10,0))
        btn_frame.columnconfigure((0, 1), weight=1)

        btn_selecionar = ttk.Button(btn_frame, text="Selecionar", command=lambda: self._adicionar_arquivos(filedialog.askopenfilenames(title=f"Selecione - {titulo}", filetypes=[("Excel 97-2003", "*.xls")]), lista_arquivos, tree))
        btn_selecionar.grid(row=0, column=0, sticky='ew', padx=(0,5))
        btn_limpar = ttk.Button(btn_frame, text="Limpar", command=lambda: self._limpar_lista(lista_arquivos, tree))
        btn_limpar.grid(row=0, column=1, sticky='ew', padx=(5,0))
        return tree
    
    def _carregar_dados_rateamento_de_arquivo(self):
        caminho_arquivo = filedialog.askopenfilename(
            title="Selecione a planilha com os dados de rateamento",
            filetypes=[("Arquivos de Planilha", "*.xlsx *.xls *.ods"), ("Todos os arquivos", "*.*")]
        )
        if not caminho_arquivo:
            return

        try:
            if caminho_arquivo.lower().endswith('.ods'):
                df = pd.read_excel(caminho_arquivo, engine='odfpy')
            else:
                df = pd.read_excel(caminho_arquivo)

            colunas_necessarias = ['Empresa', 'Garagem_ID', 'Litros_Garagem', 'Litros_Companhia', 'Posto', 'CNPJ', 'Inscrição_Estadual']
            if not all(col in df.columns for col in colunas_necessarias):
                messagebox.showerror("Erro de Formato", f"O arquivo selecionado não contém todas as colunas necessárias.\n\nNecessárias: {', '.join(colunas_necessarias)}", parent=self)
                return

            empresas_no_arquivo = df['Empresa'].unique()
            for empresa in empresas_no_arquivo:
                if empresa in self.rateamento_empresa_frames:
                    df_empresa = df[df['Empresa'] == empresa].copy()
                    self._reconstruir_widgets_empresa(empresa, df_empresa)
            
            self.label_status_rateamento.config(text="Status: Dados carregados do arquivo com sucesso! Clique em 'Salvar' para confirmar.", foreground="blue")
            messagebox.showinfo("Sucesso", "Dados de rateamento carregados e preenchidos no formulário!", parent=self)

        except Exception as e:
            messagebox.showerror("Erro ao Ler Arquivo", f"Não foi possível processar o arquivo de rateamento.\n\nDetalhes: {e}\n{traceback.format_exc()}", parent=self)
            self.label_status_rateamento.config(text="Status: Falha ao carregar dados do arquivo.", foreground="red")

    def _parse_valor(self, valor_str):
        if valor_str is None or valor_str == '':
            return None
        try:
            return int(float(str(valor_str).replace('.', '')))
        except (ValueError, TypeError):
            messagebox.showerror("Erro de Entrada", f"O valor '{valor_str}' não é um número válido. Por favor, insira apenas números.")
            return "ERRO"

    def _salvar_dados_rateamento(self):
        dados_coletados = {}
        erro = False

        for empresa, garagens_data in self.rateamento_widgets.items():
            if erro: break
            dados_empresa_atual = []
            
            for garagem_id, garagem_widgets in garagens_data.get('garagens', {}).items():
                widgets = garagem_widgets['widgets']
                
                total_val = self._parse_valor(widgets['total'].get())
                cnpj_val = widgets['cnpj'].get()
                inscricao_val = widgets['inscricao'].get()

                if total_val == "ERRO":
                    erro = True
                    break

                primeira_companhia = True
                for comp_entry, posto_entry in widgets['companhias']:
                    comp_val = self._parse_valor(comp_entry.get())
                    posto_val = posto_entry.get()

                    if comp_val == "ERRO":
                        erro = True
                        break
                    
                    if primeira_companhia:
                        dados_empresa_atual.append((total_val, comp_val, posto_val, cnpj_val, inscricao_val))
                        primeira_companhia = False
                    else:
                        dados_empresa_atual.append((None, comp_val, posto_val, cnpj_val, inscricao_val))

                if erro: break
            
            dados_coletados[empresa] = dados_empresa_atual
        
        if not erro:
            self.dados_rateamento = dados_coletados
            self.label_status_rateamento.config(text="Status: Dados coletados com sucesso! Pronto para gerar.", foreground="green")
            if hasattr(self, 'botao_gerar') and self.botao_gerar['state'] == 'disabled':
                pass
            else:
                messagebox.showinfo("Sucesso", "Dados de rateamento salvos com sucesso!", parent=self)
        else:
            self.label_status_rateamento.config(text="Status: Erro na validação dos dados. Verifique os campos.", foreground="red")

    def _adicionar_arquivos(self, paths, lista_alvo, treeview):
        novos_arquivos = False
        for path in paths:
            if path.lower().endswith('.xls') and path not in lista_alvo:
                lista_alvo.append(path)
                novos_arquivos = True
        if novos_arquivos:
            lista_alvo.sort(key=obter_chave_ordenacao)
            self._atualizar_treeview(treeview, lista_alvo)

    def _limpar_lista(self, lista_alvo, treeview):
        lista_alvo.clear()
        self._atualizar_treeview(treeview, lista_alvo)

    def _atualizar_treeview(self, treeview, lista_arquivos):
        for i in treeview.get_children():
            treeview.delete(i)
        for path in lista_arquivos:
            treeview.insert("", tk.END, values=(os.path.basename(path),))

    def selecionar_arquivo_generico(self, title, label_widget, attr_caminho):
        caminho = filedialog.askopenfilename(title=title, filetypes=[("Arquivos Excel 97-2003", "*.xls")])
        if caminho:
            label_widget.config(text=os.path.basename(caminho))
            setattr(self, attr_caminho, caminho)
        else:
            label_widget.config(text="Nenhum arquivo selecionado.")
            setattr(self, attr_caminho, "")

    def selecionar_cut2_file(self): self.selecionar_arquivo_generico("Selecione o arquivo de Rendimento PCO", self.label_cut2, '_caminho_completo_cut2')
    def selecionar_cno_file(self): self.selecionar_arquivo_generico("Selecione o arquivo do Cálculo CNO", self.label_cno, '_caminho_completo_cno')
    def selecionar_mob_file(self): self.selecionar_arquivo_generico("Selecione o arquivo do Cálculo MOB", self.label_mob, '_caminho_completo_mob')
    
    def selecionar_pasta_destino(self):
        caminho = filedialog.askdirectory(title="Selecione a pasta para salvar o relatório")
        if caminho:
            self.pasta_destino.set(caminho)
            self.label_pasta.config(text=caminho)

    def limpar_todas_selecoes(self):
        self._limpar_lista(self.arquivos_passado, self.lista_passado)
        self._limpar_lista(self.arquivos_atual, self.lista_atual)
        self.mes_cut2.set(""); self.quinzena_cut2.set("")
        self.ano_pco.set("") # <-- ADICIONADO
        self.dias_dut.set('0'); self.dias_sab.set('0'); self.dias_dom.set('0')
        for var in [self.reducao_boa, self.reducao_cax, self.reducao_cno, self.reducao_csr, self.reducao_eme, self.reducao_glo, self.reducao_mob, self.reducao_sjt, self.reducao_vml]:
            var.set("0,00%")
        self.pasta_destino.set("")
        self._caminho_completo_cut2, self._caminho_completo_cno, self._caminho_completo_mob = "", "", ""
        self.label_cut2.config(text="Nenhum arquivo selecionado.")
        self.label_cno.config(text="Nenhum arquivo selecionado.")
        self.label_mob.config(text="Nenhum arquivo selecionado.")
        self.label_pasta.config(text="Nenhuma pasta selecionada.")
        self.dados_rateamento = {}
        
        for empresa in self.rateamento_empresa_frames.keys():
            self._reconstruir_widgets_empresa(empresa)

        self.label_status_rateamento.config(text="Status: Nenhum dado coletado.", foreground="gray")
        self._setar_nome_padrao_arquivo()
        messagebox.showinfo("Limpeza", "Todos os campos foram limpos.", parent=self)
        
    def _salvar_sessao(self): 
        config_data = {
            'arquivos_passado': self.arquivos_passado,
            'arquivos_atual': self.arquivos_atual,
            'caminho_pco': self._caminho_completo_cut2,
            'mes_pco': self.mes_cut2.get(),
            'quinzena_pco': self.quinzena_cut2.get(),
            'ano_pco': self.ano_pco.get(), # <-- ADICIONADO
            'caminho_cno': self._caminho_completo_cno,
            'caminho_mob': self._caminho_completo_mob,
            'dias_dut': self.dias_dut.get(),
            'dias_sab': self.dias_sab.get(),
            'dias_dom': self.dias_dom.get(),
            'reducao_boa': self.reducao_boa.get(), 'reducao_cax': self.reducao_cax.get(), 'reducao_cno': self.reducao_cno.get(),
            'reducao_csr': self.reducao_csr.get(), 'reducao_eme': self.reducao_eme.get(), 'reducao_glo': self.reducao_glo.get(),
            'reducao_mob': self.reducao_mob.get(), 'reducao_sjt': self.reducao_sjt.get(), 'reducao_vml': self.reducao_vml.get(),
            'pasta_destino': self.pasta_destino.get(),
            'nome_arquivo_saida': self.nome_arquivo_saida.get(),
            'rateamento': self.dados_rateamento,
        }
        try:
            os.makedirs('instance', exist_ok=True)
            with open(self.SESSAO_FILE, 'w', encoding='utf-8') as f:
                json.dump(config_data, f, indent=4)
        except Exception as e:
            print(f"AVISO: Não foi possível salvar a sessão: {e}")

    def _carregar_sessao(self): 
        if not os.path.exists(self.SESSAO_FILE):
            messagebox.showinfo("Informação", "Nenhuma sessão anterior foi encontrada.", parent=self)
            return
        
        try:
            with open(self.SESSAO_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)

            self.limpar_todas_selecoes()

            arquivos_nao_encontrados = []
            
            arquivos_passado_verificados = [p for p in data.get('arquivos_passado', []) if os.path.exists(p)]
            arquivos_nao_encontrados.extend([os.path.basename(p) for p in data.get('arquivos_passado', []) if not os.path.exists(p)])
            self._adicionar_arquivos(arquivos_passado_verificados, self.arquivos_passado, self.lista_passado)
            
            arquivos_atual_verificados = [p for p in data.get('arquivos_atual', []) if os.path.exists(p)]
            arquivos_nao_encontrados.extend([os.path.basename(p) for p in data.get('arquivos_atual', []) if not os.path.exists(p)])
            self._adicionar_arquivos(arquivos_atual_verificados, self.arquivos_atual, self.lista_atual)

            if arquivos_nao_encontrados:
                messagebox.showwarning("Aviso", "Alguns arquivos da sessão anterior não foram encontrados e foram ignorados:\n\n" + "\n".join(arquivos_nao_encontrados), parent=self)

            if data.get('caminho_pco') and os.path.exists(data['caminho_pco']):
                self._caminho_completo_cut2 = data['caminho_pco']
                self.label_cut2.config(text=os.path.basename(self._caminho_completo_cut2))
            self.mes_cut2.set(data.get('mes_pco', ''))
            self.quinzena_cut2.set(data.get('quinzena_pco', ''))
            self.ano_pco.set(data.get('ano_pco', '')) # <-- ADICIONADO

            if data.get('caminho_cno') and os.path.exists(data['caminho_cno']):
                self._caminho_completo_cno = data['caminho_cno']
                self.label_cno.config(text=os.path.basename(self._caminho_completo_cno))

            if data.get('caminho_mob') and os.path.exists(data['caminho_mob']):
                self._caminho_completo_mob = data['caminho_mob']
                self.label_mob.config(text=os.path.basename(self._caminho_completo_mob))
            
            self.dias_dut.set(data.get('dias_dut', '0'))
            self.dias_sab.set(data.get('dias_sab', '0'))
            self.dias_dom.set(data.get('dias_dom', '0'))
            self.reducao_boa.set(data.get('reducao_boa', '0,00%'))
            self.reducao_cax.set(data.get('reducao_cax', '0,00%'))
            self.reducao_cno.set(data.get('reducao_cno', '0,00%'))
            self.reducao_csr.set(data.get('reducao_csr', '0,00%'))
            self.reducao_eme.set(data.get('reducao_eme', '0,00%'))
            self.reducao_glo.set(data.get('reducao_glo', '0,00%'))
            self.reducao_mob.set(data.get('reducao_mob', '0,00%'))
            self.reducao_sjt.set(data.get('reducao_sjt', '0,00%'))
            self.reducao_vml.set(data.get('reducao_vml', '0,00%'))
            

            if data.get('pasta_destino') and os.path.isdir(data['pasta_destino']):
                self.pasta_destino.set(data['pasta_destino'])
                self.label_pasta.config(text=data['pasta_destino'])
            self.nome_arquivo_saida.set(data.get('nome_arquivo_saida', ''))


            dados_rateamento_salvos = data.get('rateamento', {})
            if dados_rateamento_salvos:
                self.dados_rateamento = dados_rateamento_salvos
                df_rateamento = self._converter_rateamento_para_df(dados_rateamento_salvos)
                for empresa in df_rateamento['Empresa'].unique():
                    if empresa in self.rateamento_empresa_frames:
                        df_empresa = df_rateamento[df_rateamento['Empresa'] == empresa].copy()
                        self._reconstruir_widgets_empresa(empresa, df_empresa)
                self.label_status_rateamento.config(text="Status: Dados da última sessão carregados.", foreground="blue")

            messagebox.showinfo("Sucesso", "Última sessão carregada com sucesso!", parent=self)

        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível carregar a sessão anterior.\n\nDetalhes: {e}", parent=self)

    def _converter_rateamento_para_df(self, dados_rateamento_dict): 
        lista_final = []
        garagem_id_counter = 0
        current_empresa = None

        for empresa, dados_lista in dados_rateamento_dict.items():
            if empresa != current_empresa:
                current_empresa = empresa
                garagem_id_counter = 0 

            current_garagem_id = 0
            for item in dados_lista:
                litros_garagem, litros_companhia, posto, cnpj, inscricao = item
                
                if litros_garagem is not None:
                    garagem_id_counter += 1
                    current_garagem_id = garagem_id_counter

                lista_final.append({
                    'Empresa': empresa,
                    'Garagem_ID': current_garagem_id,
                    'Litros_Garagem': litros_garagem,
                    'Litros_Companhia': litros_companhia,
                    'Posto': posto,
                    'CNPJ': cnpj,
                    'Inscrição_Estadual': inscricao,
                })
        
        df = pd.DataFrame(lista_final)
        
        df['Litros_Garagem'] = df.groupby(['Empresa', 'Garagem_ID'])['Litros_Garagem'].transform('first')
        df['CNPJ'] = df.groupby(['Empresa', 'Garagem_ID'])['CNPJ'].transform('first')
        df['Inscrição_Estadual'] = df.groupby(['Empresa', 'Garagem_ID'])['Inscrição_Estadual'].transform('first')
        
        df = df[df['Litros_Companhia'].notna() | df['Litros_Garagem'].notna()]
        
        return df

    def iniciar_processamento(self):
        if not self.arquivos_atual: return messagebox.showwarning("Atenção", "É necessário selecionar arquivos para o 'Mês de Referência'.", parent=self)
        if not self.arquivos_passado and not messagebox.askyesno("Atenção", "Nenhum arquivo do 'Mês Passado' foi selecionado. Deseja continuar sem a comparação?", parent=self): return
        
        # --- LÓGICA DE VALIDAÇÃO MODIFICADA ---
        if self._caminho_completo_cut2:
            if not self.mes_cut2.get() or not self.quinzena_cut2.get() or not self.ano_pco.get():
                return messagebox.showwarning("Atenção", "Para o relatório de Rendimento PCO, selecione o Mês, a Quinzena e preencha o Ano.", parent=self)
            try:
                int(self.ano_pco.get()) # Verifica se o ano é um número
            except ValueError:
                return messagebox.showwarning("Atenção", "O Ano do PCO deve ser um número (ex: 2024).", parent=self)
        # --- FIM DA MODIFICAÇÃO ---

        try:
            self.contagem_dias_dict = {'DUT': int(self.dias_dut.get() or 0), 'SAB': int(self.dias_sab.get() or 0), 'DOM': int(self.dias_dom.get() or 0)}
        except ValueError:
            return messagebox.showerror("Erro de Entrada", "Os valores para a contagem de dias devem ser números inteiros.", parent=self)
        
        # --- PARSE DAS REDUÇÕES ---
        def parse_pct(val_str):
            try:
                clean_str = val_str.replace('%', '').replace(',', '.').strip()
                if not clean_str: return 0.0
                return float(clean_str) / 100
            except ValueError:
                return None

        self.reducoes_processadas = {
            'BOA': parse_pct(self.reducao_boa.get()), 'CAX': parse_pct(self.reducao_cax.get()), 'CNO': parse_pct(self.reducao_cno.get()),
            'CSR': parse_pct(self.reducao_csr.get()), 'EME': parse_pct(self.reducao_eme.get()), 'GLO': parse_pct(self.reducao_glo.get()),
            'MOB': parse_pct(self.reducao_mob.get()), 'SJT': parse_pct(self.reducao_sjt.get()), 'VML': parse_pct(self.reducao_vml.get())
        }

        if any(v is None for v in self.reducoes_processadas.values()):
             return messagebox.showerror("Erro de Entrada", "Os valores de Redução/Acréscimo devem ser números válidos (ex: -5, 5% ou -0,05).", parent=self)
        # --------------------------
        
        if (self._caminho_completo_cno or self._caminho_completo_mob) and sum(self.contagem_dias_dict.values()) == 0:
            messagebox.showwarning("Atenção", "Nenhum dia foi informado. Os cálculos de cota final para CNO/MOB serão zero.", parent=self)

        pasta_destino, nome_arquivo = self.pasta_destino.get(), self.nome_arquivo_saida.get()
        if not pasta_destino or not nome_arquivo: return messagebox.showwarning("Atenção", "Selecione uma pasta de destino e defina um nome para o arquivo de saída.", parent=self)
        
        self._salvar_dados_rateamento()

        nome_arquivo = nome_arquivo if nome_arquivo.lower().endswith('.xlsx') else nome_arquivo + '.xlsx'
        self.output_path = os.path.join(pasta_destino, nome_arquivo)

        self.botao_gerar.config(state="disabled", text="Processando...")
        self.botao_limpar.config(state="disabled")
        self.botao_carregar.config(state="disabled") 
        self.update_idletasks()
        
        threading.Thread(target=executar_processamento, args=(self,)).start()
        
    def finalizar_processamento_gui(self, mensagem, sucesso):
        self.botao_gerar.config(state="normal", text="Processar e Gerar Relatório")
        self.botao_limpar.config(state="normal")
        self.botao_carregar.config(state="normal") 
        if sucesso:
            self._salvar_sessao() 
            messagebox.showinfo("Sucesso!", f"Relatório final gerado com sucesso em:\n{mensagem}", parent=self)
            self.abrir_pasta_destino()
        else:
            messagebox.showerror("Erro no Processamento", f"Ocorreu um erro inesperado.\n\nDetalhes: {mensagem}", parent=self)

    def abrir_pasta_destino(self):
        pasta = self.pasta_destino.get()
        if not pasta: return
        try:
            if sys.platform == "win32": os.startfile(pasta)
            elif sys.platform == "darwin": subprocess.run(["open", pasta])
            else: subprocess.run(["xdg-open", pasta])
        except Exception as e:
            messagebox.showwarning("Aviso", f"Não foi possível abrir a pasta de destino automaticamente.\n\nCaminho: {pasta}\nErro: {e}", parent=self)
            
    def center_window(self):
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        
    def _setar_nome_padrao_arquivo(self):
        hoje = datetime.date.today()
        meses_pt = ['JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO', 'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']
        nome_arquivo = f"Cota de {meses_pt[hoje.month-1]} {hoje.year} RCB"
        self.nome_arquivo_saida.set(nome_arquivo)

if __name__ == "__main__":
    try:
        import xlrd
    except ImportError:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Dependência Faltando", "A biblioteca 'xlrd' não foi encontrada. Por favor, instale-a com 'pip install xlrd'.")
        sys.exit(1)
        
    app = CotaDieselGUI()
    app.mainloop()