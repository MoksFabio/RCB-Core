import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime, date
import os
import sys
import threading
import traceback
import subprocess

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

# Imports para a nova GUI
from tkinterdnd2 import DND_FILES, TkinterDnD
from PIL import Image, ImageTk
import sv_ttk

# --- Constantes e Configurações Globais ---
PERMISSIONARIAS_EMPRESAS = ["BOA", "CAX", "CSR", "EME", "GLO", "SJT", "VML"]
COLUNAS_PARA_EXCEL_PERMISSIONARIAS = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO', 'NMEFETPASST']

CONCESSIONARIAS_EMPRESAS = ["CNO", "MOB"]
COLUNAS_PARA_EXCEL_CONCESSIONARIAS = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO', 'NMPASSTOTAL']

STPP_RMR_EMPRESAS_TODAS = sorted(list(set(PERMISSIONARIAS_EMPRESAS + CONCESSIONARIAS_EMPRESAS)))
COLUNAS_PARA_EXCEL_STPP_RMR = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO', 'NMEFETPASST']


# --- Funções de Utilitários e Processamento de Dados ---
def obter_pasta_downloads_padrao():
    if os.name == 'nt':
        try:
            import winreg
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
            downloads_path = winreg.QueryValueEx(key, "{374DE290-123F-4565-9164-39C4925E467B}")[0]
            winreg.CloseKey(key)
            return downloads_path
        except Exception:
            return os.path.join(os.path.expanduser('~'), 'Downloads')
    else:
        return os.path.join(os.path.expanduser('~'), 'Downloads')

def carregar_e_filtrar_txt(caminho_arquivo_txt, config_usuario, colunas_desejadas, coluna_passageiros):
    print(f"\nLendo o arquivo TXT de: {caminho_arquivo_txt}")
    df = None
    configs_leitura = [('\t', 'latin1'), ('\t', 'utf-8'), ('\t', 'utf-8-sig'), (';', 'latin1'), (';', 'utf-8'), (';', 'utf-8-sig'), (',', 'latin1'), (',', 'utf-8'), (',', 'utf-8-sig')]
    leitura_bem_sucedida = False
    for delim, enc in configs_leitura:
        try:
            preview_df = pd.read_csv(caminho_arquivo_txt, sep=delim, dtype=str, encoding=enc, on_bad_lines='warn', nrows=5)
            if preview_df is not None and len(preview_df.columns) > 1 and not any(delim in str(col_name) for col_name in preview_df.columns):
                df_temp = pd.read_csv(caminho_arquivo_txt, sep=delim, dtype=str, encoding=enc, on_bad_lines='skip')
                if df_temp is not None and not df_temp.empty:
                    df = df_temp
                    print(f"Lido com sucesso usando delimitador '{repr(delim)}' e codificação '{enc}'.")
                    leitura_bem_sucedida = True
                    break
        except (pd.errors.EmptyDataError, Exception):
            continue
    if not leitura_bem_sucedida:
        print("Erro Crítico: Não foi possível ler o arquivo TXT com as configurações testadas.")
        return None

    df.columns = [str(col).strip().upper() for col in df.columns]
    colunas_desejadas_upper = [col.upper() for col in colunas_desejadas]
    coluna_passageiros_upper = coluna_passageiros.upper()
    if any(col not in df.columns for col in colunas_desejadas_upper):
        faltantes = [col for col in colunas_desejadas_upper if col not in df.columns]
        print(f"Erro: Colunas obrigatórias não encontradas no arquivo: {', '.join(faltantes)}. Colunas encontradas: {df.columns.tolist()}")
        return None
    
    df_processado = df[colunas_desejadas_upper].copy()
    dt_operacao_col, cd_linha_col, cd_operador_col = 'DTOPERACAO', 'CDLINHA', 'CDOPERADOR'
    df_processado.loc[:, 'DTOPERACAO_temp_dt'] = pd.to_datetime(df_processado[dt_operacao_col], errors='coerce', dayfirst=True)
    df_processado.dropna(subset=['DTOPERACAO_temp_dt'], inplace=True)
    if df_processado.empty: return pd.DataFrame(columns=df_processado.columns)

    dfs_filtrados_por_periodo = []
    if config_usuario['periodos_selecionados_objs']:
        for periodo in config_usuario['periodos_selecionados_objs']:
            df_chunk = df_processado[(df_processado['DTOPERACAO_temp_dt'].dt.date >= periodo['inicio'].date()) & (df_processado['DTOPERACAO_temp_dt'].dt.date <= periodo['fim'].date())].copy()
            if not df_chunk.empty: dfs_filtrados_por_periodo.append(df_chunk)
        if not dfs_filtrados_por_periodo: return pd.DataFrame(columns=df_processado.columns)
        df_processado = pd.concat(dfs_filtrados_por_periodo).drop_duplicates().reset_index(drop=True)
    else:
        return pd.DataFrame(columns=df_processado.columns)
    if df_processado.empty: return pd.DataFrame(columns=df_processado.columns)

    df_processado.loc[:, dt_operacao_col] = df_processado['DTOPERACAO_temp_dt'].dt.date
    df_processado = df_processado.drop(columns=['DTOPERACAO_temp_dt'])

    if config_usuario['cod_linhas_str'].upper() != "TODAS":
        linhas_para_filtrar = [l.strip() for l in config_usuario['cod_linhas_str'].split(',') if l.strip()]
        df_processado.loc[:, cd_linha_col] = df_processado[cd_linha_col].astype(str).str.strip()
        df_processado = df_processado[df_processado[cd_linha_col].isin(linhas_para_filtrar)]
    if df_processado.empty: return pd.DataFrame(columns=df_processado.columns)

    df_processado.loc[:, cd_operador_col] = df_processado[cd_operador_col].astype(str).str.strip().str.upper()
    df_processado = df_processado[df_processado[cd_operador_col].isin(config_usuario['empresas_selecionadas'])]
    if df_processado.empty: return pd.DataFrame(columns=df_processado.columns)

    if coluna_passageiros_upper in df_processado.columns and not df_processado.empty:
        pass_col_temp = df_processado[coluna_passageiros_upper].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
        df_processado.loc[:, coluna_passageiros_upper] = pd.to_numeric(pass_col_temp, errors='coerce')
    return df_processado

def sanitize_sheet_name(name_str):
    return re.sub(r'[\\/*?:\[\]]', '_', name_str)[:31]

def exportar_para_excel_formatado(df_all_data, base_titulo_para_sheets, caminho_arquivo_excel, colunas_para_exportar, periodos_objs):
    wb = Workbook()
    if wb.active: wb.remove(wb.active)
    
    colunas_para_exportar_upper = [col.upper() for col in colunas_para_exportar]
    cor_fundo_titulo_excel, fonte_titulo_obj = '1F4E78', Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    alinhamento_titulo, preenchimento_titulo = Alignment(horizontal='center', vertical='center', wrap_text=True), PatternFill(fill_type='solid', start_color=cor_fundo_titulo_excel)
    center_alignment_dados = Alignment(horizontal='center', vertical='center', wrap_text=True)
    borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_font_dados, header_fill_dados = Font(name='Calibri', size=11, bold=True, color='FFFFFF'), PatternFill(fill_type="solid", start_color="4A86E8")
    data_font_dados = Font(name='Calibri', size=10)
    linha_inicio_dados_excel = 3

    for periodo_idx, periodo in enumerate(periodos_objs):
        dt_inicio, dt_fim = periodo['inicio'].date(), periodo['fim'].date()
        df_sheet_data = df_all_data[(df_all_data['DTOPERACAO'] >= dt_inicio) & (df_all_data['DTOPERACAO'] <= dt_fim)].copy() if not df_all_data.empty and 'DTOPERACAO' in df_all_data.columns else pd.DataFrame()
        
        colunas_existentes = [col for col in colunas_para_exportar_upper if col in df_sheet_data.columns]
        df_sheet_data = df_sheet_data[colunas_existentes] if colunas_existentes else pd.DataFrame()
        
        nome_aba = sanitize_sheet_name(f"{dt_inicio.strftime('%d%m%y')}_{dt_fim.strftime('%d%m%y')}_{periodo_idx}")
        ws = wb.create_sheet(title=nome_aba)

        titulo_aba_texto = f"{base_titulo_para_sheets} PERÍODO: {dt_inicio.strftime('%d/%m/%Y')} A {dt_fim.strftime('%d/%m/%Y')}"
        num_colunas_aba = len(df_sheet_data.columns) if not df_sheet_data.empty else len(colunas_para_exportar_upper) or 1
        
        ws.merge_cells(f'A1:{get_column_letter(num_colunas_aba)}2')
        celula_titulo_aba = ws['A1']
        celula_titulo_aba.value, celula_titulo_aba.font, celula_titulo_aba.alignment, celula_titulo_aba.fill = titulo_aba_texto, fonte_titulo_obj, alinhamento_titulo, preenchimento_titulo
        ws.row_dimensions[1].height = ws.row_dimensions[2].height = 22
        
        if df_sheet_data.empty:
            ws.cell(linha_inicio_dados_excel, 1, "Nenhum dado encontrado para este período com os filtros aplicados.").font = data_font_dados
            continue
        
        for r_idx, row_data in enumerate(dataframe_to_rows(df_sheet_data, index=False, header=True)):
            for c_idx, value in enumerate(row_data, 1):
                ws.cell(linha_inicio_dados_excel + r_idx, c_idx, value)

        dt_operacao_col_index = -1
        cols = df_sheet_data.columns.tolist()
        if 'DTOPERACAO' in cols:
            dt_operacao_col_index = cols.index('DTOPERACAO') + 1

        for row in ws.iter_rows(min_row=linha_inicio_dados_excel, max_row=ws.max_row, min_col=1, max_col=num_colunas_aba):
            for cell in row:
                cell.alignment = center_alignment_dados
                cell.border = borda_fina
                if cell.row == linha_inicio_dados_excel:
                    cell.font = header_font_dados
                    cell.fill = header_fill_dados
                else:
                    cell.font = data_font_dados
                    if cell.column == dt_operacao_col_index and isinstance(cell.value, (datetime, date)):
                        cell.number_format = 'DD/MM/YYYY'
        
        for c_idx, col_name in enumerate(cols, 1):
            max_len = max(len(str(v)) for v in df_sheet_data[col_name].tolist() + [col_name])
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max(max_len + 5, 12), 40)
    
    try:
        wb.save(caminho_arquivo_excel)
    except PermissionError:
        raise PermissionError(f"Permissão negada para salvar em '{caminho_arquivo_excel}'. O arquivo pode estar aberto.")
    except Exception as e:
        raise Exception(f"Erro inesperado ao salvar Excel: {e}")

def executar_processamento_principal(gui_instance):
    try:
        config_usuario = gui_instance.config_processamento
        df_final = pd.DataFrame()

        if gui_instance.process_type == 'stpp_rmr':
            df_perm = carregar_e_filtrar_txt(gui_instance.caminho_txt_perm_var.get(), config_usuario, COLUNAS_PARA_EXCEL_PERMISSIONARIAS, 'NMEFETPASST')
            df_conc = carregar_e_filtrar_txt(gui_instance.caminho_txt_conc_var.get(), config_usuario, COLUNAS_PARA_EXCEL_CONCESSIONARIAS, 'NMPASSTOTAL')
            if df_conc is not None and not df_conc.empty: df_conc = df_conc.rename(columns={'NMPASSTOTAL': 'NMEFETPASST'})
            
            dfs = [df for df in [df_perm, df_conc] if df is not None and not df.empty]
            if dfs: df_final = pd.concat(dfs, ignore_index=True)
        else:
            caminho_txt = gui_instance.caminho_txt_perm_var.get()
            df_final = carregar_e_filtrar_txt(caminho_txt, config_usuario, gui_instance.colunas_excel, gui_instance.coluna_passageiros)

        if df_final is None:
            gui_instance.finalizar_processamento(f"Falha crítica ao carregar/processar o arquivo TXT.", sucesso=False)
            return

        if not df_final.empty:
            if 'CDLINHA' in df_final.columns: df_final['CDLINHA'] = pd.to_numeric(df_final['CDLINHA'], errors='coerce')
            df_final.dropna(subset=['CDLINHA', 'DTOPERACAO'], inplace=True)
            df_final = df_final.sort_values(by=['CDOPERADOR', 'CDLINHA', 'DTOPERACAO']).reset_index(drop=True)
        else:
            print("AVISO: Nenhum dado encontrado com os filtros. Excel será gerado com abas vazias.")

        exportar_para_excel_formatado(
            df_final,
            config_usuario['base_titulo_excel'],
            config_usuario['nome_arquivo_excel'],
            gui_instance.colunas_excel,
            config_usuario['periodos_selecionados_objs']
        )
        gui_instance.finalizar_processamento(config_usuario['nome_arquivo_excel'], sucesso=True)
    except Exception as e:
        gui_instance.finalizar_processamento(f"Erro em executar_processamento_principal: {e}\n{traceback.format_exc()}", sucesso=False)


# --- Classes da Interface Gráfica ---
class DemandaLinhaGUI(TkinterDnD.Tk):
    def __init__(self, process_type):
        super().__init__()
        self.process_type = process_type
        self.withdraw() 
        
        self.voltar_solicitado = False 
        
        self._setup_config()
        self.title(self.titulo_janela)
        self.geometry("950x950")
        
        # Define o tamanho mínimo que a janela pode ter para evitar quebras de layout
        self.minsize(800, 650)
        
        self.center_window()
        sv_ttk.set_theme("light")

        self.LOGO_IMAGE_PATH = "static/images/rcb_logo.png"
        self._setup_variables()
        self._create_widgets()
        
        self.deiconify() 

    def _setup_config(self):
        configs = {
            "permissionarias": ("Demanda - Permissionárias", PERMISSIONARIAS_EMPRESAS, COLUNAS_PARA_EXCEL_PERMISSIONARIAS, 'NMEFETPASST', "DEMANDA PERMISSIONÁRIAS"),
            "concessionarias": ("Demanda - Concessionárias", CONCESSIONARIAS_EMPRESAS, COLUNAS_PARA_EXCEL_CONCESSIONARIAS, 'NMPASSTOTAL', "DEMANDA CONCESSIONÁRIAS"),
            "stpp_rmr": ("Demanda - STPP/RMR", STPP_RMR_EMPRESAS_TODAS, COLUNAS_PARA_EXCEL_STPP_RMR, 'NMEFETPASST', "DEMANDA STPP/RMR"),
        }
        self.titulo_janela, self.empresas_validas, self.colunas_excel, self.coluna_passageiros, self.titulo_base_excel = configs[self.process_type]

    def _setup_variables(self):
        self.caminho_txt_perm_var = tk.StringVar()
        self.caminho_txt_conc_var = tk.StringVar()
        self.cod_linhas_var = tk.StringVar()
        self.empresas_var = tk.StringVar()
        self.num_ouvidoria_var = tk.StringVar(value="Resp_Ouvidoria_")
        self.local_salvar_var = tk.StringVar(value=obter_pasta_downloads_padrao())
        self.data_inicio_var = tk.StringVar()
        self.data_fim_var = tk.StringVar()
        self.lista_periodos_objs = []

    def _create_widgets(self):
        self.style = ttk.Style(self)
        self.style.configure('Accent.TButton', font=('Segoe UI', 10, 'bold'))
        self.style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'))
        self.style.configure('Header.TLabel', font=('Segoe UI', 11, 'bold'))

        header_frame = ttk.Frame(self)
        header_frame.pack(fill=tk.X, padx=20, pady=(15, 10))
        try:
            logo_original = Image.open(self.LOGO_IMAGE_PATH).convert("RGBA")
            logo_resized = logo_original.resize((int(logo_original.width*0.35), int(logo_original.height*0.35)), Image.Resampling.LANCZOS)
            self.logo_image = ImageTk.PhotoImage(logo_resized)
            ttk.Label(header_frame, image=self.logo_image).pack(side=tk.LEFT, padx=(0, 15))
        except Exception: print("AVISO: Logo não encontrado em 'static/images/rcb_logo.png'")

        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(title_frame, text="Assistente de Demanda de Linha", style='Title.TLabel').pack(anchor='w')
        ttk.Label(title_frame, text=self.titulo_janela).pack(anchor='w')
        
        ttk.Checkbutton(header_frame, text="Tema", style="Switch.TCheckbutton", command=sv_ttk.toggle_theme).pack(side=tk.RIGHT)

        notebook = ttk.Notebook(self, padding=(0, 10))
        notebook.pack(fill=tk.BOTH, expand=True, padx=20)
        
        tab1 = ttk.Frame(notebook, padding=15)
        tab2 = ttk.Frame(notebook, padding=15)
        notebook.add(tab1, text='  Arquivos e Filtros  ')
        notebook.add(tab2, text='  Configuração de Saída  ')

        self._criar_aba_filtros(tab1)
        self._criar_aba_saida(tab2)
        
        button_area = ttk.Frame(self, padding=(20, 10, 20, 20))
        button_area.pack(fill=tk.X)
        
        self.botao_voltar = ttk.Button(button_area, text="Voltar", command=self._voltar_para_selecao)
        self.botao_voltar.pack(side=tk.LEFT)
        self.botao_limpar = ttk.Button(button_area, text="Limpar Tudo", command=self.limpar_tudo)
        self.botao_limpar.pack(side=tk.LEFT, padx=(10, 0))
        
        self.botao_gerar = ttk.Button(button_area, text="Processar e Gerar Relatório", style='Accent.TButton', command=self.iniciar_processamento)
        self.botao_gerar.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))

    def _criar_aba_filtros(self, parent):
        parent.columnconfigure(0, weight=1)
        
        frame_arquivos = ttk.LabelFrame(parent, text=" Arquivo(s) de Entrada (.txt) ", padding=10)
        frame_arquivos.grid(row=0, column=0, sticky="ew", pady=(0, 15))
        frame_arquivos.columnconfigure(0, weight=1)
        
        if self.process_type == "stpp_rmr":
            self._criar_seletor_arquivo(frame_arquivos, "Permissionárias", self.caminho_txt_perm_var, 0)
            self._criar_seletor_arquivo(frame_arquivos, "Concessionárias", self.caminho_txt_conc_var, 1)
        else:
            self._criar_seletor_arquivo(frame_arquivos, self.process_type.capitalize(), self.caminho_txt_perm_var, 0)

        frame_filtros = ttk.LabelFrame(parent, text=" Filtros de Dados ", padding=10)
        frame_filtros.grid(row=1, column=0, sticky="ew", pady=(0, 15))
        frame_filtros.columnconfigure(1, weight=1)

        ttk.Label(frame_filtros, text="Códigos das Linhas:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(frame_filtros, textvariable=self.cod_linhas_var).grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        ttk.Label(frame_filtros, text="Empresa(s) Operadora(s):").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(frame_filtros, textvariable=self.empresas_var).grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        ttk.Label(frame_filtros, text=f"Válidas: {', '.join(self.empresas_validas)}", wraplength=700, justify=tk.LEFT).grid(row=2, column=1, sticky="w", padx=5)

        frame_periodos = ttk.LabelFrame(parent, text=" Períodos de Análise ", padding=10)
        frame_periodos.grid(row=2, column=0, sticky="nsew")
        parent.rowconfigure(2, weight=1)
        frame_periodos.columnconfigure(0, weight=1)
        frame_periodos.rowconfigure(1, weight=1)

        input_periodo_frame = ttk.Frame(frame_periodos)
        input_periodo_frame.grid(row=0, column=0, sticky="ew", pady=5)
        
        # Configura as colunas do frame de input para expandirem de forma proporcional
        input_periodo_frame.columnconfigure(1, weight=2) 
        input_periodo_frame.columnconfigure(3, weight=2) 
        input_periodo_frame.columnconfigure(4, weight=3) 

        ttk.Label(input_periodo_frame, text="Início (DD/MM/AAAA):").grid(row=0, column=0, sticky="w", padx=(0, 5))
        ttk.Entry(input_periodo_frame, textvariable=self.data_inicio_var, width=12).grid(row=0, column=1, sticky="ew", padx=(0, 10))
        
        ttk.Label(input_periodo_frame, text="Fim (DD/MM/AAAA):").grid(row=0, column=2, sticky="w", padx=(10, 5))
        ttk.Entry(input_periodo_frame, textvariable=self.data_fim_var, width=12).grid(row=0, column=3, sticky="ew", padx=(0, 10))
        
        ttk.Button(input_periodo_frame, text="Adicionar Período", command=self._adicionar_periodo).grid(row=0, column=4, sticky="ew")

        list_frame = ttk.Frame(frame_periodos)
        list_frame.grid(row=1, column=0, sticky="nsew", pady=5)
        list_frame.rowconfigure(0, weight=1)
        list_frame.columnconfigure(0, weight=1)
        self.listbox_periodos = tk.Listbox(list_frame, height=4, exportselection=False)
        self.listbox_periodos.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.listbox_periodos.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.listbox_periodos.config(yscrollcommand=scrollbar.set)
        
        ttk.Button(frame_periodos, text="Remover Selecionado", command=self._remover_periodo).grid(row=2, column=0, sticky="e", pady=5)

    def _criar_seletor_arquivo(self, parent, label_text, var, row_num):
        sub_frame = ttk.Frame(parent)
        sub_frame.grid(row=row_num, column=0, sticky="ew", pady=5)
        sub_frame.columnconfigure(0, weight=1)
        
        label = ttk.Label(sub_frame, text=f" Arraste o arquivo TXT {label_text} aqui", anchor='center', relief="solid", padding=20)
        label.grid(row=0, column=0, sticky="ew")
        
        label.drop_target_register(DND_FILES)
        label.dnd_bind('<<Drop>>', lambda e, v=var, l=label: self._on_drop(e, v, l))
        
        ttk.Button(sub_frame, text=f"Ou Selecione...", command=lambda v=var, l=label: self._selecionar_arquivo(v, l)).grid(row=0, column=1, padx=10)

    def _on_drop(self, event, var_alvo, label_alvo):
        caminho = self.tk.splitlist(event.data)[0]
        if caminho.lower().endswith('.txt'):
            var_alvo.set(caminho)
            label_alvo.config(text=os.path.basename(caminho), relief="sunken", foreground="green")
        else:
            messagebox.showwarning("Formato Inválido", "Por favor, solte apenas arquivos .txt", parent=self)

    def _selecionar_arquivo(self, var_alvo, label_alvo):
        caminho = filedialog.askopenfilename(parent=self, title="Selecione o arquivo TXT", filetypes=[("Arquivos de Texto", "*.txt"), ("Todos os arquivos", "*.*")])
        if caminho:
            var_alvo.set(caminho)
            label_alvo.config(text=os.path.basename(caminho), relief="sunken", foreground="green")

    def _criar_aba_saida(self, parent):
        parent.columnconfigure(0, weight=1)
        frame_saida = ttk.LabelFrame(parent, text=" Configuração do Arquivo de Saída ", padding=15)
        frame_saida.pack(fill=tk.X)
        frame_saida.columnconfigure(1, weight=1)
        
        ttk.Label(frame_saida, text="Nº Ouvidoria/ID:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(frame_saida, textvariable=self.num_ouvidoria_var).grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        
        ttk.Label(frame_saida, text="Pasta de Destino:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        path_frame = ttk.Frame(frame_saida)
        path_frame.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        path_frame.columnconfigure(0, weight=1)
        entry_path = ttk.Entry(path_frame, textvariable=self.local_salvar_var, state="readonly")
        entry_path.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        ttk.Button(path_frame, text="Selecionar Pasta...", command=self._selecionar_pasta_destino).pack(side=tk.LEFT)

    def _adicionar_periodo(self):
        inicio_str, fim_str = self.data_inicio_var.get(), self.data_fim_var.get()
        try:
            inicio_obj = datetime.strptime(inicio_str, "%d/%m/%Y")
            fim_obj = datetime.strptime(fim_str, "%d/%m/%Y")
            if fim_obj < inicio_obj:
                messagebox.showerror("Erro de Validação", "A 'Data Fim' deve ser maior ou igual à 'Data Início'.", parent=self)
                return
            periodo_display = f"{inicio_str} - {fim_str}"
            if any(p['display'] == periodo_display for p in self.lista_periodos_objs):
                messagebox.showwarning("Aviso", "Este período já foi adicionado.", parent=self)
                return
            self.lista_periodos_objs.append({'inicio': inicio_obj, 'fim': fim_obj, 'display': periodo_display})
            self.listbox_periodos.insert(tk.END, periodo_display)
            self.data_inicio_var.set(""), self.data_fim_var.set("")
        except ValueError:
            messagebox.showerror("Erro de Validação", "Formato de data inválido. Use DD/MM/AAAA.", parent=self)

    def _remover_periodo(self):
        selecionados = self.listbox_periodos.curselection()
        if not selecionados:
            messagebox.showwarning("Aviso", "Nenhum período selecionado para remover.", parent=self)
            return
        del self.lista_periodos_objs[selecionados[0]]
        self.listbox_periodos.delete(selecionados[0])

    def _selecionar_pasta_destino(self):
        pasta = filedialog.askdirectory(parent=self, title="Selecione a pasta para salvar")
        if pasta: self.local_salvar_var.set(pasta)
        
    def _voltar_para_selecao(self):
        self.voltar_solicitado = True
        self.destroy()

    def limpar_tudo(self):
        self.caminho_txt_perm_var.set(""), self.caminho_txt_conc_var.set("")
        self.cod_linhas_var.set(""), self.empresas_var.set("")
        self.num_ouvidoria_var.set("Resp_Ouvidoria_")
        self.local_salvar_var.set(obter_pasta_downloads_padrao())
        self.data_inicio_var.set(""), self.data_fim_var.set("")
        self.lista_periodos_objs.clear()
        self.listbox_periodos.delete(0, tk.END)
        for child in self.winfo_children():
            if isinstance(child, ttk.Notebook):
                tab1 = child.winfo_children()[0]
                for frame in tab1.winfo_children():
                    if isinstance(frame, ttk.LabelFrame):
                        for sub_frame in frame.winfo_children():
                            if isinstance(sub_frame, ttk.Frame):
                                for label in sub_frame.winfo_children():
                                    if isinstance(label, ttk.Label) and hasattr(label, 'dnd_bindings'):
                                        label.config(text=f" Arraste o arquivo aqui", relief="solid", foreground="")
        messagebox.showinfo("Limpeza", "Todos os campos foram redefinidos.", parent=self)

    def iniciar_processamento(self):
        if self.process_type == 'stpp_rmr' and (not self.caminho_txt_perm_var.get() or not self.caminho_txt_conc_var.get()):
            return messagebox.showerror("Erro", "Para STPP/RMR, ambos os arquivos TXT são obrigatórios.", parent=self)
        if self.process_type != 'stpp_rmr' and not self.caminho_txt_perm_var.get():
            return messagebox.showerror("Erro", "O arquivo TXT de entrada é obrigatório.", parent=self)
        if not self.cod_linhas_var.get().strip(): return messagebox.showerror("Erro", "'Códigos das Linhas' é obrigatório.", parent=self)
        if not self.empresas_var.get().strip(): return messagebox.showerror("Erro", "'Empresa(s)' é obrigatório.", parent=self)
        if not self.lista_periodos_objs: return messagebox.showerror("Erro", "Adicione pelo menos um período de análise.", parent=self)
        if not self.num_ouvidoria_var.get().strip(): return messagebox.showerror("Erro", "'Nº Ouvidoria/ID' é obrigatório.", parent=self)
        
        empresas_input_str = self.empresas_var.get().strip().upper()
        if empresas_input_str == "TODAS":
            empresas_validadas = self.empresas_validas
        else:
            empresas_lista_input = [e.strip() for e in empresas_input_str.split(',')]
            invalidas = [e for e in empresas_lista_input if e not in self.empresas_validas]
            if invalidas: return messagebox.showerror("Erro", f"Empresa(s) inválida(s): {', '.join(invalidas)}", parent=self)
            empresas_validadas = empresas_lista_input
        
        nome_arquivo = self.num_ouvidoria_var.get().strip() + '.xlsx'
        
        self.config_processamento = {
            "base_titulo_excel": self.titulo_base_excel,
            "cod_linhas_str": self.cod_linhas_var.get().strip(),
            "periodos_selecionados_objs": self.lista_periodos_objs,
            "empresas_selecionadas": empresas_validadas,
            "nome_arquivo_excel": os.path.join(self.local_salvar_var.get(), nome_arquivo),
        }
        
        self.botao_gerar.config(state="disabled", text="Processando...")
        self.botao_limpar.config(state="disabled")
        self.botao_voltar.config(state="disabled")
        self.update_idletasks()
        
        threading.Thread(target=executar_processamento_principal, args=(self,)).start()

    def finalizar_processamento(self, mensagem, sucesso):
        self.botao_gerar.config(state="normal", text="Processar e Gerar Relatório")
        self.botao_limpar.config(state="normal")
        self.botao_voltar.config(state="normal")

        if sucesso:
            if messagebox.askyesno("Sucesso!", f"Relatório gerado com sucesso em:\n{mensagem}\n\nDeseja abrir a pasta do arquivo?", parent=self):
                self.abrir_pasta_destino(os.path.dirname(mensagem))
        else:
            messagebox.showerror("Erro no Processamento", f"Ocorreu um erro inesperado.\n\nDetalhes: {mensagem}", parent=self)

    def abrir_pasta_destino(self, pasta):
        if not pasta: return
        try:
            if sys.platform == "win32": os.startfile(pasta)
            elif sys.platform == "darwin": subprocess.run(["open", pasta])
            else: subprocess.run(["xdg-open", pasta])
        except Exception as e:
            messagebox.showwarning("Aviso", f"Não foi possível abrir a pasta.\nCaminho: {pasta}\nErro: {e}", parent=self)

    def center_window(self):
        self.update_idletasks()
        width, height = self.winfo_width(), self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')

def selecionar_tipo_processamento_gui():
    escolha = {"tipo": None}
    
    janela_selecao = tk.Tk()
    janela_selecao.title("Selecionar Tipo")
    janela_selecao.geometry("400x350")
    janela_selecao.resizable(False, False)
    sv_ttk.set_theme("light")
    
    def on_select(tipo_selecionado):
        escolha["tipo"] = tipo_selecionado
        janela_selecao.destroy()

    janela_selecao.update_idletasks()
    x = (janela_selecao.winfo_screenwidth() // 2) - (janela_selecao.winfo_width() // 2)
    y = (janela_selecao.winfo_screenheight() // 2) - (janela_selecao.winfo_height() // 2)
    janela_selecao.geometry(f'+{x}+{y}')

    main_frame = ttk.Frame(janela_selecao, padding=25)
    main_frame.pack(fill=tk.BOTH, expand=True)
    
    ttk.Label(main_frame, text="Assistente de Demanda", font=('Segoe UI', 16, 'bold')).pack(pady=(0, 5))
    ttk.Label(main_frame, text="Selecione o tipo de dados para processar:", font=('Segoe UI', 10)).pack(pady=(0, 25))
    
    ttk.Button(main_frame, text="Permissionárias", style='Accent.TButton', command=lambda: on_select("permissionarias")).pack(fill=tk.X, pady=7, ipady=8)
    ttk.Button(main_frame, text="Concessionárias", style='Accent.TButton', command=lambda: on_select("concessionarias")).pack(fill=tk.X, pady=7, ipady=8)
    ttk.Button(main_frame, text="STPP/RMR (Combinado)", style='Accent.TButton', command=lambda: on_select("stpp_rmr")).pack(fill=tk.X, pady=7, ipady=8)
    
    janela_selecao.protocol("WM_DELETE_WINDOW", lambda: on_select(None))
    janela_selecao.mainloop()
    return escolha["tipo"]


def iniciar_aplicacao():
    """
    Controla o fluxo da aplicação, permitindo voltar para a tela de seleção.
    """
    while True:
        tipo_escolhido = selecionar_tipo_processamento_gui()

        if not tipo_escolhido:
            print("\nOperação cancelada. Programa encerrado.")
            break

        app = DemandaLinhaGUI(process_type=tipo_escolhido)
        app.mainloop()

        if app.voltar_solicitado:
            continue
        else:
            print("\n--- ASSISTENTE ENCERRADO ---")
            break


# --- Ponto de Entrada Principal ---
if __name__ == "__main__":
    # Verifica dependências
    try:
        from tkinterdnd2 import DND_FILES, TkinterDnD
        from PIL import Image, ImageTk
        import sv_ttk
    except ImportError as e:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Dependência Faltando", f"Uma biblioteca necessária não foi encontrada: {e.name}.\n\nPor favor, instale as dependências com:\npip install sv-ttk tkinterdnd2 Pillow")
        sys.exit(1)

    # Inicia o loop principal da aplicação
    iniciar_aplicacao()