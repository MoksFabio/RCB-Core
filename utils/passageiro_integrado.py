# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import filedialog, messagebox, PhotoImage, ttk
import pandas as pd
import openpyxl
import os
import sys
import subprocess
import traceback
import re
from datetime import datetime
import tempfile
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from PIL import Image, ImageTk
import sv_ttk
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, BaseDocTemplate, Frame, PageTemplate
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader

# --- DEFINIÇÃO DO CAMINHO BASE DO PROJETO ---
try:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    BASE_DIR = os.path.dirname(script_dir)
except NameError:
    BASE_DIR = os.getcwd()

# --- CONSTANTES E MAPEAMENTOS GLOBAIS ---

operator_mapping = {
    'BOA': 'BORBOREMA IMPERIAL TRANSPORTES LTDA',
    'CAX': 'RODOVIÁRIA CAXANGÁ S/A',
    'EME': 'EMPRESA METROPOLITANA S/A',
    'CNO': 'CONORTE',
    'CSR': 'CONSÓRCIO RECIFE DE TRANSPORTE',
    'GLO': 'TRANSPORTADORA GLOBO LTDA',
    'MOB': 'MOBI BRASIL',
    'VML': 'VIAÇÃO MIRIM LTDA',
    'SJT': 'AUTO VIAÇÃO SÃO JUDAS TADEU',
}

month_mapping = {
    '01': 'Janeiro', '02': 'Fevereiro', '03': 'Março', '04': 'Abril',
    '05': 'Maio', '06': 'Junho', '07': 'Julho', '08': 'Agosto',
    '09': 'Setembro', '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro'
}

# --- FUNÇÕES DE LÓGICA (DO SCRIPT 1 - TXT -> EXCEL) ---

def ajustar_largura_coluna_automaticamente(sheet):
    dim_holder = DimensionHolder(worksheet=sheet)
    max_rows_to_scan = sheet.max_row
    max_cols_to_scan = sheet.max_column

    if max_rows_to_scan == 0 or max_cols_to_scan == 0:
        return

    for col_idx in range(1, max_cols_to_scan + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        rows_to_scan = min(max_rows_to_scan, 500)
        indices_to_scan = list(range(1, min(rows_to_scan + 1, max_rows_to_scan + 1)))
        if max_rows_to_scan > rows_to_scan:
            if max_rows_to_scan not in indices_to_scan:
                indices_to_scan.append(max_rows_to_scan)

        for row_idx in indices_to_scan:
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell is None or cell.value is None: continue
            try:
                is_merged = False
                merged_value = None
                for merged_range_str in sheet.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = merged_range_str.bounds
                    if min_col <= col_idx <= max_col and min_row <= row_idx <= max_row:
                        is_merged = True
                        if row_idx == min_row and col_idx == min_col:
                            top_left_cell = sheet.cell(row=min_row, column=min_col)
                            merged_value = top_left_cell.value
                            if merged_value is not None:
                                cell_length = len(str(merged_value))
                                max_length = max(max_length, cell_length)
                        break

                if not is_merged:
                    value = cell.value
                    if value is not None:
                        if isinstance(value, datetime):
                            try: cell_length = len(value.strftime('%d/%m/%Y %H:%M:%S'))
                            except ValueError: cell_length = len(value.strftime('%d/%m/%Y'))
                        else: cell_length = len(str(value))
                        max_length = max(max_length, cell_length)
            except Exception: pass

        adjusted_width = max(5, max_length + 3)
        adjusted_width = min(50, adjusted_width)

        try:
            if col_letter not in dim_holder:
                dim_holder[col_letter] = ColumnDimension(sheet, min=col_idx, max=col_idx, width=adjusted_width)
            else:
                dim_holder[col_letter].width = adjusted_width
        except Exception:
            pass

    sheet.column_dimensions = dim_holder

def extrair_codigo_linha(descricao):
    if not isinstance(descricao, str):
        return None
    match = re.match(r'^\s*(\d+)\s*-', descricao)
    if match:
        return match.group(1)
    return None

def aplicar_formatacao_excel(arquivo_xlsx, df_original_context, numero_arquivo, add_custom_header=False, custom_header_text=None, period_header_text=None):
    try:
        workbook = load_workbook(arquivo_xlsx)
        sheet = workbook.active
        secondary_merged_coords = set()

        if sheet.max_row is None or sheet.max_row <= 1:
            is_effectively_empty = True
            if sheet.max_row == 1 and sheet.max_column > 0:
                if any(sheet.cell(row=1, column=c).value is not None for c in range(1, sheet.max_column + 1)):
                    is_effectively_empty = False
            if is_effectively_empty:
                try:
                    workbook.save(arquivo_xlsx)
                except Exception:
                    pass
                return True

        header_row_idx = 1
        data_start_row_idx = 2
        col_n_idx_original = 14
        col_n_idx = col_n_idx_original
        cols_to_insert = 0

        if numero_arquivo == 2 and add_custom_header:
            header_row_idx = 5
            data_start_row_idx = 6
            col_n_idx = col_n_idx_original + cols_to_insert

        try:
            secondary_merged_coords.clear()
            for merged_range in sheet.merged_cells.ranges:
                if merged_range.max_row > merged_range.min_row or merged_range.max_col > merged_range.min_col:
                    for row in range(merged_range.min_row, merged_range.max_row + 1):
                        for col in range(merged_range.min_col, merged_range.max_col + 1):
                            if not (row == merged_range.min_row and col == merged_range.min_col):
                                try: secondary_merged_coords.add(sheet.cell(row=row, column=col).coordinate)
                                except IndexError: pass
        except Exception:
            pass

        if numero_arquivo == 2 and add_custom_header:
            try:
                text_l1_a = "Grande Recife Consórcio de Transporte"; text_l2_a = "Sistema de Remuneração, Custos e Bilhetagem"; text_l4_a = custom_header_text
                cell_a1 = sheet.cell(row=1, column=1); cell_a2 = sheet.cell(row=2, column=1); cell_a4 = sheet.cell(row=4, column=1)
                cell_a1.value = text_l1_a; cell_a2.value = text_l2_a; cell_a4.value = text_l4_a
                bold_font = Font(bold=True, size=11)
                cell_a1.font = bold_font; cell_a2.font = bold_font; cell_a4.font = bold_font

                text_l1_n = "Relatório de Passageiros Integrados"; text_l2_n = period_header_text
                try:
                    if sheet.max_column >= col_n_idx:
                        cell_n1 = sheet.cell(row=1, column=col_n_idx); cell_n2 = sheet.cell(row=2, column=col_n_idx)
                        cell_n1.value = text_l1_n; cell_n1.font = bold_font
                        if text_l2_n: cell_n2.value = text_l2_n; cell_n2.font = bold_font
                    else:
                        pass
                except Exception:
                    pass

            except Exception:
                traceback.print_exc()

        is_df_empty = (df_original_context is None or df_original_context.empty)
        if is_df_empty and not (numero_arquivo == 2 and add_custom_header):
            larguras_desejadas_vazio = {}
            if numero_arquivo == 1:
                larguras_desejadas_vazio = { 'A': 12, 'B': 10, 'C': 12, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15, 'I': 15 }
            elif numero_arquivo == 2:
                larguras_desejadas_vazio = { 'A': 12, 'B': 8, 'C': 8, 'D': 10, 'E': 20, 'F': 9.29 }

            if sheet.max_column > 0:
                for col_letra, largura in larguras_desejadas_vazio.items():
                    try:
                        col_idx_test = column_index_from_string(col_letra)
                        if col_idx_test <= sheet.max_column:
                            if col_letra not in sheet.column_dimensions: sheet.column_dimensions[col_letra] = ColumnDimension(sheet, min=col_idx_test, max=col_idx_test, width=largura)
                            else: sheet.column_dimensions[col_letra].width = largura
                    except Exception:
                        pass
            else: pass

            try: workbook.save(arquivo_xlsx)
            except Exception: return False
            return True

        colunas_para_centralizar_nomes = []
        colunas_para_centralizar_indices = set()

        if numero_arquivo == 1:
            colunas_para_centralizar_nomes = ['CDOPERADOR', 'CDLINHA', 'NMINTEGESTUDANTE', 'NMINTEGESTUDANTEPCR', 'NMINTEGVALE', 'NMINTEGVALECOMUM', 'NMINTEGPASSELIVRE', 'TOTAL']
        elif numero_arquivo == 2:
            if add_custom_header:
                start_col_align_idx = 1 + cols_to_insert + 1
                if sheet.max_column >= start_col_align_idx:
                    colunas_para_centralizar_nomes = [get_column_letter(i) for i in range(start_col_align_idx, sheet.max_column + 1)]
                else:
                    colunas_para_centralizar_nomes = []
            else:
                colunas_para_centralizar_nomes = ['A', 'B', 'C', 'D', 'F']

        sheet_header = {}
        if header_row_idx <= sheet.max_row:
            try:
                if numero_arquivo == 2 and add_custom_header:
                    if df_original_context is not None and not df_original_context.empty:
                        if len(df_original_context.columns) > 0:
                            sheet_header[df_original_context.columns[0]] = 1

                df_col_index = 1
                start_data_col_idx = 1 + cols_to_insert + 1
                for col_idx in range(start_data_col_idx, sheet.max_column + 1):
                    if df_original_context is not None and df_col_index < len(df_original_context.columns):
                        cell_value = sheet.cell(row=header_row_idx, column=col_idx).value
                        original_df_col_name = df_original_context.columns[df_col_index]
                        header_key = str(cell_value).strip() if cell_value else original_df_col_name
                        sheet_header[header_key] = col_idx
                        sheet_header[original_df_col_name] = col_idx
                        df_col_index += 1
                else:
                    for col_idx in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=header_row_idx, column=col_idx)
                        if cell.value is not None:
                            header_text = str(cell.value).strip()
                            sheet_header[header_text] = col_idx
            except Exception:
                pass

        for nome_ou_letra in colunas_para_centralizar_nomes:
            idx = -1
            try:
                if nome_ou_letra.isalpha() and len(nome_ou_letra) <= 3:
                    idx = column_index_from_string(nome_ou_letra)
                elif nome_ou_letra in sheet_header:
                    idx = sheet_header[nome_ou_letra]
                else:
                    found = False
                    for header_key, header_idx in sheet_header.items():
                        if str(header_key).upper() == nome_ou_letra.upper():
                            idx = header_idx; found = True; break
                    if not found:
                        continue

                if idx != -1 and 1 <= idx <= sheet.max_column:
                    colunas_para_centralizar_indices.add(idx)
                elif idx != -1: pass

            except ValueError:
                pass
            except Exception:
                pass

        if not colunas_para_centralizar_indices: pass
        else:
            indices_ordenados = sorted(list(colunas_para_centralizar_indices))
            letras_formatar = [get_column_letter(i) for i in indices_ordenados]
            center_alignment = Alignment(horizontal='center', vertical='center')

            first_align_error_logged = False
            for row_idx_align in range(data_start_row_idx, sheet.max_row + 1):
                for col_idx_align in colunas_para_centralizar_indices:
                    try:
                        cell = sheet.cell(row=row_idx_align, column=col_idx_align)
                        if cell.coordinate not in secondary_merged_coords:
                            cell.alignment = center_alignment
                    except IndexError: continue
                    except Exception:
                        if not first_align_error_logged: first_align_error_logged = True

            if header_row_idx <= sheet.max_row:
                for col_idx_align in colunas_para_centralizar_indices:
                    try:
                        cell = sheet.cell(row=header_row_idx, column=col_idx_align)
                        if cell.coordinate not in secondary_merged_coords:
                            cell.alignment = center_alignment
                    except Exception: pass

        header_font_no_bold = Font(bold=False, size=11)
        if header_row_idx <= sheet.max_row:
            try:
                for cell in sheet[header_row_idx]:
                    if cell.value is not None:
                        is_custom_header_A_cell = (numero_arquivo == 2 and add_custom_header and cell.column == 1)
                        if not is_custom_header_A_cell and cell.coordinate not in secondary_merged_coords:
                            try: cell.font = header_font_no_bold
                            except Exception:
                                pass
            except Exception:
                pass
        else: pass

        if numero_arquivo == 1:
            if df_original_context is not None and not df_original_context.empty:
                date_format = 'DD/MM/YYYY'; dt_col_name = 'DTOPERACAO'; dt_col_idx = -1
                if dt_col_name in sheet_header: dt_col_idx = sheet_header[dt_col_name]

                if dt_col_idx != -1 and dt_col_idx <= sheet.max_column:
                    dt_col_letter = get_column_letter(dt_col_idx)
                    for row_idx in range(data_start_row_idx, sheet.max_row + 1):
                        cell = sheet.cell(row=row_idx, column=dt_col_idx)
                        if cell.value is not None and not isinstance(cell.value, str) and cell.coordinate not in secondary_merged_coords:
                            try: cell.number_format = date_format
                            except Exception:
                                pass
                    try:
                        header_cell = sheet.cell(row=header_row_idx, column=dt_col_idx)
                        if header_cell.coordinate not in secondary_merged_coords: header_cell.number_format = date_format
                    except: pass
                elif dt_col_name in df_original_context.columns: pass
                else: pass
            else: pass

        ajustar_largura_coluna_automaticamente(sheet)

        larguras_desejadas_finais = {}
        total_col_letter_final = None; total_col_idx_final = -1
        total_col_letter_adjusted = None; total_col_idx_adjusted = -1

        if numero_arquivo == 2:
            if add_custom_header:
                if df_original_context is not None and 'TOTAL' in df_original_context.columns:
                    try:
                        total_col_idx_final = df_original_context.columns.get_loc('TOTAL') + 1
                        if total_col_idx_final > 0 :
                            total_col_letter_final = get_column_letter(total_col_idx_final)
                            total_col_idx_adjusted = total_col_idx_final + cols_to_insert
                            total_col_letter_adjusted = get_column_letter(total_col_idx_adjusted)
                        else: total_col_idx_final = -1
                    except KeyError: total_col_idx_final = -1

                larguras_desejadas_finais['A'] = 40.86

                start_range_dias_idx = 1 + cols_to_insert + 1
                fim_range_dias_idx = sheet.max_column
                if total_col_idx_adjusted != -1 and total_col_idx_adjusted <= sheet.max_column:
                    fim_range_dias_idx = total_col_idx_adjusted - 1

                if fim_range_dias_idx >= start_range_dias_idx:
                    col_letra_inicio_dias = get_column_letter(start_range_dias_idx)
                    col_letra_fim_dias = get_column_letter(fim_range_dias_idx)
                    for i in range(start_range_dias_idx, fim_range_dias_idx + 1):
                        larguras_desejadas_finais[get_column_letter(i)] = 7
                else: pass

                if total_col_letter_adjusted and total_col_idx_adjusted <= sheet.max_column:
                    larguras_desejadas_finais[total_col_letter_adjusted] = 15

            else:
                larguras_desejadas_finais = { 'A': 12, 'B': 8, 'C': 8, 'D': 10, 'E': 20, 'F': 9.29 }

            if larguras_desejadas_finais:
                for col_letra, largura in larguras_desejadas_finais.items():
                    try:
                        col_idx_test = column_index_from_string(col_letra)
                        if col_idx_test <= sheet.max_column:
                            if col_letra not in sheet.column_dimensions: sheet.column_dimensions[col_letra] = ColumnDimension(sheet, min=col_idx_test, max=col_idx_test, width=largura)
                            else: sheet.column_dimensions[col_letra].width = largura
                    except Exception:
                        pass
            else: pass

            coluna_n_letra_fisica = 'N'; coluna_n_indice_fisico = 14; largura_n_desejada = 6.29
            try:
                is_col_n_fisica_total_ajustada = (add_custom_header and total_col_letter_adjusted == coluna_n_letra_fisica)
                if sheet.max_column >= coluna_n_indice_fisico and not is_col_n_fisica_total_ajustada:
                    if coluna_n_letra_fisica not in sheet.column_dimensions: sheet.column_dimensions[coluna_n_letra_fisica] = ColumnDimension(sheet, min=coluna_n_indice_fisico, max=coluna_n_indice_fisico, width=largura_n_desejada)
                    else: sheet.column_dimensions[coluna_n_letra_fisica].width = largura_n_desejada
                elif is_col_n_fisica_total_ajustada: pass
                elif sheet.max_column < coluna_n_indice_fisico: pass
            except Exception:
                pass

        elif numero_arquivo == 1:
            larguras_arq1_final = { 'A': 12, 'B': 10, 'C': 12, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15, 'I': 15 }
            for col_letra, largura in larguras_arq1_final.items():
                try:
                    col_idx_test = column_index_from_string(col_letra)
                    if col_idx_test <= sheet.max_column:
                        if col_letra not in sheet.column_dimensions: sheet.column_dimensions[col_letra] = ColumnDimension(sheet, min=col_idx_test, max=col_idx_test, width=largura)
                        else: sheet.column_dimensions[col_letra].width = largura
                except Exception:
                    pass

        if numero_arquivo == 2 and df_original_context is not None:
            last_row_label_df = None
            if not df_original_context.empty:
                try:
                    if len(df_original_context.columns) > 0:
                        last_row_label_df = df_original_context.iloc[-1, 0]
                    else:
                        pass
                except IndexError: pass
                except Exception:
                    pass

            if last_row_label_df == "Total Geral":
                last_row_idx_sheet = sheet.max_row
                if last_row_idx_sheet >= data_start_row_idx:
                    bold_font_total = Font(bold=True, size=11)
                    try:
                        for cell in sheet[last_row_idx_sheet]:
                            if cell.value is not None and cell.coordinate not in secondary_merged_coords:
                                cell.font = bold_font_total
                    except Exception:
                        pass
                else: pass
            elif df_original_context is not None and not df_original_context.empty :
                pass

        workbook.save(arquivo_xlsx)
        return True

    except FileNotFoundError:
        return False
    except Exception:
        traceback.print_exc()
        try:
            if 'workbook' in locals() and workbook is not None:
                workbook.save(arquivo_xlsx)
            else: pass
        except Exception:
            pass
        return False

def processar_efetivo(arquivo_txt):
    df = None
    encoding_esperado = 'utf-8'; read_header_idx = 0
    colunas_somar_concessionarias = ['NMINTEGESTUDANTE', 'NMINTEGESTUDANTEPCR', 'NMINTEGVALE', 'NMINTEGVALECOMUM', 'NMINTEGPASSELIVRE']
    colunas_somar_empresas = ['NMINTEGESTUDANTE', 'NMINTEGESTUDANTEPCR', 'NMINTEGVALE', 'NMINTEGVALECOMUM', 'NMINTEGPASSELIVRE']
    colunas_para_somar = list(set(colunas_somar_concessionarias + colunas_somar_empresas))
    
    colunas_para_manter = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO'] + colunas_para_somar
    
    expected_date_format = '%d/%m/%Y'

    operadoras_validas = list(operator_mapping.keys())

    try:
        common_params = {'sep': r'\s*[\t;]\s*', 'engine': 'python', 'quotechar': '"', 'skip_blank_lines': True, 'on_bad_lines': 'warn', 'dtype': str}
        try:
            df = pd.read_csv(arquivo_txt, encoding=encoding_esperado, header=read_header_idx, **common_params)
        except UnicodeDecodeError:
            encoding_esperado = 'latin-1'
            df = pd.read_csv(arquivo_txt, encoding=encoding_esperado, header=read_header_idx, **common_params)
        except pd.errors.EmptyDataError:
            return pd.DataFrame()
        except Exception:
            traceback.print_exc(); return None

        if df is not None and not df.empty:
            df.columns = df.columns.str.strip(); df.dropna(how='all', inplace=True); df.replace('', pd.NA, inplace=True); df.dropna(how='all', inplace=True)
            if df.empty: return pd.DataFrame()
        elif df is not None and df.empty: return pd.DataFrame()
        else: return None

        colunas_existentes_orig = df.columns.tolist()
        colunas_realmente_manter = [col for col in colunas_para_manter if col in colunas_existentes_orig]
        colunas_faltantes = [col for col in colunas_para_manter if col not in colunas_realmente_manter]
        if colunas_faltantes: pass
        if 'DTOPERACAO' not in colunas_realmente_manter: return None
        if not any(col in colunas_realmente_manter for col in colunas_para_somar): pass

        df = df[colunas_realmente_manter].copy()

        if 'CDOPERADOR' in df.columns:
            linhas_antes = len(df)
            df = df[df['CDOPERADOR'].isin(operadoras_validas)].copy()
            linhas_depois = len(df)
        else:
            pass

        if 'DTOPERACAO' in df.columns:
            try:
                df['DTOPERACAO_cleaned'] = df['DTOPERACAO'].astype(str).str.strip().str.lstrip("'")
                df['DTOPERACAO'] = pd.to_datetime(df['DTOPERACAO_cleaned'], format=expected_date_format, errors='coerce')
                nat_count = df['DTOPERACAO'].isnull().sum(); df.drop(columns=['DTOPERACAO_cleaned'], inplace=True, errors='ignore')
                if nat_count > 0: pass
            except Exception:
                pass


        colunas_numericas_presentes = [col for col in colunas_para_somar if col in df.columns]
        if not colunas_numericas_presentes: pass
        else:
            for col in colunas_numericas_presentes:
                try:
                    series_limpa = df[col].astype(str).str.strip().str.lstrip("'").str.replace(',', '.', regex=False)
                    df[col] = pd.to_numeric(series_limpa, errors='coerce')
                except Exception:
                    pass

        # --- MODIFICAÇÃO 3: Lógica de soma condicional ---
        # Identifica as colunas que realmente existem no DataFrame para cada grupo
        cols_sum_concessionarias_actual = [col for col in colunas_somar_concessionarias if col in df.columns]
        cols_sum_empresas_actual = [col for col in colunas_somar_empresas if col in df.columns]

        # Inicializa a coluna TOTAL
        df['TOTAL'] = 0.0 

        if 'CDOPERADOR' not in df.columns:
            # Se a coluna CDOPERADOR não existir, aplica a soma padrão (das empresas)
            if cols_sum_empresas_actual:
                 df['TOTAL'] = df[cols_sum_empresas_actual].sum(axis=1, skipna=True)
        else:
            # Lógica condicional
            # Máscara para CNO e MOB (Concessionárias)
            is_concessionaria = df['CDOPERADOR'].isin(['CNO', 'MOB'])
            # Máscara para todas as outras (Empresas)
            is_empresa = ~is_concessionaria
            
            # Calcula o total para CNO e MOB (Concessionárias)
            if cols_sum_concessionarias_actual:
                df.loc[is_concessionaria, 'TOTAL'] = df.loc[is_concessionaria, cols_sum_concessionarias_actual].sum(axis=1, skipna=True)
                
            # Calcula o total para as outras (Empresas)
            if cols_sum_empresas_actual:
                df.loc[is_empresa, 'TOTAL'] = df.loc[is_empresa, cols_sum_empresas_actual].sum(axis=1, skipna=True)
        
        # O restante do código para reordenar as colunas permanece o mesmo
        if 'TOTAL' in df.columns:
            final_column_order = [c for c in colunas_realmente_manter if c in df.columns] + ['TOTAL']
            final_column_order_existing = [c for c in final_column_order if c in df.columns]
            df = df[final_column_order_existing]
        else: pass

        return df

    except Exception:
        traceback.print_exc(); return None

def criar_mapa_totais_efetivo(df_efetivo):
    mapa_totais = {}
    required_cols = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO', 'TOTAL']

    if df_efetivo is None or df_efetivo.empty: return mapa_totais
    if not all(col in df_efetivo.columns for col in required_cols): return mapa_totais

    erros_data = 0; linhas_processadas = 0; chaves_criadas = 0
    for index, row in df_efetivo.iterrows():
        linhas_processadas += 1
        try:
            data = row['DTOPERACAO']
            if pd.isna(data):
                erros_data += 1
                continue
            dia = data.day
            operador = str(row['CDOPERADOR']).strip()
            linha = str(row['CDLINHA']).strip()
            total = row['TOTAL']
            if operador and linha:
                chave = (operador, linha, dia)
                mapa_totais[chave] = total
                chaves_criadas += 1
        except AttributeError:
            erros_data += 1
            continue
        except Exception:
            continue

    if erros_data > 0: pass
    if linhas_processadas > 0 and chaves_criadas == 0: pass
    return mapa_totais

def processar_viagem_frota(arquivo_txt, operator_mapping, month_mapping):
    df = None; encoding_esperado = 'latin-1'; read_header_idx = 1; add_custom_header = False; custom_header_text = None; period_header_text = None; sigla_arquivo = None

    nome_base = os.path.basename(arquivo_txt)
    if len(nome_base) >= 3:
        sigla_arquivo = nome_base[:3].upper()
        if sigla_arquivo in operator_mapping:
            company_name = operator_mapping[sigla_arquivo]
            custom_header_text = f"[{sigla_arquivo}] {company_name}"
            add_custom_header = True
            read_header_idx = 1
            match = re.search(r'^[A-Z]{3}[-_]?(\d{2})(\d{2})(\d{2})[-_]?', nome_base, re.IGNORECASE)
            if match:
                yy, mm, qq = match.groups()
                try:
                    year_full = int(yy)
                    current_century = datetime.now().year // 100
                    year_prefix = current_century if year_full <= (datetime.now().year % 100 + 10) else current_century - 1
                    year_str = str(year_prefix * 100 + year_full)
                    month_name = month_mapping.get(mm)
                    fortnight_str = {"01": "1ª", "02": "2ª"}.get(qq)
                    if month_name and fortnight_str:
                        period_header_text = f"{fortnight_str} Quinzena de {month_name} de {year_str}"
                    else: pass
                except ValueError: pass
                except Exception: pass
            else: pass
        else:
            add_custom_header = False; read_header_idx = 0
    else:
        add_custom_header = False; read_header_idx = 0

    common_params = {'sep': r'\s*[\t;]\s*', 'engine': 'python', 'quotechar': '"', 'skip_blank_lines': True, 'on_bad_lines': 'warn', 'dtype': str}
    try:
        df = pd.read_csv(arquivo_txt, encoding=encoding_esperado, header=read_header_idx, **common_params)
    except UnicodeDecodeError:
        encoding_esperado = 'utf-8'
        try: df = pd.read_csv(arquivo_txt, encoding=encoding_esperado, header=read_header_idx, **common_params)
        except Exception: traceback.print_exc(); return None, False, None, None, None
    except pd.errors.EmptyDataError:
        return pd.DataFrame(), add_custom_header, custom_header_text, period_header_text, sigla_arquivo
    except Exception:
        traceback.print_exc(); return None, False, None, None, None

    if df is not None and not df.empty:
        df.columns = df.columns.str.strip().str.lstrip("'"); df.dropna(how='all', inplace=True); df.replace('', pd.NA, inplace=True); df.dropna(how='all', inplace=True)
        if df.empty: return pd.DataFrame(), add_custom_header, custom_header_text, period_header_text, sigla_arquivo

        if add_custom_header:
            colunas_removidas_log = []
            if 'Operador' in df.columns:
                try: df = df.drop(columns=['Operador'], errors='ignore'); colunas_removidas_log.append('Operador');
                except Exception: pass
            cols_to_remove_fixed = ['DUT', 'SAB', 'DOM', 'Descrição']
            found_to_remove = [col for col in cols_to_remove_fixed if col in df.columns]
            if found_to_remove:
                try: df = df.drop(columns=found_to_remove, errors='ignore'); colunas_removidas_log.extend(found_to_remove);
                except Exception: pass

            if len(df.columns) >= 3:
                cols_to_drop_first_3 = df.columns[:3].tolist()
                try: df = df.drop(columns=cols_to_drop_first_3, errors='ignore'); colunas_removidas_log.extend(cols_to_drop_first_3);
                except Exception:
                    pass
            else: pass

            if colunas_removidas_log: pass
            else: pass

            if len(df.columns) > 1:
                cols_to_clear_names = df.columns[1:].tolist()
                for col_name in cols_to_clear_names: df[col_name] = pd.NA
            else: pass

            if len(df.columns) > 0:
                original_first_col_name = df.columns[0]
                if original_first_col_name != 'Linha': df.rename(columns={original_first_col_name: 'Linha'}, inplace=True)
                original_rows = len(df); df['Linha'] = df['Linha'].astype(str).str.strip(); df.drop_duplicates(subset=['Linha'], keep='first', inplace=True); removed_rows = original_rows - len(df)
                if removed_rows > 0: pass
                else: pass
                original_rows_filter = len(df); col_as_str = df['Linha'].astype(str).str.strip(); condition1 = col_as_str.ne('Descrição'); condition2 = ~col_as_str.str.startswith('0 -', na=False)
                df = df[condition1 & condition2]; removed_rows_filter = original_rows_filter - len(df)
                if removed_rows_filter > 0: pass
                else: pass
            else: pass
        else:
            pass

        target_col_name = None
        if not df.empty and len(df.columns) > 0:
            if 'Linha' in df.columns:
                target_col_name = 'Linha'
            else:
                target_col_name = df.columns[0]

        if target_col_name and target_col_name in df.columns:
            correcoes_ortograficas = {
                'HEMET\x90RIO': 'HEMETÉRIO', 'JOS\x90': 'JOSÉ', 'BONIFµCIO': 'BONIFÁCIO',
                'µGUA': 'ÁGUA', 'µGUAS': 'ÁGUAS', "D'µGUA": "D'ÁGUA", 'TI XAMBµ': 'TI XAMBÁ',
                'CABUGµ': 'CABUGÁ', 'CAXANGµ': 'CAXANGÁ', 'AERONµUTICA': 'AERONÁUTICA',
                'BEL\x90M': 'BELÉM', 'PRÖNCIPE': 'PRÍNCIPE', 'GETéLIO': 'GETÚLIO',
                'SÖTIO': 'SÍTIO', 'BRASÇ\x8dLIA': 'BRASÍLIA', 'SETéBAL': 'SETÚBAL',
                'INTEGRA\x80¶O': 'INTEGRAÇÃO', 'CONCEI\x80ÇO': 'CONCEIÇÃO',
                'BONAN\x80A': 'BONANÇA', 'JaboatÆo': 'Jaboatão', '(PR\x90-EMBARQUE)': '(PRÉ-EMBARQUE)',
                '(Pr\x82-Embarque)': '(Pré-Embarque)', 'Pr¡ncipe': 'Príncipe', 'TORRåES': 'TORRÕES',
                'TRÒS': 'TRÊS', '(N. SRA. DO à)': '(N. SRA. DO Ó)', '(N. Sra. do à)': '(N. Sra. do Ó)',
                'CHÇO': 'CHÃO', 'CAPITÇO': 'CAPITÃO', 'CORREGO': 'CÓRREGO', 'CàRREGO': 'CÓRREGO',
                'SETUBAL': 'SETÚBAL', 'CONJ. CATAMARA': 'CONJ. CATAMARÃ', 'JORDÇO': 'JORDÃO',
                'JABOATÇO': 'JABOATÃO', 'TOTà': 'TOTÓ', 'INµCIO': 'INÍCIO',
                'DOIS IRMÇOS': 'DOIS IRMÃOS', 'Dois IrmÆos': 'Dois Irmãos', 'REFéGIO': 'REFÚGIO',
                'SÇO PAULO': 'SÃO PAULO', 'GETULIO VARGAS': 'GETÚLIO VARGAS',
                'JOÇO DE BARROS': 'JOÃO DE BARROS', 'MARACANÇ': 'MARACANÃ',
                'PTE. DOS CARVALHOS': 'PONTE DOS CARVALHOS',
                '(V rzea)': '(Várzea)', 'TAMANDARÉ■': 'TAMANDARÉ', 'TAMANDAR': 'TAMANDARÉ',
                '(Cruz Cabug )': '(Cruz Cabugá)', 'Cabug ': 'Cabugá',
                'ESPERAN A': 'ESPERANÇA', 'Col nia': 'Colônia',
                '(COR. DO JOAQUIM)': '(CÓRREGO DO JOAQUIM)'
            }

            replaced_count = 0
            for errado, correto in correcoes_ortograficas.items():
                ocorrencias_antes = df[target_col_name].astype(str).str.contains(errado, regex=False).sum()
                if ocorrencias_antes > 0:
                    df[target_col_name] = df[target_col_name].astype(str).str.replace(errado, correto, regex=False)
                    replaced_count += ocorrencias_antes
            if replaced_count > 0:
                pass
            else:
                pass
        else:
            pass

    elif df is not None and df.empty: pass
    else: return None, False, None, None, None

    return df, add_custom_header, custom_header_text, period_header_text, sigla_arquivo

def popular_totais_viagem_frota(df_viagem, mapa_totais_efetivo, sigla_operador):
    if df_viagem is None or df_viagem.empty: return df_viagem
    if not mapa_totais_efetivo: return df_viagem
    if not sigla_operador: return df_viagem
    if len(df_viagem.columns) <= 1: return df_viagem

    linha_col_name = df_viagem.columns[0];
    day_cols_names = df_viagem.columns[1:].tolist();
    day_cols_map = {}

    for col_name in day_cols_names:
        try: day_cols_map[col_name] = int(col_name);
        except (ValueError, TypeError): pass

    if not day_cols_map: return df_viagem

    updates_count = 0; errors_count = 0; keys_not_found = 0
    for row in df_viagem.itertuples(index=True):
        idx = row.Index;
        descricao_linha = getattr(row, linha_col_name, None);
        codigo_linha = extrair_codigo_linha(descricao_linha)

        if codigo_linha is None: continue

        for dia_str, dia_int in day_cols_map.items():
            chave_busca = (sigla_operador, codigo_linha, dia_int)
            if chave_busca in mapa_totais_efetivo:
                total_efetivo = mapa_totais_efetivo[chave_busca]
                try:
                    try: total_efetivo_num = pd.to_numeric(total_efetivo)
                    except (ValueError, TypeError): total_efetivo_num = total_efetivo
                    df_viagem.loc[idx, dia_str] = total_efetivo_num;
                    updates_count += 1
                except Exception:
                    if errors_count < 5: errors_count +=1
            else:
                keys_not_found += 1

    if errors_count > 0: pass
    if updates_count == 0 and not df_viagem.empty and len(mapa_totais_efetivo) > 0: pass

    return df_viagem


# --- FUNÇÕES DE LÓGICA (DO SCRIPT 2 - EXCEL -> PDF) ---

def extrair_infos_relatorio(nome_arquivo_base):
    sigla = None
    nome_operador_completo = "Nome não encontrado"
    periodo_texto = "Período não encontrado"
    cnpj_e_nome_formatado = "Operador não identificado"
    nome_operador_sem_cnpj = "Operador não identificado"
    periodo_codigo_raw = None

    match_sigla = re.match(r'^([A-Z]{3})[-_ ]', nome_arquivo_base, re.IGNORECASE)
    if match_sigla:
        sigla_cand = match_sigla.group(1).upper()
        if sigla_cand in operator_mapping:
            sigla = sigla_cand
            nome_operador_completo = operator_mapping[sigla]
            cnpj_e_nome_formatado = nome_operador_completo
            nome_operador_sem_cnpj = nome_operador_completo
        else:
            sigla = sigla_cand
            nome_operador_completo = f"Operador {sigla_cand} (Não mapeado)"
            cnpj_e_nome_formatado = nome_operador_completo
            nome_operador_sem_cnpj = nome_operador_completo
    else:
        match_sigla_anywhere = re.search(r'([A-Z]{3})', nome_arquivo_base, re.IGNORECASE)
        if match_sigla_anywhere:
            sigla_cand = match_sigla_anywhere.group(1).upper()
            if sigla_cand in operator_mapping:
                sigla = sigla_cand
                nome_operador_completo = operator_mapping[sigla]
                cnpj_e_nome_formatado = nome_operador_completo
                nome_operador_sem_cnpj = nome_operador_completo

    match_periodo = re.search(r'(\d{2})(\d{2})(\d{2})', nome_arquivo_base)
    if match_periodo:
        yy, mm, qq = match_periodo.groups()
        periodo_codigo_raw = f"{yy}{mm}{qq}"
        try:
            year_int = int(yy)
            current_year_last_two_digits = datetime.now().year % 100
            if year_int <= (current_year_last_two_digits + 10) or year_int >= 90 :
                year_str = f"20{yy}" if year_int <= (current_year_last_two_digits + 10) else f"19{yy}"
            else:
                year_str = f"20{yy}"

            month_name = month_mapping.get(mm)
            fortnight_str = {"01": "1a.", "02": "2a."}.get(qq)
            if month_name and fortnight_str:
                periodo_texto = f"{fortnight_str} Quinzena de {month_name} de {year_str}"
            else:
                periodo_texto = f"Mês ({mm}) ou Quinzena ({qq}) inválido(s)"
        except Exception:
            periodo_texto = "Erro na formatação do período"
    else:
        periodo_texto = "Padrão de período (YYMMQQ) não encontrado no nome do arquivo"

    if nome_operador_sem_cnpj == "Operador não identificado" and sigla is None:
        nome_operador_sem_cnpj = "Operador não identificado (sem sigla)"

    return {
        "sigla": sigla,
        "nome_operador_completo": nome_operador_completo,
        "periodo": periodo_texto,
        "cnpj_e_nome": cnpj_e_nome_formatado,
        "nome_operador_sem_cnpj": nome_operador_sem_cnpj,
        "periodo_codigo": periodo_codigo_raw
    }

def get_data_atual_formatada():
    now = datetime.now()
    return now.strftime('%d/%m/%Y')

def format_brazilian_number(x):
    if pd.isna(x):
        return "0"
    try:
        num = float(x)
        if num == int(num):
            return f"{int(num):,}".replace(',', '.')
        else:
            return f"{num:,.2f}".replace(',', 'v').replace('.', ',').replace('v', '.')
    except (ValueError, TypeError):
        return str(x)

def gerar_pdf_final_com_contagem(df_formatado, infos, data_hoje, operador_com_sigla, output_filename, known_total_pages):
    doc = BaseDocTemplate(output_filename,
                          pagesize=landscape(A4),
                          leftMargin=2*cm, rightMargin=2*cm,
                          topMargin=2*cm, bottomMargin=2.5*cm,
                          title=os.path.basename(output_filename).replace('.pdf', ''),
                          author="Grande Recife Consórcio")

    frame_largura = doc.width
    frame_altura = doc.height
    main_frame = Frame(doc.leftMargin, doc.bottomMargin, frame_largura, frame_altura,
                       id='main_frame', leftPadding=0, bottomPadding=0,
                       rightPadding=0, topPadding=0)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='HeaderTitleMain', parent=styles['Normal'], fontSize=12, fontName='Helvetica-Bold', textColor=colors.HexColor('#D95F02'), alignment=TA_LEFT))
    styles.add(ParagraphStyle(name='HeaderSubtitle', parent=styles['Normal'], fontSize=10, fontName='Helvetica', textColor=colors.HexColor('#D95F02'), alignment=TA_LEFT))
    styles.add(ParagraphStyle(name='ReportTitle', parent=styles['Normal'], fontSize=12, fontName='Helvetica', textColor=colors.HexColor('#D95F02'), alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name='PeriodInfo', parent=styles['Normal'], fontSize=10, fontName='Helvetica', textColor=colors.HexColor('#D95F02'), alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name='SubHeader', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', textColor=colors.HexColor('#FD8D3C'), alignment=TA_LEFT, spaceBefore=0, spaceAfter=15))
    styles.add(ParagraphStyle(name='FooterTextLeft', parent=styles['Normal'], fontSize=8, fontName='Helvetica', textColor=colors.HexColor('#D95F02'), alignment=TA_LEFT))
    styles.add(ParagraphStyle(name='FooterTextRight', parent=styles['Normal'], fontSize=8, fontName='Helvetica', textColor=colors.HexColor('#D95F02'), alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name='TableCellData', parent=styles['Normal'], fontSize=6, fontName='Helvetica', alignment=TA_CENTER, textColor=colors.black))
    styles.add(ParagraphStyle(name='TableCellDataLeft', parent=styles['Normal'], fontSize=6, fontName='Helvetica', alignment=TA_LEFT, textColor=colors.black))
    styles.add(ParagraphStyle(name='TotalGeralText', parent=styles['Normal'], fontSize=6, fontName='Helvetica', alignment=TA_LEFT, textColor=colors.HexColor('#FD8D3C')))
    styles.add(ParagraphStyle(name='TableHeader', parent=styles['Normal'], fontSize=6, fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=colors.HexColor('#FD8D3C')))
    styles.add(ParagraphStyle(name='TableHeaderLeft', parent=styles['Normal'], fontSize=6, fontName='Helvetica-Bold', alignment=TA_LEFT, textColor=colors.HexColor('#FD8D3C')))

    WATERMARK_IMAGE_PATH_LOCAL = os.path.join(BASE_DIR, "static", "images", "rcb_logo_pdf.png")


    def header_footer_inner(canvas, doc):
        canvas.saveState()
        page_width = doc.pagesize[0]
        page_height = doc.pagesize[1]

        if os.path.exists(WATERMARK_IMAGE_PATH_LOCAL):
            try:
                img_reader = ImageReader(WATERMARK_IMAGE_PATH_LOCAL)
                img_width, img_height = img_reader.getSize()
                aspect = img_height / float(img_width)
                display_width = 15 * cm
                display_height = display_width * aspect
                x_centered = (page_width - display_width) / 2
                y_centered = (page_height - display_height) / 2
                canvas.setFillAlpha(0.15)
                canvas.drawImage(img_reader, x_centered, y_centered, width=display_width, height=display_height)
                canvas.setFillAlpha(1.0)
            except Exception:
                pass

        header_content_left = [Paragraph("Grande Recife Consórcio de Transporte", styles['HeaderTitleMain']), Paragraph("Sistema de Remuneração, Custos e Bilhetagem", styles['HeaderSubtitle'])]
        header_content_right = [Paragraph("Relatório de Passageiros Integrados", styles['ReportTitle']), Paragraph(infos['periodo'], styles['PeriodInfo'])]
        header_table = Table([[header_content_left, header_content_right]], colWidths=[doc.width * 0.6, doc.width * 0.4])
        header_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP'), ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0)]))
        header_height_needed = header_table.wrap(doc.width, doc.topMargin)[1]
        header_table.drawOn(canvas, doc.leftMargin, page_height - doc.topMargin + (doc.topMargin - header_height_needed - 0.5*cm))

        footer_y_line = doc.bottomMargin
        footer_y_text = footer_y_line - 0.7*cm
        canvas.setStrokeColor(colors.HexColor('#FD8D3C'))
        canvas.setLineWidth(1)
        canvas.line(doc.leftMargin, footer_y_line, doc.leftMargin + doc.width, footer_y_line)

        p_left = Paragraph(f"Processamento: {data_hoje}", styles['FooterTextLeft'])
        page_num = canvas.getPageNumber()
        total_pages_str = str(known_total_pages)
        footer_right_text = f"Página {page_num} de {total_pages_str}"
        p_right = Paragraph(footer_right_text, styles['FooterTextRight'])

        footer_table = Table([[p_left, p_right]], colWidths=[doc.width/2, doc.width/2])
        footer_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
        footer_table.wrap(doc.width, 0.5*cm)[1]
        footer_table.drawOn(canvas, doc.leftMargin, footer_y_text)

        canvas.restoreState()

    main_template = PageTemplate(id='main', frames=[main_frame], onPage=header_footer_inner)
    doc.addPageTemplates([main_template])

    header_styled = []
    for i, col_name in enumerate(df_formatado.columns):
        style = styles['TableHeaderLeft'] if i == 0 else styles['TableHeader']
        header_styled.append(Paragraph(str(col_name), style))

    data_list = []
    total_geral_df_index = -1
    first_col_name = df_formatado.columns[0]

    for index, row in df_formatado.iterrows():
        row_styled = []
        is_total_geral_row = (str(row[first_col_name]).strip() == 'Total Geral')
        if is_total_geral_row:
            total_geral_df_index = index

        for i, value in enumerate(row):
            cell_value = str(value)
            if i == 0:
                style = styles['TotalGeralText'] if is_total_geral_row else styles['TableCellDataLeft']
                row_styled.append(Paragraph(cell_value, style))
            else:
                row_styled.append(Paragraph(cell_value, styles['TableCellData']))
        data_list.append(row_styled)

    table_data = [header_styled] + data_list

    table_style_commands = [
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4), ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
        ('LINEABOVE', (0, 0), (-1, 0), 1, colors.HexColor('#FD8D3C')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#FD8D3C')),
    ]
    if total_geral_df_index != -1:
        pdf_table_total_row_index = total_geral_df_index + 1
        if 0 < pdf_table_total_row_index < len(table_data):
            table_style_commands.append(('LINEABOVE', (0, pdf_table_total_row_index), (-1, pdf_table_total_row_index), 1, colors.HexColor('#FD8D3C')))

    table_style = TableStyle(table_style_commands)

    num_cols = len(df_formatado.columns)
    if num_cols > 0:
        first_col_width = 4.5 * cm
        available_width = doc.width - first_col_width
        other_cols_count = num_cols - 1
        if other_cols_count > 0:
            other_col_width = available_width / other_cols_count
            if other_col_width < 0.5 * cm: other_col_width = 0.5 * cm
        else: other_col_width = 0
        col_widths = [first_col_width] + [other_col_width] * other_cols_count
    else: col_widths = None

    if table_data and col_widths:
        report_table = Table(table_data, colWidths=col_widths, style=table_style, repeatRows=1)
    else:
        report_table = Paragraph("Erro: Não há dados para exibir na tabela.", styles['Normal'])

    story = []
    story.append(Paragraph(operador_com_sigla, styles['SubHeader']))
    story.append(report_table)

    try:
        doc.build(story)
        return True
    except Exception:
        traceback.print_exc()
        return False

def build_pdf_two_pass(df_formatado, infos, data_hoje, operador_com_sigla, output_filename):
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='SubHeader', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', textColor=colors.HexColor('#FD8D3C'), alignment=TA_LEFT, spaceBefore=0, spaceAfter=15))
    styles.add(ParagraphStyle(name='TableCellData', parent=styles['Normal'], fontSize=6, fontName='Helvetica', alignment=TA_CENTER, textColor=colors.black))
    styles.add(ParagraphStyle(name='TableCellDataLeft', parent=styles['Normal'], fontSize=6, fontName='Helvetica', alignment=TA_LEFT, textColor=colors.black))
    styles.add(ParagraphStyle(name='TotalGeralText', parent=styles['Normal'], fontSize=6, fontName='Helvetica', alignment=TA_LEFT, textColor=colors.HexColor('#FD8D3C')))
    styles.add(ParagraphStyle(name='TableHeader', parent=styles['Normal'], fontSize=6, fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=colors.HexColor('#FD8D3C')))
    styles.add(ParagraphStyle(name='TableHeaderLeft', parent=styles['Normal'], fontSize=6, fontName='Helvetica-Bold', alignment=TA_LEFT, textColor=colors.HexColor('#FD8D3C')))

    header_styled = []
    for i, col_name in enumerate(df_formatado.columns):
        style = styles['TableHeaderLeft'] if i == 0 else styles['TableHeader']
        header_styled.append(Paragraph(str(col_name), style))

    data_list = []
    total_geral_df_index = -1
    first_col_name = df_formatado.columns[0]
    for index, row in df_formatado.iterrows():
        row_styled = []
        is_total_geral_row = (str(row[first_col_name]).strip() == 'Total Geral')
        if is_total_geral_row: total_geral_df_index = index
        for i, value in enumerate(row):
            cell_value = str(value)
            if i == 0:
                style = styles['TotalGeralText'] if is_total_geral_row else styles['TableCellDataLeft']
                row_styled.append(Paragraph(cell_value, style))
            else:
                row_styled.append(Paragraph(cell_value, styles['TableCellData']))
        data_list.append(row_styled)
    table_data = [header_styled] + data_list

    table_style_commands = [
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4), ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
        ('LINEABOVE', (0, 0), (-1, 0), 1, colors.HexColor('#FD8D3C')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#FD8D3C')),
    ]
    if total_geral_df_index != -1:
        pdf_table_total_row_index = total_geral_df_index + 1
        if 0 < pdf_table_total_row_index < len(table_data):
            table_style_commands.append(('LINEABOVE', (0, pdf_table_total_row_index), (-1, pdf_table_total_row_index), 1, colors.HexColor('#FD8D3C')))
    table_style = TableStyle(table_style_commands)

    num_cols = len(df_formatado.columns)
    if num_cols > 0:
        first_col_width = 4.5 * cm
        doc_width_temp = landscape(A4)[0] - 4*cm
        available_width = doc_width_temp - first_col_width
        other_cols_count = num_cols - 1
        if other_cols_count > 0:
            other_col_width = available_width / other_cols_count
            if other_col_width < 0.5 * cm: other_col_width = 0.5 * cm
        else: other_col_width = 0
        col_widths = [first_col_width] + [other_col_width] * other_cols_count
    else: col_widths = None

    if table_data and col_widths:
        report_table = Table(table_data, colWidths=col_widths, style=table_style, repeatRows=1)
    else:
        report_table = Paragraph("Erro: Não há dados para exibir na tabela.", styles['Normal'])

    story = []
    story.append(Paragraph(operador_com_sigla, styles['SubHeader']))
    story.append(report_table)

    total_pages = 0
    temp_pdf_path = None
    try:
        temp_pdf_fd, temp_pdf_path = tempfile.mkstemp(suffix=".pdf")
        os.close(temp_pdf_fd)

        doc_pass1 = BaseDocTemplate(temp_pdf_path,
                                      pagesize=landscape(A4),
                                      leftMargin=2*cm, rightMargin=2*cm,
                                      topMargin=2*cm, bottomMargin=2.5*cm)

        frame_pass1 = Frame(doc_pass1.leftMargin, doc_pass1.bottomMargin, doc_pass1.width, doc_pass1.height, id='frame_pass1')
        def onPage_dummy(canvas, doc): pass
        template_pass1 = PageTemplate(id='pass1', frames=[frame_pass1], onPage=onPage_dummy)
        doc_pass1.addPageTemplates([template_pass1])

        doc_pass1.build(story)
        total_pages = doc_pass1.page

    except Exception:
        traceback.print_exc()
        raise
    finally:
        if temp_pdf_path and os.path.exists(temp_pdf_path):
            try:
                os.remove(temp_pdf_path)
            except OSError:
                pass

    if total_pages == 0:
        raise ValueError("A contagem de páginas resultou em 0. Verifique o conteúdo do arquivo de entrada.")

    success_pass2 = gerar_pdf_final_com_contagem(
        df_formatado, infos, data_hoje, operador_com_sigla, output_filename, total_pages
    )

    return success_pass2


# --- CLASSE PRINCIPAL DA GUI UNIFICADA ---

class PassageiroIntegradoGUI(tk.Tk):
    def __init__(self, title="Ferramenta de Integração de Passageiros", size=(1000, 800)):
        super().__init__()
        self.title(title)
        self.geometry(f'{size[0]}x{size[1]}')
        self.resizable(True, True)

        # Configuração de Estilo e Tema
        self.style = ttk.Style(self)
        self.style.configure('Accent.TButton', font=('Segoe UI', 10, 'bold'))
        self.style.configure('TLabel', font=('Segoe UI', 10))
        self.style.configure('TLabelframe.Label', font=('Segoe UI', 11, 'bold'))
        self.style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'))
        self.style.configure('Header.TLabel', font=('Segoe UI', 12, 'bold'))
        self.pasta_destino_excel = tk.StringVar()
        self.caminho_efetivo = tk.StringVar()
        self.caminhos_viagem = tk.StringVar()
        self._full_path_efetivo = ""
        self._full_paths_viagem = []

        self.operadoras_vars = {
            sigla: tk.BooleanVar(value=True)
            for sigla in operator_mapping.keys()
        }

        # Aba 2: Gerador PDF
        self.caminhos_excel_pdf = tk.StringVar()
        self.pasta_saida_pdf = tk.StringVar()
        self._full_paths_excel_pdf = []

        self.create_widgets()

    def center_window(self):
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')

    def toggle_theme(self):
        sv_ttk.toggle_theme()

    def toggle_size(self):
        if not hasattr(self, 'is_compact') or self.is_compact:
            self.geometry('1000x800')
            self.is_compact = False
        else:
            self.geometry('850x680')
            self.is_compact = True
        self.center_window()

    def create_widgets(self):
        try:
            bg_image_pil = Image.open(self.BACKGROUND_IMAGE_PATH)
            bg_image_pil = bg_image_pil.resize((1000, 800), Image.Resampling.LANCZOS)
            self.background_image = ImageTk.PhotoImage(bg_image_pil)
            background_label = tk.Label(self, image=self.background_image)
            background_label.place(x=0, y=0, relwidth=1, relheight=1)
        except Exception:
            pass

        main_frame = ttk.Frame(self, padding=(20, 15))
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Cabeçalho da Aplicação ---
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 20))

        try:
            logo_original = Image.open(self.LOGO_IMAGE_PATH).convert("RGBA")
            logo_resized = logo_original.resize((int(logo_original.width * 0.4), int(logo_original.height * 0.4)), Image.Resampling.LANCZOS)
            self.logo_image = ImageTk.PhotoImage(logo_resized)
            logo_label = ttk.Label(header_frame, image=self.logo_image)
            logo_label.pack(side=tk.LEFT, padx=(0, 15))
        except Exception:
            pass

        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(title_frame, text="Integrador de Passageiros", style="Title.TLabel").pack(anchor='w')
        ttk.Label(title_frame, text="Ferramenta para consolidar e gerar relatórios", style="TLabel").pack(anchor='w')

        theme_switch = ttk.Checkbutton(header_frame, text="Tema", style="Switch.TCheckbutton", command=self.toggle_theme)
        theme_switch.pack(side=tk.RIGHT, padx=10)
        resize_button = ttk.Button(header_frame, text="Ajustar Janela", command=self.toggle_size)
        resize_button.pack(side=tk.RIGHT, padx=5)

        # --- Notebook com as Abas ---
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=10)

        tab1 = ttk.Frame(notebook, padding=10)
        tab2 = ttk.Frame(notebook, padding=10)

        notebook.add(tab1, text='    1. Integrar Relatórios (TXT → Excel)    ')
        notebook.add(tab2, text='    2. Gerar Relatório (Excel → PDF)    ')

        self._criar_aba_integrador(tab1)
        self._criar_aba_pdf(tab2)

    def _criar_aba_integrador(self, parent):
        parent.columnconfigure(0, weight=1)

        frame_entrada = ttk.LabelFrame(parent, text=" Arquivos de Entrada (.txt) ", padding=15)
        frame_entrada.pack(fill=tk.X, pady=(5, 10))
        frame_entrada.columnconfigure(1, weight=1)

        ttk.Label(frame_entrada, text="Arquivo de Efetivo (Geral):").grid(row=0, column=0, sticky='w', padx=(0,10), pady=(5,2))
        ttk.Entry(frame_entrada, textvariable=self.caminho_efetivo, state='readonly').grid(row=0, column=1, sticky='ew')
        ttk.Button(frame_entrada, text="Selecionar", command=self.selecionar_arquivo_efetivo).grid(row=0, column=2, padx=(10,0))

        ttk.Label(frame_entrada, text="Arquivos de Viagem/Frota:").grid(row=1, column=0, sticky='w', padx=(0,10), pady=(5,2))
        ttk.Entry(frame_entrada, textvariable=self.caminhos_viagem, state='readonly').grid(row=1, column=1, sticky='ew')
        ttk.Button(frame_entrada, text="Selecionar", command=self.selecionar_arquivos_viagem).grid(row=1, column=2, padx=(10,0))

        frame_operadoras = ttk.LabelFrame(parent, text=" Operadoras para Processar ", padding=15)
        frame_operadoras.pack(fill=tk.X, pady=(5, 10))
        
        operadoras = sorted(self.operadoras_vars.keys())
        cols = 4 
        for i, sigla in enumerate(operadoras):
            row, col = divmod(i, cols)
            ttk.Checkbutton(frame_operadoras, text=sigla, variable=self.operadoras_vars[sigla]).grid(row=row, column=col, sticky='w', padx=5, pady=2)

        frame_saida = ttk.LabelFrame(parent, text=" Local de Saída do Excel ", padding=15)
        frame_saida.pack(fill=tk.X, pady=(5, 10))
        frame_saida.columnconfigure(1, weight=1)

        ttk.Button(frame_saida, text="Selecionar Pasta de Destino", command=self.selecionar_pasta_destino_excel).grid(row=0, column=0, padx=(0,10), pady=5, sticky='w')
        self.label_pasta_excel = ttk.Label(frame_saida, text="Nenhuma pasta selecionada.")
        self.label_pasta_excel.grid(row=0, column=1, pady=5, sticky='ew')

        button_area = ttk.Frame(parent, padding=(0, 20, 0, 0))
        button_area.pack(fill=tk.X, side=tk.BOTTOM)
        self.botao_limpar_tab1 = ttk.Button(button_area, text="Limpar Campos", command=self.limpar_campos_tab1)
        self.botao_limpar_tab1.pack(side=tk.LEFT, padx=(0, 10))
        self.botao_gerar_tab1 = ttk.Button(button_area, text="Integrar e Gerar Excel", style='Accent.TButton', command=self.iniciar_integracao)
        self.botao_gerar_tab1.pack(side=tk.RIGHT, fill=tk.X, expand=True)

    def _criar_aba_pdf(self, parent):
        parent.columnconfigure(0, weight=1)

        frame_entrada = ttk.LabelFrame(parent, text=" Arquivos de Entrada (.xlsx) ", padding=15)
        frame_entrada.pack(fill=tk.X, pady=(5, 10))
        frame_entrada.columnconfigure(1, weight=1)

        ttk.Label(frame_entrada, text="Arquivos Excel Processados:").grid(row=0, column=0, sticky='w', padx=(0,10), pady=(5,2))
        ttk.Entry(frame_entrada, textvariable=self.caminhos_excel_pdf, state='readonly').grid(row=0, column=1, sticky='ew')
        ttk.Button(frame_entrada, text="Selecionar", command=self.selecionar_arquivos_excel_pdf).grid(row=0, column=2, padx=(10,0))

        frame_saida = ttk.LabelFrame(parent, text=" Local de Saída dos PDFs ", padding=15)
        frame_saida.pack(fill=tk.X, pady=(5, 10))
        frame_saida.columnconfigure(1, weight=1)

        ttk.Button(frame_saida, text="Selecionar Pasta de Destino", command=self.definir_pasta_saida_pdf).grid(row=0, column=0, padx=(0,10), pady=5, sticky='w')
        self.entry_pdf_saida = ttk.Entry(frame_saida, textvariable=self.pasta_saida_pdf, state='readonly')
        self.entry_pdf_saida.grid(row=0, column=1, pady=5, sticky='ew')

        button_area = ttk.Frame(parent, padding=(0, 20, 0, 0))
        button_area.pack(fill=tk.X, side=tk.BOTTOM)
        self.botao_limpar_tab2 = ttk.Button(button_area, text="Limpar Campos", command=self.limpar_campos_tab2)
        self.botao_limpar_tab2.pack(side=tk.LEFT, padx=(0, 10))
        self.botao_gerar_tab2 = ttk.Button(button_area, text="Gerar Relatórios PDF", style='Accent.TButton', command=self.iniciar_geracao_pdf)
        self.botao_gerar_tab2.pack(side=tk.RIGHT, fill=tk.X, expand=True)

    # --- MÉTODOS DA ABA 1: INTEGRADOR TXT -> EXCEL ---

    def selecionar_arquivo_efetivo(self):
        arquivo = filedialog.askopenfilename(title="Selecione o TXT de Efetivo", filetypes=[("Arquivos de Texto", "*.txt")])
        if arquivo:
            self.caminho_efetivo.set(os.path.basename(arquivo))
            self._full_path_efetivo = arquivo

    def selecionar_arquivos_viagem(self):
        arquivos = filedialog.askopenfilenames(title="Selecione os arquivos TXT de Viagem e Frota", filetypes=[("Arquivos de Texto", "*.txt")])
        if arquivos:
            self._full_paths_viagem = arquivos
            self.caminhos_viagem.set(f"{len(arquivos)} arquivos selecionados")

    def selecionar_pasta_destino_excel(self):
        caminho = filedialog.askdirectory(title="Selecione a pasta para salvar os relatórios Excel")
        if caminho:
            self.pasta_destino_excel.set(caminho)
            self.label_pasta_excel.config(text=caminho)

    def limpar_campos_tab1(self):
        self.caminho_efetivo.set("")
        self.caminhos_viagem.set("")
        self._full_path_efetivo = ""
        self._full_paths_viagem = []
        self.pasta_destino_excel.set("")
        self.label_pasta_excel.config(text="Nenhuma pasta selecionada.")
        for var in self.operadoras_vars.values():
            var.set(True)
        messagebox.showinfo("Limpeza", "Todos os campos da Aba 1 foram limpos.", parent=self)

    def abrir_pasta(self, caminho_pasta):
        try:
            if sys.platform == "win32":
                os.startfile(caminho_pasta)
            elif sys.platform == "darwin":
                subprocess.run(["open", caminho_pasta])
            else:
                subprocess.run(["xdg-open", caminho_pasta])
        except Exception:
            messagebox.showinfo("Aviso", "O processo foi concluído, mas não foi possível abrir a pasta de destino automaticamente.", parent=self)

    def iniciar_integracao(self):
        if not self._full_path_efetivo:
            messagebox.showerror("Erro", "É necessário selecionar o arquivo TXT de Efetivo.", parent=self)
            return
        if not self._full_paths_viagem:
            messagebox.showerror("Erro", "É necessário selecionar pelo menos um arquivo TXT de Viagem/Frota.", parent=self)
            return
        pasta_destino = self.pasta_destino_excel.get()
        if not pasta_destino:
            messagebox.showerror("Erro", "É necessário selecionar uma pasta de destino.", parent=self)
            return

        self.botao_gerar_tab1.config(state="disabled", text="Processando...")
        self.botao_limpar_tab1.config(state="disabled")
        self.update_idletasks()

        operadoras_selecionadas = [sigla for sigla, var in self.operadoras_vars.items() if var.get()]
        if not operadoras_selecionadas:
            messagebox.showerror("Erro", "Nenhuma operadora foi selecionada para o processamento.", parent=self)
            self.botao_gerar_tab1.config(state="normal", text="Integrar e Gerar Excel")
            self.botao_limpar_tab1.config(state="normal")
            return

        try:
            df_efetivo = processar_efetivo(self._full_path_efetivo)
            if df_efetivo is None:
                raise ValueError("Falha crítica ao ler o arquivo de Efetivo. O resultado está vazio.")
            mapa_totais = criar_mapa_totais_efetivo(df_efetivo)
        except Exception as e:
            messagebox.showerror("Erro Crítico no Efetivo", f"Não foi possível processar o arquivo de Efetivo:\n\n{e}", parent=self)
            self.botao_gerar_tab1.config(state="normal", text="Integrar e Gerar Excel")
            self.botao_limpar_tab1.config(state="normal")
            return

        sucesso_count = 0
        falha_count = 0
        pulado_count = 0
        falhas_detalhes = []

        for caminho_viagem in self._full_paths_viagem:
            try:
                df_viagem, add_hdr, custom_txt, period_txt, sigla_operador = processar_viagem_frota(caminho_viagem, operator_mapping, month_mapping)

                if not sigla_operador or sigla_operador not in operadoras_selecionadas:
                    pulado_count += 1
                    continue

                if df_viagem is None:
                    raise ValueError("Falha na leitura do arquivo de viagem.")

                if add_hdr and df_viagem is not None and not df_viagem.empty and mapa_totais and sigla_operador:
                    df_viagem = popular_totais_viagem_frota(df_viagem, mapa_totais, sigla_operador)

                if df_viagem is not None and not df_viagem.empty and len(df_viagem.columns) > 1:
                    cols_to_sum_viagem = df_viagem.columns[1:]
                    if not cols_to_sum_viagem.empty:
                        df_viagem_copy = df_viagem.copy()
                        for col in cols_to_sum_viagem: df_viagem_copy[col] = pd.to_numeric(df_viagem_copy[col], errors='coerce')
                        df_viagem_copy['TOTAL'] = df_viagem_copy[cols_to_sum_viagem].sum(axis=1, skipna=True)
                        df_viagem = df_viagem_copy

                if df_viagem is not None and not df_viagem.empty and len(df_viagem.columns) > 1:
                    numeric_cols_to_sum = df_viagem.columns[1:]
                    df_numeric_copy = df_viagem[numeric_cols_to_sum].apply(pd.to_numeric, errors='coerce')
                    grand_totals = df_numeric_copy.sum(skipna=True)
                    total_geral_dict = {df_viagem.columns[0]: "Total Geral"}
                    total_geral_dict.update(grand_totals.to_dict())
                    total_geral_df = pd.DataFrame([total_geral_dict])
                    df_viagem = pd.concat([df_viagem, total_geral_df], ignore_index=True)

                if df_viagem is not None and not df_viagem.empty:
                    nome_arquivo_2 = f"{os.path.splitext(os.path.basename(caminho_viagem))[0]}_processado.xlsx"
                    caminho_completo_2 = os.path.join(pasta_destino, nome_arquivo_2)
                    start_row_excel = 4 if add_hdr else 0
                    df_viagem.to_excel(caminho_completo_2, index=False, header=True, engine='openpyxl', startrow=start_row_excel)
                    aplicar_formatacao_excel(caminho_completo_2, df_viagem, 2, add_hdr, custom_txt, period_txt)
                
                sucesso_count += 1
            except Exception as e:
                falha_count += 1
                nome_arquivo_falha = os.path.basename(caminho_viagem)
                falhas_detalhes.append(f"- {nome_arquivo_falha}: {str(e)}")
                traceback.print_exc()

        mensagem_final = f"Processamento em lote concluído!\n\n"
        mensagem_final += f"Sucesso: {sucesso_count} arquivo(s)\n"
        mensagem_final += f"Falha: {falha_count} arquivo(s)\n"
        mensagem_final += f"Pulados: {pulado_count} arquivo(s) (não selecionados)\n"

        if falhas_detalhes:
            mensagem_final += "\nDetalhes das falhas:\n" + "\n".join(falhas_detalhes)

        messagebox.showinfo("Resultado do Processamento", mensagem_final, parent=self)
        if sucesso_count > 0:
            self.abrir_pasta(pasta_destino)
            
        self.botao_gerar_tab1.config(state="normal", text="Integrar e Gerar Excel")
        self.botao_limpar_tab1.config(state="normal")

    # --- MÉTODOS DA ABA 2: GERADOR PDF ---

    def selecionar_arquivos_excel_pdf(self):
        arquivos = filedialog.askopenfilenames(title="Selecione os arquivos XLSX gerados", filetypes=[("Arquivos Excel", "*.xlsx")])
        if arquivos:
            self._full_paths_excel_pdf = arquivos
            self.caminhos_excel_pdf.set(f"{len(arquivos)} arquivos selecionados")

    def definir_pasta_saida_pdf(self):
        caminho = filedialog.askdirectory(title="Selecione a pasta para salvar os relatórios PDF")
        if caminho:
            self.pasta_saida_pdf.set(caminho)

    def limpar_campos_tab2(self):
        self.caminhos_excel_pdf.set("")
        self.pasta_saida_pdf.set("")
        self._full_paths_excel_pdf = []
        messagebox.showinfo("Limpeza", "Todos os campos da Aba 2 foram limpos.", parent=self)

    def iniciar_geracao_pdf(self):
        arquivos_xlsx = self._full_paths_excel_pdf
        pasta_saida = self.pasta_saida_pdf.get()

        if not arquivos_xlsx:
            messagebox.showerror("Erro", "É necessário selecionar pelo menos um arquivo Excel de entrada.", parent=self)
            return
        if not pasta_saida:
            messagebox.showerror("Erro", "É necessário selecionar uma pasta de destino para os PDFs.", parent=self)
            return

        self.botao_gerar_tab2.config(state="disabled", text="Gerando...")
        self.botao_limpar_tab2.config(state="disabled")
        self.update_idletasks()

        sucesso_count = 0
        falha_count = 0
        falhas_detalhes = []

        for arquivo_xlsx in arquivos_xlsx:
            try:
                nome_base_excel = os.path.splitext(os.path.basename(arquivo_xlsx))[0]
                infos = extrair_infos_relatorio(nome_base_excel)
                
                nome_sugerido_pdf = f"{infos['sigla']}-{infos['periodo_codigo']}-Passageiro Integrado.pdf" if infos['sigla'] and infos.get('periodo_codigo') else f"{nome_base_excel}.pdf"
                caminho_saida_pdf = os.path.join(pasta_saida, nome_sugerido_pdf)

                operador_com_sigla = f"[{infos['sigla']}] {infos['nome_operador_sem_cnpj']}" if infos['sigla'] else infos['nome_operador_sem_cnpj']

                df = None
                try:
                    df = pd.read_excel(arquivo_xlsx, sheet_name=0, header=4)
                    if not isinstance(df.columns, pd.Index) or not str(df.columns[0]).strip().upper().startswith('LINHA'): raise ValueError()
                except Exception:
                    df = pd.read_excel(arquivo_xlsx, sheet_name=0, header=0)

                if df is None or df.empty: raise ValueError("Não foi possível ler os dados do Excel.")

                primeira_coluna_nome = df.columns[0]
                new_columns = [primeira_coluna_nome] + [f"{int(float(c)):02d}" if isinstance(c, (float, int)) and re.fullmatch(r'\d+(\.0)?', str(c)) else str(c).strip() for c in df.columns[1:]]
                df.columns = new_columns
                df[df.columns[1:]] = df[df.columns[1:]].fillna(0)
                df[primeira_coluna_nome] = df[primeira_coluna_nome].fillna('')

                df_filled = df.apply(lambda col: col.apply(format_brazilian_number) if col.name != primeira_coluna_nome else col)

                sucesso_build = build_pdf_two_pass(df_filled, infos, get_data_atual_formatada(), operador_com_sigla, caminho_saida_pdf)
                if not sucesso_build: raise Exception("Falha na função build_pdf_two_pass.")

                sucesso_count += 1
            except Exception as e:
                falha_count += 1
                nome_arquivo_falha = os.path.basename(arquivo_xlsx)
                falhas_detalhes.append(f"- {nome_arquivo_falha}: {str(e)}")
                traceback.print_exc()

        mensagem_final = f"Geração de PDF em lote concluída!\n\n"
        mensagem_final += f"Sucesso: {sucesso_count} arquivo(s)\n"
        mensagem_final += f"Falha: {falha_count} arquivo(s)\n"

        if falhas_detalhes:
            mensagem_final += "\nDetalhes das falhas:\n" + "\n".join(falhas_detalhes)

        messagebox.showinfo("Resultado da Geração de PDF", mensagem_final, parent=self)
        if sucesso_count > 0:
            self.abrir_pasta(pasta_saida)

        self.botao_gerar_tab2.config(state="normal", text="Gerar Relatórios PDF")
        self.botao_limpar_tab2.config(state="normal")


if __name__ == "__main__":
    app = PassageiroIntegradoGUI()
    app.mainloop()