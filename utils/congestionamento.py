# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import os
import sys
import subprocess
import traceback
import threading
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, PhotoImage, ttk
from tkinterdnd2 import DND_FILES, TkinterDnD
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, PatternFill
from openpyxl.styles.borders import Border, Side
import sv_ttk
from io import StringIO

# =====================================================================================
# --- BLOCO 1: LÓGICA DE PROCESSAMENTO (ANÁLISE DE CONGESTIONAMENTO) ---
# =====================================================================================

VALORES_A_REMOVER_DESCRICAO = {
    "Viagem Realizada", "Frota Programada", "Frota Realizada", "Frota Admitida",
    "Extensão Útil", "Extensão Morta", "KM Útil Informada", "KM Morta Informada",
    "KM Total Informada", "KM Útil Programada", "KM Morta Programada",
    "KM Total Programada", "KM Total Admitida", "Viagens Sistema Programadas",
    "Viagens Sistema Realizadas", "Viagens Sistema Admitidas", "Viagens SEI Programadas",
    "Viagens SEI Realizadas", "Viagens SEI Admitidas", "Viagens Convencional Programadas",
    "Viagens Convencional Realizadas", "Viagens Convencional Admitidas",
    "Viagens Opcional Programadas", "Viagens Opcional Realizadas", "Viagens Opcional Admitidas",
    "Viagens Micro Programadas", "Viagens Micro Realizadas", "Viagens Micro Admitidas",
    "Viagens BRT Programadas", "Viagens BRT Realizadas", "Viagens BRT Admitidas",
    "Frota Sistema Programada", "Frota Sistema Realizada", "Frota Sistema Admitida",
    "Frota SEI Programada", "Frota SEI Realizada", "Frota SEI Admitida",
    "Frota Convencional Programada", "Frota Convencional Realizada", "Frota Convencional Admitida",
    "Frota Opcional Programada", "Frota Opcional Realizada", "Frota Opcional Admitida",
    "Frota Micro Programada", "Frota Micro Realizada", "Frota Micro Admitida",
    "Frota BRT Programada", "Frota BRT Realizada", "Frota BRT Admitida"
}
VALORES_A_REMOVER_SECUNDARIA = {"Relatório de Viagem, Quilometragem e Frota"}
ORDEM_EMPRESAS_GLOBAL = ['BOA', 'CAX', 'CSR', 'EME', 'GLO', 'SJT', 'VML']


def processar_arquivos_txt_congestionamento(arquivos_txt):
    if not arquivos_txt:
        print("Nenhum arquivo .txt foi fornecido para processamento (Congestionamento).")
        return pd.DataFrame()

    lista_df_saldo = []
    for txt_file in arquivos_txt:
        nome_empresa = os.path.basename(txt_file).split("-")[0]
        print(f"[{nome_empresa}] Processando {os.path.basename(txt_file)}...")
        try:
            df = pd.read_csv(txt_file, sep='\t', skiprows=1, encoding='latin1', dtype=str)
            df = df[~df['Descrição'].isin(VALORES_A_REMOVER_DESCRICAO)]
            df = df[~df.iloc[:, 0].isin(VALORES_A_REMOVER_SECUNDARIA)]
            df['Linha'] = df['Linha'].str.split(" - ").str[0]

            id_vars = ['Operador', 'Ano', 'Mês', 'Quinzena', 'Linha', 'Descrição']
            value_vars = [col for col in df.columns if col not in id_vars]

            df_long = df.melt(id_vars=id_vars, value_vars=value_vars, var_name='Dia', value_name='Valor')
            df_long['Valor'] = pd.to_numeric(df_long['Valor'].str.replace(',', '.'), errors='coerce').fillna(0)

            df_pivot = df_long.pivot_table(index=['Operador', 'Ano', 'Mês', 'Quinzena', 'Linha', 'Dia'], columns='Descrição', values='Valor', fill_value=0).reset_index()

            if 'Viagem Programada' in df_pivot.columns and 'Viagem Admitida' in df_pivot.columns:
                df_pivot['Saldo'] = df_pivot['Viagem Programada'] - df_pivot['Viagem Admitida']

                df_final_saldo = df_pivot.pivot_table(index=['Operador', 'Ano', 'Mês', 'Quinzena', 'Linha'], columns='Dia', values='Saldo').reset_index()

                df_final_saldo.insert(5, 'Descrição', 'Saldo')
                lista_df_saldo.append(df_final_saldo)
                print(f"[{nome_empresa}] Processamento concluído.")
        except Exception as e:
            print(f"Erro ao processar o arquivo {txt_file}: {e}")
            messagebox.showerror("Erro em Arquivo TXT", f"Falha ao processar:\n{os.path.basename(txt_file)}\n\nErro: {e}")
            return pd.DataFrame()

    if not lista_df_saldo:
        return pd.DataFrame()

    df_concatenado = pd.concat(lista_df_saldo, ignore_index=True)

    print("\nAplicando ordenação no resultado dos arquivos TXT...")
    df_concatenado['Operador'] = pd.Categorical(df_concatenado['Operador'], categories=ORDEM_EMPRESAS_GLOBAL, ordered=True)
    df_concatenado['Linha'] = pd.to_numeric(df_concatenado['Linha'], errors='coerce')
    df_concatenado.dropna(subset=['Linha', 'Operador'], inplace=True)
    df_concatenado.sort_values(by=['Operador', 'Linha'], inplace=True)
    df_concatenado.columns = [str(col).split('.')[0] if isinstance(col, str) and col.replace('.', '', 1).isdigit() else col for col in df_concatenado.columns]

    meta_cols = [col for col in df_concatenado.columns if not col.isdigit()]
    day_cols = [col for col in df_concatenado.columns if col.isdigit()]
    
    sorted_day_cols = sorted(day_cols, key=int)
    
    df_concatenado = df_concatenado[meta_cols + sorted_day_cols]

    return df_concatenado


def processar_arquivos_xls_congestionamento(arquivos_xls):
    if not arquivos_xls:
        print("Nenhum arquivo .xls foi fornecido para processamento (Congestionamento).")
        return pd.DataFrame()

    dfs_xls_unificados = []
    for xls_path in arquivos_xls:
        try:
            nome_base = os.path.splitext(os.path.basename(xls_path))[0]
            sigla = nome_base[:3]
            print(f"Processando: {nome_base} -> Sigla: {sigla}")

            df = pd.read_excel(xls_path, header=11, dtype=str)

            df.dropna(how='all', axis=0, inplace=True)
            df.dropna(how='all', axis=1, inplace=True)

            cols_to_delete = ['Unnamed: 0', 'Unnamed: 6', 'Unnamed: 7', 'Unnamed: 8', 'Unnamed: 18', 'Unnamed: 19']
            df.drop(columns=[col for col in cols_to_delete if col in df.columns], inplace=True)

            df.insert(0, "Operador", sigla)
            dfs_xls_unificados.append(df)
        except Exception as e:
            print(f"Erro ao processar o arquivo {xls_path}: {e}")
            messagebox.showerror("Erro em Arquivo XLS", f"Falha ao processar:\n{os.path.basename(xls_path)}\n\nErro: {e}")
            return pd.DataFrame()

    if not dfs_xls_unificados:
        return pd.DataFrame()

    return pd.concat(dfs_xls_unificados, ignore_index=True)


def comparar_e_gerar_final_congestionamento(df_txt, df_xls):
    if df_txt is None or df_xls is None or df_txt.empty or df_xls.empty:
        print("Um dos DataFrames de entrada (Congestionamento) está vazio. Não é possível continuar a comparação.")
        return None, None

    planilha1 = df_txt.copy()
    planilha1.columns = [str(col).strip() for col in planilha1.columns]
    colunas_dia_a_dia = {col: f"{int(col):02d}" for col in planilha1.columns if col.isdigit()}
    planilha1.rename(columns=colunas_dia_a_dia, inplace=True)

    planilha2 = df_xls.copy()
    planilha2.columns = [str(col).strip() if col is not None else "" for col in planilha2.columns]
    planilha2.rename(columns={"Linhas": "Linha"}, inplace=True)
    planilha2['Linha'] = pd.to_numeric(planilha2['Linha'], errors='coerce')

    chaves_juncao = ['Operador', 'Linha']
    df_merged = pd.merge(planilha1, planilha2, on=chaves_juncao, how='inner', suffixes=('_p1', '_p2'))

    dados_integrados = df_merged[['Operador', 'Linha', 'Ano', 'Mês', 'Quinzena', 'Descrição']].copy()
    dias_comuns = sorted(list(colunas_dia_a_dia.values()))

    for dia in dias_comuns:
        if dia + '_p1' in df_merged.columns and dia + '_p2' in df_merged.columns:
            v1 = pd.to_numeric(df_merged[dia + '_p1'], errors='coerce').fillna(0)
            v2 = pd.to_numeric(df_merged[dia + '_p2'], errors='coerce').fillna(0)
            dados_integrados[dia] = np.where(v1 == v2, -v1, np.where(v1 < v2, -v1, -v2))
        else:
            dados_integrados[dia] = 0.0

    dados_integrados.sort_values(by=['Operador', 'Linha'], inplace=True)

    id_vars = ["Operador", "Linha", "Ano", "Mês"]
    value_vars = [col for col in dados_integrados.columns if col not in id_vars + ['Quinzena', 'Descrição']]
    df_transformado = pd.melt(dados_integrados, id_vars=id_vars, value_vars=value_vars, var_name='Dia', value_name='NMQTDVIAGENSMETA')
    df_transformado = df_transformado[pd.to_numeric(df_transformado['NMQTDVIAGENSMETA'], errors='coerce').fillna(0) != 0]

    df_transformado['Mês'] = df_transformado['Mês'].astype(str).str.zfill(2)
    df_transformado['Dia'] = df_transformado['Dia'].astype(str).str.zfill(2)
    df_transformado['DTOPERACAO'] = df_transformado['Dia'] + '/' + df_transformado['Mês'] + '/' + df_transformado['Ano'].astype(str)
    df_transformado.rename(columns={"Operador": "CDOPERADOR", "Linha": "CDLINHA"}, inplace=True)
    df_transformado["NMQTDVIAGENSMETA"] = df_transformado["NMQTDVIAGENSMETA"].astype(float).map("{:.1f}".format).str.replace(".", ",", regex=False)

    df_transformado['CDOPERADOR'] = pd.Categorical(df_transformado['CDOPERADOR'], categories=ORDEM_EMPRESAS_GLOBAL, ordered=True)
    df_transformado['CDLINHA'] = pd.to_numeric(df_transformado['CDLINHA'], errors='coerce')
    df_transformado.dropna(subset=['CDLINHA'], inplace=True)
    df_transformado.sort_values(by=['CDOPERADOR', 'CDLINHA'], inplace=True)
    df_transformado['CDLINHA'] = df_transformado['CDLINHA'].astype(int).astype(str)

    colunas_finais = ["CDOPERADOR", "CDLINHA", "DTOPERACAO", "NMQTDVIAGENSMETA", "NMFROTAMETA", "NMEXTUTILMETA", "NMEXTMORTAMETA", "NMQTDVIAGENSREF", "NMFROTAREF", "NMEXTUTILREF", "NMEXTMORTAREF", "DSMOTIVO"]
    for col in colunas_finais:
        if col not in df_transformado.columns:
            df_transformado[col] = ""
    df_transformado["DSMOTIVO"] = "Con"

    return dados_integrados, df_transformado[colunas_finais]


def executar_processamento_congestionamento(gui):
    try:
        print("\n--- INICIANDO PROCESSAMENTO (ANÁLISE DE CONGESTIONAMENTO) ---")
        df_resultado_txt = processar_arquivos_txt_congestionamento(gui.arquivos_txt_cong)
        if df_resultado_txt.empty:
            raise ValueError("O processamento dos arquivos TXT não gerou dados.")

        df_resultado_xls = processar_arquivos_xls_congestionamento(gui.arquivos_xls_cong)
        if df_resultado_xls.empty:
            raise ValueError("O processamento dos arquivos XLS não gerou dados.")

        df_integrado, df_final = comparar_e_gerar_final_congestionamento(df_resultado_txt, df_resultado_xls)
        if df_integrado is None or df_final is None:
            raise ValueError("A comparação entre os dados TXT e XLS falhou.")

        pasta_destino = gui.pasta_destino_cong.get()

        caminhos = {
            "txt": os.path.join(pasta_destino, "resultado_final_txt.xlsx"),
            "xls": os.path.join(pasta_destino, "resultado_final_xls_unificado.xlsx"),
            "integrado": os.path.join(pasta_destino, "resultado_final_Integrado.xlsx"),
            "final": os.path.join(pasta_destino, "AlteracoesProgramacao.txt")
        }

        df_resultado_txt['Linha'] = df_resultado_txt['Linha'].astype(int).astype(str)
        df_resultado_txt.to_excel(caminhos["txt"], index=False)
        print(f"Resultado TXT salvo em: {caminhos['txt']}")

        df_resultado_xls.to_excel(caminhos["xls"], index=False)
        print(f"Resultado XLS salvo em: {caminhos['xls']}")

        df_integrado.to_excel(caminhos["integrado"], index=False)
        print(f"Resultado Integrado salvo em: {caminhos['integrado']}")

        df_final.to_csv(caminhos["final"], index=False, sep='\t', encoding='utf-8')
        print(f"Arquivo final salvo em: {caminhos['final']}")

        gui.finalizar_processamento_gui(pasta_destino, sucesso=True)

    except Exception as e:
        print(f"** ERRO CRÍTICO DURANTE A EXECUÇÃO (Congestionamento): {e}")
        traceback.print_exc()
        gui.finalizar_processamento_gui(str(e), sucesso=False)


# =====================================================================================
# --- BLOCO 2: LÓGICA DE PROCESSAMENTO (AJUSTE DE VIAGENS) ---
# =====================================================================================

COLUNA_DATA_BASE = 'DTOPERACAO'
COLUNA_OPERADOR_BASE = 'CDOPERADOR'
COLUNA_LINHA_BASE = 'CDLINHA'
COLUNA_VALOR_BASE = 'NMQTDVIAGENSMETA'
COLUNA_DATA_ALT = 'DATA_ALT'
COLUNA_EMPRESA_ALT = 'EMPRESA_ALT'
COLUNA_LINHA_ALT = 'LINHA_ALT'
COLUNA_VALOR_ALT = 'VALOR_ALT'
NOME_COLUNA_MOTIVO = 'DSMOTIVO'
NOME_ABA_REDUCOES = 'REDUÇÃO_DE_SERVIÇOS'
NOME_ABA_EXTRAS = '_VIAGENS_EXTRAS'
VF_OPERADOR = 'Operador'
VF_LINHA = 'Linha'

def fase0_ler_limites_xlsx(arquivo_limites_xlsx):
    print("--- FASE 0: Lendo Arquivo de Limites (resultado_final_txt.xlsx) ---")
    if not arquivo_limites_xlsx:
        raise ValueError("FASE 0: Nenhum arquivo de Limites selecionado.")

    print(f"   Lendo arquivo: '{os.path.basename(arquivo_limites_xlsx)}'...")
    df_limites = pd.read_excel(arquivo_limites_xlsx, header=0)
    
    df_limites.rename(columns={"Operador": VF_OPERADOR, "Linha": VF_LINHA}, inplace=True)
    
    df_limites.dropna(subset=[VF_OPERADOR, VF_LINHA], inplace=True)
    df_limites[VF_OPERADOR] = df_limites[VF_OPERADOR].astype(str).str.strip().str.upper()
    df_limites[VF_LINHA] = pd.to_numeric(df_limites[VF_LINHA], errors='coerce').astype('Int64').astype(str)

    colunas_dias = [col for col in df_limites.columns if str(col).isnumeric()]
    for col in colunas_dias:
        df_limites[col] = pd.to_numeric(df_limites[col], errors='coerce').fillna(0)
        df_limites[col] = df_limites[col] * -1
        df_limites.rename(columns={col: str(col)}, inplace=True)

    df_limites.set_index([VF_OPERADOR, VF_LINHA], inplace=True)
    
    print(f"--- SUCESSO - FASE 0 --- (Limites lidos para {len(df_limites)} combinações)")
    return df_limites


def fase0_5_ler_unificado(arquivo_unificado_xlsx):
    print("\n--- FASE 0.5: Lendo Arquivo de Verificação 'Unificado' ---")
    if not arquivo_unificado_xlsx:
        print("   FASE 0.5: Nenhum arquivo 'Unificado' selecionado. O processo continuará sem esta verificação.")
        return None

    print(f"   Lendo arquivo: '{os.path.basename(arquivo_unificado_xlsx)}'...")
    df_unificado = pd.read_excel(arquivo_unificado_xlsx, header=0)
    df_unificado.rename(columns={df_unificado.columns[0]: VF_OPERADOR, df_unificado.columns[1]: VF_LINHA}, inplace=True)

    df_unificado[VF_OPERADOR] = df_unificado[VF_OPERADOR].astype(str).str.strip().str.upper()
    df_unificado[VF_LINHA] = df_unificado[VF_LINHA].astype(str).str.extract(r'^(\d+)', expand=False)
    df_unificado.dropna(subset=[VF_OPERADOR, VF_LINHA], inplace=True)
    
    colunas_dias_unif = [col for col in df_unificado.columns if str(col).isnumeric()]
    for col in colunas_dias_unif:
        df_unificado[col] = pd.to_numeric(df_unificado[col], errors='coerce').fillna(0)
        df_unificado.rename(columns={col: str(col)}, inplace=True)

    df_unificado[VF_OPERADOR] = pd.Categorical(df_unificado[VF_OPERADOR], categories=ORDEM_EMPRESAS_GLOBAL, ordered=True)
    df_unificado[VF_LINHA] = pd.to_numeric(df_unificado[VF_LINHA], errors='coerce')
    df_unificado.sort_values(by=[VF_OPERADOR, VF_LINHA], inplace=True)

    df_unificado.set_index([VF_OPERADOR, VF_LINHA], inplace=True)
    
    print(f"--- SUCESSO - FASE 0.5 --- (Arquivo 'Unificado' lido com {len(df_unificado)} linhas)")
    return df_unificado


def fase1_converter_txt_para_excel(arquivo_txt_entrada):
    print("\n--- FASE 1: Iniciando Conversão TXT para XLSX ---")
    if not arquivo_txt_entrada:
        raise ValueError("FASE 1: Nenhum arquivo TXT base selecionado.")

    diretorio, nome_arquivo = os.path.split(arquivo_txt_entrada)
    nome_base, _ = os.path.splitext(nome_arquivo)
    arquivo_xlsx_saida = os.path.join(diretorio, f"{nome_base}.xlsx")

    df = pd.read_csv(arquivo_txt_entrada, sep='\t', header=0, engine='python', encoding='latin-1', decimal=',')
    if COLUNA_DATA_BASE in df.columns:
        df[COLUNA_DATA_BASE] = pd.to_datetime(df[COLUNA_DATA_BASE], errors='coerce', dayfirst=True)
    if COLUNA_VALOR_BASE in df.columns:
        df[COLUNA_VALOR_BASE] = pd.to_numeric(df[COLUNA_VALOR_BASE], errors='coerce').fillna(0.0)

    with pd.ExcelWriter(arquivo_xlsx_saida, engine='openpyxl', datetime_format='dd/mm/yyyy', date_format='dd/mm/yyyy') as writer:
        df.to_excel(writer, index=False, sheet_name='Dados')
        worksheet = writer.sheets['Dados']
        for row_idx, row in enumerate(df.itertuples(index=False), 1):
            for col_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_idx + 1, column=col_idx)
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
    
    print(f"--- SUCESSO - FASE 1 --- (Arquivo '{os.path.basename(arquivo_xlsx_saida)}' criado)")
    return arquivo_xlsx_saida


def tratar_valor_negativo(valor):
    if not isinstance(valor, str): return valor
    valor_str = valor.strip()
    tem_negativo = '-' in valor_str
    numero_limpo_str = valor_str.replace('-', '').replace(',', '.')
    try:
        numero_final = float(numero_limpo_str)
    except (ValueError, TypeError):
        return np.nan
    return -abs(numero_final) if tem_negativo else numero_final


def fase2_aplicar_alteracoes(caminho_xlsx_base, arquivo_alteracoes_ods, limites_reducao_df, df_unificado):
    print("\n--- FASE 2: Aplicar Alterações (com Verificação Unificada) ---")
    if not arquivo_alteracoes_ods:
        raise ValueError("FASE 2: Nenhum arquivo de alterações (.ods) selecionado.")

    log_data = [] # Lista para armazenar os dados do log

    reducoes_df_raw = pd.read_excel(arquivo_alteracoes_ods, engine='odf', sheet_name=NOME_ABA_REDUCOES, skiprows=11, header=None, usecols=[0, 1, 2, 3])
    extras_df_raw = pd.read_excel(arquivo_alteracoes_ods, engine='odf', sheet_name=NOME_ABA_EXTRAS, skiprows=10, header=None, usecols=[0, 1, 2, 3])

    def limpar_dados_ods(df):
        cols = [COLUNA_DATA_ALT, COLUNA_EMPRESA_ALT, COLUNA_LINHA_ALT, COLUNA_VALOR_ALT]
        if df is None or df.empty: return pd.DataFrame(columns=cols)
        df.columns = cols
        df[COLUNA_EMPRESA_ALT] = df[COLUNA_EMPRESA_ALT].astype(str).str.strip().str.upper()
        df[COLUNA_DATA_ALT] = pd.to_datetime(df[COLUNA_DATA_ALT], errors='coerce', dayfirst=True)
        df[COLUNA_VALOR_ALT] = df[COLUNA_VALOR_ALT].astype(str).apply(tratar_valor_negativo).astype(float)
        df.dropna(subset=[COLUNA_DATA_ALT, COLUNA_EMPRESA_ALT, COLUNA_LINHA_ALT, COLUNA_VALOR_ALT], inplace=True)
        df = df[df[COLUNA_EMPRESA_ALT] != '']
        df[COLUNA_LINHA_ALT] = pd.to_numeric(df[COLUNA_LINHA_ALT], errors='coerce').fillna(0).astype(int).astype(str)
        return df

    reducoes_df = limpar_dados_ods(reducoes_df_raw.copy())
    extras_df = limpar_dados_ods(extras_df_raw.copy())
        
    for df, nome_aba in [(reducoes_df, NOME_ABA_REDUCOES), (extras_df, NOME_ABA_EXTRAS)]:
        colunas_chave = [COLUNA_DATA_ALT, COLUNA_EMPRESA_ALT, COLUNA_LINHA_ALT]
        duplicados = df[df.duplicated(subset=colunas_chave, keep=False)]

        if not duplicados.empty:
            duplicados_unicos = duplicados[colunas_chave].drop_duplicates().sort_values(by=colunas_chave)
            mensagem = (f"Erro: Foram encontradas entradas duplicadas com a mesma DATA, "
                        f"EMPRESA e LINHA na aba '{nome_aba}' do arquivo de alterações.\n\n"
                        "Por favor, corrija os seguintes itens:\n")
            for _, row in duplicados_unicos.iterrows():
                data_str = row[COLUNA_DATA_ALT].strftime('%d/%m/%Y')
                mensagem += f"- Data: {data_str}, Empresa: {row[COLUNA_EMPRESA_ALT]}, Linha: {row[COLUNA_LINHA_ALT]}\n"
            
            raise ValueError(mensagem)

    reducoes_agg = reducoes_df.groupby([COLUNA_DATA_ALT, COLUNA_EMPRESA_ALT, COLUNA_LINHA_ALT])[COLUNA_VALOR_ALT].sum().reset_index()
    extras_agg = extras_df.groupby([COLUNA_DATA_ALT, COLUNA_EMPRESA_ALT, COLUNA_LINHA_ALT])[COLUNA_VALOR_ALT].sum().reset_index()

    base_df = pd.read_excel(caminho_xlsx_base, sheet_name='Dados')
    base_df[COLUNA_DATA_BASE] = pd.to_datetime(base_df[COLUNA_DATA_BASE], errors='coerce')
    base_df[COLUNA_OPERADOR_BASE] = base_df[COLUNA_OPERADOR_BASE].astype(str).str.strip().str.upper()
    base_df[COLUNA_LINHA_BASE] = base_df[COLUNA_LINHA_BASE].astype(str).str.extract(r'^(\d+)', expand=False)
    base_df[COLUNA_LINHA_BASE] = pd.to_numeric(base_df[COLUNA_LINHA_BASE], errors='coerce').fillna(0).astype(int).astype(str)
    colunas_originais_base = list(base_df.columns)

    df_final_list = []
    chaves_reducoes = set(tuple(x) for x in reducoes_agg[[COLUNA_DATA_ALT, COLUNA_EMPRESA_ALT, COLUNA_LINHA_ALT]].to_numpy())
    chaves_extras = set(tuple(x) for x in extras_agg[[COLUNA_DATA_ALT, COLUNA_EMPRESA_ALT, COLUNA_LINHA_ALT]].to_numpy())
    chaves_base = set(tuple(x) for x in base_df[[COLUNA_DATA_BASE, COLUNA_OPERADOR_BASE, COLUNA_LINHA_BASE]].to_numpy())
    todas_chaves = sorted(list(chaves_base.union(chaves_reducoes).union(chaves_extras)))

    for data, emp, lin in todas_chaves:
        dia_str = str(data.day)
        
        linha_original = base_df[(base_df[COLUNA_DATA_BASE] == data) & (base_df[COLUNA_OPERADOR_BASE] == emp) & (base_df[COLUNA_LINHA_BASE] == lin)]
        valor_base = 0
        if not linha_original.empty:
            valor_base = linha_original.iloc[0][COLUNA_VALOR_BASE]

        valor_unificado = 0
        if df_unificado is not None:
            try:
                valor_unificado = df_unificado.loc[(emp, lin), dia_str]
            except (KeyError, IndexError):
                valor_unificado = 0
        
        valor_extra = extras_agg[(extras_agg[COLUNA_DATA_ALT] == data) & (extras_agg[COLUNA_EMPRESA_ALT] == emp) & (extras_agg[COLUNA_LINHA_ALT] == lin)][COLUNA_VALOR_ALT].sum()
        valor_reducao = reducoes_agg[(reducoes_agg[COLUNA_DATA_ALT] == data) & (reducoes_agg[COLUNA_EMPRESA_ALT] == emp) & (reducoes_agg[COLUNA_LINHA_ALT] == lin)][COLUNA_VALOR_ALT].sum()
        
        valor_inicial_ajuste = valor_base
        if df_unificado is not None and valor_unificado != 0:
            valor_inicial_ajuste = -abs(valor_unificado)
        
        valor_potencial = valor_inicial_ajuste + valor_extra + valor_reducao

        try:
            limite = limites_reducao_df.loc[(emp, lin), dia_str]
        except (KeyError, IndexError):
            limite = float('-inf')
            if valor_reducao != 0:
                  print(f" --> AVISO: Redução manual para {emp}-{lin} no dia {dia_str} "
                        f"não encontrou dados de limite correspondentes no arquivo de limites. "
                        f"A redução de {valor_reducao} será aplicada integralmente.")
        
        limite_ajustado = min(0, limite)
        
        valor_final = max(valor_potencial, limite_ajustado)
        
        observacao = "Aplicado Integralmente"
        if valor_potencial < limite_ajustado:
            observacao = "Limitado pelo saldo"
            print(f" --> ALERTA: Redução manual para {emp}-{lin} no dia {dia_str} foi limitada. (Limite: {limite_ajustado:.2f}, Redução: {valor_reducao:.2f})")
        elif valor_extra == 0 and valor_reducao == 0:
            observacao = "Sem alteração manual"
            
        log_entry = {
            "Data": data,
            "Empresa": emp,
            "Linha": lin,
            "Valor Base/Unificado": valor_inicial_ajuste,
            "Redução ODS": valor_reducao,
            "Extra ODS": valor_extra,
            "Valor Potencial": valor_potencial,
            "Limite Disponível": limite_ajustado,
            "Valor Final Calculado": valor_final,
            "Observação": observacao
        }
        log_data.append(log_entry)

        if not linha_original.empty:
            nova_linha = linha_original.iloc[0].to_dict()
            nova_linha[COLUNA_VALOR_BASE] = valor_final
        else:
            nova_linha = {c: pd.NA for c in colunas_originais_base}
            nova_linha.update({
                COLUNA_DATA_BASE: data, COLUNA_OPERADOR_BASE: emp,
                COLUNA_LINHA_BASE: lin, COLUNA_VALOR_BASE: valor_final,
                NOME_COLUNA_MOTIVO: 'Con'
            })
        df_final_list.append(nova_linha)

    df_final = pd.DataFrame(df_final_list)
    df_final[COLUNA_LINHA_BASE] = pd.to_numeric(df_final[COLUNA_LINHA_BASE], errors='coerce')
    df_final[COLUNA_OPERADOR_BASE] = pd.Categorical(df_final[COLUNA_OPERADOR_BASE], categories=ORDEM_EMPRESAS_GLOBAL, ordered=True)
    df_final.sort_values(by=[COLUNA_OPERADOR_BASE, COLUNA_LINHA_BASE, COLUNA_DATA_BASE], inplace=True, na_position='last')
    df_final = df_final[df_final[COLUNA_VALOR_BASE] != 0].copy()

    with pd.ExcelWriter(caminho_xlsx_base, engine='openpyxl', datetime_format='dd/mm/yyyy', date_format='dd/mm/yyyy') as writer:
        df_final.to_excel(writer, index=False, sheet_name='Dados')
        
    log_df = pd.DataFrame(log_data)

    print("--- SUCESSO - FASE 2 ---")
    return caminho_xlsx_base, log_df


def fase3_converter_xlsx_para_txt(caminho_xlsx_final, diretorio_saida):
    if not caminho_xlsx_final: raise ValueError("FASE 3: Caminho do arquivo Excel final não fornecido.")
    nome_base_xlsx = os.path.splitext(os.path.basename(caminho_xlsx_final))[0]
    arquivo_txt_saida = os.path.join(diretorio_saida, f"{nome_base_xlsx}_AJUSTADO.txt")

    df_final = pd.read_excel(caminho_xlsx_final, sheet_name='Dados')
    if COLUNA_LINHA_BASE in df_final.columns:
        df_final[COLUNA_LINHA_BASE] = pd.to_numeric(df_final[COLUNA_LINHA_BASE], errors='coerce').astype('Int64')
    if COLUNA_DATA_BASE in df_final.columns:
        df_final[COLUNA_DATA_BASE] = pd.to_datetime(df_final[COLUNA_DATA_BASE]).dt.strftime('%d/%m/%Y')
    
    df_final.to_csv(arquivo_txt_saida, sep='\t', index=False, header=True, encoding='utf-8', decimal=',', na_rep='')
    print(f"\n--- SUCESSO - FASE 3 --- (Arquivo TXT Final salvo em: {arquivo_txt_saida})")
    return arquivo_txt_saida


def executar_processamento_ajustes(gui):
    try:
        limites = fase0_ler_limites_xlsx(gui.ajuste_arquivo_limites_xlsx)
        unificado_df = fase0_5_ler_unificado(gui.ajuste_arquivo_unificado)
        xlsx_fase1 = fase1_converter_txt_para_excel(gui.ajuste_arquivo_base_txt)
        
        xlsx_fase2, log_df = fase2_aplicar_alteracoes(xlsx_fase1, gui.ajuste_arquivo_alteracoes_ods, limites, unificado_df)
        
        pasta_destino = gui.pasta_destino_ajuste.get()
        caminho_final_txt = fase3_converter_xlsx_para_txt(xlsx_fase2, pasta_destino)

        if not log_df.empty:
            log_path = os.path.join(pasta_destino, "log_detalhado_ajustes.xlsx")
            log_df['Data'] = log_df['Data'].dt.strftime('%d/%m/%Y')
            log_df.sort_values(by=["Empresa", "Linha", "Data"], inplace=True)
            log_df.to_excel(log_path, index=False)
            print(f"\n--- SUCESSO - LOG --- (Log de processamento detalhado salvo em: {log_path})")

        gui.finalizar_processamento_gui(os.path.dirname(caminho_final_txt), sucesso=True, modulo="Ajuste")

    except Exception as e:
        print(f"** ERRO CRÍTICO DURANTE A EXECUÇÃO (Ajustes): {e}")
        traceback.print_exc()
        gui.finalizar_processamento_gui(str(e), sucesso=False, modulo="Ajuste")

# =====================================================================================
# --- BLOCO 3: LÓGICA DE PROCESSAMENTO (E.P. - VIAGEM E FROTA) ---
# =====================================================================================

EP_OPERADORES = ["BOA", "CAX", "CSR", "EME", "GLO", "SJT", "VML", "Todas"]
EP_ORDEM_EMPRESAS = ["BOA", "CAX", "CSR", "EME", "GLO", "SJT", "VML"]
EP_COLUNAS_SAIDA = [
    "CDOPERADOR", "CDLINHA", "DTOPERACAO", "NMQTDVIAGENSMETA",
    "NMFROTAMETA", "NMEXTUTILMETA", "NMEXTMORTAMETA",
    "NMQTDVIAGENSREF", "NMFROTAREF", "NMEXTUTILREF",
    "NMEXTMORTAREF", "DSMOTIVO"
]

def ep_convert_to_number(valor):
    if isinstance(valor, str):
        try:
            return float(valor.replace(',', '.', 1))
        except ValueError:
            return valor
    return valor

def ep_extrair_codigo_linha(texto_coluna_linha):
    if pd.isna(texto_coluna_linha) or texto_coluna_linha == "":
        return ""
    return str(texto_coluna_linha).split(" - ")[0].strip()

def ep_processar_dados(arquivos_txt_viagem_frota, operador_selecionado, linhas_selecionadas_str,
                       dias_selecionados, arquivo_saida_path, gui_ref=None):
    try:
        if not dias_selecionados:
            messagebox.showerror("Erro de Entrada", "Nenhum dia foi selecionado para o processamento.", parent=gui_ref)
            return

        colunas_dias_processar = [str(d) for d in dias_selecionados]
        dados_finais_para_txt = []
        
        for idx, arquivo_txt in enumerate(arquivos_txt_viagem_frota):
            print(f"Processando {idx+1}/{len(arquivos_txt_viagem_frota)}: {os.path.basename(arquivo_txt)}")
            try:
                nome_base = os.path.basename(arquivo_txt)
                partes_nome = nome_base.split('-')
                op = partes_nome[0].strip().upper()
                data_part = partes_nome[1].strip()
                ano = "20" + data_part[0:2]
                mes = data_part[2:4]
                if not (op in EP_OPERADORES and len(op) == 3 and ano.isdigit() and mes.isdigit()):
                    raise ValueError("Formato do nome do arquivo inválido.")
            except (IndexError, ValueError):
                messagebox.showwarning("Nome de Arquivo Inválido", f"O nome do arquivo '{os.path.basename(arquivo_txt)}' não segue o padrão 'EMPRESA-AAMMQ-Texto.txt'.\n\nPulando este arquivo.", parent=gui_ref)
                continue

            if operador_selecionado != "Todas" and op != operador_selecionado:
                continue

            try:
                with open(arquivo_txt, 'r', encoding='latin1') as f:
                    linhas_arquivo = f.readlines()
                indice_cabecalho = -1
                for i, linha_str in enumerate(linhas_arquivo):
                    if "Linha" in linha_str and "Descrição" in linha_str:
                        indice_cabecalho = i
                        break
                if indice_cabecalho == -1:
                    messagebox.showwarning("Aviso", f"Cabeçalho não encontrado em {os.path.basename(arquivo_txt)}. Pulando.", parent=gui_ref)
                    continue
                dados_csv = StringIO("".join(linhas_arquivo[indice_cabecalho:]))
                df = pd.read_csv(dados_csv, sep='\t', dtype=str)
            except Exception as e:
                messagebox.showerror("Erro ao Ler Arquivo", f"Erro ao ler {os.path.basename(arquivo_txt)}: {e}", parent=gui_ref)
                continue

            df.columns = [str(col).strip() for col in df.columns]
            df['CodigoLinhaExtraido'] = df['Linha'].apply(ep_extrair_codigo_linha)

            if linhas_selecionadas_str.upper() != "TODAS":
                linhas_para_filtrar = [linha.strip() for linha in linhas_selecionadas_str.split(',') if linha.strip()]
                df = df[df['CodigoLinhaExtraido'].isin(linhas_para_filtrar)]
            if df.empty: continue

            colunas_dias_presentes_df = [d for d in colunas_dias_processar if d in df.columns]
            for dia_col in colunas_dias_presentes_df:
                df[dia_col] = df[dia_col].apply(ep_convert_to_number)
                df[dia_col] = pd.to_numeric(df[dia_col], errors='coerce').fillna(0.0)

            for codigo_linha, group_df in df.groupby('CodigoLinhaExtraido'):
                if not codigo_linha: continue
                
                viagem_prog_row = group_df[group_df['Descrição'].str.contains("Viagem Programada", case=False, na=False)]
                viagem_real_row = group_df[group_df['Descrição'].str.contains("Viagem Realizada", case=False, na=False)]
                frota_prog_row = group_df[group_df['Descrição'].str.contains("Frota Programada", case=False, na=False)]
                frota_real_row = group_df[group_df['Descrição'].str.contains("Frota Realizada", case=False, na=False)]

                if viagem_prog_row.empty or viagem_real_row.empty: 
                    continue
                
                viagem_prog_data = viagem_prog_row.iloc[0]
                viagem_real_data = viagem_real_row.iloc[0]
                frota_prog_data = frota_prog_row.iloc[0] if not frota_prog_row.empty else None
                frota_real_data = frota_real_row.iloc[0] if not frota_real_row.empty else None

                for dia_str in colunas_dias_presentes_df:
                    val_viagem_prog = viagem_prog_data.get(dia_str, 0.0)
                    val_viagem_real = viagem_real_data.get(dia_str, 0.0)
                    diferenca_viagem = val_viagem_real - val_viagem_prog
                    
                    if diferenca_viagem >= 0:
                        continue

                    nmqtdviagensmeta_formatado = "{:.2f}".format(diferenca_viagem).replace('.', ',')
                    
                    val_frota_prog = frota_prog_data.get(dia_str, 0.0) if frota_prog_data is not None else 0.0
                    val_frota_real = frota_real_data.get(dia_str, 0.0) if frota_real_data is not None else 0.0
                    diferenca_frota = val_frota_real - val_frota_prog

                    nmfrotameta_formatado = ""
                    if diferenca_frota < 0:
                        nmfrotameta_formatado = str(int(diferenca_frota))
                    
                    dt_operacao = f"{str(dia_str).zfill(2)}/{mes}/{ano}"
                    dados_finais_para_txt.append({
                        "CDOPERADOR": op, "CDLINHA": codigo_linha, "DTOPERACAO": dt_operacao,
                        "NMQTDVIAGENSMETA": nmqtdviagensmeta_formatado, 
                        "NMFROTAMETA": nmfrotameta_formatado,
                        "NMEXTUTILMETA": "", "NMEXTMORTAMETA": "", "NMQTDVIAGENSREF": "", 
                        "NMFROTAREF": "", "NMEXTUTILREF": "", "NMEXTMORTAREF": "", 
                        "DSMOTIVO": "Con"
                    })

        if not dados_finais_para_txt:
            messagebox.showinfo("Concluído", "Nenhum dado encontrado para gerar o arquivo com os critérios atuais.", parent=gui_ref)
            return

        df_saida = pd.DataFrame(dados_finais_para_txt, columns=EP_COLUNAS_SAIDA)

        if not df_saida.empty:
            df_saida['CDOPERADOR'] = pd.Categorical(df_saida['CDOPERADOR'], categories=EP_ORDEM_EMPRESAS, ordered=True)
            df_saida['CDLINHA'] = pd.to_numeric(df_saida['CDLINHA'], errors='coerce')
            df_saida['DTOPERACAO_SORT'] = pd.to_datetime(df_saida['DTOPERACAO'], format='%d/%m/%Y', errors='coerce')
            df_saida = df_saida.sort_values(by=['CDOPERADOR', 'CDLINHA', 'DTOPERACAO_SORT'])
            df_saida = df_saida.drop(columns=['DTOPERACAO_SORT'])

        df_saida.to_csv(arquivo_saida_path, sep='\t', index=False, encoding='utf-8', lineterminator='\n')
        messagebox.showinfo("Sucesso", f"Arquivo '{os.path.basename(arquivo_saida_path)}' gerado com sucesso em:\n{arquivo_saida_path}", parent=gui_ref)

    except Exception as e:
        messagebox.showerror("Erro Inesperado", f"Ocorreu um erro no processamento E.P.: {e}", parent=gui_ref)
        traceback.print_exc()

# =====================================================================================
# --- BLOCO 4: INTERFACE GRÁFICA UNIFICADA (GUI) ---
# =====================================================================================

class EP_DialogoSelecaoDias(tk.Toplevel):
    def __init__(self, parent, dias_pre_selecionados):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        self.title("Selecionar Dias")
        self.resultado = None

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(expand=True, fill=tk.BOTH)
        
        ttk.Label(main_frame, text="Marque os dias que deseja processar:").pack(pady=(0, 10))
        grid_frame = ttk.Frame(main_frame)
        grid_frame.pack()

        self.vars_dias = []
        for i in range(1, 32):
            var = tk.BooleanVar(value=(i in dias_pre_selecionados))
            chk = ttk.Checkbutton(grid_frame, text=str(i), variable=var)
            row, col = divmod(i - 1, 7)
            chk.grid(row=row, column=col, padx=5, pady=5, sticky="w")
            self.vars_dias.append(var)

        botoes_frame = ttk.Frame(self, padding=(10, 15))
        botoes_frame.pack(fill=tk.X)
        botoes_frame.columnconfigure(0, weight=1)
        
        btn_ok = ttk.Button(botoes_frame, text="OK", command=self.on_ok, style='Accent.TButton')
        btn_ok.pack(side=tk.RIGHT, padx=5)
        
        btn_cancel = ttk.Button(botoes_frame, text="Cancelar", command=self.on_cancel)
        btn_cancel.pack(side=tk.RIGHT)
        
        self.wait_window(self)

    def on_ok(self):
        self.resultado = [i + 1 for i, var in enumerate(self.vars_dias) if var.get()]
        self.destroy()

    def on_cancel(self):
        self.destroy()

class AppGUI(TkinterDnD.Tk):
    def __init__(self, title="Gerenciador de Congestionamento e Ajustes", size=(1000, 800)):
        super().__init__()
        self.title(title)
        self.geometry(f'{size[0]}x{size[1]}')
        self.resizable(True, True)

        self.style = ttk.Style(self)
        self.style.configure('Accent.TButton', font=('Segoe UI', 10, 'bold'))
        self.style.configure('TLabel', font=('Segoe UI', 10))
        self.style.configure('TLabelframe.Label', font=('Segoe UI', 11, 'bold'))
        self.style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'))
        
        self.center_window()

        self.LOGO_IMAGE_PATH = "static/images/rcb_logo.png"
        
        self.arquivos_txt_cong = []
        self.arquivos_xls_cong = []
        self.pasta_destino_cong = tk.StringVar()

        self.ajuste_arquivo_limites_xlsx = ""
        self.ajuste_arquivo_unificado = ""
        self.ajuste_arquivo_base_txt = ""
        self.ajuste_arquivo_alteracoes_ods = ""
        self.pasta_destino_ajuste = tk.StringVar()
        
        self.ep_arquivos_txt_selecionados = []
        self.ep_dias_selecionados = list(range(1, 16))
        self.ep_operador_var = tk.StringVar(value=EP_OPERADORES[-1])
        self.ep_linhas_var = tk.StringVar(value="Todas")
        self.ep_nome_arquivo_saida_var = tk.StringVar(value="AlteracaoProgramacao")
        self.ep_local_salvar_var = tk.StringVar(value=os.getcwd())


        self.create_widgets()
        sv_ttk.set_theme("light")

    def toggle_theme(self):
        sv_ttk.toggle_theme()

    def create_widgets(self):
        main_frame = ttk.Frame(self, padding=(20, 15))
        main_frame.pack(fill=tk.BOTH, expand=True)

        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 20))

        try:
            logo_original = Image.open(self.LOGO_IMAGE_PATH).convert("RGBA")
            logo_resized = logo_original.resize((int(logo_original.width * 0.4), int(logo_original.height * 0.4)), Image.Resampling.LANCZOS)
            self.logo_image = ImageTk.PhotoImage(logo_resized)
            ttk.Label(header_frame, image=self.logo_image).pack(side=tk.LEFT, padx=(0, 15))
        except Exception as e:
            print(f"AVISO: Não foi possível carregar a imagem do logo: {e}")
        
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(title_frame, text="Gerenciador de Congestionamento e Ajustes", style="Title.TLabel").pack(anchor='w')
        ttk.Label(title_frame, text="Ferramenta para análise e ajuste de dados de operação", style="TLabel").pack(anchor='w')
        
        theme_switch = ttk.Checkbutton(header_frame, text="Tema", style="Switch.TCheckbutton", command=self.toggle_theme)
        theme_switch.pack(side=tk.RIGHT, padx=10)

        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=10)

        tab_congestionamento = ttk.Frame(notebook, padding=10)
        tab_ajuste = ttk.Frame(notebook, padding=10)
        tab_ep = ttk.Frame(notebook, padding=10)
        
        notebook.add(tab_congestionamento, text='   Análise de Congestionamento   ')
        notebook.add(tab_ajuste, text='   Ajuste de Viagens   ')
        notebook.add(tab_ep, text='   E.P. (Viagem e Frota)   ')

        self._criar_aba_congestionamento(tab_congestionamento)
        self._criar_aba_ajuste_viagens(tab_ajuste)
        self._criar_aba_ep(tab_ep)

    def _criar_aba_congestionamento(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)

        button_area = ttk.Frame(parent, padding=(0, 10, 0, 0))
        button_area.pack(fill=tk.X, side=tk.BOTTOM)
        self.botao_limpar_cong = ttk.Button(button_area, text="Limpar Tudo", command=self.limpar_selecoes_congestionamento)
        self.botao_limpar_cong.pack(side=tk.LEFT, padx=(0, 10))
        self.botao_gerar_cong = ttk.Button(button_area, text="Processar e Gerar Relatórios", style='Accent.TButton', command=self.iniciar_processamento_congestionamento)
        self.botao_gerar_cong.pack(side=tk.RIGHT, fill=tk.X, expand=True)

        notebook_interno = ttk.Notebook(parent)
        notebook_interno.pack(fill=tk.BOTH, expand=True, pady=5)
        
        tab_entrada = ttk.Frame(notebook_interno, padding=10)
        tab_saida = ttk.Frame(notebook_interno, padding=10)
        notebook_interno.add(tab_entrada, text='   Arquivos de Entrada   ')
        notebook_interno.add(tab_saida, text='   Saída e Geração   ')

        tab_entrada.columnconfigure((0, 1), weight=1, uniform="group1")
        tab_entrada.rowconfigure(0, weight=1)
        self.lista_txt_tree_cong = self._criar_painel_arquivos(
            tab_entrada, "Arquivos de Saldo (.txt)", self.arquivos_txt_cong, 0, [("Arquivos TXT", "*.txt")], "cong_txt"
        )
        self.lista_xls_tree_cong = self._criar_painel_arquivos(
            tab_entrada, "Arquivos de Referência (.xls)", self.arquivos_xls_cong, 1, [("Arquivos XLS", "*.xls")], "cong_xls"
        )

        tab_saida.columnconfigure(0, weight=1)
        frame_saida = ttk.LabelFrame(tab_saida, text=" Configuração do Arquivo de Saída ", padding=15)
        frame_saida.pack(fill=tk.BOTH, expand=True)
        frame_saida.columnconfigure(1, weight=1)
        
        ttk.Button(frame_saida, text="Selecionar Pasta de Destino", command=lambda: self.selecionar_pasta_destino("cong")).grid(row=0, column=0, padx=(0,10), pady=5, sticky='w')
        self.label_pasta_cong = ttk.Label(frame_saida, text="Nenhuma pasta selecionada.")
        self.label_pasta_cong.grid(row=0, column=1, pady=5, sticky='ew')
        
        info_label = ttk.Label(frame_saida, text="\nOs seguintes arquivos serão gerados na pasta de destino:\n\n• resultado_final_txt.xlsx\n• resultado_final_xls_unificado.xlsx\n• resultado_final_Integrado.xlsx\n• AlteracoesProgramacao.txt", justify=tk.LEFT)
        info_label.grid(row=1, column=0, columnspan=2, sticky='w', pady=10)

    def _criar_painel_arquivos(self, parent, titulo, lista_arquivos, col, filetypes, tipo):
        frame = ttk.LabelFrame(parent, text=f" {titulo} ", padding=10)
        frame.grid(row=0, column=col, padx=10, pady=5, sticky='nsew')
        frame.grid_rowconfigure(1, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        ttk.Label(frame, text="Arraste e solte arquivos aqui ou use os botões.", justify=tk.CENTER).grid(row=0, column=0, columnspan=2, pady=(0, 10))
        
        tree = ttk.Treeview(frame, columns=('filename',), show='headings', height=8)
        tree.heading('filename', text='Arquivos Selecionados')
        tree.grid(row=1, column=0, columnspan=2, sticky='nsew')
        
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=1, column=2, sticky='ns')

        tree.drop_target_register(DND_FILES)
        tree.dnd_bind('<<Drop>>', lambda e: self._adicionar_arquivos(self.tk.splitlist(e.data), lista_arquivos, tree, filetypes[0][1].split('*')[1]))
        
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=2, column=0, columnspan=2, sticky='ew', pady=(10,0))
        btn_frame.columnconfigure((0, 1), weight=1)

        btn_selecionar = ttk.Button(btn_frame, text="Selecionar", command=lambda: self._adicionar_arquivos(filedialog.askopenfilenames(title=f"Selecione - {titulo}", filetypes=filetypes), lista_arquivos, tree, filetypes[0][1].split('*')[1]))
        btn_selecionar.grid(row=0, column=0, sticky='ew', padx=(0,5))
        btn_limpar = ttk.Button(btn_frame, text="Limpar", command=lambda: self._limpar_lista(lista_arquivos, tree))
        btn_limpar.grid(row=0, column=1, sticky='ew', padx=(5,0))
        return tree

    def _criar_aba_ajuste_viagens(self, parent):
        parent.columnconfigure(0, weight=1)

        frame = ttk.LabelFrame(parent, text=" Fluxo de Trabalho para Ajuste de Viagens ", padding=15)
        frame.pack(fill=tk.X, pady=5)
        frame.columnconfigure(1, weight=1)
        
        ttk.Label(frame, text="Fase 0: Arquivo de Limites (.xlsx)").grid(row=0, column=0, sticky='w', padx=5, pady=8)
        self.label_ajuste_limites = ttk.Label(frame, text="Nenhum arquivo (resultado_final_txt.xlsx) selecionado.")
        self.label_ajuste_limites.grid(row=0, column=1, sticky='ew', padx=5, pady=8)
        ttk.Button(frame, text="Selecionar...", command=lambda: self.selecionar_arquivos_ajuste("limites_xlsx")).grid(row=0, column=2, sticky='e', padx=5, pady=8)
        
        ttk.Label(frame, text="Fase 0.5: Arquivo Unificado (.xlsx)").grid(row=1, column=0, sticky='w', padx=5, pady=8)
        self.label_ajuste_unificado = ttk.Label(frame, text="Nenhum arquivo selecionado.")
        self.label_ajuste_unificado.grid(row=1, column=1, sticky='ew', padx=5, pady=8)
        ttk.Button(frame, text="Selecionar...", command=lambda: self.selecionar_arquivos_ajuste("unificado")).grid(row=1, column=2, sticky='e', padx=5, pady=8)
        
        ttk.Label(frame, text="Fase 1: TXT Base de Operações").grid(row=2, column=0, sticky='w', padx=5, pady=8)
        self.label_ajuste_base_txt = ttk.Label(frame, text="Nenhum arquivo selecionado.")
        self.label_ajuste_base_txt.grid(row=2, column=1, sticky='ew', padx=5, pady=8)
        ttk.Button(frame, text="Selecionar...", command=lambda: self.selecionar_arquivos_ajuste("base_txt")).grid(row=2, column=2, sticky='e', padx=5, pady=8)
        
        ttk.Label(frame, text="Fase 2: Arquivo de Alterações (.ods)").grid(row=3, column=0, sticky='w', padx=5, pady=8)
        self.label_ajuste_alteracoes_ods = ttk.Label(frame, text="Nenhum arquivo selecionado.")
        self.label_ajuste_alteracoes_ods.grid(row=3, column=1, sticky='ew', padx=5, pady=8)
        ttk.Button(frame, text="Selecionar...", command=lambda: self.selecionar_arquivos_ajuste("alteracoes_ods")).grid(row=3, column=2, sticky='e', padx=5, pady=8)
        
        ttk.Label(frame, text="Pasta de Destino Final").grid(row=4, column=0, sticky='w', padx=5, pady=8)
        self.label_pasta_ajuste = ttk.Label(frame, text="Nenhuma pasta selecionada.")
        self.label_pasta_ajuste.grid(row=4, column=1, sticky='ew', padx=5, pady=8)
        ttk.Button(frame, text="Selecionar...", command=lambda: self.selecionar_pasta_destino("ajuste")).grid(row=4, column=2, sticky='e', padx=5, pady=8)
        
        button_area = ttk.Frame(parent, padding=(0, 15, 0, 0))
        button_area.pack(fill=tk.X)
        self.botao_limpar_ajuste = ttk.Button(button_area, text="Limpar Tudo", command=self.limpar_selecoes_ajuste)
        self.botao_limpar_ajuste.pack(side=tk.LEFT, padx=(0, 10))
        self.botao_gerar_ajuste = ttk.Button(button_area, text="Iniciar Processo de Ajuste", style='Accent.TButton', command=self.iniciar_processamento_ajustes)
        self.botao_gerar_ajuste.pack(side=tk.RIGHT, fill=tk.X, expand=True)

    def _criar_aba_ep(self, parent):
        parent.columnconfigure(0, weight=1)

        button_area = ttk.Frame(parent, padding=(0, 15, 0, 0))
        button_area.pack(fill=tk.X, side=tk.BOTTOM)
        self.botao_limpar_ep = ttk.Button(button_area, text="Limpar Tudo", command=self.ep_limpar_formulario)
        self.botao_limpar_ep.pack(side=tk.LEFT, padx=(0, 10))
        self.botao_gerar_ep = ttk.Button(button_area, text="Processar Arquivos", style='Accent.TButton', command=self.iniciar_processamento_ep)
        self.botao_gerar_ep.pack(side=tk.RIGHT, fill=tk.X, expand=True)

        content_frame = ttk.Frame(parent)
        content_frame.pack(fill=tk.BOTH, expand=True)
        content_frame.columnconfigure(0, weight=1)
        
        dias_frame = ttk.LabelFrame(content_frame, text="Dias a Processar", padding=10)
        dias_frame.pack(fill=tk.X, pady=5, padx=3)
        ttk.Button(dias_frame, text="Selecionar Dias...", command=self.ep_abrir_dialogo_selecao_dias).pack(side=tk.LEFT, padx=(0,10))
        self.ep_lbl_dias_selecionados = ttk.Label(dias_frame, text="", wraplength=700)
        self.ep_lbl_dias_selecionados.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.ep_atualizar_label_dias()

        entrada_frame = ttk.LabelFrame(content_frame, text="Arquivo(s) TXT de Viagem e Frota", padding=10)
        entrada_frame.pack(fill=tk.X, pady=5, padx=3)
        self.ep_lbl_arquivos_selecionados = ttk.Label(entrada_frame, text="Nenhum arquivo selecionado", wraplength=800)
        self.ep_lbl_arquivos_selecionados.pack(pady=(3,5), fill=tk.X, padx=3)
        ttk.Button(entrada_frame, text="Selecionar Arquivo(s) TXT", command=self.ep_selecionar_arquivos).pack(pady=3, padx=3)

        filtros_frame = ttk.LabelFrame(content_frame, text="Filtros", padding=10)
        filtros_frame.pack(fill=tk.X, pady=5, padx=3)
        filtros_frame.columnconfigure(1, weight=1)
        ttk.Label(filtros_frame, text="Operador:").grid(row=0, column=0, padx=(0,5), pady=5, sticky=tk.W)
        operador_dropdown = ttk.Combobox(filtros_frame, textvariable=self.ep_operador_var, values=EP_OPERADORES, state="readonly", width=15)
        operador_dropdown.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        ttk.Label(filtros_frame, text="Linhas (sep. por vírgula, ou 'Todas'):").grid(row=1, column=0, padx=(0,5), pady=5, sticky=tk.W)
        ttk.Entry(filtros_frame, textvariable=self.ep_linhas_var).grid(row=1, column=1, padx=5, pady=5, sticky=tk.EW)

        saida_frame = ttk.LabelFrame(content_frame, text="Arquivo de Saída", padding=10)
        saida_frame.pack(fill=tk.X, pady=5, padx=3)
        saida_frame.columnconfigure(1, weight=1)
        ttk.Label(saida_frame, text="Nome do Arquivo:").grid(row=0, column=0, padx=(0,5), pady=5, sticky=tk.W)
        entry_nome_saida = ttk.Entry(saida_frame, textvariable=self.ep_nome_arquivo_saida_var)
        entry_nome_saida.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Label(saida_frame, text=".txt").grid(row=0, column=2, padx=(0,5), pady=5, sticky=tk.W)
        
        ttk.Label(saida_frame, text="Local para Salvar:").grid(row=1, column=0, padx=(0,5), pady=5, sticky=tk.W)
        local_entry_frame = ttk.Frame(saida_frame)
        local_entry_frame.grid(row=1, column=1, columnspan=2, sticky=tk.EW, padx=5, pady=5)
        local_entry_frame.columnconfigure(0, weight=1)
        local_entry = ttk.Entry(local_entry_frame, textvariable=self.ep_local_salvar_var, state="readonly")
        local_entry.grid(row=0, column=0, sticky=tk.EW, padx=(0,5))
        ttk.Button(local_entry_frame, text="...", command=self.ep_selecionar_local_salvamento, width=4).grid(row=0, column=1)

    def _adicionar_arquivos(self, paths, lista_alvo, treeview, extension):
        novos_arquivos = False
        for path in paths:
            if path.lower().endswith(extension) and path not in lista_alvo:
                lista_alvo.append(path)
                novos_arquivos = True
        if novos_arquivos:
            self._atualizar_treeview(treeview, lista_alvo)

    def _limpar_lista(self, lista_alvo, treeview):
        lista_alvo.clear()
        self._atualizar_treeview(treeview, lista_alvo)

    def _atualizar_treeview(self, treeview, lista_arquivos):
        for i in treeview.get_children():
            treeview.delete(i)
        for path in sorted(lista_arquivos):
            treeview.insert("", tk.END, values=(os.path.basename(path),))
    
    def selecionar_arquivos_ajuste(self, tipo):
        if tipo == "limites_xlsx":
            path = filedialog.askopenfilename(title="Selecione o arquivo de Limites (.xlsx)", filetypes=[("Arquivos Excel", "*.xlsx")])
            if path:
                self.ajuste_arquivo_limites_xlsx = path
                self.label_ajuste_limites.config(text=os.path.basename(path))
        elif tipo == "unificado":
            path = filedialog.askopenfilename(title="Selecione o arquivo Unificado (.xlsx)", filetypes=[("Arquivos Excel", "*.xlsx")])
            if path:
                self.ajuste_arquivo_unificado = path
                self.label_ajuste_unificado.config(text=os.path.basename(path))
        elif tipo == "base_txt":
            path = filedialog.askopenfilename(title="Selecione o TXT Base de Operações", filetypes=[("Arquivos TXT", "*.txt")])
            if path:
                self.ajuste_arquivo_base_txt = path
                self.label_ajuste_base_txt.config(text=os.path.basename(path))
        elif tipo == "alteracoes_ods":
            path = filedialog.askopenfilename(title="Selecione o arquivo de Alterações (.ods)", filetypes=[("Planilha ODS", "*.ods")])
            if path:
                self.ajuste_arquivo_alteracoes_ods = path
                self.label_ajuste_alteracoes_ods.config(text=os.path.basename(path))

    def selecionar_pasta_destino(self, modulo):
        caminho = filedialog.askdirectory(title="Selecione a pasta para salvar os relatórios")
        if caminho:
            if modulo == "cong":
                self.pasta_destino_cong.set(caminho)
                self.label_pasta_cong.config(text=caminho)
            elif modulo == "ajuste":
                self.pasta_destino_ajuste.set(caminho)
                self.label_pasta_ajuste.config(text=caminho)
    
    def limpar_selecoes_congestionamento(self):
        self._limpar_lista(self.arquivos_txt_cong, self.lista_txt_tree_cong)
        self._limpar_lista(self.arquivos_xls_cong, self.lista_xls_tree_cong)
        self.pasta_destino_cong.set("")
        self.label_pasta_cong.config(text="Nenhuma pasta selecionada.")
        messagebox.showinfo("Limpeza", "Todos os campos da Análise de Congestionamento foram limpos.", parent=self)

    def limpar_selecoes_ajuste(self):
        self.ajuste_arquivo_limites_xlsx = ""
        self.ajuste_arquivo_unificado = ""
        self.ajuste_arquivo_base_txt = ""
        self.ajuste_arquivo_alteracoes_ods = ""
        self.pasta_destino_ajuste.set("")
        self.label_ajuste_limites.config(text="Nenhum arquivo (resultado_final_txt.xlsx) selecionado.")
        self.label_ajuste_unificado.config(text="Nenhum arquivo selecionado.")
        self.label_ajuste_base_txt.config(text="Nenhum arquivo selecionado.")
        self.label_ajuste_alteracoes_ods.config(text="Nenhum arquivo selecionado.")
        self.label_pasta_ajuste.config(text="Nenhuma pasta selecionada.")
        messagebox.showinfo("Limpeza", "Todos os campos do Ajuste de Viagens foram limpos.", parent=self)
        
    def iniciar_processamento_congestionamento(self):
        if not self.arquivos_txt_cong or not self.arquivos_xls_cong:
            messagebox.showwarning("Atenção", "É necessário selecionar arquivos .txt e .xls para continuar.", parent=self)
            return
        if not self.pasta_destino_cong.get():
            messagebox.showwarning("Atenção", "Selecione uma pasta de destino para salvar os relatórios.", parent=self)
            return

        self.botao_gerar_cong.config(state="disabled", text="Processando...")
        self.botao_limpar_cong.config(state="disabled")
        self.update_idletasks()
        
        threading.Thread(target=executar_processamento_congestionamento, args=(self,)).start()
        
    def iniciar_processamento_ajustes(self):
        if not all([self.ajuste_arquivo_limites_xlsx, self.ajuste_arquivo_base_txt, self.ajuste_arquivo_alteracoes_ods, self.pasta_destino_ajuste.get()]):
            messagebox.showwarning("Atenção", "Os campos 'Arquivo de Limites', 'TXT Base', 'Alterações' e 'Pasta de Destino' devem ser preenchidos.", parent=self)
            return

        self.botao_gerar_ajuste.config(state="disabled", text="Processando...")
        self.botao_limpar_ajuste.config(state="disabled")
        self.update_idletasks()

        threading.Thread(target=executar_processamento_ajustes, args=(self,)).start()
    
    def finalizar_processamento_gui(self, mensagem, sucesso, modulo="Congestionamento"):
        if modulo == "Congestionamento":
            self.botao_gerar_cong.config(state="normal", text="Processar e Gerar Relatórios")
            self.botao_limpar_cong.config(state="normal")
        elif modulo == "Ajuste":
            self.botao_gerar_ajuste.config(state="normal", text="Iniciar Processo de Ajuste")
            self.botao_limpar_ajuste.config(state="normal")
        elif modulo == "EP":
            self.botao_gerar_ep.config(state="normal", text="Processar Arquivos")
            self.botao_limpar_ep.config(state="normal")


        if sucesso:
            if "dummy" in modulo: return
            messagebox.showinfo("Sucesso!", f"Módulo '{modulo}':\n\nRelatórios gerados com sucesso na pasta:\n{mensagem}", parent=self)
            self.abrir_pasta_destino(mensagem)
        else:
            messagebox.showerror("Erro no Processamento", f"Módulo '{modulo}':\n\nOcorreu um erro.\n\nDetalhes: {mensagem}", parent=self)

    def abrir_pasta_destino(self, pasta):
        if not pasta: return
        try:
            if sys.platform == "win32": os.startfile(pasta)
            elif sys.platform == "darwin": subprocess.run(["open", pasta])
            else: subprocess.run(["xdg-open", pasta])
        except Exception as e:
            messagebox.showwarning("Aviso", f"Não foi possível abrir a pasta de destino automaticamente.\n\nCaminho: {pasta}\nErro: {e}", parent=self)
            
    def center_window(self):
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')

    def ep_atualizar_label_dias(self):
        if not self.ep_dias_selecionados:
            self.ep_lbl_dias_selecionados.config(text="Nenhum dia selecionado.")
        else:
            dias_str = ", ".join(map(str, self.ep_dias_selecionados))
            self.ep_lbl_dias_selecionados.config(text=f"Dias: {dias_str}")
    
    def ep_abrir_dialogo_selecao_dias(self):
        dialogo = EP_DialogoSelecaoDias(self, self.ep_dias_selecionados)
        if dialogo.resultado is not None:
            self.ep_dias_selecionados = sorted(dialogo.resultado)
            self.ep_atualizar_label_dias()

    def ep_selecionar_arquivos(self):
        arquivos = filedialog.askopenfilenames(
            parent=self,
            title="Selecione o(s) arquivo(s) TXT de Viagem e Frota",
            filetypes=[("Arquivos TXT", "*.txt"), ("Todos os arquivos", "*.*")]
        )
        if arquivos:
            self.ep_arquivos_txt_selecionados = list(arquivos)
            nomes = [os.path.basename(arq) for arq in self.ep_arquivos_txt_selecionados]
            self.ep_lbl_arquivos_selecionados.config(text=", ".join(nomes))
        elif not self.ep_arquivos_txt_selecionados:
            self.ep_lbl_arquivos_selecionados.config(text="Nenhum arquivo selecionado")

    def ep_selecionar_local_salvamento(self):
        diretorio = filedialog.askdirectory(parent=self, title="Selecione o Local para Salvar o Arquivo Final")
        if diretorio:
            self.ep_local_salvar_var.set(diretorio)

    def ep_limpar_formulario(self):
        self.ep_arquivos_txt_selecionados = []
        self.ep_lbl_arquivos_selecionados.config(text="Nenhum arquivo selecionado")
        self.ep_dias_selecionados = []
        self.ep_atualizar_label_dias()
        self.ep_operador_var.set(EP_OPERADORES[-1])
        self.ep_linhas_var.set("Todas")
        self.ep_nome_arquivo_saida_var.set("AlteracaoProgramacao")
        self.ep_local_salvar_var.set(os.getcwd())
        messagebox.showinfo("Limpeza", "Todos os campos do E.P. foram limpos.", parent=self)

    def iniciar_processamento_ep(self):
        if not self.ep_arquivos_txt_selecionados:
            messagebox.showerror("Erro", "Nenhum arquivo TXT selecionado.", parent=self)
            return
        if not self.ep_dias_selecionados:
            messagebox.showerror("Erro", "Nenhum dia foi selecionado. Clique em 'Selecionar Dias...'", parent=self)
            return

        nome_arquivo_final = self.ep_nome_arquivo_saida_var.get().strip()
        if not nome_arquivo_final:
            messagebox.showerror("Erro", "O nome do arquivo de saída não pode estar vazio.", parent=self)
            return
        if not nome_arquivo_final.lower().endswith(".txt"):
            nome_arquivo_final += ".txt"

        local_salvar = self.ep_local_salvar_var.get()
        if not local_salvar or not os.path.isdir(local_salvar):
            messagebox.showerror("Erro", "Local para salvar inválido.", parent=self)
            return

        caminho_completo_saida = os.path.join(local_salvar, nome_arquivo_final)
        
        self.botao_gerar_ep.config(state="disabled", text="Processando...")
        self.botao_limpar_ep.config(state="disabled")
        self.update_idletasks()
        
        try:
            ep_processar_dados(
                self.ep_arquivos_txt_selecionados,
                self.ep_operador_var.get(),
                self.ep_linhas_var.get(),
                self.ep_dias_selecionados,
                caminho_completo_saida,
                gui_ref=self 
            )
        finally:
            self.finalizar_processamento_gui("", sucesso=True, modulo="EP_dummy") 
            self.botao_gerar_ep.config(state="normal", text="Processar Arquivos")
            self.botao_limpar_ep.config(state="normal")


if __name__ == "__main__":
    try:
        import tkinterdnd2
    except ImportError:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Dependência Faltando", "A biblioteca 'tkinterdnd2' não foi encontrada. Por favor, instale-a com 'pip install tkinterdnd2'.")
        sys.exit(1)

    app = AppGUI()
    app.mainloop()