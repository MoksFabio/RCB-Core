import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime, date, timedelta
import os
import re
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PERMISSIONARIAS_EMPRESAS = ["BOA", "CAX", "CSR", "EME", "GLO", "SJT", "VML"]
COLS_PERMISSIONARIAS_INPUT = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO', 'NMEFETPASST', 'DSDIATIPO']
CONCESSIONARIAS_EMPRESAS = ["CNO", "MOB"]
COLS_CONCESSIONARIAS_INPUT = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO', 'NMPASSTOTAL', 'NMPASSEQUIVALENTE']
STPP_RMR_EMPRESAS_TODAS = sorted(list(set(PERMISSIONARIAS_EMPRESAS + CONCESSIONARIAS_EMPRESAS)))
COLS_STPP_RMR_PROCESSADO = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO', 'NMEFETPASST', 'NMEFETPASSEQUIV', 'DSDIATIPO']
COLS_OUTPUT_EXCEL_DIARIO = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO', 'MÉDIA NMPASSTOTAL', 'MÉDIA NMPASSEQUIVALENTE']
TIPOS_DIA_VALIDOS = ["DUT", "SAB", "DOM", "TODOS"]

def configurar_estilos_ttk():
    style = ttk.Style()
    try: style.theme_use('clam')
    except tk.TclError:
        try: style.theme_use('alt')
        except tk.TclError: style.theme_use('default')
    cor_fundo_janela = "#2B2B2B"; cor_fundo_frame = "#3C3F41"; cor_texto_principal = "#E0E0E0"
    cor_texto_secundario = "#BDBDBD"; cor_borda = "#4A4A4A"; cor_primaria = "#0D6EFD"
    cor_primaria_hover = "#0A58CA"; cor_secundaria_botao_bg = "#6C757D"; cor_secundaria_botao_fg = "#FFFFFF"
    cor_secundaria_botao_bg_hover = "#5A6268"; cor_erro = "#DC3545"; cor_entry_fundo = "#495057"
    cor_entry_texto = cor_texto_principal; cor_entry_borda_foco = cor_primaria
    fonte_padrao = ("Segoe UI", 10); fonte_titulo_label = ("Segoe UI", 12, "bold")
    fonte_botao = ("Segoe UI", 10, "bold"); fonte_label_info = ("Segoe UI", 9)
    style.configure('TFrame', background=cor_fundo_janela)
    style.configure('Form.TFrame', background=cor_fundo_frame, borderwidth=1, relief='solid', bordercolor=cor_borda)
    style.configure('NoBorder.TFrame', background=cor_fundo_frame)
    style.configure('ButtonFrame.TFrame', background=cor_fundo_frame)
    style.configure('TLabel', background=cor_fundo_frame, foreground=cor_texto_principal, font=fonte_padrao, padding=5)
    style.configure('Header.TLabel', background=cor_fundo_janela, foreground=cor_primaria, font=("Segoe UI", 16, "bold"), padding=(10, 15))
    style.configure('SubHeader.TLabel', background=cor_fundo_janela, foreground=cor_texto_principal, font=("Segoe UI", 13, "bold"), padding=(5,10))
    style.configure('Info.TLabel', background=cor_fundo_frame, foreground=cor_texto_secundario, font=fonte_label_info)
    style.configure('Error.TLabel', background=cor_fundo_frame, foreground=cor_erro, font=fonte_padrao)
    style.configure('TButton', font=fonte_botao, padding=(10, 8), relief="raised", borderwidth=1)
    style.configure('Primary.TButton', background=cor_primaria, foreground=cor_texto_principal, bordercolor=cor_primaria)
    style.map('Primary.TButton', background=[('active', cor_primaria_hover), ('disabled', '#4A4A4A')], foreground=[('disabled', cor_texto_secundario)], bordercolor=[('active', cor_primaria_hover)])
    style.configure('Secondary.TButton', background=cor_secundaria_botao_bg, foreground=cor_secundaria_botao_fg, bordercolor=cor_secundaria_botao_bg)
    style.map('Secondary.TButton', background=[('active', cor_secundaria_botao_bg_hover), ('disabled', '#4A4A4A')], foreground=[('disabled', cor_texto_secundario)])
    style.configure('TEntry', font=fonte_padrao, padding=6, relief="flat", borderwidth=2, foreground=cor_entry_texto, fieldbackground=cor_entry_fundo)
    style.map('TEntry', bordercolor=[('focus', cor_entry_borda_foco), ('!focus', cor_borda)], foreground=[('disabled', cor_texto_secundario)], fieldbackground=[('readonly', cor_fundo_frame), ('disabled', '#333333')])
    style.configure('TCombobox', font=fonte_padrao, padding=6, relief="flat", borderwidth=2, arrowcolor=cor_texto_principal, fieldbackground=cor_entry_fundo, foreground=cor_entry_texto)
    style.map('TCombobox', bordercolor=[('focus', cor_entry_borda_foco), ('!focus', cor_borda)], foreground=[('disabled', cor_texto_secundario), ('readonly', cor_entry_texto)], fieldbackground=[('readonly', cor_entry_fundo), ('disabled', '#333333')])
    style.configure('TSeparator', background=cor_borda)
    style.configure('TLabelframe', background=cor_fundo_frame, bordercolor=cor_borda, padding=10)
    style.configure('TLabelframe.Label', background=cor_fundo_frame, foreground=cor_texto_principal, font=fonte_titulo_label)
    style.configure('TCheckbutton', background=cor_fundo_frame, foreground=cor_texto_principal, font=fonte_padrao)
    style.map('TCheckbutton', indicatorcolor=[('selected', cor_primaria), ('active', cor_primaria_hover), ('!selected', cor_entry_fundo)])
    return style

def obter_pasta_downloads_padrao():
    if os.name == 'nt':
        try:
            import winreg
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
            downloads_path = winreg.QueryValueEx(key, "{374DE290-123F-4565-9164-39C4925E467B}")[0]
            winreg.CloseKey(key)
            return downloads_path
        except Exception:
            return os.path.join(os.path.expanduser('~'), 'Downloads')
    else:
        return os.path.join(os.path.expanduser('~'), 'Downloads')

def centralizar_janela(janela):
    janela.update_idletasks()
    width = janela.winfo_width(); height = janela.winfo_height()
    if width <= 1 or height <= 1:
        width = janela.winfo_reqwidth(); height = janela.winfo_reqheight()
    x_pos = (janela.winfo_screenwidth() // 2) - (width // 2)
    y_pos = (janela.winfo_screenheight() // 2) - (height // 2)
    janela.geometry(f'{width}x{height}+{x_pos}+{y_pos}')

def titulo_aba_excel_safe(titulo_original):
    if not isinstance(titulo_original, str): titulo_original = str(titulo_original)
    titulo_limpo = re.sub(r'[\\/*?:\[\]]', '', titulo_original)
    return titulo_limpo[:31]

def contar_dias_no_periodo_por_tipo(lista_periodos_obj, tipo_dia_alvo):
    total_dias_contados = 0
    if not lista_periodos_obj: return 0
    tipo_dia_alvo_upper = tipo_dia_alvo.upper()
    for periodo in lista_periodos_obj:
        data_corrente = periodo['inicio'].date() if isinstance(periodo['inicio'], datetime) else periodo['inicio']
        data_fim_periodo = periodo['fim'].date() if isinstance(periodo['fim'], datetime) else periodo['fim']
        while data_corrente <= data_fim_periodo:
            weekday = data_corrente.weekday()
            if tipo_dia_alvo_upper == "TODOS": total_dias_contados += 1
            elif tipo_dia_alvo_upper == "DUT" and 0 <= weekday <= 4: total_dias_contados += 1
            elif tipo_dia_alvo_upper == "SAB" and weekday == 5: total_dias_contados += 1
            elif tipo_dia_alvo_upper == "DOM" and weekday == 6: total_dias_contados += 1
            data_corrente += timedelta(days=1)
    return total_dias_contados

def get_concessionaria_dut_divisor_mensal(year, month):
    days_in_month = calendar.monthrange(year, month)[1]
    num_weekdays = 0
    for day_num in range(1, days_in_month + 1):
        if date(year, month, day_num).weekday() < 5:
            num_weekdays += 1
    return num_weekdays

def selecionar_tipo_processamento_gui_media():
    escolha_feita = {"tipo": None}
    janela_selecao = tk.Tk()
    janela_selecao.title("Demanda Média - Selecionar Tipo")
    current_style = configurar_estilos_ttk()
    janela_selecao.configure(background=current_style.lookup('TFrame', 'background'))
    janela_selecao.resizable(False, False)
    def on_permissionarias_click(): escolha_feita["tipo"] = "permissionarias"; janela_selecao.destroy()
    def on_concessionarias_click(): escolha_feita["tipo"] = "concessionarias"; janela_selecao.destroy()
    def on_stpp_rmr_click(): escolha_feita["tipo"] = "stpp_rmr"; janela_selecao.destroy()
    def on_cancelar_click(): escolha_feita["tipo"] = None; janela_selecao.destroy()
    frame_conteudo = ttk.Frame(janela_selecao, padding="25 25 25 25", style='TFrame')
    frame_conteudo.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
    janela_selecao.columnconfigure(0, weight=1); janela_selecao.rowconfigure(0, weight=1)
    ttk.Label(frame_conteudo, text="Assistente de Demanda Média RCB", style='Header.TLabel').grid(row=0, column=0, columnspan=2, pady=(0, 15), sticky="ew")
    ttk.Label(frame_conteudo, text="Selecione o tipo de dados para calcular a média:", style='SubHeader.TLabel').grid(row=1, column=0, columnspan=2, pady=(0, 25), sticky="ew")
    btn_permissionarias = ttk.Button(frame_conteudo, text="Permissionárias", command=on_permissionarias_click, style='Primary.TButton', width=35)
    btn_permissionarias.grid(row=2, column=0, columnspan=2, pady=10, ipady=10, sticky="ew")
    btn_concessionarias = ttk.Button(frame_conteudo, text="Concessionárias", command=on_concessionarias_click, style='Primary.TButton', width=35)
    btn_concessionarias.grid(row=3, column=0, columnspan=2, pady=10, ipady=10, sticky="ew")
    btn_stpp_rmr = ttk.Button(frame_conteudo, text="STPP/RMR (Combinado)", command=on_stpp_rmr_click, style='Primary.TButton', width=35)
    btn_stpp_rmr.grid(row=4, column=0, columnspan=2, pady=10, ipady=10, sticky="ew")
    ttk.Separator(frame_conteudo, orient='horizontal').grid(row=5, column=0, columnspan=2, sticky='ew', pady=(25,20))
    btn_cancelar = ttk.Button(frame_conteudo, text="Cancelar", command=on_cancelar_click, style='Secondary.TButton', width=20)
    btn_cancelar.grid(row=6, column=0, columnspan=2, pady=(15,0))
    janela_selecao.protocol("WM_DELETE_WINDOW", on_cancelar_click)
    btn_permissionarias.focus_set();
    janela_selecao.update_idletasks()
    centralizar_janela(janela_selecao)
    janela_selecao.mainloop()
    return escolha_feita["tipo"]

def obter_entrada_gui_media(titulo_janela, empresas_validas, tipo_relatorio_raw):
    resultados_gui = {}
    janela_formulario = tk.Tk()
    janela_formulario.title(titulo_janela)
    current_style = configurar_estilos_ttk()
    janela_formulario.configure(background=current_style.lookup('TFrame', 'background'))
    janela_formulario.resizable(False, False)
    cod_linhas_var = tk.StringVar()
    data_inicio_var = tk.StringVar()
    data_fim_var = tk.StringVar()
    meses_especificos_var = tk.StringVar()
    multiplos_meses_check_var = tk.BooleanVar(value=False)
    empresas_var = tk.StringVar()
    tipo_dia_var = tk.StringVar()
    processar_equivalente_var = tk.BooleanVar(value=False)
    num_id_var = tk.StringVar(value="Resp_Ouvidoria_")
    local_salvar_var = tk.StringVar(value=obter_pasta_downloads_padrao())

    def selecionar_pasta_destino():
        pasta_selecionada = filedialog.askdirectory(parent=janela_formulario, title="Selecione a pasta para salvar o arquivo Excel")
        if pasta_selecionada: local_salvar_var.set(pasta_selecionada)

    def submeter_formulario_gui():
        nonlocal resultados_gui
        is_multi_month_mode_active = multiplos_meses_check_var.get()
        lista_periodos_obj_validados = []
        periodos_str_para_nome_arquivo = []
        if is_multi_month_mode_active:
            meses_gui_str = meses_especificos_var.get().strip()
            if not meses_gui_str: messagebox.showerror("Erro Validação", "Campo 'Meses Específicos' é obrigatório.", parent=janela_formulario); return
            meses_individuais_str = [m.strip() for m in meses_gui_str.split(',') if m.strip()]
            if not meses_individuais_str: messagebox.showerror("Erro Validação", "Nenhum mês válido fornecido.", parent=janela_formulario); return
            for mes_str_item in meses_individuais_str:
                try:
                    mes_dt_obj = datetime.strptime(mes_str_item, "%m/%Y")
                    primeiro_dia = mes_dt_obj.replace(day=1)
                    _, num_dias_mes = calendar.monthrange(primeiro_dia.year, primeiro_dia.month)
                    ultimo_dia = primeiro_dia.replace(day=num_dias_mes)
                    lista_periodos_obj_validados.append({'inicio': primeiro_dia, 'fim': ultimo_dia, 'str_original': mes_str_item})
                    periodos_str_para_nome_arquivo.append(mes_str_item.replace('/', ''))
                except ValueError:
                    messagebox.showerror("Erro Validação", f"Formato de mês '{mes_str_item}' inválido. Use MM/AAAA.", parent=janela_formulario); return
        else:
            data_inicio_str_gui = data_inicio_var.get().strip()
            data_fim_str_gui = data_fim_var.get().strip()
            if not data_inicio_str_gui or not data_fim_str_gui: messagebox.showerror("Erro Validação", "Campos 'Data Início' e 'Data Fim' são obrigatórios.", parent=janela_formulario); return
            try:
                dt_ini = datetime.strptime(data_inicio_str_gui, "%d/%m/%Y")
                dt_fim = datetime.strptime(data_fim_str_gui, "%d/%m/%Y")
                if dt_fim < dt_ini: messagebox.showerror("Erro Validação", "Data Fim anterior à Data Início.", parent=janela_formulario); return
                lista_periodos_obj_validados.append({'inicio': dt_ini, 'fim': dt_fim, 'str_original': f"{data_inicio_str_gui}-{data_fim_str_gui}"})
                periodos_str_para_nome_arquivo.append(f"{dt_ini.strftime('%d%m%y')}-{dt_fim.strftime('%d%m%y')}")
            except ValueError:
                messagebox.showerror("Erro Validação", "Formato de Data Início/Fim inválido. Use DD/MM/AAAA.", parent=janela_formulario); return
        if not lista_periodos_obj_validados: messagebox.showerror("Erro Validação", "Nenhum período válido processado.", parent=janela_formulario); return
        
        exibir_formato_mensal = False
        if lista_periodos_obj_validados:
            min_date = min(p['inicio'] for p in lista_periodos_obj_validados)
            max_date = max(p['fim'] for p in lista_periodos_obj_validados)
            if min_date.year != max_date.year or min_date.month != max_date.month:
                exibir_formato_mensal = True
        
        cod_linhas_str_val = cod_linhas_var.get().strip()
        empresas_input_str_val = empresas_var.get().strip().upper()
        tipo_dia_selecionado_val = tipo_dia_var.get().strip().upper()
        id_arquivo_val = num_id_var.get().strip()
        local_salvar_final_val = local_salvar_var.get().strip()
        processar_equivalente_val = processar_equivalente_var.get()
        if not cod_linhas_str_val: messagebox.showerror("Erro Validação", "'Códigos das Linhas' é obrigatório.", parent=janela_formulario); return
        
        valid_day_types_for_current_report = TIPOS_DIA_VALIDOS
        if tipo_relatorio_raw == "stpp_rmr": 
            valid_day_types_for_current_report = ["DUT", "TODOS"]
        
        if not tipo_dia_selecionado_val or tipo_dia_selecionado_val not in valid_day_types_for_current_report:
            messagebox.showerror("Erro Validação", f"Selecione um 'Tipo de Dia' válido para este relatório: {', '.join(valid_day_types_for_current_report)}.", parent=janela_formulario); return
        empresas_selecionadas_validadas = []
        if not empresas_input_str_val: messagebox.showerror("Erro Validação", "'Empresa(s) Operadora(s)' é obrigatório.", parent=janela_formulario); return
        elif empresas_input_str_val == "TODAS": empresas_selecionadas_validadas = empresas_validas
        else:
            empresas_lista_input = [emp.strip() for emp in empresas_input_str_val.split(',') if emp.strip()]
            if not empresas_lista_input: messagebox.showerror("Erro Validação", "Nenhuma empresa válida fornecida.", parent=janela_formulario); return
            for emp_input in empresas_lista_input:
                if emp_input not in empresas_validas: messagebox.showerror("Erro Validação", f"Empresa '{emp_input}' inválida. Válidas: {', '.join(empresas_validas)}", parent=janela_formulario); return
                empresas_selecionadas_validadas.append(emp_input)
        if not id_arquivo_val: messagebox.showerror("Erro Validação", "'ID para Nome do Arquivo' é obrigatório.", parent=janela_formulario); return
        if not local_salvar_final_val or not os.path.isdir(local_salvar_final_val): messagebox.showerror("Erro Validação", f"Caminho '{local_salvar_final_val}' inválido.", parent=janela_formulario); return
        tipo_dia_fmt = "TODOS OS DIAS" if tipo_dia_selecionado_val == "TODOS" else tipo_dia_selecionado_val
        periodos_desc_titulo_fmt = "; ".join([p['str_original'] for p in lista_periodos_obj_validados])
        if len(periodos_desc_titulo_fmt) > 60 : periodos_desc_titulo_fmt = f"{len(lista_periodos_obj_validados)} período(s)" if not is_multi_month_mode_active else f"{len(lista_periodos_obj_validados)} meses"
        titulo_rel_principal = f"PASSAGEIROS POR DIA ({tipo_dia_fmt}) "
        if is_multi_month_mode_active: titulo_rel_principal += f"MESES: {periodos_desc_titulo_fmt}"
        else: titulo_rel_principal += f"PERÍODO: {periodos_desc_titulo_fmt}"
        titulo_rel_principal += f" - {tipo_relatorio_raw.upper().replace('/', '-')}"
        datas_nome_arq_fmt = "_E_".join(periodos_str_para_nome_arquivo)
        if is_multi_month_mode_active: datas_nome_arq_fmt = "_".join(periodos_str_para_nome_arquivo)
        if len(datas_nome_arq_fmt) > 50: datas_nome_arq_fmt = f"MULTIPERIODOS_{len(periodos_str_para_nome_arquivo)}" if not is_multi_month_mode_active else f"MULTIMESES_{len(periodos_str_para_nome_arquivo)}"
        nome_arq_final = f"{id_arquivo_val}.xlsx"
        resultados_gui.update({
            "titulo_relatorio_principal": titulo_rel_principal,
            "cod_linhas_str": cod_linhas_str_val,
            "lista_periodos_obj": lista_periodos_obj_validados,
            "exibir_formato_mensal": exibir_formato_mensal,
            "tipo_dia_selecionado": tipo_dia_selecionado_val,
            "empresas_selecionadas": empresas_selecionadas_validadas,
            "nome_arquivo_excel": os.path.join(local_salvar_final_val, nome_arq_final),
            "tipo_relatorio_raw": tipo_relatorio_raw,
            "processar_equivalente": processar_equivalente_val
        })
        janela_formulario.quit()

    frame_principal = ttk.Frame(janela_formulario, padding="20", style='Form.TFrame')
    frame_principal.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=10)
    janela_formulario.columnconfigure(0, weight=1); janela_formulario.rowconfigure(0, weight=1)
    current_row = 0
    ttk.Label(frame_principal, text="Códigos das Linhas (ex: 101,102 ou 'Todas'):").grid(row=current_row, column=0, columnspan=2, sticky=tk.W, pady=(0,2), padx=5); current_row += 1
    entry_cod_linhas = ttk.Entry(frame_principal, textvariable=cod_linhas_var, width=60)
    entry_cod_linhas.grid(row=current_row, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0,10), padx=5); current_row += 1
    check_multiplos_meses = ttk.Checkbutton(frame_principal, text="Selecionar Meses Específicos (em vez de intervalo)", variable=multiplos_meses_check_var, style='TCheckbutton')
    check_multiplos_meses.grid(row=current_row, column=0, columnspan=2, sticky=tk.W, pady=(5,2), padx=5); current_row += 1
    frame_periodo_unico = ttk.Frame(frame_principal, style='NoBorder.TFrame')
    label_data_inicio = ttk.Label(frame_periodo_unico, text="Data Início (DD/MM/AAAA):")
    label_data_inicio.grid(row=0, column=0, sticky=tk.W, pady=(0,2), padx=(5,5))
    entry_data_inicio = ttk.Entry(frame_periodo_unico, textvariable=data_inicio_var, width=15)
    entry_data_inicio.grid(row=0, column=1, sticky=tk.W, pady=(0,2), padx=0)
    label_data_fim = ttk.Label(frame_periodo_unico, text="Data Fim (DD/MM/AAAA):")
    label_data_fim.grid(row=0, column=2, sticky=tk.W, pady=(0,2), padx=(10,5))
    entry_data_fim = ttk.Entry(frame_periodo_unico, textvariable=data_fim_var, width=15)
    entry_data_fim.grid(row=0, column=3, sticky=tk.W, pady=(0,2), padx=0)
    frame_periodo_unico.columnconfigure(1, weight=1)
    frame_periodo_unico.columnconfigure(3, weight=1)
    frame_multi_meses = ttk.Frame(frame_principal, style='NoBorder.TFrame')
    label_meses_especificos = ttk.Label(frame_multi_meses, text="Meses Específicos (MM/AAAA, MM/AAAA, ...):")
    label_meses_especificos.grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0,2), padx=5)
    entry_meses_especificos = ttk.Entry(frame_multi_meses, textvariable=meses_especificos_var, width=58)
    entry_meses_especificos.grid(row=1, column=0, columnspan=2, sticky=(tk.W,tk.E), pady=(0,2), padx=5)
    ttk.Label(frame_multi_meses, text="Use vírgula ',' para separar. Ex: 01/2025, 03/2025", style="Info.TLabel").grid(row=2, column=0,columnspan=2, sticky=tk.W, pady=(0,2), padx=5)
    row_idx_data_placeholder = current_row
    current_row +=1

    def toggle_periodo_mode_actual(*args):
        grid_info_unico = frame_periodo_unico.grid_info()
        grid_info_multi = frame_multi_meses.grid_info()
        target_row = row_idx_data_placeholder
        if multiplos_meses_check_var.get():
            frame_periodo_unico.grid_remove()
            if not grid_info_multi or not frame_multi_meses.winfo_ismapped():
                frame_multi_meses.grid(row=target_row, column=0, columnspan=2, sticky=tk.W+tk.E, pady=(0,10), padx=0)
            entry_data_inicio.configure(state=tk.DISABLED)
            entry_data_fim.configure(state=tk.DISABLED)
            entry_meses_especificos.configure(state=tk.NORMAL)
            entry_meses_especificos.focus()
        else:
            frame_multi_meses.grid_remove()
            if not grid_info_unico or not frame_periodo_unico.winfo_ismapped():
                frame_periodo_unico.grid(row=target_row, column=0, columnspan=2, sticky=tk.W+tk.E, pady=(0,10), padx=0)
            entry_data_inicio.configure(state=tk.NORMAL)
            entry_data_fim.configure(state=tk.NORMAL)
            entry_meses_especificos.configure(state=tk.DISABLED)
            entry_data_inicio.focus()
    multiplos_meses_check_var.trace_add("write", toggle_periodo_mode_actual)
    ttk.Label(frame_principal, text="Tipo de Dia para Média:").grid(row=current_row, column=0, columnspan=2, sticky=tk.W, pady=(0,2), padx=5); current_row += 1
    
    opcoes_tipo_dia_atuais = TIPOS_DIA_VALIDOS
    if tipo_relatorio_raw == "stpp_rmr":
        opcoes_tipo_dia_atuais = ["DUT", "TODOS"]
        if tipo_dia_var.get().upper() in ["SAB", "DOM"] : 
            tipo_dia_var.set("DUT")

    combo_tipo_dia = ttk.Combobox(frame_principal, textvariable=tipo_dia_var, values=opcoes_tipo_dia_atuais, state="readonly", width=57)
    combo_tipo_dia.grid(row=current_row, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0,10), padx=5); current_row += 1
    if opcoes_tipo_dia_atuais: 
        if tipo_dia_var.get() not in opcoes_tipo_dia_atuais: 
            tipo_dia_var.set(opcoes_tipo_dia_atuais[0]) 
        elif not tipo_dia_var.get(): 
            tipo_dia_var.set(opcoes_tipo_dia_atuais[0])
    ttk.Label(frame_principal, text="Empresa(s) Operadora(s) (ex: BOA,GLO ou 'Todas'):").grid(row=current_row, column=0, columnspan=2, sticky=tk.W, pady=(0,2), padx=5); current_row += 1
    entry_empresas = ttk.Entry(frame_principal, textvariable=empresas_var, width=60)
    entry_empresas.grid(row=current_row, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0,2), padx=5); current_row += 1
    ttk.Label(frame_principal, text=f"Válidas: {', '.join(empresas_validas)}", style="Info.TLabel", wraplength=450, justify=tk.LEFT).grid(row=current_row, column=0, columnspan=2, sticky=tk.W, padx=5, pady=(0,10)); current_row += 1
    
    if tipo_relatorio_raw in ["concessionarias", "stpp_rmr"]:
        check_equivalente = ttk.Checkbutton(frame_principal, text="Incluir cálculo de Passageiros Equivalentes (NMPASSEQUIVALENTE)", variable=processar_equivalente_var, style='TCheckbutton')
        check_equivalente.grid(row=current_row, column=0, columnspan=2, sticky=tk.W, pady=(0, 10), padx=5)
        current_row += 1
    
    ttk.Label(frame_principal, text="ID para Nome do Arquivo:").grid(row=current_row, column=0, columnspan=2, sticky=tk.W, pady=(0,2), padx=5); current_row += 1
    ttk.Entry(frame_principal, textvariable=num_id_var, width=60).grid(row=current_row, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0,10), padx=5); current_row += 1
    ttk.Label(frame_principal, text="Salvar Arquivo Em:").grid(row=current_row, column=0, columnspan=2, sticky=tk.W, pady=(0,2), padx=5); current_row += 1
    path_frame = ttk.Frame(frame_principal, style='NoBorder.TFrame')
    path_frame.grid(row=current_row, column=0, columnspan=2, sticky=tk.W+tk.E, pady=(0,10), padx=0); current_row +=1
    entry_local_salvar = ttk.Entry(path_frame, textvariable=local_salvar_var)
    entry_local_salvar.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(5,5))
    btn_selecionar_pasta = ttk.Button(path_frame, text="Selecionar...", command=selecionar_pasta_destino, style='Secondary.TButton')
    btn_selecionar_pasta.pack(side=tk.LEFT, padx=(0,5))
    frame_botoes = ttk.Frame(frame_principal, style='ButtonFrame.TFrame')
    frame_botoes.grid(row=current_row, column=0, columnspan=2, pady=(20, 0), sticky=tk.E)
    ttk.Button(frame_botoes, text="Processar", command=submeter_formulario_gui, style='Primary.TButton').pack(side=tk.LEFT, padx=(0,10))
    ttk.Button(frame_botoes, text="Cancelar", command=janela_formulario.destroy, style='Secondary.TButton').pack(side=tk.LEFT)
    toggle_periodo_mode_actual() 
    if opcoes_tipo_dia_atuais:
        if not tipo_dia_var.get() or tipo_dia_var.get() not in opcoes_tipo_dia_atuais:
            tipo_dia_var.set(opcoes_tipo_dia_atuais[0])
    janela_formulario.update_idletasks()
    width1 = janela_formulario.winfo_reqwidth(); height1 = janela_formulario.winfo_reqheight()
    multiplos_meses_check_var.set(True); janela_formulario.update_idletasks()
    width2 = janela_formulario.winfo_reqwidth(); height2 = janela_formulario.winfo_reqheight()
    multiplos_meses_check_var.set(False); janela_formulario.update_idletasks()
    final_width = max(width1, width2) + 20; final_height = max(height1, height2) + 20 
    janela_formulario.geometry(f'{final_width}x{final_height}')
    centralizar_janela(janela_formulario)
    entry_cod_linhas.focus_set()
    janela_formulario.protocol("WM_DELETE_WINDOW", janela_formulario.destroy)
    janela_formulario.mainloop()
    try: janela_formulario.destroy()
    except tk.TclError: pass
    return resultados_gui if resultados_gui else None

def carregar_dados_txt_base(caminho_arquivo_txt, colunas_desejadas_originais, config_usuario_filtros):
    df = None
    configs_leitura = [('\t', 'latin1'), ('\t', 'utf-8'), ('\t', 'utf-8-sig'), (';', 'latin1'), (';', 'utf-8'), (';', 'utf-8-sig'), (',', 'latin1'), (',', 'utf-8'), (',', 'utf-8-sig')]
    leitura_bem_sucedida = False
    for delim, enc in configs_leitura:
        try:
            preview_df = pd.read_csv(caminho_arquivo_txt, sep=delim, dtype=str, encoding=enc, on_bad_lines='warn', nrows=5)
            if preview_df is not None and len(preview_df.columns) > 1 and not any(delim in str(col_name) for col_name in preview_df.columns):
                df_temp = pd.read_csv(caminho_arquivo_txt, sep=delim, dtype=str, encoding=enc, on_bad_lines='skip')
                if df_temp is not None and not df_temp.empty: df = df_temp; leitura_bem_sucedida = True; break
        except Exception: pass
    if not leitura_bem_sucedida: return None
    df.columns = [str(col).strip().upper() for col in df.columns]
    colunas_desejadas_upper = [col.upper() for col in colunas_desejadas_originais]
    colunas_presentes_no_df = [col for col in colunas_desejadas_upper if col in df.columns]
    
    colunas_obrigatorias_faltantes = [
        col for col in colunas_desejadas_upper 
        if col not in df.columns and col != 'NMPASSEQUIVALENTE'
    ]
    if colunas_obrigatorias_faltantes: return None

    df_processado = df[colunas_presentes_no_df].copy()
    if 'DTOPERACAO' in df_processado.columns:
        df_processado.loc[:, 'DTOPERACAO_temp_dt'] = pd.to_datetime(df_processado['DTOPERACAO'], errors='coerce', dayfirst=True)
        df_processado.dropna(subset=['DTOPERACAO_temp_dt'], inplace=True)
        if df_processado.empty: return df_processado
    if config_usuario_filtros:
        df_para_filtrar = df_processado.copy()
        if 'DTOPERACAO_temp_dt' in df_para_filtrar.columns and 'lista_periodos_obj' in config_usuario_filtros:
            lista_periodos = config_usuario_filtros['lista_periodos_obj']
            if lista_periodos:
                condicoes_periodos = []
                for p_obj in lista_periodos:
                    dt_ini_periodo = p_obj['inicio'].date() if isinstance(p_obj['inicio'], datetime) else p_obj['inicio']
                    dt_fim_periodo = p_obj['fim'].date() if isinstance(p_obj['fim'], datetime) else p_obj['fim']
                    cond = ((df_para_filtrar['DTOPERACAO_temp_dt'].dt.date >= dt_ini_periodo) & \
                            (df_para_filtrar['DTOPERACAO_temp_dt'].dt.date <= dt_fim_periodo))
                    condicoes_periodos.append(cond)
                if condicoes_periodos:
                    filtro_datas_combinado = pd.DataFrame(condicoes_periodos).transpose().any(axis=1)
                    if filtro_datas_combinado.any():
                        df_para_filtrar = df_para_filtrar[filtro_datas_combinado]
                    else:
                        df_para_filtrar = df_para_filtrar.iloc[0:0].copy()
        df_processado = df_para_filtrar
        if not df_processado.empty and 'CDLINHA' in df_processado.columns and config_usuario_filtros.get('cod_linhas_str','TODAS').upper() != "TODAS":
            linhas_para_filtrar = [str(l).strip() for l in config_usuario_filtros['cod_linhas_str'].split(',') if l.strip()]
            df_processado = df_processado[df_processado['CDLINHA'].astype(str).str.strip().isin(linhas_para_filtrar)]
        if not df_processado.empty and 'CDOPERADOR' in df_processado.columns and 'empresas_selecionadas' in config_usuario_filtros:
            empresas_a_filtrar = config_usuario_filtros['empresas_selecionadas']
            if isinstance(empresas_a_filtrar, list) and empresas_a_filtrar:
                df_processado = df_processado[df_processado['CDOPERADOR'].astype(str).str.strip().str.upper().isin(empresas_a_filtrar)]
    if 'DTOPERACAO_temp_dt' in df_processado.columns :
        if not df_processado.empty:
            df_processado.loc[:, 'DTOPERACAO'] = df_processado['DTOPERACAO_temp_dt'].dt.date
        elif 'DTOPERACAO_temp_dt' in df_processado:
            df_processado.loc[:, 'DTOPERACAO'] = pd.Series(dtype='object')
        df_processado.drop(columns=['DTOPERACAO_temp_dt'], errors='ignore', inplace=True)
    if df_processado.empty: return df_processado
    cols_to_convert_to_numeric = ['NMEFETPASST', 'NMPASSTOTAL', 'NMPASSEQUIVALENTE']
    for col in cols_to_convert_to_numeric:
        if col in df_processado.columns:
            try:
                if not pd.api.types.is_string_dtype(df_processado[col]):
                    df_processado[col] = df_processado[col].astype(str)
                pass_col_temp = df_processado[col].str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
                df_processado.loc[:, col] = pd.to_numeric(pass_col_temp, errors='coerce')
            except Exception:
                pass
    if 'DSDIATIPO' in df_processado.columns: df_processado.loc[:, 'DSDIATIPO'] = df_processado['DSDIATIPO'].astype(str).str.strip().str.upper()
    return df_processado

def popular_uma_planilha_excel(ws, df_media, titulo_relatorio_na_planilha, config_usuario):
    ws.title = titulo_aba_excel_safe(ws.title)
    
    df_para_exibir = df_media.copy() if df_media is not None else pd.DataFrame()
    cols_para_exibir = []
    
    base_cols = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO']
    
    if not df_para_exibir.empty:
        if config_usuario and config_usuario.get('processar_equivalente') and 'MÉDIA NMPASSEQUIVALENTE' in df_para_exibir.columns:
            cols_para_exibir = base_cols + ['MÉDIA NMPASSTOTAL', 'MÉDIA NMPASSEQUIVALENTE']
        else:
            if 'MÉDIA NMPASSTOTAL' in df_para_exibir.columns:
                df_para_exibir.rename(columns={'MÉDIA NMPASSTOTAL': 'MÉDIA'}, inplace=True)
            cols_para_exibir = base_cols + ['MÉDIA']
        
        cols_para_exibir = [col for col in cols_para_exibir if col in df_para_exibir.columns]
    else:
        if config_usuario and config_usuario.get('processar_equivalente'):
            cols_para_exibir = base_cols + ['MÉDIA NMPASSTOTAL', 'MÉDIA NMPASSEQUIVALENTE']
        else:
            cols_para_exibir = base_cols + ['MÉDIA']

    num_colunas_df = len(cols_para_exibir)
    col_final_titulo_letra = get_column_letter(num_colunas_df if num_colunas_df > 0 else 1)
    ws.merge_cells(f'A1:{col_final_titulo_letra}2')
    celula_titulo = ws['A1']
    celula_titulo.value = titulo_relatorio_na_planilha
    celula_titulo.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    celula_titulo.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    celula_titulo.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws.row_dimensions[1].height = 20; ws.row_dimensions[2].height = 20
    linha_inicio_dados = 3
    for c_idx, col_name in enumerate(cols_para_exibir, start=1):
        cell = ws.cell(row=linha_inicio_dados, column=c_idx, value=col_name)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF'); cell.fill = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
    
    if not df_para_exibir.empty:
        df_media_output = df_para_exibir.reindex(columns=cols_para_exibir)
        for r_idx_df, row_tuple in enumerate(df_media_output.itertuples(index=False)):
            excel_row_num = linha_inicio_dados + 1 + r_idx_df
            for c_idx, value in enumerate(row_tuple, start=1):
                cell = ws.cell(row=excel_row_num, column=c_idx, value=value)
                cell.font = Font(name='Calibri', size=10); cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
                if cols_para_exibir[c_idx-1] in ['MÉDIA', 'MÉDIA NMPASSTOTAL', 'MÉDIA NMPASSEQUIVALENTE']: cell.number_format = '#,##0'
                if cols_para_exibir[c_idx-1] == 'DTOPERACAO' and isinstance(value, (date, datetime)): cell.number_format = 'DD/MM/YYYY'
        for col_idx_excel, col_name in enumerate(cols_para_exibir, start=1):
            column_letter = get_column_letter(col_idx_excel)
            if col_name in ['MÉDIA', 'MÉDIA NMPASSTOTAL', 'MÉDIA NMPASSEQUIVALENTE']: 
                ws.column_dimensions[column_letter].width = 28
            else:
                max_len = len(str(col_name))
                for row_num in range(linha_inicio_dados + 1, linha_inicio_dados + 1 + len(df_media_output)):
                    cell_value = ws.cell(row=row_num, column=col_idx_excel).value
                    if cell_value is not None: max_len = max(max_len, len(str(cell_value)))
                ws.column_dimensions[column_letter].width = max_len + 5
    else:
        msg_sem_dados = "Nenhum dado para este período/mês."
        if num_colunas_df > 0:
            ws.cell(row=linha_inicio_dados + 1, column=1, value=msg_sem_dados).alignment = Alignment(horizontal='center')
            if num_colunas_df > 1 : ws.merge_cells(start_row=linha_inicio_dados + 1, start_column=1, end_row=linha_inicio_dados + 1, end_column=num_colunas_df)
        else:
            ws.cell(row=linha_inicio_dados + 1, column=1, value=msg_sem_dados)

def exportar_excel_com_multiplas_abas(df_media_principal, titulo_principal_no_excel, nome_aba_principal, lista_dados_abas_secundarias, caminho_arquivo_excel, config_usuario):
    wb = Workbook()
    if wb.sheetnames and len(wb.sheetnames) == 1 and wb.sheetnames[0] == 'Sheet':
        default_sheet = wb.active
        wb.remove(default_sheet)
    
    abas_adicionadas = False

    if df_media_principal is not None and not df_media_principal.empty and nome_aba_principal:
        ws_principal = wb.create_sheet(title=titulo_aba_excel_safe(nome_aba_principal))
        popular_uma_planilha_excel(ws_principal, df_media_principal, titulo_principal_no_excel, config_usuario)
        abas_adicionadas = True

    for dados_aba in lista_dados_abas_secundarias:
        nome_aba_sec = dados_aba['nome_aba']
        df_media_aba = dados_aba['df_media']
        titulo_no_excel_aba = dados_aba['titulo_no_excel']
        ws_secundaria = wb.create_sheet(title=titulo_aba_excel_safe(nome_aba_sec))
        popular_uma_planilha_excel(ws_secundaria, df_media_aba, titulo_no_excel_aba, config_usuario)
        abas_adicionadas = True
        
    if not abas_adicionadas:
        ws_info = wb.create_sheet(title="Informação")
        ws_info['A1'] = "Nenhum dado disponível para gerar o relatório."

    if wb.sheetnames: 
        wb.active = 0 
    else: 
        if not abas_adicionadas:
            ws_info_fallback = wb.create_sheet(title="Info")
            ws_info_fallback['A1'] = "Relatório vazio."
            wb.active = 0
    
    try:
        wb.save(caminho_arquivo_excel)
    except PermissionError:
        messagebox.showerror("Erro Permissão", f"Permissão negada para salvar em '{caminho_arquivo_excel}'. Verifique se o arquivo está aberto ou se você tem permissão de escrita na pasta.")
    except Exception as e:
        messagebox.showerror("Erro ao Salvar", f"Erro desconhecido ao salvar Excel: {e}")

def calcular_media_demanda(df_passageiros_filtrado_por_tipo_dia, dias_para_divisao_no_periodo):
    if df_passageiros_filtrado_por_tipo_dia is None or df_passageiros_filtrado_por_tipo_dia.empty:
        return pd.DataFrame(columns=COLS_OUTPUT_EXCEL_DIARIO)
    
    df_calc = df_passageiros_filtrado_por_tipo_dia.copy()
    
    required_cols = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO', 'NMEFETPASST']
    if any(col not in df_calc.columns for col in required_cols):
        return pd.DataFrame(columns=COLS_OUTPUT_EXCEL_DIARIO)
        
    if 'DTOPERACAO' in df_calc.columns and not df_calc['DTOPERACAO'].empty:
        if not df_calc['DTOPERACAO'].apply(lambda x: isinstance(x, date)).all():
            df_calc.loc[:, 'DTOPERACAO'] = pd.to_datetime(df_calc['DTOPERACAO'], errors='coerce').dt.date
            df_calc.dropna(subset=['DTOPERACAO'], inplace=True)
            
    df_calc.drop_duplicates(subset=['CDOPERADOR', 'CDLINHA', 'DTOPERACAO'], keep='first', inplace=True)
    if df_calc.empty:
        return pd.DataFrame(columns=COLS_OUTPUT_EXCEL_DIARIO)
        
    df_final = df_calc[['CDOPERADOR', 'CDLINHA', 'DTOPERACAO']].copy()
    
    df_final['MÉDIA NMPASSTOTAL'] = pd.to_numeric(df_calc['NMEFETPASST'], errors='coerce').fillna(0).round(0).astype(int)
    
    if 'NMEFETPASSEQUIV' in df_calc.columns:
        df_final['MÉDIA NMPASSEQUIVALENTE'] = pd.to_numeric(df_calc['NMEFETPASSEQUIV'], errors='coerce').fillna(0).round(0).astype(int)
    else:
        df_final['MÉDIA NMPASSEQUIVALENTE'] = 0

    df_final.sort_values(by=['CDOPERADOR', 'CDLINHA', 'DTOPERACAO'], inplace=True)
    
    return df_final.reindex(columns=COLS_OUTPUT_EXCEL_DIARIO)

def calcular_media_mensal_demanda(df_passageiros_filtrado, config_usuario):
    if df_passageiros_filtrado is None or df_passageiros_filtrado.empty:
        return pd.DataFrame(columns=COLS_OUTPUT_EXCEL_DIARIO)

    df_calc = df_passageiros_filtrado.copy()
    
    if 'DTOPERACAO' in df_calc.columns and not df_calc['DTOPERACAO'].empty:
        if not df_calc['DTOPERACAO'].apply(lambda x: isinstance(x, date)).all():
            df_calc.loc[:, 'DTOPERACAO'] = pd.to_datetime(df_calc['DTOPERACAO'], errors='coerce').dt.date
            df_calc.dropna(subset=['DTOPERACAO'], inplace=True)

    if df_calc.empty:
        return pd.DataFrame(columns=COLS_OUTPUT_EXCEL_DIARIO)
        
    df_calc['DTOPERACAO_MESANO'] = pd.to_datetime(df_calc['DTOPERACAO']).dt.strftime('%m/%Y')
    
    aggregation_functions = {'NMEFETPASST': 'sum'}
    if config_usuario.get('processar_equivalente') and 'NMEFETPASSEQUIV' in df_calc.columns:
        aggregation_functions['NMEFETPASSEQUIV'] = 'sum'
        
    grouping_keys = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO_MESANO']
    dias_operados = df_calc.groupby(grouping_keys)['DTOPERACAO'].nunique().reset_index(name='DIAS_OPERADOS')
    df_summed = df_calc.groupby(grouping_keys).agg(aggregation_functions).reset_index()
    df_merged = pd.merge(df_summed, dias_operados, on=grouping_keys)
    
    df_merged['MÉDIA NMPASSTOTAL'] = (df_merged['NMEFETPASST'] / df_merged['DIAS_OPERADOS']).round(0).astype(int)
    
    if config_usuario.get('processar_equivalente') and 'NMEFETPASSEQUIV' in df_merged.columns:
        df_merged['MÉDIA NMPASSEQUIVALENTE'] = (df_merged['NMEFETPASSEQUIV'] / df_merged['DIAS_OPERADOS']).round(0).astype(int)
    else:
        df_merged['MÉDIA NMPASSEQUIVALENTE'] = 0

    df_merged.rename(columns={'DTOPERACAO_MESANO': 'DTOPERACAO'}, inplace=True)
    
    final_cols = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO', 'MÉDIA NMPASSTOTAL', 'MÉDIA NMPASSEQUIVALENTE']
    df_final = df_merged[[col for col in final_cols if col in df_merged.columns]]
    
    df_final.sort_values(by=['CDOPERADOR', 'CDLINHA', 'DTOPERACAO'], inplace=True)
    
    return df_final.reindex(columns=COLS_OUTPUT_EXCEL_DIARIO)

def contar_dias_operantes_reais_por_dsdiatipo(lista_periodos_obj, tipo_dia_alvo, df_lookup_dsdiatipo_mapa):
    tipo_dia_alvo_upper = tipo_dia_alvo.upper()
    if tipo_dia_alvo_upper == "TODOS":
        return contar_dias_no_periodo_por_tipo(lista_periodos_obj, "TODOS")
    if df_lookup_dsdiatipo_mapa is None or df_lookup_dsdiatipo_mapa.empty:
        return contar_dias_no_periodo_por_tipo(lista_periodos_obj, tipo_dia_alvo_upper)
    if 'DTOPERACAO' not in df_lookup_dsdiatipo_mapa.columns or 'DSDIATIPO' not in df_lookup_dsdiatipo_mapa.columns:
        return contar_dias_no_periodo_por_tipo(lista_periodos_obj, tipo_dia_alvo_upper)
    df_lookup_copy = df_lookup_dsdiatipo_mapa.copy()
    if not df_lookup_copy['DTOPERACAO'].empty:
        non_null_dates = df_lookup_copy['DTOPERACAO'].dropna()
        if not non_null_dates.empty and not isinstance(non_null_dates.iloc[0], date):
            df_lookup_copy.loc[:, 'DTOPERACAO'] = pd.to_datetime(df_lookup_copy['DTOPERACAO'], errors='coerce').dt.date
            df_lookup_copy.dropna(subset=['DTOPERACAO'], inplace=True)
    else: return 0
    datas_operantes_do_tipo_especifico = set()
    for periodo in lista_periodos_obj:
        data_inicio_periodo = periodo['inicio'].date() if isinstance(periodo['inicio'], datetime) else periodo['inicio']
        data_fim_periodo = periodo['fim'].date() if isinstance(periodo['fim'], datetime) else periodo['fim']
        dias_filtrados_mapa = df_lookup_copy[
            (df_lookup_copy['DTOPERACAO'] >= data_inicio_periodo) &
            (df_lookup_copy['DTOPERACAO'] <= data_fim_periodo) &
            (df_lookup_copy['DSDIATIPO'].astype(str).str.upper() == tipo_dia_alvo_upper)
        ]
        datas_operantes_do_tipo_especifico.update(dias_filtrados_mapa['DTOPERACAO'].unique())
    return len(datas_operantes_do_tipo_especifico)

def _processar_e_coletar_dados_para_excel(df_base_filtrado_periodos_usuario, config_usuario, df_dsdiatipo_map_externo=None):
    df_media_total_calculada = pd.DataFrame(columns=COLS_OUTPUT_EXCEL_DIARIO)
    lista_dados_abas_secundarias = []
    if df_base_filtrado_periodos_usuario is None or df_base_filtrado_periodos_usuario.empty:
        return df_media_total_calculada, lista_dados_abas_secundarias
    df_base_para_processar = df_base_filtrado_periodos_usuario.copy()
    if 'DTOPERACAO' in df_base_para_processar.columns and not df_base_para_processar['DTOPERACAO'].empty:
        if not df_base_para_processar['DTOPERACAO'].apply(lambda x: isinstance(x, date)).all():
            df_base_para_processar.loc[:, 'DTOPERACAO'] = pd.to_datetime(df_base_para_processar['DTOPERACAO'], errors='coerce').dt.date
            df_base_para_processar.dropna(subset=['DTOPERACAO'], inplace=True)
        if df_base_para_processar.empty:
            return df_media_total_calculada, lista_dados_abas_secundarias
    tipo_dia_selecionado_usuario = config_usuario['tipo_dia_selecionado'].upper()
    df_passageiros_a_somar = pd.DataFrame() 

    if config_usuario['tipo_relatorio_raw'] == "concessionarias":
        if tipo_dia_selecionado_usuario == "TODOS":
            df_passageiros_a_somar = df_base_para_processar.copy()
        else:
            if 'DSDIATIPO' in df_base_para_processar.columns and not df_base_para_processar['DSDIATIPO'].isnull().all():
                df_temp_dsdiatipo_conc = df_base_para_processar.dropna(subset=['DSDIATIPO']).copy()
                df_temp_dsdiatipo_conc.loc[:, 'DSDIATIPO'] = df_temp_dsdiatipo_conc['DSDIATIPO'].astype(str).str.upper()
                df_passageiros_a_somar = df_temp_dsdiatipo_conc[df_temp_dsdiatipo_conc['DSDIATIPO'] == tipo_dia_selecionado_usuario].copy()
            else:
                messagebox.showwarning("Aviso DSDIATIPO (Concessionárias)", f"Coluna DSDIATIPO não disponível ou sem valores para filtrar Concessionárias por '{tipo_dia_selecionado_usuario}'. A média será 0.")
                df_passageiros_a_somar = pd.DataFrame()
    elif config_usuario['tipo_relatorio_raw'] == "permissionarias":
        if tipo_dia_selecionado_usuario == "TODOS":
            df_passageiros_a_somar = df_base_para_processar.copy()
        else: 
            if 'DSDIATIPO' in df_base_para_processar.columns:
                df_temp_dsdiatipo = df_base_para_processar.dropna(subset=['DSDIATIPO']).copy()
                df_temp_dsdiatipo.loc[:, 'DSDIATIPO'] = df_temp_dsdiatipo['DSDIATIPO'].astype(str).str.upper()
                df_passageiros_a_somar = df_temp_dsdiatipo[df_temp_dsdiatipo['DSDIATIPO'] == tipo_dia_selecionado_usuario].copy()
            else: 
                messagebox.showwarning("Aviso DSDIATIPO", f"Coluna DSDIATIPO ausente em {config_usuario['tipo_relatorio_raw']} para filtrar passageiros por '{tipo_dia_selecionado_usuario}'. Somatório considerará todos os dados.")
                df_passageiros_a_somar = df_base_para_processar.copy()
    if df_passageiros_a_somar.empty and tipo_dia_selecionado_usuario != "TODOS" and config_usuario['tipo_relatorio_raw'] != "stpp_rmr":
        pass 
    if config_usuario['tipo_relatorio_raw'] == "stpp_rmr":
        df_perm_component = df_base_para_processar[df_base_para_processar['CDOPERADOR'].isin(PERMISSIONARIAS_EMPRESAS)].copy()
        df_conc_component = df_base_para_processar[df_base_para_processar['CDOPERADOR'].isin(CONCESSIONARIAS_EMPRESAS)].copy()
        df_para_calculo_final = pd.DataFrame()
        
        lista_componentes = []
        if not df_perm_component.empty:
            df_pass_perm_stpp = pd.DataFrame()
            if tipo_dia_selecionado_usuario == "TODOS": df_pass_perm_stpp = df_perm_component
            elif 'DSDIATIPO' in df_perm_component.columns: 
                df_temp_p = df_perm_component.dropna(subset=['DSDIATIPO']).copy()
                df_temp_p.loc[:, 'DSDIATIPO'] = df_temp_p['DSDIATIPO'].astype(str).str.upper()
                df_pass_perm_stpp = df_temp_p[df_temp_p['DSDIATIPO'] == tipo_dia_selecionado_usuario]
            else: df_pass_perm_stpp = df_perm_component
            if not df_pass_perm_stpp.empty: lista_componentes.append(df_pass_perm_stpp)
            
        if not df_conc_component.empty:
            df_pass_conc_stpp = pd.DataFrame()
            if tipo_dia_selecionado_usuario == "TODOS":
                df_pass_conc_stpp = df_conc_component.copy()
            else:
                if 'DSDIATIPO' in df_conc_component.columns and not df_conc_component['DSDIATIPO'].isnull().all():
                    df_temp_c = df_conc_component.dropna(subset=['DSDIATIPO']).copy()
                    df_temp_c.loc[:, 'DSDIATIPO'] = df_temp_c['DSDIATIPO'].astype(str).str.upper()
                    df_pass_conc_stpp = df_temp_c[df_temp_c['DSDIATIPO'] == tipo_dia_selecionado_usuario].copy()
                else:
                    messagebox.showwarning("Aviso DSDIATIPO (STPP-Conc)", f"Coluna DSDIATIPO não disponível no componente Concessionária para filtrar por '{tipo_dia_selecionado_usuario}'.")
                    df_pass_conc_stpp = pd.DataFrame()
            if not df_pass_conc_stpp.empty: lista_componentes.append(df_pass_conc_stpp)
        
        if lista_componentes:
            df_para_calculo_final = pd.concat(lista_componentes, ignore_index=True)
        else:
            df_para_calculo_final = pd.DataFrame()
            
        df_passageiros_a_somar = df_para_calculo_final

    if config_usuario.get('exibir_formato_mensal'):
        df_media_total_calculada = calcular_media_mensal_demanda(df_passageiros_a_somar, config_usuario)
    else:
        df_media_total_calculada = calcular_media_demanda(df_passageiros_a_somar, 1)
        
    if not df_media_total_calculada.empty:
        df_media_total_calculada.sort_values(by=['CDOPERADOR', 'CDLINHA', 'DTOPERACAO'], ascending=[True, True, True], inplace=True)

    return df_media_total_calculada, lista_dados_abas_secundarias

def run_demanda_media_permissionarias():
    root_dialogo = tk.Tk(); root_dialogo.withdraw()
    caminho_arquivo_txt = filedialog.askopenfilename(parent=None, title="Selecione o TXT das Permissionárias (com DSDIATIPO)", filetypes=(("Arquivos de Texto", "*.txt"), ("Todos os arquivos", "*.*")))
    root_dialogo.destroy()
    if not caminho_arquivo_txt: messagebox.showwarning("Seleção de Arquivo", "Nenhum TXT selecionado."); return
    config_usuario = obter_entrada_gui_media("Demanda Média - Permissionárias", PERMISSIONARIAS_EMPRESAS, "permissionarias")
    if config_usuario is None: return
    df_perm_base_filtrado_periodos_usuario = carregar_dados_txt_base(caminho_arquivo_txt, COLS_PERMISSIONARIAS_INPUT, config_usuario)
    if df_perm_base_filtrado_periodos_usuario is None or df_perm_base_filtrado_periodos_usuario.empty: 
        messagebox.showerror("Erro Processamento", "Falha ao carregar/processar TXT ou nenhum dado encontrado após filtros iniciais."); 
        return
    
    df_resultado_final, _ = _processar_e_coletar_dados_para_excel(df_perm_base_filtrado_periodos_usuario, config_usuario, df_dsdiatipo_map_externo=None)
    
    if df_resultado_final is None or df_resultado_final.empty:
        messagebox.showinfo("Resultado", "Nenhum dado para gerar o relatório Excel.")
        return

    exportar_excel_com_multiplas_abas(
        df_resultado_final,
        config_usuario['titulo_relatorio_principal'],
        "Relatório de Demanda",
        [],
        config_usuario['nome_arquivo_excel'],
        config_usuario
    )
    if os.path.exists(config_usuario['nome_arquivo_excel']):
        messagebox.showinfo("Exportação Concluída", f"Relatório gerado!\n\nLocal: {config_usuario['nome_arquivo_excel']}")

def run_demanda_media_concessionarias():
    root_dialogo = tk.Tk(); root_dialogo.withdraw()
    caminho_txt_conc_dados = filedialog.askopenfilename(parent=None, title="CONC: Selecione o TXT de CONCESSIONÁRIAS (dados)", filetypes=(("Arquivos de Texto", "*.txt"), ("Todos os arquivos", "*.*")))
    if not caminho_txt_conc_dados: messagebox.showwarning("Seleção de Arquivo", "Arquivo de Concessionárias (dados) não selecionado."); root_dialogo.destroy(); return
    
    caminho_txt_perm_for_dsdiatipo = filedialog.askopenfilename(parent=None, title="CONC: Selecione o TXT de PERMISSIONÁRIAS (para DSDIATIPO)", filetypes=(("Arquivos de Texto", "*.txt"), ("Todos os arquivos", "*.*")))
    root_dialogo.destroy()
    if not caminho_txt_perm_for_dsdiatipo: messagebox.showwarning("Seleção de Arquivo", "Arquivo de Permissionárias (para DSDIATIPO) não selecionado."); return

    config_usuario = obter_entrada_gui_media("Demanda Média - Concessionárias", CONCESSIONARIAS_EMPRESAS, "concessionarias")
    if config_usuario is None: return

    config_lookup_conc_dsdiatipo = {'lista_periodos_obj': config_usuario['lista_periodos_obj'], 'cod_linhas_str': 'TODAS', 'empresas_selecionadas': PERMISSIONARIAS_EMPRESAS}
    df_perm_map_lookup_conc_raw = carregar_dados_txt_base(caminho_txt_perm_for_dsdiatipo, ['DTOPERACAO', 'DSDIATIPO'], config_lookup_conc_dsdiatipo)
    df_dsdiatipo_map_concessionarias_global = pd.DataFrame()
    if df_perm_map_lookup_conc_raw is not None and not df_perm_map_lookup_conc_raw.empty and 'DSDIATIPO' in df_perm_map_lookup_conc_raw.columns:
        df_map_copy_conc = df_perm_map_lookup_conc_raw.copy()
        if 'DTOPERACAO' in df_map_copy_conc.columns and not df_map_copy_conc['DTOPERACAO'].empty:
            if not df_map_copy_conc['DTOPERACAO'].apply(lambda x: isinstance(x, date)).all():
                df_map_copy_conc.loc[:, 'DTOPERACAO'] = pd.to_datetime(df_map_copy_conc['DTOPERACAO'], errors='coerce').dt.date
                df_map_copy_conc.dropna(subset=['DTOPERACAO'], inplace=True)
        df_dsdiatipo_map_concessionarias_global = df_map_copy_conc[['DTOPERACAO', 'DSDIATIPO']].drop_duplicates().dropna().reset_index(drop=True)
    if df_dsdiatipo_map_concessionarias_global.empty:
        messagebox.showwarning("Aviso Lookup (Concessionárias)", "Mapa DSDIATIPO para Concessionárias vazio ou não pôde ser criado. O cálculo pode ser afetado.");

    cols_conc_input = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO', 'NMPASSTOTAL']
    if config_usuario.get('processar_equivalente'):
        cols_conc_input.append('NMPASSEQUIVALENTE')
        
    df_conc_raw_filtrado_periodos = carregar_dados_txt_base(caminho_txt_conc_dados, cols_conc_input, config_usuario)
    if df_conc_raw_filtrado_periodos is None or df_conc_raw_filtrado_periodos.empty:
        messagebox.showerror("Erro Processamento", "Falha ao carregar dados de Concessionárias ou nenhum dado encontrado após filtros."); 
        return
    
    rename_dict = {'NMPASSTOTAL': 'NMEFETPASST'}
    if config_usuario.get('processar_equivalente') and 'NMPASSEQUIVALENTE' in df_conc_raw_filtrado_periodos.columns:
        rename_dict['NMPASSEQUIVALENTE'] = 'NMEFETPASSEQUIV'
    df_conc_raw_filtrado_periodos.rename(columns=rename_dict, inplace=True)

    if not df_dsdiatipo_map_concessionarias_global.empty and not df_conc_raw_filtrado_periodos.empty:
        if 'DTOPERACAO' in df_conc_raw_filtrado_periodos.columns and not df_conc_raw_filtrado_periodos['DTOPERACAO'].empty:
            if not df_conc_raw_filtrado_periodos['DTOPERACAO'].apply(lambda x: isinstance(x, date)).all():
                df_conc_raw_filtrado_periodos.loc[:, 'DTOPERACAO'] = pd.to_datetime(df_conc_raw_filtrado_periodos['DTOPERACAO'], errors='coerce').dt.date
                df_conc_raw_filtrado_periodos.dropna(subset=['DTOPERACAO'], inplace=True)
        df_conc_raw_filtrado_periodos = pd.merge(df_conc_raw_filtrado_periodos, df_dsdiatipo_map_concessionarias_global, on='DTOPERACAO', how='left')
    elif not df_conc_raw_filtrado_periodos.empty: 
        df_conc_raw_filtrado_periodos['DSDIATIPO'] = pd.NA
    
    df_resultado_final, _ = _processar_e_coletar_dados_para_excel(df_conc_raw_filtrado_periodos, config_usuario, df_dsdiatipo_map_externo=df_dsdiatipo_map_concessionarias_global)
    
    if df_resultado_final is None or df_resultado_final.empty:
        messagebox.showinfo("Resultado", "Nenhum dado para gerar o relatório Excel.")
        return
        
    exportar_excel_com_multiplas_abas(
        df_resultado_final,
        config_usuario['titulo_relatorio_principal'],
        "Relatório de Demanda",
        [],
        config_usuario['nome_arquivo_excel'],
        config_usuario
    )
    if os.path.exists(config_usuario['nome_arquivo_excel']):
        messagebox.showinfo("Exportação Concluída", f"Relatório gerado!\n\nLocal: {config_usuario['nome_arquivo_excel']}")

def run_demanda_media_stpp_rmr():
    root_dialogo = tk.Tk(); root_dialogo.withdraw()
    caminho_txt_perm = filedialog.askopenfilename(parent=None, title="STPP/RMR: TXT PERMISSIONÁRIAS (dados e DSDIATIPO)", filetypes=(("Arquivos de Texto", "*.txt"), ("Todos os arquivos", "*.*")))
    if not caminho_txt_perm: messagebox.showwarning("Seleção de Arquivo", "Arquivo de Permissionárias não selecionado."); root_dialogo.destroy(); return
    caminho_txt_conc = filedialog.askopenfilename(parent=None, title="STPP/RMR: TXT CONCESSIONÁRIAS (dados)", filetypes=(("Arquivos de Texto", "*.txt"), ("Todos os arquivos", "*.*")))
    root_dialogo.destroy()
    if not caminho_txt_conc: messagebox.showwarning("Seleção de Arquivo", "Arquivo de Concessionárias não selecionado."); return
    config_usuario = obter_entrada_gui_media("Demanda Média - STPP/RMR", STPP_RMR_EMPRESAS_TODAS, "stpp_rmr")
    if config_usuario is None: return
    df_list_para_concatenar = []
    empresas_selecionadas_pelo_usuario = config_usuario.get('empresas_selecionadas', [])
    todas_empresas_selecionadas = "TODAS" in empresas_selecionadas_pelo_usuario or not empresas_selecionadas_pelo_usuario
    
    config_lookup_stpp_temp = {'lista_periodos_obj': config_usuario['lista_periodos_obj'], 'cod_linhas_str': 'TODAS', 'empresas_selecionadas': PERMISSIONARIAS_EMPRESAS}
    df_perm_map_lookup_stpp_raw = carregar_dados_txt_base(caminho_txt_perm, ['DTOPERACAO', 'DSDIATIPO'], config_lookup_stpp_temp)
    df_dsdiatipo_map_stpp_global = pd.DataFrame()
    if df_perm_map_lookup_stpp_raw is None or df_perm_map_lookup_stpp_raw.empty or 'DSDIATIPO' not in df_perm_map_lookup_stpp_raw.columns:
        messagebox.showwarning("Aviso Lookup (STPP)", "Falha ao criar mapa DSDIATIPO. Divisor pode usar contagem de calendário.");
    else:
        df_map_copy = df_perm_map_lookup_stpp_raw.copy()
        if 'DTOPERACAO' in df_map_copy.columns and not df_map_copy['DTOPERACAO'].empty:
            non_null_dates_map_stpp = df_map_copy['DTOPERACAO'].dropna()
            if not non_null_dates_map_stpp.empty: 
                if not isinstance(non_null_dates_map_stpp.iloc[0], date) and not pd.api.types.is_datetime64_any_dtype(non_null_dates_map_stpp):
                    df_map_copy.loc[:, 'DTOPERACAO'] = pd.to_datetime(df_map_copy['DTOPERACAO'], errors='coerce').dt.date
        df_dsdiatipo_map_stpp_global = df_map_copy[['DTOPERACAO', 'DSDIATIPO']].drop_duplicates().dropna().reset_index(drop=True)
        if df_dsdiatipo_map_stpp_global.empty: messagebox.showwarning("Aviso Lookup (STPP)", "Mapa DSDIATIPO (STPP) vazio após processamento.");
    
    cols_stpp_rmr_processado_dinamico = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO', 'NMEFETPASST', 'DSDIATIPO']
    if config_usuario.get('processar_equivalente'):
        cols_stpp_rmr_processado_dinamico.insert(4, 'NMEFETPASSEQUIV')

    empresas_perm_a_processar = PERMISSIONARIAS_EMPRESAS if todas_empresas_selecionadas else [emp for emp in empresas_selecionadas_pelo_usuario if emp in PERMISSIONARIAS_EMPRESAS]
    if empresas_perm_a_processar:
        config_perm_stpp_temp = config_usuario.copy(); config_perm_stpp_temp['empresas_selecionadas'] = empresas_perm_a_processar
        df_perm_raw_stpp = carregar_dados_txt_base(caminho_txt_perm, COLS_PERMISSIONARIAS_INPUT, config_perm_stpp_temp)
        if df_perm_raw_stpp is not None and not df_perm_raw_stpp.empty:
            df_list_para_concatenar.append(df_perm_raw_stpp.reindex(columns=cols_stpp_rmr_processado_dinamico))

    empresas_conc_a_processar = CONCESSIONARIAS_EMPRESAS if todas_empresas_selecionadas else [emp for emp in empresas_selecionadas_pelo_usuario if emp in CONCESSIONARIAS_EMPRESAS]
    if empresas_conc_a_processar:
        config_conc_stpp_temp = config_usuario.copy(); config_conc_stpp_temp['empresas_selecionadas'] = empresas_conc_a_processar
        
        cols_conc_input_stpp = ['CDOPERADOR', 'CDLINHA', 'DTOPERACAO', 'NMPASSTOTAL']
        if config_usuario.get('processar_equivalente'):
            cols_conc_input_stpp.append('NMPASSEQUIVALENTE')

        df_conc_raw_stpp = carregar_dados_txt_base(caminho_txt_conc, cols_conc_input_stpp, config_conc_stpp_temp)
        if df_conc_raw_stpp is not None and not df_conc_raw_stpp.empty:
            rename_dict_stpp = {'NMPASSTOTAL': 'NMEFETPASST'}
            if config_usuario.get('processar_equivalente') and 'NMPASSEQUIVALENTE' in df_conc_raw_stpp.columns:
                rename_dict_stpp['NMPASSEQUIVALENTE'] = 'NMEFETPASSEQUIV'
            df_conc_raw_stpp.rename(columns=rename_dict_stpp, inplace=True)

            df_conc_processed_stpp = df_conc_raw_stpp.copy()
            if 'DTOPERACAO' in df_conc_processed_stpp.columns and not df_conc_processed_stpp['DTOPERACAO'].empty:
                non_null_dates_conc_stpp = df_conc_processed_stpp['DTOPERACAO'].dropna()
                if not non_null_dates_conc_stpp.empty : 
                    if not isinstance(non_null_dates_conc_stpp.iloc[0], date) and not pd.api.types.is_datetime64_any_dtype(non_null_dates_conc_stpp): 
                        df_conc_processed_stpp.loc[:, 'DTOPERACAO'] = pd.to_datetime(df_conc_processed_stpp['DTOPERACAO'], errors='coerce').dt.date
                        df_conc_processed_stpp.dropna(subset=['DTOPERACAO'], inplace=True)
            if not df_dsdiatipo_map_stpp_global.empty and not df_conc_processed_stpp.empty:
                df_conc_merged_stpp = pd.merge(df_conc_processed_stpp, df_dsdiatipo_map_stpp_global, on='DTOPERACAO', how='left')
            else:
                df_conc_merged_stpp = df_conc_processed_stpp
                if 'DSDIATIPO' not in df_conc_merged_stpp.columns: df_conc_merged_stpp['DSDIATIPO'] = pd.NA
                if df_dsdiatipo_map_stpp_global.empty : pass
            if 'DSDIATIPO' in df_conc_merged_stpp and df_conc_merged_stpp['DSDIATIPO'].isnull().all() and config_usuario.get('tipo_dia_selecionado').upper() != 'TODOS':
                messagebox.showwarning("Aviso Merge DSDIATIPO (STPP-Conc)", "Nenhuma correspondência de DTOPERACAO para preencher DSDIATIPO nos dados de Concessionária (STPP).");
            df_list_para_concatenar.append(df_conc_merged_stpp.reindex(columns=cols_stpp_rmr_processado_dinamico))

    if not df_list_para_concatenar:
        messagebox.showinfo("Resultado STPP/RMR", "Nenhum dado encontrado para Permissionárias ou Concessionárias com os filtros aplicados."); 
        return
    df_stpp_rmr_base_completo = pd.concat(df_list_para_concatenar, ignore_index=True, sort=False)
    for col in cols_stpp_rmr_processado_dinamico:
        if col not in df_stpp_rmr_base_completo.columns: df_stpp_rmr_base_completo[col] = pd.NA
    df_stpp_rmr_base_completo = df_stpp_rmr_base_completo.reindex(columns=cols_stpp_rmr_processado_dinamico)
    if df_stpp_rmr_base_completo.empty: 
        messagebox.showinfo("Resultado STPP/RMR", "DataFrame combinado STPP/RMR vazio após processamento."); 
        return
    df_resultado_final, _ = _processar_e_coletar_dados_para_excel(df_stpp_rmr_base_completo, config_usuario, df_dsdiatipo_map_externo=df_dsdiatipo_map_stpp_global)
    
    if df_resultado_final is None or df_resultado_final.empty:
        messagebox.showinfo("Resultado", "Nenhum dado para gerar o relatório Excel.")
        return
        
    exportar_excel_com_multiplas_abas(
        df_resultado_final,
        config_usuario['titulo_relatorio_principal'],
        "Relatório de Demanda",
        [],
        config_usuario['nome_arquivo_excel'],
        config_usuario
    )
    if os.path.exists(config_usuario['nome_arquivo_excel']):
        messagebox.showinfo("Exportação Concluída", f"Relatório gerado!\n\nLocal: {config_usuario['nome_arquivo_excel']}")

if __name__ == "__main__":
    tipo_processamento_escolhido = selecionar_tipo_processamento_gui_media()
    if tipo_processamento_escolhido == "permissionarias": run_demanda_media_permissionarias()
    elif tipo_processamento_escolhido == "concessionarias": run_demanda_media_concessionarias()
    elif tipo_processamento_escolhido == "stpp_rmr": run_demanda_media_stpp_rmr()
    elif tipo_processamento_escolhido is None:
        pass
    else:
        root_msg = tk.Tk(); root_msg.withdraw()
        messagebox.showwarning("Opção Inválida", f"Tipo de processamento '{tipo_processamento_escolhido}' não reconhecido.")
        root_msg.destroy()