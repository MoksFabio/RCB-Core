import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import traceback
import pandas as pd
from simpledbf import Dbf5
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
import datetime
import subprocess
import sys
import threading
import copy
from PIL import Image, ImageTk
from tkinterdnd2 import DND_FILES, TkinterDnD
import sv_ttk

class Config:
    NOMES_EMPRESAS = {
        "BOA": "BOA - Borborema Imperial Transportes Ltda", "CAX": "CAX - Rodoviária Caxangá Ltda",
        "CSR": "CSR - Consórcio Recife de Transporte",
        "EME": "EME - Expresso Metropolitana Ltda", "GLO": "GLO - Transportadora Globo Ltda",
        "SJT": "SJT - São Judas Tadeu", "VML": "VML - Viação Mirim Ltda",
    }
    FONT_TITULO_PRINCIPAL = Font(bold=True, size=16, color="FFFFFF")
    FILL_TITULO_PRINCIPAL = PatternFill(start_color="2F4F4F", fill_type="solid")
    FONT_HEADER_TABELA = Font(bold=True, color="000000")
    FILL_HEADER_TABELA = PatternFill(start_color="DDEBF7", fill_type="solid")
    FONT_TOTAL = Font(bold=True)
    FILL_TOTAL = PatternFill(start_color="F2F2F2", fill_type="solid")
    FONT_STATUS_OK = Font(color="006100", bold=True)
    FONT_STATUS_DIFF = Font(color="9C0006", bold=True)
    DXF_DIFF = DifferentialStyle(font=Font(color="FF0000"))
    ZEBRA_FILL = PatternFill(start_color="F5F5F5", fill_type="solid")
    BORDER_THIN = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

def adicionar_marca_dagua(ws, anexo_celula='P1'):
    caminho_logo = "static/images/rcb_logo.png"
    try:
        img = OpenpyxlImage(caminho_logo)
        img.height = 60
        img.width = 76
        ws.add_image(img, anexo_celula)
    except FileNotFoundError:
        print(f"AVISO: Arquivo de imagem para marca d'água não encontrado em '{caminho_logo}'.")
    except Exception as e:
        print(f"AVISO: Não foi possível adicionar a marca d'água. Erro: {e}")

def auto_ajustar_colunas(ws):
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

def aplicar_wrap_text_em_todas_abas(workbook):
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                new_alignment = copy.copy(cell.alignment)
                new_alignment.wrap_text = True
                cell.alignment = new_alignment

def processar_arquivos_e_gerar_relatorio(gui_instance):
    try:
        wb_destino = openpyxl.Workbook()
        ws_sabe = wb_destino.active
        ws_sabe.title = "SABE - Total de Passageiros"

        empresas_interesse = ['BOA', 'CAX', 'CSR', 'EME', 'GLO', 'SJT', 'VML']

        df_remuneracao_totais = pd.DataFrame()
        if gui_instance.arquivo_txt_remuneracao:
            try:
                df_original_txt = pd.read_csv(gui_instance.arquivo_txt_remuneracao, sep='\t', header=0, encoding='utf-8')

                coluna_operador_txt = 'CDOPERADOR'
                coluna_nme_efet_passt = 'NMEFETPASST'

                if coluna_operador_txt not in df_original_txt.columns or coluna_nme_efet_passt not in df_original_txt.columns:
                    raise KeyError(f"Colunas esperadas no TXT ('{coluna_operador_txt}', '{coluna_nme_efet_passt}') não encontradas. Colunas disponíveis: {df_original_txt.columns.tolist()}")

                def converter_para_float(valor):
                    try:
                        if pd.isna(valor) or str(valor).strip() == '':
                            return 0.0
                        clean_valor = ''.join(c for c in str(valor) if c.isdigit() or c == ',' or c == '.')
                        return float(clean_valor.replace(',', '.'))
                    except ValueError:
                        return 0.0

                df_original_txt[coluna_nme_efet_passt] = df_original_txt[coluna_nme_efet_passt].apply(converter_para_float)

                df_filtrado_txt = df_original_txt[df_original_txt[coluna_operador_txt].isin(empresas_interesse)].copy()
                
                df_remuneracao_totais = df_filtrado_txt.groupby(coluna_operador_txt)[coluna_nme_efet_passt].sum().reset_index()
                
                df_remuneracao_totais.rename(columns={
                    coluna_operador_txt: 'Empresa',
                    coluna_nme_efet_passt: 'Total CTM' 
                }, inplace=True)

                df_empresas_base = pd.DataFrame({'Empresa': empresas_interesse})
                df_remuneracao_totais = pd.merge(df_empresas_base, df_remuneracao_totais, on='Empresa', how='left')
                df_remuneracao_totais['Total CTM'] = df_remuneracao_totais['Total CTM'].fillna(0.0)
                
            except FileNotFoundError:
                messagebox.showerror("Erro (TXT)", f"O arquivo TXT '{gui_instance.arquivo_txt_remuneracao}' não foi encontrado.", parent=gui_instance)
                gui_instance.finalizar_processamento_gui("Erro ao processar TXT.", sucesso=False)
                return
            except KeyError as ke:
                messagebox.showerror("Erro (TXT)", f"Uma coluna esperada para o TXT não foi encontrada. Detalhe do erro: {ke}", parent=gui_instance)
                gui_instance.finalizar_processamento_gui("Erro ao processar TXT (coluna ausente).", sucesso=False)
                return
            except Exception as e:
                messagebox.showerror("Erro (TXT)", f"ERRO INESPERADO durante o processamento do TXT: {e}", parent=gui_instance)
                gui_instance.finalizar_processamento_gui("Erro inesperado no TXT.", sucesso=False)
                return
        else:
            df_remuneracao_totais = pd.DataFrame({'Empresa': empresas_interesse, 'Total CTM': 0.0})

        df_catraca_totais = pd.DataFrame()
        if gui_instance.arquivo_dbf_conferencia:
            try:
                dbf = Dbf5(gui_instance.arquivo_dbf_conferencia, codec='latin-1') 
                df_dbf = dbf.to_dataframe() 

                coluna_operador_dbf = 'OPERADORA' 
                coluna_catraca_fi = 'CATRACA_FI' 
                coluna_linha_dbf = 'LINHA' 

                if coluna_operador_dbf not in df_dbf.columns or \
                   coluna_catraca_fi not in df_dbf.columns or \
                   coluna_linha_dbf not in df_dbf.columns:
                    raise KeyError(f"Colunas esperadas no DBF ('{coluna_operador_dbf}', '{coluna_catraca_fi}', '{coluna_linha_dbf}') não encontradas. Colunas disponíveis: {df_dbf.columns.tolist()}")
                
                df_dbf_filtrado_por_linha = df_dbf[df_dbf[coluna_linha_dbf].astype(str).str.strip().str[0].str.isdigit()].copy()
                
                df_dbf_filtrado_por_linha[coluna_catraca_fi] = pd.to_numeric(df_dbf_filtrado_por_linha[coluna_catraca_fi].astype(str).str.replace(',', '.'), errors='coerce').fillna(0.0)
                
                df_filtrado_dbf_por_empresa = df_dbf_filtrado_por_linha[df_dbf_filtrado_por_linha[coluna_operador_dbf].isin(empresas_interesse)].copy()

                df_catraca_totais = df_filtrado_dbf_por_empresa.groupby(coluna_operador_dbf)[coluna_catraca_fi].sum().reset_index()

                df_catraca_totais.rename(columns={
                    coluna_operador_dbf: 'Empresa',
                    coluna_catraca_fi: 'Total Urbana' 
                }, inplace=True)

                df_empresas_base = pd.DataFrame({'Empresa': empresas_interesse})
                df_catraca_totais = pd.merge(df_empresas_base, df_catraca_totais, on='Empresa', how='left')
                df_catraca_totais['Total Urbana'] = df_catraca_totais['Total Urbana'].fillna(0.0)
                
            except FileNotFoundError:
                messagebox.showerror("Erro (DBF)", f"O arquivo DBF '{gui_instance.arquivo_dbf_conferencia}' não foi encontrado.", parent=gui_instance)
                gui_instance.finalizar_processamento_gui("Erro ao processar DBF.", sucesso=False)
                return
            except KeyError as ke:
                messagebox.showerror("Erro (DBF)", f"Uma coluna esperada para o DBF não foi encontrada. Detalhe do erro: {ke}", parent=gui_instance)
                gui_instance.finalizar_processamento_gui("Erro ao processar DBF (coluna ausente).", sucesso=False)
                return
            except Exception as e:
                messagebox.showerror("Erro (DBF)", f"ERRO INESPERADO durante o processamento do DBF: {e}", parent=gui_instance)
                gui_instance.finalizar_processamento_gui("Erro inesperado no DBF.", sucesso=False)
                return
        else:
            df_catraca_totais = pd.DataFrame({'Empresa': empresas_interesse, 'Total Urbana': 0.0})

        df_comparacao_final = pd.merge(
            df_remuneracao_totais,
            df_catraca_totais,
            on='Empresa',
            how='outer'
        )
        
        df_comparacao_final = df_comparacao_final.fillna(0.0)
        df_comparacao_final['Diferença'] = df_comparacao_final['Total CTM'] - df_comparacao_final['Total Urbana']

        df_comparacao_final['Empresa_Order'] = pd.Categorical(df_comparacao_final['Empresa'], categories=empresas_interesse, ordered=True)
        df_comparacao_final = df_comparacao_final.sort_values('Empresa_Order').drop('Empresa_Order', axis=1)

        df_comparacao_final['Status'] = df_comparacao_final['Diferença'].apply(lambda x: "✓ OK" if abs(x) < 0.01 else "✗ Diferença")

        df_comparacao_final = df_comparacao_final[['Status', 'Empresa', 'Total CTM', 'Total Urbana', 'Diferença']]

        current_row = 1

        ws_sabe.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=df_comparacao_final.shape[1])
        cell_titulo = ws_sabe.cell(row=current_row, column=1, value="SABE - CONFERÊNCIA TOTAL DE PASSAGEIROS")
        cell_titulo.font = Config.FONT_TITULO_PRINCIPAL
        cell_titulo.fill = Config.FILL_TITULO_PRINCIPAL
        cell_titulo.alignment = Config.ALIGN_CENTER
        current_row += 1

        start_data_row = current_row + 1
        for col_idx, col_name in enumerate(df_comparacao_final.columns, 1):
            cell = ws_sabe.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = Config.FONT_HEADER_TABELA
            cell.fill = Config.FILL_HEADER_TABELA
            cell.alignment = Config.ALIGN_CENTER
            cell.border = Config.BORDER_THIN
        current_row += 1

        for r_idx, row_data in df_comparacao_final.iterrows():
            for c_idx, value in enumerate(row_data, 1):
                cell = ws_sabe.cell(row=current_row, column=c_idx, value=value)
                cell.border = Config.BORDER_THIN
                if r_idx % 2 != 0:
                    cell.fill = Config.ZEBRA_FILL
                
                if c_idx in [3, 4, 5]:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Config.ALIGN_RIGHT
                elif c_idx == 1:
                    cell.alignment = Config.ALIGN_CENTER
                    if row_data['Status'] == "✓ OK":
                        cell.font = Config.FONT_STATUS_OK
                    else:
                        cell.font = Config.FONT_STATUS_DIFF
                else:
                    cell.alignment = Config.ALIGN_LEFT
            current_row += 1

        total_geral_ctm = df_comparacao_final['Total CTM'].sum()
        total_geral_urbana = df_comparacao_final['Total Urbana'].sum()
        total_geral_diferenca = df_comparacao_final['Diferença'].sum()

        ws_sabe.cell(row=current_row, column=2, value="TOTAL GERAL").font = Config.FONT_TOTAL
        ws_sabe.cell(row=current_row, column=2).alignment = Config.ALIGN_LEFT
        ws_sabe.cell(row=current_row, column=3, value=total_geral_ctm).number_format = '#,##0.00'
        ws_sabe.cell(row=current_row, column=4, value=total_geral_urbana).number_format = '#,##0.00'
        ws_sabe.cell(row=current_row, column=5, value=total_geral_diferenca).number_format = '#,##0.00'

        for col_idx in range(1, df_comparacao_final.shape[1] + 1):
            cell = ws_sabe.cell(row=current_row, column=col_idx)
            cell.font = Config.FONT_TOTAL
            cell.fill = Config.FILL_TOTAL
            cell.border = Config.BORDER_THIN
            if col_idx in [3, 4, 5]: cell.alignment = Config.ALIGN_RIGHT
            elif col_idx == 1: cell.value = ""

        ws_sabe.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        
        if df_comparacao_final.shape[0] > 0:
            col_diferenca_letra = get_column_letter(df_comparacao_final.columns.get_loc('Diferença') + 1)
            range_str = f"{col_diferenca_letra}{start_data_row}:{col_diferenca_letra}{current_row - 1}"
            rule = Rule(type="expression", dxf=Config.DXF_DIFF, formula=[f"ABS({col_diferenca_letra}{start_data_row})>=0.01"])
            ws_sabe.conditional_formatting.add(range_str, rule)

        adicionar_marca_dagua(ws_sabe, anexo_celula='F1')
        auto_ajustar_colunas(ws_sabe)
        aplicar_wrap_text_em_todas_abas(wb_destino)
        
        wb_destino.save(gui_instance.output_path)
        gui_instance.finalizar_processamento_gui(gui_instance.output_path, sucesso=True)

    except Exception as e:
        gui_instance.finalizar_processamento_gui(f"Erro no processamento principal: {e}\n{traceback.format_exc()}", sucesso=False)

class SABEGUI(TkinterDnD.Tk):
    def __init__(self, title="Comparador SABE - Total de Passageiros", size=(750, 450)):
        super().__init__()
        self.title(title)
        self.geometry(f'{size[0]}x{size[1]}')
        self.resizable(False, False)

        self.style = ttk.Style(self)
        self.style.configure('Accent.TButton', font=('Segoe UI', 10, 'bold'))
        self.style.configure('TLabel', font=('Segoe UI', 10))
        self._setar_nome_padrao_arquivo()
        self.create_widgets()
        sv_ttk.set_theme("dark")

    def toggle_theme(self):
        sv_ttk.toggle_theme()

    def create_widgets(self):
        try:
            if os.path.exists(self.BACKGROUND_IMAGE_PATH):
                bg_image_pil = Image.open(self.BACKGROUND_IMAGE_PATH)
                bg_image_pil = bg_image_pil.resize((self.winfo_width(), self.winfo_height()), Image.Resampling.LANCZOS)
                self.background_image = ImageTk.PhotoImage(bg_image_pil)
                background_label = tk.Label(self, image=self.background_image)
                background_label.place(x=0, y=0, relwidth=1, relheight=1)
            else:
                print(f"AVISO: Imagem de fundo não encontrada em '{self.BACKGROUND_IMAGE_PATH}'.")
        except Exception as e:
            print(f"AVISO: Não foi possível carregar a imagem de fundo: {e}.")

        main_frame = ttk.Frame(self, padding=(20, 15))
        main_frame.pack(fill=tk.BOTH, expand=True)

        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 20))

        try:
            if os.path.exists(self.LOGO_IMAGE_PATH):
                logo_original = Image.open(self.LOGO_IMAGE_PATH).convert("RGBA")
                logo_resized = logo_original.resize((int(logo_original.width * 0.4), int(logo_original.height * 0.4)), Image.Resampling.LANCZOS)
                self.logo_image = ImageTk.PhotoImage(logo_resized)
                logo_label = ttk.Label(header_frame, image=self.logo_image)
                logo_label.pack(side=tk.LEFT, padx=(0, 15))
            else:
                print(f"AVISO: Imagem do logo não encontrada em '{self.LOGO_IMAGE_PATH}'.")
        except Exception:
            print(f"AVISO: Não foi possível carregar a imagem do logo.")
            
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(title_frame, text="Comparador SABE - Total de Passageiros", style="Title.TLabel").pack(anchor='w')
        ttk.Label(title_frame, text="Compare facilmente dados de remuneração TXT e conferência DBF.", style="TLabel").pack(anchor='w')
        
        theme_switch = ttk.Checkbutton(header_frame, text="Tema", style="Switch.TCheckbutton", command=self.toggle_theme)
        theme_switch.pack(side=tk.RIGHT, padx=10)
        
        files_frame = ttk.LabelFrame(main_frame, text=" Seleção de Arquivos ", padding=10)
        files_frame.pack(fill=tk.X, pady=(0, 10))
        files_frame.columnconfigure(1, weight=1)

        ttk.Label(files_frame, text="Arquivo TXT (Remuneração):").grid(row=0, column=0, padx=(0, 10), pady=5, sticky='w')
        self.label_txt = ttk.Label(files_frame, text="Nenhum arquivo selecionado.", anchor='w')
        self.label_txt.grid(row=0, column=1, sticky='ew', padx=5)
        ttk.Button(files_frame, text="Selecionar TXT", command=self.selecionar_txt_file).grid(row=0, column=2, padx=(5, 0), sticky='e')

        ttk.Label(files_frame, text="Arquivo DBF (Conferência):").grid(row=1, column=0, padx=(0, 10), pady=5, sticky='w')
        self.label_dbf = ttk.Label(files_frame, text="Nenhum arquivo selecionado.", anchor='w')
        self.label_dbf.grid(row=1, column=1, sticky='ew', padx=5)
        ttk.Button(files_frame, text="Selecionar DBF", command=self.selecionar_dbf_file).grid(row=1, column=2, padx=(5, 0), sticky='e')

        output_frame = ttk.LabelFrame(main_frame, text=" Configuração de Saída ", padding=10)
        output_frame.pack(fill=tk.X, pady=(0, 10))
        output_frame.columnconfigure(1, weight=1)

        ttk.Label(output_frame, text="Nome do Arquivo de Saída:").grid(row=0, column=0, padx=(0, 10), pady=5, sticky='w')
        ttk.Entry(output_frame, textvariable=self.nome_arquivo_saida).grid(row=0, column=1, pady=5, sticky='ew')
        
        ttk.Label(output_frame, text="Pasta de Destino:").grid(row=1, column=0, padx=(0, 10), pady=5, sticky='w')
        self.label_pasta = ttk.Label(output_frame, text="Nenhuma pasta selecionada.", anchor='w')
        self.label_pasta.grid(row=1, column=1, sticky='ew', padx=5)
        ttk.Button(output_frame, text="Selecionar Pasta", command=self.selecionar_pasta_destino).grid(row=1, column=2, padx=(5, 0), sticky='e')

        button_area = ttk.Frame(main_frame, padding=(0, 10, 0, 0))
        button_area.pack(fill=tk.X)
        self.botao_limpar = ttk.Button(button_area, text="Limpar Seleções", command=self.limpar_todas_selecoes)
        self.botao_limpar.pack(side=tk.LEFT, padx=(0, 10))
        self.botao_gerar = ttk.Button(button_area, text="Gerar Relatório de Comparação", style='Accent.TButton', command=self.iniciar_processamento)
        self.botao_gerar.pack(side=tk.RIGHT, fill=tk.X, expand=True)

    def selecionar_txt_file(self):
        caminho = filedialog.askopenfilename(title="Selecione o arquivo RemuneraçãoCCTLinhaDia.txt", filetypes=[("Arquivos de Texto", "*.txt")])
        if caminho:
            self.label_txt.config(text=os.path.basename(caminho))
            self.arquivo_txt_remuneracao = caminho
        else:
            self.label_txt.config(text="Nenhum arquivo selecionado.")
            self.arquivo_txt_remuneracao = ""

    def selecionar_dbf_file(self):
        caminho = filedialog.askopenfilename(title="Selecione o arquivo DBF de Conferência", filetypes=[("Arquivos DBF", "*.dbf"), ("Todos os arquivos", "*.*")])
        if caminho:
            self.label_dbf.config(text=os.path.basename(caminho))
            self.arquivo_dbf_conferencia = caminho
        else:
            self.label_dbf.config(text="Nenhum arquivo selecionado.")
            self.arquivo_dbf_conferencia = ""

    def selecionar_pasta_destino(self):
        caminho = filedialog.askdirectory(title="Selecione a pasta para salvar o relatório")
        if caminho:
            self.pasta_destino.set(caminho)
            self.label_pasta.config(text=caminho)

    def limpar_todas_selecoes(self):
        self.arquivo_txt_remuneracao = ""
        self.arquivo_dbf_conferencia = ""
        self.label_txt.config(text="Nenhum arquivo selecionado.")
        self.label_dbf.config(text="Nenhum arquivo selecionado.")
        self.pasta_destino.set("")
        self.label_pasta.config(text="Nenhuma pasta selecionada.")
        self._setar_nome_padrao_arquivo()
        messagebox.showinfo("Limpeza", "Todos os campos foram limpos.", parent=self)

    def iniciar_processamento(self):
        if not self.arquivo_txt_remuneracao:
            return messagebox.showwarning("Atenção", "Por favor, selecione o **Arquivo TXT (Remuneração)**.", parent=self)
        if not self.arquivo_dbf_conferencia:
            return messagebox.showwarning("Atenção", "Por favor, selecione o **Arquivo DBF (Conferência)**.", parent=self)
        
        pasta_destino, nome_arquivo = self.pasta_destino.get(), self.nome_arquivo_saida.get()
        if not pasta_destino or not nome_arquivo:
            return messagebox.showwarning("Atenção", "Selecione uma **pasta de destino** e defina um **nome** para o arquivo de saída.", parent=self)

        nome_arquivo = nome_arquivo if nome_arquivo.lower().endswith('.xlsx') else nome_arquivo + '.xlsx'
        self.output_path = os.path.join(pasta_destino, nome_arquivo)

        self.botao_gerar.config(state="disabled", text="Processando...")
        self.botao_limpar.config(state="disabled")
        self.update_idletasks()
        
        threading.Thread(target=processar_arquivos_e_gerar_relatorio, args=(self,)).start()
        
    def finalizar_processamento_gui(self, mensagem, sucesso):
        self.botao_gerar.config(state="normal", text="Gerar Relatório de Comparação")
        self.botao_limpar.config(state="normal")
        if sucesso:
            messagebox.showinfo("Sucesso!", f"Relatório final gerado com sucesso em:\n{mensagem}", parent=self)
            self.abrir_pasta_destino()
        else:
            messagebox.showerror("Erro no Processamento", f"Ocorreu um erro inesperado.\n\nDetalhes: {mensagem}", parent=self)

    def abrir_pasta_destino(self):
        pasta = self.pasta_destino.get()
        if not pasta: return
        try:
            if sys.platform == "win32": os.startfile(pasta)
            elif sys.platform == "darwin": subprocess.run(["open", pasta])
            else: subprocess.run(["xdg-open", pasta])
        except Exception as e:
            messagebox.showwarning("Aviso", f"Não foi possível abrir a pasta de destino automaticamente.\n\nCaminho: {pasta}\nErro: {e}", parent=self)
            
    def center_window(self):
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        
    def _setar_nome_padrao_arquivo(self):
        hoje = datetime.date.today()
        meses_pt = ['JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO', 'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']
        nome_arquivo = f"RELATÓRIO SABE"
        self.nome_arquivo_saida.set(nome_arquivo)

if __name__ == "__main__":
    app = SABEGUI()
    app.mainloop()