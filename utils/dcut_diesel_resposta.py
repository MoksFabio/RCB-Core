import tkinter as tk
from tkinter import filedialog, messagebox, Listbox, PhotoImage, ttk
import os
import re
import sys
import threading
import datetime
import subprocess
import traceback
import copy
import math
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from tkinterdnd2 import DND_FILES, TkinterDnD
from PIL import Image, ImageTk
import sv_ttk

class Config:
    RENDIMENTO_DISPEL = {
        "Micro Urbano  s/ar": 2.610, "Midi Medio Urbano  s/ar": 2.610,
        "Básico Méd. urb s/ar": 2.610, "Básico Méd. urb c/ar": 0,
        "Padron 12  Pés s/ar": 2.610, "Padron 12  Pés c/ar": 2.219,
        "Padron 13 Pés s /ar": 2.420, "Padron 13 Pés c /ar": 2.057,
        "Padron 14 Pés s /ar": 2.381, "Padron 14 Pés c /ar": 2.218,
        "Padron 15 Pés s /ar": 2.381, "Padron 15 Pés C /ar": 1.832,
        "Artic.Ext.Pés s/ar": 1.750, "Artic.Ext.Pés c/ar": 1.200,
        "BRT 1 Art.Ext Pés C/ar": 1.200, "RODVIÁRIOS P. 13 C/AR": 1.750,
        "RODVIÁRIOS S/AR": 0, "RODVIÁRIOS C/AR": 1.750,
    }
    NOMES_EMPRESAS = {
        "BOA": "BOA - Borborema Imperial Transportes Ltda", "CAX": "CAX - Rodoviária Caxangá Ltda",
        "CSR": "CSR - Consórcio Recife de Transporte", "CNO": "CNO – CONSÓRCIO CONORTE",
        "EME": "EME - Expresso Metropolitana Ltda", "GLO": "GLO - Transportadora Globo Ltda",
        "MOB": "MOB – Mobibrasil Expresso S.A", "SJT": "SJT - São Judas Tadeu", "VML": "VML - Viação Mirim Ltda",
        "CTC": "CTC - Companhia de Transp. e Comunicação"
    }
    # Lista de empresas a serem incluídas no relatório final
    EMPRESAS_VALIDAS = ["BOA", "CAX", "CNO", "CSR", "EME", "GLO", "MOB", "SJT", "VML"]

def auto_ajustar_colunas(ws):
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

def obter_chave_ordenacao(filepath):
    filename = os.path.basename(filepath).upper()
    if "DUT" in filename: return 0
    if "SÁB" in filename or "SAB" in filename: return 1
    if "DOM" in filename: return 2
    return 3

def processar_quinzena(caminho_arquivo, mes, quinzena):
    intervalo = {'1ª': (7, 19, 8), '2ª': (29, 41, 30)}
    linha_inicio, linha_fim, linha_cabecalho = intervalo[quinzena]

    try:
        df_mes = pd.read_excel(caminho_arquivo, sheet_name=mes, header=None)
        bloco = df_mes.iloc[linha_inicio:linha_fim, :21].copy()
        headers = bloco.iloc[linha_cabecalho - linha_inicio].tolist()
        headers[0] = "EMPRESA"

        for c in ["Total CCT", "Total Geladinho"]:
            if c in headers:
                idx = headers.index(c)
                bloco.drop(bloco.columns[idx], axis=1, inplace=True)
                headers.pop(idx)

        bloco.columns = headers
        bloco = bloco.drop(bloco.index[linha_cabecalho - linha_inicio])
        bloco = bloco.dropna(how='all')

        dados_finais = []
        for _, linha in bloco.iterrows():
            empresa = str(linha['EMPRESA']).strip()
            if empresa.upper() in ["TOTAL STPP", "TOTAL", ""] or pd.isna(empresa):
                continue
            for disel, valor in linha.items():
                if disel == 'EMPRESA' or pd.isna(valor):
                    continue
                try:
                    valor_num = int(float(valor))
                    rendimento = Config.RENDIMENTO_DISPEL.get(disel, "")
                    dados_finais.append({
                        'EMPRESA': empresa, 'Disel': disel, 'Valor': valor_num,
                        'Rendimento pela Licitação': rendimento
                    })
                except Exception:
                    continue

        df = pd.DataFrame(dados_finais)
        if not df.empty:
            df["Qte X Rend"] = pd.to_numeric(df["Valor"]) * pd.to_numeric(df["Rendimento pela Licitação"])
        return df

    except Exception as e:
        messagebox.showerror("Erro ao processar planilha", f"Erro na aba '{mes}':\n{e}")
        return pd.DataFrame()

def adicionar_aba_rendimento_simplificada(wb, df):
    ws = wb.create_sheet("Rendimento PCO")
    ws.append(["Empresa", "Média Geral de Rendimento da Empresa"])
    
    for empresa in sorted(df["EMPRESA"].unique()):
        df_emp = df[df["EMPRESA"] == empresa].copy()
        df_emp = df_emp[df_emp["Disel"].astype(str).str.upper() != "TOTAL CADASTRO"]
        if df_emp.empty: continue

        df_emp["Qte X Rend"] = pd.to_numeric(df_emp["Qte X Rend"], errors="coerce").fillna(0.0)
        df_emp["Valor"] = pd.to_numeric(df_emp["Valor"], errors="coerce").fillna(0).astype(int)
        
        total_valor = df_emp["Valor"].sum()
        total_qte_rend = df_emp["Qte X Rend"].sum()
        rend_total = total_qte_rend / total_valor if total_valor else 0
        
        nome_completo = Config.NOMES_EMPRESAS.get(empresa, empresa)
        
        nova_linha = ws.max_row + 1
        ws.cell(row=nova_linha, column=1, value=nome_completo)
        cell_rendimento = ws.cell(row=nova_linha, column=2, value=rend_total)
        cell_rendimento.number_format = '0.00'

def get_rendimento(vehicle_type, empresa):
    if not isinstance(vehicle_type, str): return 0
    v_type = vehicle_type.upper()
    rules_mob = {"MICRO": 2.2185, "PADRON COM AR": 1.8315}
    common_rules = {"PESADO COM AR": 2.2185, "BRT": 1.2, "MIDI": 2.61, "PESADO": 2.61}
    
    empresa_rules = rules_mob if empresa.upper() == "MOB" else {}
    
    for key, value in {**empresa_rules, **common_rules}.items():
        if key in v_type:
            return value
    return 0

def _adicionar_aba_calculo_simplificada(wb_destino, caminho_arquivo, df_km_prog_atual, contagem_dias, nome_aba, empresa_id):
    if not caminho_arquivo: return
    try:
        df_source = pd.read_excel(caminho_arquivo, header=None)
        colunas_a_manter = min(df_source.shape[1], 12)
        df_source = df_source.iloc[:, :colunas_a_manter]

        mapa_km = {}
        if not df_km_prog_atual.empty:
            df_km_prog_atual['Dia Tipo'] = df_km_prog_atual['Dia Tipo'].str.upper().str.strip()
            df_km_prog_atual['Linha'] = df_km_prog_atual['Linha'].astype(str).str.strip()
            dia_tipo_map = {'DIA ÚTIL': 'DUT', 'SABADO': 'SAB', 'DOMINGO': 'DOM'}
            df_km_prog_atual['Dia Tipo'] = df_km_prog_atual['Dia Tipo'].replace(dia_tipo_map)
            for _, row in df_km_prog_atual.iterrows():
                chave = (row['Linha'], row['Dia Tipo'])
                mapa_km[chave] = row['Km Total']
            
        col_cod, col_km_dut, col_km_sab, col_km_dom = 0, 9, 10, 11
        for i in range(4, len(df_source)):
            try:
                cod_val = df_source.iloc[i, col_cod]
                if pd.isna(cod_val) or not str(cod_val).strip(): continue
                cod_str = str(int(cod_val))
                df_source.iat[i, col_km_dut] = mapa_km.get((cod_str, 'DUT'), df_source.iloc[i, col_km_dut])
                df_source.iat[i, col_km_sab] = mapa_km.get((cod_str, 'SAB'), df_source.iloc[i, col_km_sab])
                df_source.iat[i, col_km_dom] = mapa_km.get((cod_str, 'DOM'), df_source.iloc[i, col_km_dom])
            except (ValueError, TypeError, IndexError): continue
        
        ws = wb_destino.create_sheet(nome_aba)
        df_data = df_source.iloc[4:].reindex(columns=range(16)).copy()
        df_data.columns = range(df_data.shape[1])
        for col in df_data.columns[3:12]: df_data[col] = pd.to_numeric(df_data[col], errors='coerce')
        
        condition = (df_data[0].astype(str).str.strip().replace('nan', '') == '') & (df_data[1].astype(str).str.strip().replace('nan', '') == '') & (df_data[2].astype(str).str.strip().replace('nan', '') == '')
        total_geral_indices = df_data[condition].index
        if not total_geral_indices.empty: df_data = df_data.drop(index=total_geral_indices)
        if not df_data.empty: df_data = df_data.iloc[:-1].copy()

        df_data[12] = df_data[2].apply(lambda x: get_rendimento(x, empresa_id))
        df_data.loc[:, 0] = df_data[0].ffill().infer_objects(copy=False)
        df_data.loc[:, 1] = df_data[1].ffill().infer_objects(copy=False)

        dados_processados = pd.DataFrame()
        for _, group in df_data.groupby(0, dropna=True):
            group_modificado = group.copy()
            unique_vehicle_type_count = group[2].dropna().astype(str).str.strip().replace('', pd.NA).nunique()
            is_multi_vehicle_group = unique_vehicle_type_count > 1

            if is_multi_vehicle_group:
                col_veic_dut, col_veic_sab, col_veic_dom = 6, 7, 8
                total_km_dut, total_km_sab, total_km_dom = group.iloc[0, col_km_dut], group.iloc[0, col_km_sab], group.iloc[0, col_km_dom]
                total_veic_dut, total_veic_sab, total_veic_dom = group[col_veic_dut].sum(), group[col_veic_sab].sum(), group[col_veic_dom].sum()
                group_modificado[col_km_dut] = group.apply(lambda r: (total_km_dut / total_veic_dut * r[col_veic_dut]) if total_veic_dut else 0, axis=1)
                group_modificado[col_km_sab] = group.apply(lambda r: (total_km_sab / total_veic_sab * r[col_veic_sab]) if total_veic_sab else 0, axis=1)
                group_modificado[col_km_dom] = group.apply(lambda r: (total_km_dom / total_veic_dom * r[col_veic_dom]) if total_veic_dom else 0, axis=1)
            dados_processados = pd.concat([dados_processados, group_modificado])

        for i, cota_col in enumerate([13, 14, 15]):
            dados_processados[cota_col] = dados_processados.apply(lambda r: r[col_km_dut + i] / r[12] if r[12] else 0, axis=1)

        if not dados_processados.empty:
            total_geral_series = pd.Series([''] * 16, index=range(16));
            for col_idx in range(3, 16):
                if col_idx != 12: total_geral_series[col_idx] = dados_processados[col_idx].sum()

            dias = {'DUT': contagem_dias.get('DUT', 0), 'SAB': contagem_dias.get('SAB', 0), 'DOM': contagem_dias.get('DOM', 0)}
            
            ws.cell(row=1, column=2, value="Qtd. Dias")
            ws.cell(row=1, column=3, value="LITROS")

            litros_diario = {'DUT': total_geral_series[13] or 0, 'SAB': total_geral_series[14] or 0, 'DOM': total_geral_series[15] or 0}
            litros_total = {k: dias[k] * litros_diario[k] for k in dias}
            
            dados_resumo = [("DUT", dias['DUT'], litros_total['DUT']), ("SAB", dias['SAB'], litros_total['SAB']), ("DOM", dias['DOM'], litros_total['DOM'])]
            linha_inicio_resumo = 2
            for i, (label, dia, litro) in enumerate(dados_resumo):
                linha = linha_inicio_resumo + i
                ws.cell(row=linha, column=1, value=label)
                ws.cell(row=linha, column=2, value=dia).number_format = '0'
                ws.cell(row=linha, column=3, value=litro).number_format = '#,##0.00'

            linha_total_resumo = linha_inicio_resumo + len(dados_resumo)
            ws.cell(row=linha_total_resumo, column=1, value="TOTAL")
            ws.cell(row=linha_total_resumo, column=2, value=f"=SUM(B{linha_inicio_resumo}:B{linha_total_resumo-1})").number_format = '0'
            ws.cell(row=linha_total_resumo, column=3, value=f"=SUM(C{linha_inicio_resumo}:C{linha_total_resumo-1})").number_format = '#,##0.00'
            
            total_litros_final = sum(litros_total.values())
            valor_arredondado = math.ceil(total_litros_final / 5000) * 5000
            ws.cell(row=linha_total_resumo, column=4, value=valor_arredondado).number_format = '#,##0'

    except Exception as e:
        messagebox.showerror(f"Erro no {nome_aba}", f"Não foi possível processar a aba '{nome_aba}'.\n\nDetalhes: {e}")

def extrair_dados_para_dataframe(lista_arquivos):
    dados_completos = []
    chaves_necessarias = {'Linha', 'Km Total', 'Empresa', 'Dia Tipo'}
    for arquivo in lista_arquivos:
        try:
            import xlrd
            book = xlrd.open_workbook(arquivo)
            sheet = book.sheet_by_index(0)
            raw_data = [sheet.row_values(i) for i in range(sheet.nrows)]
            indices, linha_header = {}, -1
            for r in range(min(15, len(raw_data))):
                for c, valor_cabecalho in enumerate(raw_data[r]):
                    if isinstance(valor_cabecalho, str):
                        header = valor_cabecalho.strip()
                        if header in chaves_necessarias and header not in indices:
                            indices[header] = c
                if len(indices) == len(chaves_necessarias):
                    linha_header = r
                    break
            if linha_header == -1: continue
            
            for r in range(linha_header + 1, len(raw_data)):
                row_data = raw_data[r]
                try:
                    linha_val = row_data[indices['Linha']]
                    # Corrigido .strip para .strip()
                    empresa_val = str(row_data[indices['Empresa']]).strip()
                    if 'total' in str(linha_val).lower() or not empresa_val: continue
                    try: linha_str = str(int(float(linha_val)))
                    except (ValueError, TypeError): linha_str = str(linha_val).strip()
                    dados_completos.append({
                        'Linha': linha_str,
                        'Empresa': empresa_val,
                        'Dia Tipo': str(row_data[indices['Dia Tipo']]),
                        'Km Total': float(row_data[indices['Km Total']])
                    })
                except (ValueError, TypeError, IndexError): continue
        except Exception: continue
    if not dados_completos: return pd.DataFrame()
    df = pd.DataFrame(dados_completos)
    df.drop_duplicates(subset=['Linha', 'Empresa', 'Dia Tipo'], keep='last', inplace=True)
    return df

def gerar_dados_resumo_km(df):
    if df.empty:
        return pd.DataFrame()
    
    # Filtro 1: Manter apenas as empresas da lista de configuração.
    df_filtrado = df[df['Empresa'].isin(Config.EMPRESAS_VALIDAS)].copy()

    # Filtro 2: Remover linhas onde 'Dia Tipo' está vazio ou contém apenas espaços para evitar duplicatas.
    df_filtrado = df_filtrado[df_filtrado['Dia Tipo'].str.strip() != ''].copy()

    if df_filtrado.empty:
        return pd.DataFrame()

    # Agrupa por empresa e tipo de dia, somando o KM Total
    resumo = df_filtrado.groupby(['Empresa', 'Dia Tipo'])['Km Total'].sum().reset_index()
    resumo.rename(columns={'Km Total': 'KM Calculado'}, inplace=True)
    return resumo

def criar_aba_km_prog_resumo(wb, dados_resumo):
    ws = wb.create_sheet("Km Prog", 0)
    ws.append(['Empresa', 'Dia Tipo', 'KM Calculado'])
    
    if not dados_resumo.empty:
        # Define a ordem personalizada para a coluna 'Dia Tipo'
        ordem_dia_tipo = ['DUT', 'SAB', 'DOM']
        dados_resumo['Dia Tipo'] = pd.Categorical(dados_resumo['Dia Tipo'], categories=ordem_dia_tipo, ordered=True)
        
        # Ordena os dados pela Empresa e pela ordem personalizada do Dia Tipo
        dados_resumo_sorted = dados_resumo.sort_values(by=['Empresa', 'Dia Tipo'])
        for _, row in dados_resumo_sorted.iterrows():
            ws.append(list(row))
    
    # Formata a coluna de KM
    for cell in ws['C']:
        if cell.row > 1: # Pula o cabeçalho
            cell.number_format = '#,##0.00'

def executar_processamento(gui_instance):
    try:
        wb_destino = openpyxl.Workbook()
        wb_destino.remove(wb_destino.active)

        # 1. Extrai os dados do mês atual para um DataFrame
        df_atual = extrair_dados_para_dataframe(gui_instance.arquivos_atual)

        # 2. Gera os dados de resumo do KM a partir do DataFrame (já com filtros aplicados)
        dados_resumo_km = gerar_dados_resumo_km(df_atual)

        # 3. Cria a aba "Km Prog" com a tabela de resumo (já ordenada)
        criar_aba_km_prog_resumo(wb_destino, dados_resumo_km)

        # 4. Processamento do Rendimento PCO (se houver)
        if all([gui_instance._caminho_completo_cut2, gui_instance.mes_cut2.get(), gui_instance.quinzena_cut2.get()]):
            df_rendimento = processar_quinzena(gui_instance._caminho_completo_cut2, gui_instance.mes_cut2.get(), gui_instance.quinzena_cut2.get())
            if not df_rendimento.empty:
                adicionar_aba_rendimento_simplificada(wb_destino, df_rendimento)
        
        # 5. Processamento dos cálculos de Cota (se houver)
        if gui_instance._caminho_completo_cno:
            _adicionar_aba_calculo_simplificada(wb_destino, gui_instance._caminho_completo_cno, df_atual, gui_instance.contagem_dias_dict, "Cálculo CNO", "CNO")
        if gui_instance._caminho_completo_mob:
            _adicionar_aba_calculo_simplificada(wb_destino, gui_instance._caminho_completo_mob, df_atual, gui_instance.contagem_dias_dict, "Cálculo MOB", "MOB")

        # Ajusta colunas e salva o arquivo
        for sheet in wb_destino.sheetnames:
            auto_ajustar_colunas(wb_destino[sheet])
        
        if "Km Prog" in wb_destino.sheetnames:
            wb_destino.active = wb_destino["Km Prog"]

        wb_destino.save(gui_instance.output_path)
        gui_instance.finalizar_processamento_gui(gui_instance.output_path, sucesso=True)
    except Exception as e:
        gui_instance.finalizar_processamento_gui(f"Erro em executar_processamento: {e}\n{traceback.format_exc()}", sucesso=False)


# A classe da GUI (CotaDieselGUI) permanece a mesma.
# Nenhuma alteração é necessária nela.

class CotaDieselGUI(TkinterDnD.Tk):
    def __init__(self, title="Cota de Óleo Diesel", size=(1000, 850)):
        super().__init__()
        self.title(title)
        self.geometry(f'{size[0]}x{size[1]}')
        self.resizable(True, True)

        self.style = ttk.Style(self)
        self.style.configure('Accent.TButton', font=('Segoe UI', 10, 'bold'))
        self.style.configure('TLabel', font=('Segoe UI', 10))
        self.style.configure('TLabelframe.Label', font=('Segoe UI', 11, 'bold'))
        
        self._setar_nome_padrao_arquivo()
        self.create_widgets()
        sv_ttk.set_theme("light")

    def toggle_theme(self):
        sv_ttk.toggle_theme()

    def toggle_size(self):
        if not hasattr(self, 'is_compact') or self.is_compact:
            self.geometry('1000x850')
            self.is_compact = False
        else:
            self.geometry('850x720')
            self.is_compact = True
        self.center_window()

    def create_widgets(self):
        try:
            bg_image_pil = Image.open(self.BACKGROUND_IMAGE_PATH)
            bg_image_pil = bg_image_pil.resize((1000, 850), Image.Resampling.LANCZOS)
            self.background_image = ImageTk.PhotoImage(bg_image_pil)
            background_label = tk.Label(self, image=self.background_image)
            background_label.place(x=0, y=0, relwidth=1, relheight=1)
        except Exception as e:
            print(f"AVISO: Não foi possível carregar a imagem de fundo: {e}.")

        main_frame = ttk.Frame(self, padding=(20, 15))
        main_frame.pack(fill=tk.BOTH, expand=True)

        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 20))

        try:
            logo_original = Image.open(self.LOGO_IMAGE_PATH).convert("RGBA")
            logo_resized = logo_original.resize((int(logo_original.width * 0.4), int(logo_original.height * 0.4)), Image.Resampling.LANCZOS)
            self.logo_image = ImageTk.PhotoImage(logo_resized)
            logo_label = ttk.Label(header_frame, image=self.logo_image)
            logo_label.pack(side=tk.LEFT, padx=(0, 15))
        except Exception:
            print(f"AVISO: Não foi possível carregar a imagem do logo.")
        
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(title_frame, text="Integrador de Cota de Óleo Diesel", style="Title.TLabel").pack(anchor='w')
        ttk.Label(title_frame, text="Ferramenta para consolidação e comparação de quilometragem", style="TLabel").pack(anchor='w')
        
        theme_switch = ttk.Checkbutton(header_frame, text="Tema", style="Switch.TCheckbutton", command=self.toggle_theme)
        theme_switch.pack(side=tk.RIGHT, padx=10)
        
        resize_button = ttk.Button(header_frame, text="Ajustar Janela", command=self.toggle_size)
        resize_button.pack(side=tk.RIGHT, padx=5)

        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=10)

        tab1 = ttk.Frame(notebook, padding=10)
        tab2 = ttk.Frame(notebook, padding=10)
        tab3 = ttk.Frame(notebook, padding=10)
        notebook.add(tab1, text='     Arquivos Principais     ')
        notebook.add(tab2, text='     Relatórios Adicionais     ')
        notebook.add(tab3, text='     Saída e Geração       ')

        self._criar_aba_arquivos(tab1)
        self._criar_aba_adicionais(tab2)
        self._criar_aba_saida(tab3)
        
        button_area = ttk.Frame(main_frame, padding=(0, 10, 0, 0))
        button_area.pack(fill=tk.X)
        self.botao_limpar = ttk.Button(button_area, text="Limpar Tudo", command=self.limpar_todas_selecoes)
        self.botao_limpar.pack(side=tk.LEFT, padx=(0, 10))
        self.botao_gerar = ttk.Button(button_area, text="Processar e Gerar Relatório", style='Accent.TButton', command=self.iniciar_processamento)
        self.botao_gerar.pack(side=tk.RIGHT, fill=tk.X, expand=True)

    def _criar_aba_arquivos(self, parent):
        parent.columnconfigure((0, 1), weight=1, uniform="group1")
        parent.rowconfigure(0, weight=1)
        self.lista_passado = self._criar_painel_arquivos(parent, "Arquivos do Mês Passado (para cálculo)", self.arquivos_passado, 0)
        self.lista_atual = self._criar_painel_arquivos(parent, "Arquivos do Mês de Referência", self.arquivos_atual, 1)

    def _criar_aba_adicionais(self, parent):
        parent.columnconfigure(1, weight=1)

        frame_pco = ttk.LabelFrame(parent, text=" Rendimento PCO (Opcional) ", padding=10)
        frame_pco.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 15))
        frame_pco.columnconfigure(1, weight=1)

        ttk.Button(frame_pco, text="Selecionar Arquivo PCO (.xls)", command=self.selecionar_cut2_file).grid(row=0, column=0, padx=(0, 10), sticky='w')
        self.label_cut2 = ttk.Label(frame_pco, text="Nenhum arquivo selecionado.", style='TLabel', anchor='w')
        self.label_cut2.grid(row=0, column=1, sticky='ew')
        
        pco_options_frame = ttk.Frame(frame_pco)
        pco_options_frame.grid(row=1, column=1, sticky='w', pady=(5,0))
        ttk.Label(pco_options_frame, text="Mês:").pack(side=tk.LEFT, padx=(0,5))
        ttk.Combobox(pco_options_frame, textvariable=self.mes_cut2, values=["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"], state="readonly", width=8).pack(side=tk.LEFT, padx=(0,15))
        ttk.Label(pco_options_frame, text="Quinzena:").pack(side=tk.LEFT, padx=(0,5))
        ttk.Combobox(pco_options_frame, textvariable=self.quinzena_cut2, values=["1ª", "2ª"], state="readonly", width=5).pack(side=tk.LEFT)

        frame_calc = ttk.LabelFrame(parent, text=" Cálculos de Cota (Opcional) ", padding=10)
        frame_calc.grid(row=1, column=0, columnspan=2, sticky="ew")
        frame_calc.columnconfigure(1, weight=1)

        ttk.Button(frame_calc, text="Cálculo CNO (.xls)", command=self.selecionar_cno_file).grid(row=0, column=0, padx=(0,10), pady=5, sticky='w')
        self.label_cno = ttk.Label(frame_calc, text="Nenhum arquivo selecionado.", anchor='w')
        self.label_cno.grid(row=0, column=1, sticky='ew', padx=5)

        ttk.Button(frame_calc, text="Cálculo MOB (.xls)", command=self.selecionar_mob_file).grid(row=1, column=0, padx=(0,10), pady=5, sticky='w')
        self.label_mob = ttk.Label(frame_calc, text="Nenhum arquivo selecionado.", anchor='w')
        self.label_mob.grid(row=1, column=1, sticky='ew', padx=5)
        
        dias_frame = ttk.Frame(frame_calc)
        dias_frame.grid(row=2, column=0, columnspan=2, sticky='w', pady=(10,0))
        ttk.Label(dias_frame, text="Contagem de dias no período:", style='Header.TLabel').pack(side=tk.LEFT, padx=(0,15))
        ttk.Label(dias_frame, text="DUT:").pack(side=tk.LEFT, padx=(0,5)); ttk.Entry(dias_frame, textvariable=self.dias_dut, width=4).pack(side=tk.LEFT, padx=(0,10))
        ttk.Label(dias_frame, text="SAB:").pack(side=tk.LEFT, padx=(0,5)); ttk.Entry(dias_frame, textvariable=self.dias_sab, width=4).pack(side=tk.LEFT, padx=(0,10))
        ttk.Label(dias_frame, text="DOM:").pack(side=tk.LEFT, padx=(0,5)); ttk.Entry(dias_frame, textvariable=self.dias_dom, width=4).pack(side=tk.LEFT)

    def _criar_aba_saida(self, parent):
        parent.columnconfigure(0, weight=1)
        
        frame = ttk.LabelFrame(parent, text=" Configuração do Arquivo de Saída ", padding=15)
        frame.pack(fill=tk.BOTH, expand=True)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text="Nome do Arquivo:").grid(row=0, column=0, padx=(0,10), pady=5, sticky='w')
        ttk.Entry(frame, textvariable=self.nome_arquivo_saida).grid(row=0, column=1, pady=5, sticky='ew')
        
        ttk.Button(frame, text="Selecionar Pasta de Destino", command=self.selecionar_pasta_destino).grid(row=1, column=0, padx=(0,10), pady=5, sticky='w')
        self.label_pasta = ttk.Label(frame, text="Nenhuma pasta selecionada.")
        self.label_pasta.grid(row=1, column=1, pady=5, sticky='ew')

    def _criar_painel_arquivos(self, parent, titulo, lista_arquivos, col):
        frame = ttk.LabelFrame(parent, text=f" {titulo} ", padding=10)
        frame.grid(row=0, column=col, padx=10, pady=5, sticky='nsew')
        frame.grid_rowconfigure(1, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        ttk.Label(frame, text="Arraste e solte arquivos .xls aqui ou use os botões abaixo.", justify=tk.CENTER).grid(row=0, column=0, columnspan=2, pady=(0, 10))
        
        tree = ttk.Treeview(frame, columns=('filename',), show='headings', height=6)
        tree.heading('filename', text='Arquivos Selecionados')
        tree.grid(row=1, column=0, columnspan=2, sticky='nsew')
        
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=1, column=2, sticky='ns')

        tree.drop_target_register(DND_FILES)
        tree.dnd_bind('<<Drop>>', lambda e: self._adicionar_arquivos(self.tk.splitlist(e.data), lista_arquivos, tree))
        
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=2, column=0, columnspan=2, sticky='ew', pady=(10,0))
        btn_frame.columnconfigure((0, 1), weight=1)

        btn_selecionar = ttk.Button(btn_frame, text="Selecionar", command=lambda: self._adicionar_arquivos(filedialog.askopenfilenames(title=f"Selecione - {titulo}", filetypes=[("Excel 97-2003", "*.xls")]), lista_arquivos, tree))
        btn_selecionar.grid(row=0, column=0, sticky='ew', padx=(0,5))
        btn_limpar = ttk.Button(btn_frame, text="Limpar", command=lambda: self._limpar_lista(lista_arquivos, tree))
        btn_limpar.grid(row=0, column=1, sticky='ew', padx=(5,0))
        return tree

    def _adicionar_arquivos(self, paths, lista_alvo, treeview):
        novos_arquivos = False
        for path in paths:
            if path.lower().endswith('.xls') and path not in lista_alvo:
                lista_alvo.append(path)
                novos_arquivos = True
        if novos_arquivos:
            lista_alvo.sort(key=obter_chave_ordenacao)
            self._atualizar_treeview(treeview, lista_alvo)

    def _limpar_lista(self, lista_alvo, treeview):
        lista_alvo.clear()
        self._atualizar_treeview(treeview, lista_alvo)

    def _atualizar_treeview(self, treeview, lista_arquivos):
        for i in treeview.get_children():
            treeview.delete(i)
        for path in lista_arquivos:
            treeview.insert("", tk.END, values=(os.path.basename(path),))

    def selecionar_arquivo_generico(self, title, label_widget, attr_caminho):
        caminho = filedialog.askopenfilename(title=title, filetypes=[("Arquivos Excel 97-2003", "*.xls")])
        if caminho:
            label_widget.config(text=os.path.basename(caminho))
            setattr(self, attr_caminho, caminho)
        else:
            label_widget.config(text="Nenhum arquivo selecionado.")
            setattr(self, attr_caminho, "")

    def selecionar_cut2_file(self): self.selecionar_arquivo_generico("Selecione o arquivo de Rendimento PCO", self.label_cut2, '_caminho_completo_cut2')
    def selecionar_cno_file(self): self.selecionar_arquivo_generico("Selecione o arquivo do Cálculo CNO", self.label_cno, '_caminho_completo_cno')
    def selecionar_mob_file(self): self.selecionar_arquivo_generico("Selecione o arquivo do Cálculo MOB", self.label_mob, '_caminho_completo_mob')
    
    def selecionar_pasta_destino(self):
        caminho = filedialog.askdirectory(title="Selecione a pasta para salvar o relatório")
        if caminho:
            self.pasta_destino.set(caminho)
            self.label_pasta.config(text=caminho)

    def limpar_todas_selecoes(self):
        self._limpar_lista(self.arquivos_passado, self.lista_passado)
        self._limpar_lista(self.arquivos_atual, self.lista_atual)
        self.mes_cut2.set(""); self.quinzena_cut2.set("")
        self.dias_dut.set('0'); self.dias_sab.set('0'); self.dias_dom.set('0')
        self.pasta_destino.set("")
        self._caminho_completo_cut2, self._caminho_completo_cno, self._caminho_completo_mob = "", "", ""
        self.label_cut2.config(text="Nenhum arquivo selecionado.")
        self.label_cno.config(text="Nenhum arquivo selecionado.")
        self.label_mob.config(text="Nenhum arquivo selecionado.")
        self.label_pasta.config(text="Nenhuma pasta selecionada.")
        self._setar_nome_padrao_arquivo()
        messagebox.showinfo("Limpeza", "Todos os campos foram limpos.", parent=self)

    def iniciar_processamento(self):
        if not self.arquivos_atual: return messagebox.showwarning("Atenção", "É necessário selecionar arquivos para o 'Mês de Referência'.", parent=self)
        if not self.arquivos_passado and (self._caminho_completo_cno or self._caminho_completo_mob):
            if not messagebox.askyesno("Atenção", "Nenhum arquivo do 'Mês Passado' foi selecionado. Os cálculos de CNO/MOB podem não ter a referência correta de KM. Deseja continuar?", parent=self):
                return
        if self._caminho_completo_cut2 and (not self.mes_cut2.get() or not self.quinzena_cut2.get()): return messagebox.showwarning("Atenção", "Para o relatório de Rendimento PCO, selecione o Mês e a Quinzena.", parent=self)

        try:
            self.contagem_dias_dict = {'DUT': int(self.dias_dut.get() or 0), 'SAB': int(self.dias_sab.get() or 0), 'DOM': int(self.dias_dom.get() or 0)}
        except ValueError:
            return messagebox.showerror("Erro de Entrada", "Os valores para a contagem de dias devem ser números inteiros.", parent=self)
        
        if (self._caminho_completo_cno or self._caminho_completo_mob) and sum(self.contagem_dias_dict.values()) == 0:
            messagebox.showwarning("Atenção", "Nenhum dia foi informado. Os cálculos de cota final para CNO/MOB serão zero.", parent=self)

        pasta_destino, nome_arquivo = self.pasta_destino.get(), self.nome_arquivo_saida.get()
        if not pasta_destino or not nome_arquivo: return messagebox.showwarning("Atenção", "Selecione uma pasta de destino e defina um nome para o arquivo de saída.", parent=self)

        nome_arquivo = nome_arquivo if nome_arquivo.lower().endswith('.xlsx') else nome_arquivo + '.xlsx'
        self.output_path = os.path.join(pasta_destino, nome_arquivo)

        self.botao_gerar.config(state="disabled", text="Processando...")
        self.botao_limpar.config(state="disabled")
        self.update_idletasks()
        
        threading.Thread(target=executar_processamento, args=(self,)).start()
        
    def finalizar_processamento_gui(self, mensagem, sucesso):
        self.botao_gerar.config(state="normal", text="Processar e Gerar Relatório")
        self.botao_limpar.config(state="normal")
        if sucesso:
            messagebox.showinfo("Sucesso!", f"Relatório final gerado com sucesso em:\n{mensagem}", parent=self)
            self.abrir_pasta_destino()
        else:
            messagebox.showerror("Erro no Processamento", f"Ocorreu um erro inesperado.\n\nDetalhes: {mensagem}", parent=self)

    def abrir_pasta_destino(self):
        pasta = self.pasta_destino.get()
        if not pasta: return
        try:
            if sys.platform == "win32": os.startfile(pasta)
            elif sys.platform == "darwin": subprocess.run(["open", pasta])
            else: subprocess.run(["xdg-open", pasta])
        except Exception as e:
            messagebox.showwarning("Aviso", f"Não foi possível abrir a pasta de destino automaticamente.\n\nCaminho: {pasta}\nErro: {e}", parent=self)
            
    def center_window(self):
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        
    def _setar_nome_padrao_arquivo(self):
        hoje = datetime.date.today()
        meses_pt = ['JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO', 'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']
        nome_arquivo = f"Cota de {meses_pt[hoje.month-1]} {hoje.year} RCB"
        self.nome_arquivo_saida.set(nome_arquivo)

if __name__ == "__main__":
    try:
        import xlrd
    except ImportError:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Dependência Faltando", "A biblioteca 'xlrd' não foi encontrada. Por favor, instale-a com 'pip install xlrd'.")
        sys.exit(1)
        
    app = CotaDieselGUI()
    app.mainloop()